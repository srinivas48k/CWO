# Decompiled with PyLingual (https://pylingual.io)
# Internal filename: OSS Chatbot 1.16.py
# Bytecode version: 3.10.0rc2 (3439)
# Source timestamp: 1970-01-01 00:00:00 UTC (0)

global ssh  # inserted
global RRSG_auto_login  # inserted
import ctypes
import glob
import pathlib
import os
import re
import socket
from io import StringIO
import sqlite3
import shutil
import secrets
import sys
import threading
import time
import tkinter as tk
import tkinter.messagebox as msgbox
from tkinter import messagebox
import customtkinter
import numpy as np
import pandas as pd
import paramiko
import random
from PyQt5.QtCore import Qt, QTimer, QDate
from PyQt5.QtGui import QFont, QPainter, QColor, QTextOption, QPixmap, QMovie
from PyQt5.QtWidgets import QApplication, QMessageBox, QDialog, QWidget, QTextEdit, QFileDialog, QLabel, QLineEdit, QPushButton, QVBoxLayout, QScrollArea, QDesktopWidget, QMainWindow, QTabWidget, QHBoxLayout, QCheckBox, QRadioButton, QSpacerItem, QSizePolicy, QComboBox, QSpinBox, QDateEdit, QButtonGroup
from tabulate import tabulate
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfTransformer
import long_responses as long
import fitz
from sentence_transformers import SentenceTransformer, util
import json
from datetime import datetime, timedelta
import requests
from PIL import Image
from googletrans import Translator
from py_topping.data_connection.sharepoint import da_tran_SP365
import warnings
import jieba
from langdetect import detect
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from gensim.parsing.preprocessing import remove_stopwords
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import schedule
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.styles import PatternFill
import openpyxl
import psutil
from spellchecker import SpellChecker
from difflib import get_close_matches
import gc
warnings.simplefilter(action='ignore', category=pd.errors.PerformanceWarning)
warnings.simplefilter('ignore', UserWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=pd.errors.PerformanceWarning)
warnings.filterwarnings('ignore')
warnings.simplefilter('ignore', category=DeprecationWarning)

checkamos_java = []
GSM_LST = []
LTE_LST = []
WCDMA_LST = []
NR_LST = []
selected_checkboxes = []
GSM_NODE_LST = []
LTE_NODE_LST = []
WCDMA_NODE_LST = []
NR_NODE_LST = []
COMMAND_DEFALT_AUTO = ['ok']
FREQUENCY_LST = []
FREQUENCY_count_LST = []
EMAIL_METHOOD_LST = []
WORKORDER_CLOSE = []
chek_exit = []
chek_first_time_rsg = []
CHECK_ISF_SEQUENCE = []
ISF_STEP_START_Task_Id_lst = []
AZURE_CREAD_FAIL = []
Azue_OTP_lst = []
establish_ssh_conn_lst = []
conversations_LST = []
KEYWORD_ADD = []
change_User = []
command_barred = []
excel_df_LST = []
amos_node = []
Buddy_request_lst = []
buddy_node = []
buddy_node_curr = []
amos_nod2 = []
ltall_nod2 = []
check_yesno = []
ASKSITID = []
cmmmddd = []
siteddd = []
checknxtsit = []
processdone = []
combined_cmd_window = []
combined_cmd_window_chk = []
checkamos = []
command = []
converstion2 = []
check_textbox = []
LOGINERROR_LST = []
chk_commad_add = []
sh_map_lst = []
CHECKISFWO = []
CHK_TO_CONTinue = []
CHK_TO_CONTinue2 = []
ROP_ACT = []
az_check = []
WCDMA_LST_RBS = []
WCDMA_LST_RNC = []
WCDMA_NODE_LST_RBS = []
WCDMA_NODE_LST_RNC = []
PROJECT_USER_LST = []
PROJECT_LST = []
SELECTED_PROJECT = []
USER_MAPPED_IN_PROJ = []
USER_MAPPED_ROLE = []
PROJECT_SME_LST = []
switch_PMX_REQ_CHECK = []
RBS_PASSLST = []
VPN_select_METHOD_F = []
VPN_RSSG_CRREDD = []
RBSPASS_CRED_CRREDD = ['rbs', 'rbs']
OUTFOLDERR_LTE = []
OUTFOLDERR_LTE = []
ROP_WISE_REPORT = []
OUTFOLDERR_WCDMA = []
PMAX_DATE = []
PMX_ROP_TIME = []
mfolder_create_lst = []

def clear_all_lists():
    exclude_lists = ['PROJECT_USER_LST', 'PROJECT_LST', 'SELECTED_PROJECT', 'USER_MAPPED_IN_PROJ', 'USER_MAPPED_ROLE', 'PROJECT_SME_LST', 'RBSPASS_CRED_CRREDD', 'COMMAND_DEFALT_AUTO', 'VPN_RSSG_CRREDD', 'VPN_select_METHOD_F', 'RBS_PASSLST', 'VPN_select_METHOD_F', 'VPN_RSSG_CRREDD', 'processdone']
    for name, value in list(globals().items()):
        try:
            if isinstance(value, list) and name not in exclude_lists:
                globals()[name].clear()
        except:
            continue
    try:
        exclude_lists.clear()
        del exclude_lists
    except:
        pass
    try:
        cmmmddd.clear()
    except:
        pass
    try:
        siteddd.clear()
    except:
        pass
    try:
        checknxtsit.clear()
    except:
        pass
    try:
        processdone.clear()
    except:
        pass
    try:
        gc.collect()
    except:
        return None
DEFAULT_RESPONSES_lst = {'Hello': 'Hi there! How can I assist you?', 'Goodbye': 'Goodbye! Have a great day!', 'Hi': 'Hello! How can I assist you?', 'Bro': 'Hi! How can I support you today?', 'Bhai': 'Hello! How can I assist you?', 'Brother': 'Hi there! How can I assist you?'}

def find_keyword(self, user_input, DEFAULT_RESPONSES_lst):
    def translate_text(user_input):
        try:
            translator = Translator()
            translated_text = translator.translate(user_input, dest='en').text
            return translated_text
        except:
            return user_input

    def get_response_keyword_find(user_input, default_responses):
        def find_best_match(user_input, model, conversations, spll):
            best_similarity = 0.6
            best_match = None
            user_input_embedding = model.encode(user_input.lower())
            for conversation in conversations:
                try:
                    conversation_embedding = model.encode(conversation.split('_')[spll].lower())
                    similarity = util.cos_sim(user_input_embedding, conversation_embedding).item()
                    if similarity >= best_similarity:
                        best_similarity = similarity
                        best_match = conversation
                except:
                    continue
            return best_match
        user_input_cleaned2 = user_input_cleaned1.replace('_', ' ')
        user_input_cleaned3 = user_input_cleaned2.split(' ')
        try:
            conversations_LST2 = conversations_LST + converstion2
            conversations_LST2 = [i for n, i in enumerate(conversations_LST2) if i not in conversations_LST2[:n]]
        except:
            conversations_LST2 = conversations_LST
        try:
            conversations_LST.clear()
            converstion2.clear()
        except:
            pass
        model_names = ['paraphrase-distilroberta-base-v1']
        best_similarity = 0.0
        best_match = None
        conversations33 = [conversations33.split('_')[0].lower() for conversations33 in conversations_LST2]
        user_input = user_input.replace('_', ' ')
        for model_name in model_names:
            model = SentenceTransformer(model_name)
            user_input_embedding = model.encode(user_input.lower())
            conv_embeddings = model.encode(conversations33)
            similarities = util.cos_sim([user_input_embedding], conv_embeddings)[0]
            for i, similarity in enumerate(similarities):
                if similarity >= 0.75:
                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_match = conversations_LST2[i]
                    break
        if best_match:
            break
        model = SentenceTransformer('paraphrase-distilroberta-base-v1')
        for user_input_cleaned in user_input_cleaned3:
            user_input_cleaned = user_input_cleaned.lower()
            if not user_input_cleaned.strip():
                user_input_cleaned = 'hi'
            if user_input_cleaned in default_responses:
                return default_responses[user_input_cleaned]
            best_match = find_best_match(user_input_cleaned, model, conversations_LST2, 0)
            if best_match:
                break
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Kindly wait as it may take a little time to provide accurate results.\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            best_match = find_best_match(user_input_cleaned, model, conversations_LST2, 1)
            if not best_match or 'none' in best_match.lower():
                best_match = None
            else:  # inserted
                break
        try:
            if 'none' in best_match.lower():
                best_match = None
            if best_match == 'health_health':
                best_match = 'audit_audit'
        except:
            best_match = None
        if best_match == None:
            user_input = correct_text(user_input)
            user_input = user_input.replace('_', ' ')
            for model_name in model_names:
                model = SentenceTransformer(model_name)
                user_input_embedding = model.encode(user_input.lower())
                conv_embeddings = model.encode(conversations33)
                similarities = util.cos_sim([user_input_embedding], conv_embeddings)[0]
                for i, similarity in enumerate(similarities):
                    if similarity >= 0.75:
                        if similarity > best_similarity:
                            best_similarity = similarity
                            best_match = conversations_LST2[i]
                        break
            if best_match:
                break
            model = SentenceTransformer('paraphrase-distilroberta-base-v1')
            for user_input_cleaned in user_input_cleaned3:
                user_input_cleaned = user_input_cleaned.lower()
                if not user_input_cleaned.strip():
                    user_input_cleaned = 'hi'
                if user_input_cleaned in default_responses:
                    return default_responses[user_input_cleaned]
                best_match = find_best_match(user_input_cleaned, model, conversations_LST2, 0)
                if best_match:
                    break
                self.CHATBUDDY_textbox4.configure(state='normal')
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Kindly wait checking again for accurate results.\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                self.CHATBUDDY_textbox4.configure(state='disabled')
                best_match = find_best_match(user_input_cleaned, model, conversations_LST2, 1)
                if not best_match or 'none' in best_match.lower():
                    best_match = None
                else:  # inserted
                    break
        try:
            if 'none' in best_match.lower():
                best_match = None
            if best_match == 'health_health':
                best_match = 'audit_audit'
        except:
            best_match = None
        print(best_match)
        try:
            conversations_LST2.clear()
            conversations33.clear()
        except:
            pass
        if best_match:
            return best_match
        response342 = ask_yes_or_no_qt('Do you want to add command? Reply with (yes/no).')
        if response342 == True:
            user_input = 'yes'
        else:  # inserted
            user_input = 'no'
        if user_input.lower() == '':
            user_input = 'stop'
        if user_input.lower() == 'yes':
            submit_node = submit_node22(self)
            if submit_node == 'None_None':
                thread = threading.Thread(target=enabled_button, args=(self,))
                thread.start()
                thread.join()
            return submit_node
        thread = threading.Thread(target=enabled_button, args=(self,))
        thread.start()
        thread.join()
        check_yesno.clear()
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Thanks! how Chat Buddy will assist you.\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        return 'close_chat'
    if 'FAIL' in establish_ssh_conn_lst:
        establish_ssh_conn_lst.clear()
        try:
            llll = RRSG_auto_login.remote_conn
        except:
            try:
                con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                del Rdf['indexxx']
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                PID_LST = Rdf['value'].tolist()
                remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
            except:
                RSG_LST_N = []
            paaaawe = OTPPP_tets()
            self.Auto_login_sc(paaaawe, RSG_LST_N)

    def correct_text(text):
        try:
            conversations_LST21 = conversations_LST + converstion2
            conversations_LST21 = list(set(conversations_LST21))
            conversations_LST21 = [conversations_LST21.split('_')[1].lower() for conversations_LST21 in conversations_LST21]
            MEANINGLESS_STRINGS = set(conversations_LST21)
            conversations_LST21.clear()
        except:
            MEANINGLESS_STRINGS = {}
            spell = SpellChecker(language=None, case_sensitive=False, distance=2)
            dictionary_path = './resources'
            if dictionary_path:
                spell.word_frequency.load_text_file(os.path.join(dictionary_path, 'en.json.gz'))
            words = text.split()
            corrected_words = []
            for word in words:
                close_matches = get_close_matches(word, MEANINGLESS_STRINGS, n=1, cutoff=0.8)
                if close_matches:
                    corrected_words.append(word)
                else:  # inserted
                    corrected_words.append(spell.correction(word))
            MEANINGLESS_STRINGS = {}
            return ' '.join(corrected_words)
        except:
            print('error spell check')
            return ' '.join(text)
    if SQL_METHOD == 'LOCAL':
        try:
            conversations_LST.clear()
            conn = sqlite3.connect('./res/conversations.db')
            cursor = conn.cursor()
            cursor.execute('SELECT text FROM conversations')
            rows = cursor.fetchall()
            conversations_L11 = [row[0] for row in rows]
            for bbn in conversations_L11:
                conversations_LST.append(bbn)
            try:
                conversations_L11.clear()
            except:
                pass
        except:
            pass
        try:
            con = sqlite3.connect(REPORT_INPATH22 + 'conv.db')
            Rdf = pd.read_sql_query(f"SELECT * FROM {'conversations'}", con)
            con.close()
            conv_ll = Rdf['text'].unique()
            for ddd in conv_ll:
                converstion2.append(ddd)
            try:
                conv_ll.clear()
            except:
                pass
        except:
            pass
    if SQL_METHOD == 'AZURE_CLOUDE':
        try:
            con = sqlite3.connect(REPORT_INPATH22 + 'conv.db')
            Rdf = pd.read_sql_query(f"SELECT * FROM {'conversations'}", con)
            con.close()
            conv_ll = Rdf['text'].unique()
            for ddd in conv_ll:
                converstion2.append(ddd)
            try:
                conv_ll.clear()
            except:
                pass
        except:
            pass
    if user_input.lower().strip() in ['exit', 'exit.', 'quit', 'quit.', 'q', 'stop', 'stop.', 'next site', 'next site.', 'refresh', 'refresh.']:
        return 'quit'
    exclude_txt = {'node', 'rnc', 'dekhna', 'tell', 'two', 'that', 'chat bot', 'hi', 'can', 'this', 'and', 'brother', 'i', 'check', 'or', 'for', 'identify', 'need', 'bata', 'perform', 'chat', 'dekhna', 'too', '!', 'help', 'node id', 'want', 'help', 'there', 'find', 'value', 'nodeid', 'need', 'help', 'there', 'need', 'there', 'need', 'need', 'there', 'need', 'need', 'need', 'need', 'need', 'need', 'there', 'need', 'find', 'value', '
    user_input_cleaned1 = ' '.join((word for word in user_input.split() if word.lower() not in exclude_txt))
    user_input_cleaned2 = user_input_cleaned1.replace('_', ' ')
    user_input_cleaned3 = user_input_cleaned2.split(' ')
    user_input = ' '.join(user_input_cleaned3).lower()
    try:
        lang = detect(user_input)
    except:
        lang = 'none'
    try:
        if lang == 'zh-cn':
            try:
                words = jieba.cut(user_input)
                split_text22 = list(words)
                split_text22 = list(filter(lambda x: x!= ' ', split_text22))
                word_lst = []
                for eew in split_text22:
                    word_lst.append(translate_text(eew))
                user_input = ' '.join(word_lst)
            except:
                pass
    except:
        pass
    DEFAULT_RESPONSES = {key.lower(): value for key, value in DEFAULT_RESPONSES_lst.items()}
    if user_input.lower() in DEFAULT_RESPONSES:
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Hello! How can I assist you?\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            else:  # inserted
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        return 'close_chat'
    if not user_input.strip():
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Hello! How can I assist you?\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            else:  # inserted
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        return 'close_chat'
    text_tokens = word_tokenize(user_input.strip())
    tokens_without_sw = [word for word in text_tokens if word not in stopwords.words()]
    if tokens_without_sw:
        tokens_cleantext = ' '.join(tokens_without_sw)
    else:  # inserted
        tokens_cleantext = user_input.strip
    try:
        word_tokens_new_clean = remove_stopwords(tokens_cleantext.strip())
    except:
        word_tokens_new_clean = user_input.lower().strip()
    response_keyword = get_response_keyword_find(word_tokens_new_clean.strip().lower(), DEFAULT_RESPONSES)
    return response_keyword

def unauthorized_cmd(self):
    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('You are not authorized to run this command.') + '\n')
    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
    start_index = '1.0'
    while True:
        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
        if not start_index:
            break
        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
        start_index = end_index
    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    self.after(0, self.textbox8.delete, '1.0', 'end')
    self.after(0, self.textbox8.insert, 'end', 'Press enter to excute command...')
    self.after(0, self.textbox8.see, 'end')

def kill_process_by_name(process_name):
    try:
        for proc in psutil.process_iter(['pid', 'name']):
            if proc.info['name'] == process_name:
                proc.kill()
                break
    except:
        return None

def SQLITE_DB_CLEAR(DATABASE):
    try:
        conn = sqlite3.connect('./res/my_database.db')
        cursor = conn.cursor()
        cursor.execute(f'DELETE FROM {DATABASE}')
        conn.commit()
        conn.close()
    except:
        return None

def SQLITE_DB_READ(DATABASE):
    conn = sqlite3.connect('./res/my_database.db')
    cursor = conn.cursor()
    cursor.execute(f'CREATE TABLE IF NOT EXISTS {DATABASE} (id INTEGER PRIMARY KEY AUTOINCREMENT, command TEXT)')
    conn.commit()
    cursor.execute(f'SELECT * FROM {DATABASE}')
    rows = cursor.fetchall()
    command_list = [row[1] for row in rows]
    conn.close()
    return command_list

def disable_button(self):
    for button in [self.main_button_shot, self.main_button_amos, self.main_button_CLEAR, self.CHATBUDDY_textbox4, self.main_button_big, self.main_button_big2, self.main_button_module, self.main_button_shot, self.switch3, self.switch_mo, self.sidebar_button_3, self.sidebar_button_1, self.textbox8, self.main_button_quit, self.switch_Secondary_pass_rbs, self.sidebar_Project, self.switch_VPPN_CASS]:
        button.configure(state='disabled')
    self.after(0, self.CHATBUDDY_textbox4.delete, '1.0', 'end')
    self.after(0, self.CHATBUDDY_textbox4.insert, 'end', 'Please wait for response..')
    self.after(0, self.CHATBUDDY_textbox4.see, 'end')
    self.CHATBUDDY_textbox4.configure(state='disabled')
    self.after(0, self.textbox8.delete, '1.0', 'end')
    self.after(0, self.textbox8.insert, 'end', 'Please wait for response..')
    self.after(0, self.textbox8.see, 'end')
    self.textbox8.configure(state='disabled')

def disable_button2(self):
    for button in [self.main_button_shot, self.main_button_amos, self.main_button_CLEAR, self.CHATBUDDY_textbox4, self.main_button_big, self.main_button_big2, self.main_button_module, self.main_button_shot, self.switch3, self.switch_mo, self.sidebar_button_3, self.sidebar_button_1, self.textbox8, self.main_button_quit, self.switch_Secondary_pass_rbs, self.sidebar_Project, self.switch_VPPN_CASS, self.reset_button1]:
        button.configure(state='disabled')
    self.after(0, self.CHATBUDDY_textbox4.delete, '1.0', 'end')
    self.after(0, self.CHATBUDDY_textbox4.insert, 'end', 'Please wait for response..')
    self.after(0, self.CHATBUDDY_textbox4.see, 'end')
    self.CHATBUDDY_textbox4.configure(state='disabled')
    self.after(0, self.textbox8.delete, '1.0', 'end')
    self.after(0, self.textbox8.insert, 'end', 'Please wait for response..')
    self.after(0, self.textbox8.see, 'end')
    self.textbox8.configure(state='disabled')

def enabled_button(self):
    for button in [self.main_button_shot, self.main_button_amos, self.main_button_CLEAR, self.CHATBUDDY_textbox4, self.main_button_big, self.main_button_big2, self.main_button_module, self.main_button_shot, self.switch3, self.switch_mo, self.sidebar_button_3, self.sidebar_button_1, self.textbox8, self.main_button_quit, self.switch_Secondary_pass_rbs, self.sidebar_Project, self.switch_VPPN_CASS, self.reset_button1]:
        button.configure(state='normal')
    self.CHATBUDDY_textbox4.configure(state='normal')
    self.after(0, self.CHATBUDDY_textbox4.delete, '1.0', 'end')
    self.after(0, self.CHATBUDDY_textbox4.insert, 'end', 'Please wait for response..')
    self.after(0, self.CHATBUDDY_textbox4.see, 'end')
    self.textbox8.configure(state='normal')
    self.after(0, self.textbox8.delete, '1.0', 'end')
    self.after(0, self.textbox8.insert, 'end', 'Please wait for response..')
    self.after(0, self.textbox8.see, 'end')

def SQLITE_DB(DATABASE, commands):
    conn = sqlite3.connect('./res/my_database.db')
    cursor = conn.cursor()
    cursor.execute(f'CREATE TABLE IF NOT EXISTS {DATABASE} (id INTEGER PRIMARY KEY AUTOINCREMENT, command TEXT)')
    conn.commit()
    for command in commands:
        cursor.execute(f'SELECT COUNT(*) FROM {DATABASE} WHERE command = ?', (command,))
        exists = cursor.fetchone()[0]
        if exists == 0:
            cursor.execute(f'INSERT INTO {DATABASE} (command) VALUES (?)', (command,))
    conn.commit()
    cursor.execute(f'SELECT * FROM {DATABASE}')
    rows = cursor.fetchall()
    command_list = [row[1] for row in rows]
    conn.close()
    return command_list

def encrypt(text, shift):
    encrypted_text = ''
    for char in text:
        if char.isalpha():
            shifted = ord(char) + shift
            if char.islower():
                if shifted > ord('z'):
                    shifted -= 26
                else:  # inserted
                    if shifted < ord('a'):
                        shifted += 26
            else:  # inserted
                if char.isupper():
                    if shifted > ord('Z'):
                        shifted -= 26
                    else:  # inserted
                        if shifted < ord('A'):
                            shifted += 26
            encrypted_text += chr(shifted)
        else:  # inserted
            if char.isnumeric():
                shifted = (int(char) + shift) % 10
                encrypted_text += str(shifted)
            else:  # inserted
                encrypted_text += char
    return encrypted_text

def decrypt(encrypted_text, shift):
    decrypted_text = ''
    for char in encrypted_text:
        if char.isalpha():
            shifted = ord(char) - shift
            if char.islower():
                if shifted > ord('z'):
                    shifted -= 26
                else:  # inserted
                    if shifted < ord('a'):
                        shifted += 26
            else:  # inserted
                if char.isupper():
                    if shifted > ord('Z'):
                        shifted -= 26
                    else:  # inserted
                        if shifted < ord('A'):
                            shifted += 26
            decrypted_text += chr(shifted)
        else:  # inserted
            if char.isnumeric():
                shifted = (int(char) - shift) % 10
                decrypted_text += str(shifted)
            else:  # inserted
                decrypted_text += char
    return decrypted_text

def greet_based_on_time(self):
    import datetime

    def greet_based_on_time2():
        current_time = datetime.datetime.now().time()
        if current_time < datetime.time(12, 0, 0):
            mmsgo = 'Good morning!'
            return mmsgo
        if current_time < datetime.time(17, 0, 0):
            mmsgo = 'Good afternoon!'
            return mmsgo
        mmsgo = 'Good evening!'
        return mmsgo
    gret_msg = greet_based_on_time2()

    def sucss_msg_greating(self):
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', gret_msg + ' How may I be of assistance to you?' + '\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search(gret_msg, start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 12 chars')
            self.CHATBUDDY_textbox3.tag_add(gret_msg, start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config(gret_msg, foreground='#43884e')
    threading.Thread(target=sucss_msg_greating, args=(self,)).start()

def showMessage(message, timeout=5000):
    root = tk.Tk()
    root.withdraw()
    root.after(timeout, root.destroy)
    msgbox.showinfo('Information', message, master=root)

def askYesNoStylishAutoClose_reset(question, timeout=10000):
    root = tk.Tk()
    root.withdraw()
    root.option_add('*Dialog.msg.font', 'Helvetica 12 bold')
    root.option_add('*Dialog.msg.width', 20)
    root.option_add('*Dialog.msg.wrapLength', '4i')
    root.option_add('*Dialog.msg.background', 'lightgray')
    root.option_add('*Dialog.msg.foreground', 'navy')
    root.option_add('*Dialog.msg.borderWidth', 5)
    root.option_add('*Dialog.msg.relief', 'raised')
    root.option_add('*Dialog.msg.highlightBackground', 'red')
    root.option_add('*Dialog.msg.highlightColor', 'green')
    root.wm_attributes('-topmost', True)
    root.withdraw()
    root.after(timeout, lambda: (root.destroy(), chek_exit.append('True')))
    response = msgbox.askyesno('Confirmation', question, master=root)
    if response:
        root.destroy()
    return response

def loggs_up():
    try:
        if LOG_UPLODER!= 'NOTRequired':
            for file_path in glob.glob('./res/CHATLOGS/*.Log'):
                if os.path.isfile(file_path):
                    FILENNNAME = os.path.basename(file_path)
                    sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                    sh_ul_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/USER_LOGS/' + FILENNNAME
                    client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                    client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                    sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                    up_path = sp.create_link('https:' + sh_ul_path + '/' + FILENNNAME)
                    sp.upload(sharepoint_location=up_path, local_location=file_path)
                    try:
                        os.remove(file_path)
                    except:
                        continue
    except:
        return None

def loggs_up_audit():
    try:
        if LOG_UPLODER!= 'NOTRequired':
            for file_path in glob.glob('./res/CHATLOGS/*.Log'):
                if os.path.isfile(file_path):
                    FILENNNAME = os.path.basename(file_path)
                    sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                    sh_ul_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/USER_LOGS/' + FILENNNAME
                    client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                    client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                    sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                    up_path = sp.create_link('https:' + sh_ul_path + '/' + FILENNNAME)
                    sp.upload(sharepoint_location=up_path, local_location=file_path)
                    try:
                        os.remove(file_path)
                    except:
                        continue
    except:
        return None

def askYesNoStylish(question):
    root = tk.Tk()
    root.withdraw()
    root.option_add('*Dialog.msg.font', 'Helvetica 12 bold')
    root.option_add('*Dialog.msg.width', 20)
    root.option_add('*Dialog.msg.wrapLength', '4i')
    root.option_add('*Dialog.msg.background', 'lightgray')
    root.option_add('*Dialog.msg.foreground', 'navy')
    root.option_add('*Dialog.msg.borderWidth', 5)
    root.option_add('*Dialog.msg.relief', 'raised')
    root.option_add('*Dialog.msg.highlightBackground', 'red')
    root.option_add('*Dialog.msg.highlightColor', 'green')
    answer = messagebox.askyesno('Confirmation', question)
    root.wm_attributes('-topmost', True)
    root.destroy()
    return answer

def delete_empty_folder(root_directory):
    def is_folder_empty(folder_path):
        return not any(os.listdir(folder_path))

    def delete_empty_folder_main(folder_path):
        if not os.path.exists(folder_path) or not os.listdir(folder_path):
            os.rmdir(folder_path)

    def delete_folder(folder_path):
        try:
            os.rmdir(folder_path)
        except:
            return None

    def delete_empty_folders(root_folder):
        for folder_name in os.listdir(root_folder):
            folder_path = os.path.join(root_folder, folder_name)
            if os.path.isdir(folder_path):
                if is_folder_empty(folder_path):
                    delete_folder(folder_path)
                else:  # inserted
                    delete_empty_folders(folder_path)
    try:
        delete_empty_folders(root_directory)
    except:
        pass
    try:
        delete_empty_folder_main(root_directory)
    except:
        return None

def Autenticate_Sh():
    myvariable = 5
    while myvariable > 0:
        try:
            try:
                os.remove(REPORT_INPATH22 + 'User_access.db')
            except:
                pass
            sh_dl_fname = 'User_access.db'
            sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
            sh_dl_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/reqfile/' + sh_dl_fname
            client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
            client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
            REPORT_INPATH = REPORT_INPATH22
            sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
            download_path = sp.create_link('https:' + sh_dl_path + '/' + sh_dl_fname)
            sp.download(sharepoint_location=download_path, local_location=REPORT_INPATH + sh_dl_fname)
            myvariable = 0
        except Exception:
            time.sleep(2)
            myvariable -= 1

def Autenticate_Sh2(REPORT_INPATH22):
    myvariable = 5
    while myvariable > 0:
        try:
            for sh_dl_fname in ['process.db', 'Audit_command_list_default.db', 'authnnnti.db', 'conv.db', 'CMD_SHORTCUT.db']:
                sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                sh_dl_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/reqfile/' + sh_dl_fname
                client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                download_path = sp.create_link('https:' + sh_dl_path + '/' + sh_dl_fname)
                sp.download(sharepoint_location=download_path, local_location=REPORT_INPATH22 + sh_dl_fname)
            myvariable = 0
        except Exception:
            time.sleep(2)
            myvariable -= 1

def PROJECT_ISF_DL(PID_ISF_FILE):
    myvariable = 5
    while myvariable > 0:
        try:
            for sh_dl_fname in ['ISFdetails.db']:
                sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                sh_dl_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/PROJECT_MAPPING/PID-' + str(SELECTED_PROJECT[0])
                client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                download_path = sp.create_link('https:' + sh_dl_path + '/' + sh_dl_fname)
                sp.download(sharepoint_location=download_path, local_location=PID_ISF_FILE + sh_dl_fname)
            myvariable = 0
        except Exception:
            time.sleep(2)
            myvariable -= 1

def Autenticate_Sh3(REPORT_INPATH22):
    myvariable = 5
    while myvariable > 0:
        try:
            for sh_dl_fname in ['authnnnti_2.db']:
                sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                sh_dl_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/reqfile/' + sh_dl_fname
                client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                download_path = sp.create_link('https:' + sh_dl_path + '/' + sh_dl_fname)
                sp.download(sharepoint_location=download_path, local_location=REPORT_INPATH22 + sh_dl_fname)
            myvariable = 0
        except Exception:
            time.sleep(2)
            myvariable -= 1

def ALL_PROJECT_MAPPING_LST2(path):
    try:
        con = sqlite3.connect(path)
        PROJECT_MAPPING_use = pd.read_sql_query(f"SELECT * FROM {'MyTable'}", con)
        PROJECT_MAPPING_use['PROJECT_USER_MAP'] = PROJECT_MAPPING_use['PROJECT ID'].astype(str) + '_' + PROJECT_MAPPING_use['Signum'].astype(str) + '_' + PROJECT_MAPPING_use['Role'].astype(str)
        PROJECT_LST_1 = PROJECT_MAPPING_use['PROJECT ID'].unique()
        PROJECT_LST_1 = list(PROJECT_LST_1)
        PROJECT_USER_LST_1 = PROJECT_MAPPING_use['PROJECT_USER_MAP'].unique()
        PROJECT_USER_LST_1 = list(PROJECT_USER_LST_1)
        for pp in PROJECT_LST_1:
            PROJECT_LST.append(str(pp))
        for pp in PROJECT_USER_LST_1:
            PROJECT_SME_LST.append(str(pp))
    except:
        return None

def PROJECT_MAPPING(REPORT_INPATH22):
    myvariable = 5
    while myvariable > 0:
        try:
            for sh_dl_fname in ['PROJECT_MAPPING_SME.db']:
                sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                sh_dl_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/reqfile/' + sh_dl_fname
                client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                download_path = sp.create_link('https:' + sh_dl_path + '/' + sh_dl_fname)
                try:
                    os.remove(REPORT_INPATH22 + 'PROJECT_MAPPING_SME.db')
                except:
                    pass
                sp.download(sharepoint_location=download_path, local_location=REPORT_INPATH22 + sh_dl_fname)
            myvariable = 0
        except Exception:
            time.sleep(2)
            myvariable -= 1

def PROJECT_MAPPING_LST2(path, PIDDD):
    try:
        con = sqlite3.connect(path)
        PROJECT_MAPPING_use = pd.read_sql_query(f"SELECT * FROM {'MyTable'}", con)
        PROJECT_MAPPING_use['PROJECT ID'] = PROJECT_MAPPING_use['PROJECT ID'].astype(str)
        PROJECT_MAPPING_use = PROJECT_MAPPING_use.loc[PROJECT_MAPPING_use['PROJECT ID'] == str(PIDDD)]
        PROJECT_USER_LST_1 = PROJECT_MAPPING_use['Signum'].unique()
        PROJECT_USER_LST_1 = list(PROJECT_USER_LST_1)
        for pp in PROJECT_USER_LST_1:
            if str(pp) not in PROJECT_USER_LST:
                PROJECT_USER_LST.append(str(pp.lower()))
        if os.getlogin().lower() not in PROJECT_USER_LST:
            PROJECT_USER_LST.insert(0, os.getlogin().lower())
    except:
        return None

def PROJECT_WISE_MAPPING(PIID, REPORT_INPATH22):
    myvariable = 5
    while myvariable > 0:
        try:
            for sh_dl_fname in [PIID + '_PROJECT_MAPPING.db']:
                sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                sh_dl_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/PROJECT_MAPPING/PID-' + str(SELECTED_PROJECT[0])
                client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                download_path = sp.create_link('https:' + sh_dl_path + '/' + sh_dl_fname)
                try:
                    os.remove(REPORT_INPATH22 + PIID + '_PROJECT_MAPPING.db')
                except:
                    pass
                sp.download(sharepoint_location=download_path, local_location=REPORT_INPATH22 + sh_dl_fname)
            myvariable = 0
        except Exception:
            time.sleep(2)
            myvariable -= 1

def conv_Sh2(REPORT_INPATH22):
    while True:
        try:
            CHK_TO_CONTinue2.append('continue')
            return
        except:
            continue

def conv_Sh2_2(REPORT_INPATH22):
    myvariable = 5
    while myvariable > 0:
        try:
            sh_dl_fname = 'conv.db'
            sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
            sh_dl_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/reqfile/' + sh_dl_fname
            client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
            client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
            sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
            download_path = sp.create_link('https:' + sh_dl_path + '/' + sh_dl_fname)
            sp.download(sharepoint_location=download_path, local_location=REPORT_INPATH22 + sh_dl_fname)
            myvariable = 0
        except Exception:
            time.sleep(2)
            myvariable -= 1

def gear_widget():
    class GearLoadingWidget(QWidget):
        def __init__(self, parent=None):
            super(GearLoadingWidget, self).__init__(parent)
            self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
            self.setAttribute(Qt.WA_TranslucentBackground)
            self.init_ui()

        def init_ui(self):
            layout = QVBoxLayout(self)
            self.loading_label = QLabel(self)
            layout.addWidget(self.loading_label, alignment=Qt.AlignCenter)
            movie = QMovie('./res/loading1.gif')
            self.loading_label.setMovie(movie)
            movie.start()
            timer = QTimer(self)
            timer.timeout.connect(self.close_application)
            timer.start(2000)
            self.setWindowTitle('Transparent Gear Loading Icon Example')
            self.setGeometry(850, 400, 300, 200)
            self.setStyleSheet('background: transparent;')
            self.setWindowFlag(Qt.WindowStaysOnTopHint)

        def close_application(self):
            self.close()
    app = QApplication(sys.argv)
    window = GearLoadingWidget()
    window.show()
    app.exec_()

def CIPRIDFFF(text):
    try:
        pattern = 'Total: (\\d+) CPRI links \\(.*\\)'
        match = re.search(pattern, text)
        if match:
            cipri_data = match.group(0)
        pattern = 'Total: (\\d+) CPRI links \\('
        match = re.search(pattern, text)
        if match:
            total_count = match.group(1)
        pattern = 'Total: \\d+ CPRI links \\((\\d+) OK,.*\\)'
        match = re.search(pattern, text)
        if match:
            ok_count = match.group(1)
        pattern = 'Total: \\d+ CPRI links \\(\\d+ OK, (\\d+) OKW,.*\\)'
        match = re.search(pattern, text)
        if match:
            OKW_count = match.group(1)
        pattern = 'Total: \\d+ CPRI links \\(\\d+ OK, \\d+ OKW, (\\d+) NOK, \\d+ NT\\)'
        match = re.search(pattern, text)
        if match:
            nok_count = match.group(1)
        cipri_data = cipri_data.replace('Total: ', '')
        CIPRIDF = pd.DataFrame({'CPRI links': [cipri_data], 'Total CPRI Link': [total_count], 'CPRI Link OK': [ok_count], 'CPRI Link OKW': [OKW_count], 'CPRI Link NOK': [nok_count]})
        return CIPRIDF
    except:
        CIPRIDF = pd.DataFrame()

def AZURE_CLOUDE_login(self):
    global my_transport  # inserted
    global remote_ssh_client  # inserted
    try:
        CHK_TO_CONTinue2.clear()
        con = sqlite3.connect('./res/AZURE_CRED.db')
        Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
        con.close()
        del Rdf['indexxx']
        Rdf = Rdf.drop_duplicates(subset=['value'], keep='first')
        ppPID_LST = Rdf['value'].tolist()
        ppPID_LST2 = []
        for ppp in ppPID_LST:
            ppPID_LST2.append(decrypt(ppp, 3))
        del Rdf['value']
        Rdf['value'] = ppPID_LST2
        PID_LST = Rdf['value'].tolist()
        remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5']
        AZ_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
        AZ_LST.insert(1, os.getlogin().lower())
        try:
            del Rdf
        except:
            pass
    except:
        AZ_LST = []
    try:
        azure_cred = 'YES'
        if AZ_LST:
            for kl in AZ_LST:
                if 'LINE' in kl:
                    azure_cred = 'NO'
                    break
        if not AZ_LST or azure_cred == 'YES':
            gateway_host = AZ_LST[0]
            remote_host = AZ_LST[3]
            remote_username = AZ_LST[4]
            remote_password = AZ_LST[5]
            gateway_user = os.getlogin().lower()
            gateway_pass = AZ_LST[2]
            gateway_Token = Azure_OTP()
            if gateway_Token!= None:
                if az_check:
                    break
                az_check.append('continue')
            if az_check:
                if gateway_Token == None:
                    thread = threading.Thread(target=enabled_button, args=(self,))
                    thread.start()
                    thread.join()
            else:  # inserted
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Sorry, Chatbuddy unable to connect to Azure Cloud at the moment. Please refresh the application and try again.') + '\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            if gateway_Token!= None:
                Azue_OTP_lst.append('az_otp')
            if gateway_Token!= None:
                self.CHATBUDDY_textbox3.configure(state='normal')
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Please wait connecting to Azure cloud.') + '\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                remote_file_path = '/home/inputdata_TIGO-TZ/Input/Ericsson/4G/RCA/Test_data/conversations.db'
                SQL_TABLE_signum = 'SELECT signum FROM change_writes'
                SQL_TABLE_barred_cmd = 'SELECT text FROM conversations_barred'

                def inter_handler(title, instructions, prompt_list):
                    resp = []
                    for pr in prompt_list:
                        if str(pr[0]).strip() == 'EGAD-Password:':
                            resp.append(gateway_pass)
                        if 'Enter Your Microsoft verification code' in str(pr[0]).strip():
                            resp.append(gateway_Token)
                    return tuple(resp)
                gateway_port = 22
                my_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                my_socket.connect((gateway_host, gateway_port))
                my_transport = paramiko.Transport(my_socket)
                my_transport.start_client(timeout=5020)
                my_transport.auth_interactive(gateway_user, inter_handler)
                channel = my_transport.open_session()
                transport = channel.get_transport()
                dest_addr = (remote_host, 22)
                local_addr = ('', 0)
                tunnel = transport.open_channel('direct-tcpip', dest_addr, local_addr)
                remote_ssh_client = paramiko.SSHClient()
                remote_ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                remote_ssh_client.connect(remote_host, username=remote_username, password=remote_password, sock=tunnel)
                connection_string = f'sqlite3 {remote_file_path} \"{SQL_TABLE_signum};\"'
                stdin, stdout, stderr = remote_ssh_client.exec_command(connection_string)
                change_User1 = stdout.readlines()
                change_User1 = [string.replace('\n', '') for string in change_User1]
                change_User1 = [element.lower() for element in change_User1]
                connection_string = f'sqlite3 {remote_file_path} \"{SQL_TABLE_barred_cmd};\"'
                stdin, stdout, stderr = remote_ssh_client.exec_command(connection_string)
                command_barred1 = stdout.readlines()
                command_barred1 = [string.replace('\n', '') for string in command_barred1]
                command_barred1 = [element.lower() for element in command_barred1]
                CONVERSION_TABLE = 'SELECT text FROM conversations'
                connection_strin22 = f'sqlite3 {remote_file_path} \"{CONVERSION_TABLE};\"'
                stdin, stdout, stderr = remote_ssh_client.exec_command(connection_strin22)
                conversations11 = stdout.readlines()
                conversations11 = [string.replace('\n', '') for string in conversations11]
                conversations11 = [element.lower() for element in conversations11]
                DF = pd.DataFrame()
                DF['text'] = conversations11
                try:
                    os.remove(REPORT_INPATH22 + 'conv.db')
                except:
                    pass
                con = sqlite3.connect(REPORT_INPATH22 + 'conv.db')
                try:
                    DF.to_sql('conversations', con, index=False, if_exists='replace')
                except:
                    pass
                try:
                    del DF
                except:
                    pass
                try:
                    conversations11.clear()
                except:
                    pass
                for nn in change_User1:
                    change_User.append(nn)
                for nn in command_barred1:
                    command_barred.append(nn)
                CHK_TO_CONTinue2.append('connected')
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Connected to Azure cloud.') + '\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    except:
        self.CHATBUDDY_textbox3.configure(state='normal')
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('connection refused please check Azure credential.') + '\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        self.CHATBUDDY_textbox3.configure(state='disabled')
        AZURE_CREAD_FAIL.append('Fail')

def ISF_STEP_CLOSE(ISF_API_Based, ISF_STEP_START_Task_Id, ISF_WO_CREATE_woId, ISF_WO_CREATE_headers):
    if ISF_WORKORDER == 'Required':
        timestrCT2 = time.strftime('%Y-%m-%d %H:%M:%S')
        URLc1 = 'https://integratedserviceflow.ericsson.net/apim/v1/externalInterface/stopTask'
        payload = {'wOID': ISF_WO_CREATE_woId, 'taskID': ISF_STEP_START_Task_Id, 'reason': 'SUCCESS', 'externalSourceName': ISF_API_Based}
        URLc3 = URLc1
        response2 = requests.post(URLc3, headers=ISF_WO_CREATE_headers, json=payload)
        return 'Please provide the valid taskID for the Step' if 'Please provide the valid taskID for the Step' in response2.text else 'sucess'
    return 'sucess'

def ISF_WO_CLOSE(ISF_API_Based, ISF_WO_CREATE_headers, ISF_WO_CREATE_woId):
    ISF_WORKORDER_con = 'Required'
    if ISF_WORKORDER == 'Required' and CHECKISFWO:
        if ASK_ISF_CLOSE == 'YES':
            response = ask_yes_or_no_qt('Do you want to close the ISF work order? Otherwise, it will auto-close in 10 minutes')
            if response == True:
                ISF_WORKORDER_con = 'Required'
            else:  # inserted
                ISF_WORKORDER_con = 'NOTRequired'
        else:  # inserted
            if False:
                pass  # postinserted
    if ISF_WORKORDER_con == 'Required':
        if ISF_WORKORDER == 'Required':
            CHECKISFWO.clear()
            timestrCT3 = time.strftime('%Y-%m-%d %H:%M:%S')
            URLw1 = 'https://integratedserviceflow.ericsson.net/apim/v1/externalInterface/closeWO'
            payload = {'wOID': ISF_WO_CREATE_woId, 'deliveryStatus': 'Success', 'externalSourceName': ISF_API_Based, 'lastModifiedBy': os.getlogin().lower()}
            URLw = URLw1
            response3 = requests.post(URLw, headers=ISF_WO_CREATE_headers, json=payload)
            if 'Please complete all the tasks of this work order to close this' in response3.text:
                CHECKISFWO.clear()
                return 'Please complete all the tasks of this work order to close this.'
            WORKORDER_CLOSE.append('eoclosed')
            clear_all_lists()
            return 'sucess'
        WORKORDER_CLOSE.append('eoclosed')
        clear_all_lists()
        return 'sucess'
    WORKORDER_CLOSE.append('eoclosed')
    clear_all_lists()
    return 'sucess'

def ISF_API(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY, ISF_STEP_START_stepID, ISF_STEP_START_Task_Id):
    CHECKISFWO.clear()
    if ISF_WORKORDER == 'Required':
        def ISF_WO_CREATE(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY):
            try:
                timestr = datetime.now() + timedelta(minutes=5)
                timestr2 = timestr.strftime('%Y-%m-%d')
                timestr1 = timestr.strftime('%H:%M:%S')
                l1 = 'https://integratedserviceflow.ericsson.net/apim/v1/externalInterface/createWorkOrderPlan/'
                urlcreateWO = l1
                Parameters = {'projectID': projectID, 'priority': priority, 'startDate': timestr2, 'startTime': timestr1, 'wOName': wOName, 'assignedTo': sig, 'lastModifiedBy': lastModifiedBy, 'listOfNode': [{'nodeNames': nodname, 'nodeType': type1}], 'comment': ISF_API_Based, 'executionPlanName': executionPlanName, 'externalSourceName': ISF_API_Based}
                ISF_WO_CREATE_headers = {'Content-Type': 'application/json', 'Apim-Subscription-Key': KEY}
                response = requests.post(urlcreateWO, data=json.dumps(Parameters), headers=ISF_WO_CREATE_headers)
                getResponse = response.json().get('responseData')
                workOrderObj = getResponse.get('WorkOrderID')
                ISF_WO_CREATE_woId = workOrderObj[0].get('woId')
                ISF_WO_CREATE.timestr = timestr
                ISF_WO_CREATE.urlcreateWO = urlcreateWO
                ISF_WO_CREATE.ISF_WO_CREATE_headers = ISF_WO_CREATE_headers
                CHECKISFWO.append(ISF_WO_CREATE_woId)
                return (urlcreateWO, timestr, ISF_WO_CREATE_headers, ISF_WO_CREATE_woId)
            except:
                return ('STOP_WO', 'STOP_WO', 'STOP_WO', 'STOP_WO')

        def ISF_STEP_START(ISF_API_Based, ISF_STEP_START_stepID, ISF_WO_CREATE_woId, ISF_WO_CREATE_headers, ISF_STEP_START_Task_Id):
            timestrCT = time.strftime('%Y-%m-%d %H:%M:%S')
            URL = 'https://integratedserviceflow.ericsson.net/apim/v1/externalInterface/startTask'
            payload = {'wOID': ISF_WO_CREATE_woId, 'stepID': ISF_STEP_START_stepID, 'externalSourceName': ISF_API_Based, 'taskID': ISF_STEP_START_Task_Id}
            URL3 = URL
            res = requests.post(URL3, headers=ISF_WO_CREATE_headers, json=payload)
            return 'Please provide the valid taskID for the Step' if 'does not exists for woID' in res.text else 'sucess'
        urlcreateWO, timestr, ISF_WO_CREATE_headers, ISF_WO_CREATE_woId = ISF_WO_CREATE(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY)
        ISF_API.headers = ISF_WO_CREATE_headers
        ISF_API.woid = ISF_WO_CREATE_woId
        ISF_API.API_Based = ISF_API_Based
        try:
            ISF_API.timestr = ISF_WO_CREATE.timestr
        except:
            ISF_API.timestr = ''
        try:
            ISF_API.urlcreateWO = ISF_WO_CREATE.urlcreateWO
        except:
            ISF_API.urlcreateWO = ''
        try:
            ISF_API.ISF_WO_CREATE_headers = ISF_WO_CREATE.ISF_WO_CREATE_headers
        except:
            ISF_API.ISF_WO_CREATE_headers = ''
        if ISF_WO_CREATE_woId!= 'STOP_WO':
            STEP_START = ISF_STEP_START(ISF_API_Based, ISF_STEP_START_stepID, ISF_WO_CREATE_woId, ISF_WO_CREATE_headers, ISF_STEP_START_Task_Id)
            return (ISF_WO_CREATE_woId, STEP_START)
        return ('Some issue in isf workflow', 'STOP_STEP_START')
    return ('222334859', 'sucess')

def ISFLoginApp_fun():
    class AZURELoginApp(QWidget):
        def __init__(self):
            super().__init__()
            self.load_credentials()
            self.setGeometry(300, 300, 600, 526)
            self.setWindowTitle('ISF DETAILS')
            self.setFixedSize(650, 720)
            self.center_window()
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.create_gui()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def load_credentials(self):
            try:
                db_path = os.path.join('./res/ISFdetails.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                con.close()
                self.PID_LST = Rdf['value'].tolist()
                while len(self.PID_LST) < 8:
                    self.PID_LST.append(f'LINE{len(self.PID_LST) + 1}')
            except:
                self.PID_LST = [f'LINE{i + 1}' for i in range(8)]
            if self.PID_LST[0] == 'LINE1':
                self.PID_LST[0] = 'Enter project id'
            if self.PID_LST[1] == 'LINE2':
                self.PID_LST[1] = 'Enter network element type'
            if self.PID_LST[2] == 'LINE3':
                self.PID_LST[2] = 'Enter network element Name'
            if self.PID_LST[3] == 'LINE4':
                self.PID_LST[3] = 'Enter WOName'
            if self.PID_LST[4] == 'LINE5':
                self.PID_LST[4] = 'Enter deliverable name'
            if self.PID_LST[5] == 'LINE6':
                self.PID_LST[5] = 'Last modified by'
            if self.PID_LST[6] == 'LINE7':
                self.PID_LST[6] = 'Enter work flow stepId'
            if self.PID_LST[7] == 'LINE8':
                self.PID_LST[7] = 'Enter work flow taskId'

        def create_gui(self):
            layout = QVBoxLayout(self)
            label = QLabel('                                  ISF DETAILS', self)
            label.setStyleSheet('font-size: 24px; color: #fe6f5e; font-weight: bold;')
            layout.addWidget(label)
            self.entries = []
            PROJECTID_label = QLabel('Project id:', self)
            PROJECTID_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(PROJECTID_label)
            PROJECTID_entry = QLineEdit(self)
            font = QFont('Calibri', 11)
            PROJECTID_entry.setFont(font)
            PROJECTID_entry.setFont(font)
            if os.getlogin().lower() == 'emshsni':
                PROJECTID_entry.setReadOnly(False)
            else:  # inserted
                PROJECTID_entry.setReadOnly(True)
            PROJECTID_entry.setFixedSize(615, 35)
            PROJECTID_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(PROJECTID_entry)
            PROJECTID_entry.setText(self.PID_LST[0])
            self.entries.append(PROJECTID_entry)
            ELEMENTTYPE_label = QLabel('Network Element type:', self)
            ELEMENTTYPE_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(ELEMENTTYPE_label)
            ELEMENTTYPE_entry = QLineEdit(self)
            font = QFont('Calibri', 11)
            ELEMENTTYPE_entry.setFont(font)
            ELEMENTTYPE_entry.setFont(font)
            ELEMENTTYPE_entry.setFixedSize(615, 35)
            ELEMENTTYPE_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(ELEMENTTYPE_entry)
            ELEMENTTYPE_entry.setText(self.PID_LST[1])
            self.entries.append(ELEMENTTYPE_entry)
            ELEMENTNAME_label = QLabel('Network Element Name:', self)
            ELEMENTNAME_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(ELEMENTNAME_label)
            ELEMENTNAME__entry = QLineEdit(self)
            font = QFont('Calibri', 11)
            ELEMENTNAME__entry.setFont(font)
            ELEMENTNAME__entry.setFont(font)
            ELEMENTNAME__entry.setFixedSize(615, 35)
            ELEMENTNAME__entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(ELEMENTNAME__entry)
            ELEMENTNAME__entry.setText(self.PID_LST[2])
            self.entries.append(ELEMENTNAME__entry)
            WONAME_LABE = QLabel('WOName:', self)
            WONAME_LABE.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(WONAME_LABE)
            WONAME_entry = QLineEdit(self)
            WONAME_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            WONAME_entry.setFont(font)
            WONAME_entry.setFont(font)
            if os.getlogin().lower() == 'emshsni':
                WONAME_entry.setReadOnly(False)
            else:  # inserted
                WONAME_entry.setReadOnly(True)
            WONAME_entry.setFixedSize(615, 35)
            WONAME_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(WONAME_entry)
            WONAME_entry.setText(self.PID_LST[3])
            self.entries.append(WONAME_entry)
            deleverable_label = QLabel('Deliverable Name:', self)
            deleverable_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(deleverable_label)
            deleverable_entry = QLineEdit(self)
            deleverable_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            deleverable_entry.setFont(font)
            deleverable_entry.setFont(font)
            if os.getlogin().lower() == 'emshsni':
                deleverable_entry.setReadOnly(False)
            else:  # inserted
                deleverable_entry.setReadOnly(True)
            deleverable_entry.setFixedSize(615, 35)
            deleverable_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(deleverable_entry)
            deleverable_entry.setText(self.PID_LST[4])
            self.entries.append(deleverable_entry)
            Last_Modified_label = QLabel('Last Modified by:', self)
            Last_Modified_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(Last_Modified_label)
            Last_Modified_entry = QLineEdit(self)
            Last_Modified_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            Last_Modified_entry.setFont(font)
            if os.getlogin().lower() == 'emshsni':
                Last_Modified_entry.setReadOnly(False)
            else:  # inserted
                Last_Modified_entry.setReadOnly(True)
            Last_Modified_entry.setFont(font)
            Last_Modified_entry.setFixedSize(615, 35)
            Last_Modified_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(Last_Modified_entry)
            Last_Modified_entry.setText(self.PID_LST[5])
            self.entries.append(Last_Modified_entry)
            WF_step_id_label = QLabel('WF stepID:', self)
            WF_step_id_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(WF_step_id_label)
            WF_step_id_entry = QLineEdit(self)
            WF_step_id_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            WF_step_id_entry.setFont(font)
            WF_step_id_entry.setFont(font)
            if os.getlogin().lower() == 'emshsni':
                WF_step_id_entry.setReadOnly(False)
            else:  # inserted
                WF_step_id_entry.setReadOnly(True)
            WF_step_id_entry.setFixedSize(615, 35)
            WF_step_id_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(WF_step_id_entry)
            WF_step_id_entry.setText(self.PID_LST[6])
            self.entries.append(WF_step_id_entry)
            WF_Task_id_label = QLabel('WF Taskid:', self)
            WF_Task_id_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(WF_Task_id_label)
            WF_Task_id_entry = QLineEdit(self)
            WF_Task_id_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            WF_Task_id_entry.setFont(font)
            WF_Task_id_entry.setFont(font)
            if os.getlogin().lower() == 'emshsni':
                WF_Task_id_entry.setReadOnly(False)
            else:  # inserted
                WF_Task_id_entry.setReadOnly(True)
            WF_Task_id_entry.setFixedSize(615, 35)
            WF_Task_id_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(WF_Task_id_entry)
            WF_Task_id_entry.setText(self.PID_LST[7])
            self.entries.append(WF_Task_id_entry)
            update_button = QPushButton('UPDATE', self)
            update_button.clicked.connect(self.login)
            update_button.setStyleSheet('padding: 8px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;')
            layout.addWidget(update_button)
            USER_ADD_button = QPushButton('ADD USERS IN PROJECT', self)
            USER_ADD_button.clicked.connect(self.UPDATE_USER)
            USER_ADD_button.setStyleSheet('padding: 8px; background-color: #d49f0f; color: white; border: none; border-radius: 4px; font-weight: bold;')
            USER_ADD_button.setFixedWidth(200)
            self.setLayout(layout)

        def UPDATE_USER(self):
            self.close()
            PROJECT_USRE_DETAILS_ADD_SME_ROLE()

        def login(self):
            values = []
            for entry in self.entries:
                values.append(encrypt(entry.text(), 3))
            DF = pd.DataFrame({'indexxx': [f'Line{i + 1}' for i in range(8)]})
            DF['value'] = values
            try:
                os.remove(os.path.join('./res/ISFdetails.db'))
            except:
                pass
            con = sqlite3.connect(os.path.join('./res/ISFdetails.db'))
            try:
                DF.to_sql('Credd', con, index=False)
            except:
                pass
            finally:  # inserted
                con.close()
            try:
                def uplode_datata_2():
                    FILENNNAME = 'ISFdetails.db'
                    sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                    sh_ul_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/PROJECT_MAPPING/PID-' + str(SELECTED_PROJECT[0])
                    client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                    client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                    sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                    up_path = sp.create_link('https:' + sh_ul_path + '/' + FILENNNAME)
                    sp.upload(sharepoint_location=up_path, local_location='./res/ISFdetails.db')
                threading.Thread(target=uplode_datata_2).start()
            except:
                pass
            self.close()
    if __name__ == '__main__':
        app = QApplication([])
        rsg_app = AZURELoginApp()
        rsg_app.show()
        app.exec_()

def combined_xl_audit(directory11, TECHNOR):
    try:
        files = os.listdir(directory11)
        files = [file for file in files if 'Audit_Output' not in file]
        merged_data = {}
        for file in files:
            if file.endswith('.xlsx'):
                sheets = pd.read_excel(os.path.join(directory11, file), sheet_name=None)
                for sheet_name, sheet_data in sheets.items():
                    if sheet_name in merged_data:
                        merged_data[sheet_name] = pd.concat([merged_data[sheet_name], sheet_data], ignore_index=True)
                    else:  # inserted
                        merged_data[sheet_name] = sheet_data.copy()
        with pd.ExcelWriter(directory11 + '\\' + str(TECHNOR) + str('_Combined_KGET_Dump') + '_' + datetime.now().strftime('%d%m%Y') + '.xlsx') as writer:
            for sheet_name, sheet_data in merged_data.items():
                sheet_data.T.reset_index().T.to_excel(writer, sheet_name=sheet_name, engine='xlsxwriter', index=False, header=None)
        try:
            writer.close()
        except:
            return
    except:
        return None

def apply_conditional_formatting(df, thresholds):
    formatted_df = df.copy()
    for index, row in thresholds.iterrows():
        try:
            condition = row['Condition']
            threshold = float(row['Threshold'])
            if condition == '<':
                formatted_df[row['KPI']] = formatted_df[row['KPI']].apply(lambda x: f'<span style=\"background-color: red; color: white;\">{x}</span>' if float(x) < threshold else x)
            else:  # inserted
                if condition == '>':
                    formatted_df[row['KPI']] = formatted_df[row['KPI']].apply(lambda x: f'<span style=\"background-color: red; color: white;\">{x}</span>' if float(x) > threshold else x)
                else:  # inserted
                    if condition == '>=':
                        formatted_df[row['KPI']] = formatted_df[row['KPI']].apply(lambda x: f'<span style=\"background-color: red; color: white;\">{x}</span>' if float(x) >= threshold else x)
                    else:  # inserted
                        if condition == '<=':
                            formatted_df[row['KPI']] = formatted_df[row['KPI']].apply(lambda x: f'<span style=\"background-color: red; color: white;\">{x}</span>' if float(x) <= threshold else x)
                        else:  # inserted
                            if condition == '==':
                                formatted_df[row['KPI']] = formatted_df[row['KPI']].apply(lambda x: f'<span style=\"background-color: red; color: white;\">{x}</span>' if float(x) == threshold else x)
                            else:  # inserted
                                if condition == '!=':
                                    formatted_df[row['KPI']] = formatted_df[row['KPI']].apply(lambda x: f'<span style=\"background-color: red; color: white;\">{x}</span>' if float(x)!= threshold else x)
        except:
            continue
    return formatted_df

def send_email(self, sender_email, sender_password, receiver_email, subject, message_body, excel_filename, sheet_name, df, thresholds, message_body2, cc_emails, receiver_emails):
    try:
        smtp_server = smtplib.SMTP('smtps.internal.ericsson.com', 587)
        smtp_server.starttls()
        smtp_server.login(sender_email, sender_password)
        df_formatted = apply_conditional_formatting(df, thresholds)
        html_table = df_formatted.to_html(index=False, escape=False)
        html_table = html_table.replace('<td>', '<td style=\"text-align: center;\">')
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ', '.join(receiver_emails)
        msg['Cc'] = ', '.join(cc_emails)
        msg['Subject'] = subject
        email_body = f'        <html>\n            <head>\n                <style>\n                    th {\n                        background-color: #cccccc; /* Change this to your desired header color */\n                    }\n                </style>\n            </head>\n            <body>\n             <p>{message_body}</p>\n                <p>{message_body2}</p>\n                {html_table}\n                <p style=\"font-weight: bold;\">Thanks</p>\n            </body>\n        </html>\n        '
        msg.attach(MIMEText(email_body, 'html'))
        with open(excel_filename, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename= {os.path.basename(excel_filename)}')
        msg.attach(part)
        smtp_server.send_message(msg)
        try:
            eyye = RRSG_auto_login.remote_conn
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : KPI Monitoring email send !\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        except:
            pass
        smtp_server.quit()
    except:
        return None

def EMAIL_INITIAL_KPI(self, Techhh, threshold_file_path, excel_file_path, sheet_name):
    try:
        import openpyxl
        FILTER_KPI = EMAIL_METHOOD_LST[0]
        db_path = os.path.join('./res/EMAIL_CRED.db')
        con = sqlite3.connect(db_path)
        cursor = con.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
        con.commit()
        Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
        con.close()
        ppPID_LST = Rdf['value'].tolist()
        ppPID_LST2 = []
        for ppp in ppPID_LST:
            ppPID_LST2.append(decrypt(ppp, 3))
        del Rdf['value']
        Rdf['value'] = ppPID_LST2
        PID_LST = Rdf['value'].tolist()
        email_to_list = PID_LST[2].split('\n') if len(PID_LST) >= 1 else []
        email_cc_list = PID_LST[3].split('\n') if len(PID_LST) >= 1 else []
        EMAIL_USER = PID_LST[0]
        EMAIL_PASS = PID_LST[1]
        sender_email = EMAIL_USER
        sender_password = EMAIL_PASS
        receiver_emails = email_to_list
        cc_emails = email_cc_list
        message_body = 'Hello All,'
        message_body2 = 'Please find Initial KPI.'
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[sheet_name]
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        NODELST = df['Node Id'].unique()
        NODELST = list(NODELST)
        if FILTER_KPI == 'Node Wise':
            for nodd in NODELST:
                df_filter = df.copy()
                df_filter = df_filter.loc[df_filter['Node Id'] == str(nodd)]
                subject = Techhh + ' ' + str(nodd) + ' Initial KPI Monitoring ' + datetime.now().strftime('%d-%B %H:%M:%S')
                thresholds_df = pd.read_excel(threshold_file_path, sheet_name='Initial KPI Threshold')
                thresholds_df = thresholds_df.loc[thresholds_df['TECH'] == Techhh]
                thresholds_df = thresholds_df[['KPI', 'Condition', 'Threshold', 'Status']]
                send_email(self, sender_email, sender_password, 'flfl', subject, message_body, excel_file_path, sheet_name, df_filter, thresholds_df, message_body2, cc_emails, receiver_emails)
        else:  # inserted
            subject = Techhh + ' Initial KPI Monitoring ' + datetime.now().strftime('%d-%B %H:%M:%S')
            thresholds_df = pd.read_excel(threshold_file_path, sheet_name='Initial KPI Threshold')
            thresholds_df = thresholds_df.loc[thresholds_df['TECH'] == Techhh]
            thresholds_df = thresholds_df[['KPI', 'Condition', 'Threshold', 'Status']]
            send_email(self, sender_email, sender_password, 'kfk', subject, message_body, excel_file_path, sheet_name, df, thresholds_df, message_body2, cc_emails, receiver_emails)
    except:
        return None

def RRSG_KPI_MONITORING_Module(self, selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST, activity_type, BOTTYPE):
    activity_type = 'KPI MONITORING'
    PATHHHH_TIME_AUDIT = time.strftime('%d%m%Y%H%M%S')
    try:
        con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
        Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
        con.close()
        del Rdf['indexxx']
        ppPID_LST = Rdf['value'].tolist()
        ppPID_LST2 = []
        for ppp in ppPID_LST:
            ppPID_LST2.append(decrypt(ppp, 3))
        del Rdf['value']
        Rdf['value'] = ppPID_LST2
        PID_LST = Rdf['value'].tolist()
        remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
        Login_comm_lst = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
    except:
        Login_comm_lst = []
    try:
        db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db')
        con = sqlite3.connect(db_path)
        cursor = con.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
        con.commit()
        Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
        con.close()
        Rdf = Rdf.drop_duplicates(keep='first')
        PID_LST = Rdf['value'].tolist()
        HOSTIP = PID_LST[0]
    except:
        HOSTIP = '148.135.15.71'
    try:
        self.LIVECMD_TEXT.delete(1.0, 'end')
    except:
        pass
    AZURE_CREAD_FAIL.clear()
    Azue_OTP_lst.clear()
    establish_ssh_conn_lst.clear()
    KEYWORD_ADD.clear()
    excel_df_LST.clear()
    amos_node.clear()
    Buddy_request_lst.clear()
    buddy_node.clear()
    buddy_node_curr.clear()
    amos_nod2.clear()
    ltall_nod2.clear()
    check_yesno.clear()
    ASKSITID.clear()
    cmmmddd.clear()
    siteddd.clear()
    checknxtsit.clear()
    processdone.clear()
    combined_cmd_window.clear()
    combined_cmd_window_chk.clear()
    checkamos.clear()
    command.clear()
    LSTTT_cmdd = []
    for check_tec in selected_checkboxes:
        if check_tec == 'GSM' and activity_type in ['KPI MONITORING']:
            Node_cmd_lst = GSM_NODE_LST
            router = HOSTIP
            username = os.getlogin().lower()
            password = 'OK'
            Save_logfile = 'YES'
            MOTASK = 'MOBATCH'
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait kpi monitoring ongoing !\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            RRSG_monitoring_process(self, Login_comm_lst, Node_cmd_lst, GSM_LST, router, username, password, Save_logfile, MOTASK, 'Volte', activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd)
            filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'KPI Monitoring ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
            initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
            EMAIL_INITIAL_KPI(self, check_tec, INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI')
            try:
                eyye = RRSG_auto_login.remote_conn
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : KPI Monitoring done !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            except:
                pass
            try:
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            except:
                pass
        if check_tec == 'WCDMA':
            WCDMA_NODE_LST_ne = []
            for nnn in WCDMA_NODE_LST:
                if nnn not in WCDMA_NODE_LST_ne:
                    WCDMA_NODE_LST_ne.append(nnn.split('@')[0])
            if activity_type in ['KPI MONITORING']:
                Node_cmd_lst = WCDMA_NODE_LST_ne
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait kpi monitoring ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_monitoring_process(self, Login_comm_lst, Node_cmd_lst, WCDMA_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd)
                filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'KPI Monitoring ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                EMAIL_INITIAL_KPI(self, check_tec, INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : KPI Monitoring done !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
        if check_tec == 'LTE' and activity_type in ['KPI MONITORING']:
            Node_cmd_lst = LTE_NODE_LST
            router = HOSTIP
            username = os.getlogin().lower()
            password = 'OK'
            Save_logfile = 'YES'
            MOTASK = 'MOBATCH'
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait kpi monitoring ongoing !\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            RRSG_monitoring_process(self, Login_comm_lst, Node_cmd_lst, LTE_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd)
            filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'KPI Monitoring ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
            initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
            EMAIL_INITIAL_KPI(self, check_tec, INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI')
            try:
                eyye = RRSG_auto_login.remote_conn
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : KPI Monitoring done !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            except:
                pass
            try:
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            except:
                pass
        if check_tec == 'NR' and activity_type in ['KPI MONITORING']:
            Node_cmd_lst = NR_NODE_LST
            router = HOSTIP
            username = os.getlogin().lower()
            password = 'OK'
            Save_logfile = 'YES'
            MOTASK = 'MOBATCH'
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait kpi monitoring ongoing !\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            RRSG_monitoring_process(self, Login_comm_lst, Node_cmd_lst, NR_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd)
            filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'KPI Monitoring ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
            initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
            EMAIL_INITIAL_KPI(self, check_tec, INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI')
            try:
                eyye = RRSG_auto_login.remote_conn
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : KPI Monitoring done !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            except:
                pass
            try:
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            except:
                continue
    if len(FREQUENCY_LST) > 1 and FREQUENCY_LST[0] == 'Not required':
        pass  # postinserted
    try:
        chk_rssg = RRSG_auto_login.remote_conn
        chk_rssg_chk = 'OK'
    except:
        chk_rssg_chk = 'NOK'
    if chk_rssg_chk == 'OK':
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Done! data has been successfully fetched. Feel free to ask for more assistance or information.') + '\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    if len(FREQUENCY_LST) > 1 and FREQUENCY_LST[0] == 'Not required':
        GSM_LST.clear()
        LTE_LST.clear()
        WCDMA_LST.clear()
        NR_LST.clear()
        GSM_NODE_LST.clear()
        LTE_NODE_LST.clear()
        WCDMA_NODE_LST.clear()
        NR_NODE_LST.clear()
    if 'cancle' not in processdone:
        try:
            ISF_STEP_CLOSE(ISF_API.API_Based, ISF_STEP_START_Task_Id_lst[0], ISF_API.woid, ISF_API.headers)
        except:
            pass
        try:
            ISF_WO_CLOSE(ISF_API.API_Based, ISF_API.headers, ISF_API.woid)
            try:
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + str(activity_type) + str(' WOID') + '__' + str(ISF_API.woid) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            except:
                pass
        except:
            pass
    if len(FREQUENCY_LST) > 1 and FREQUENCY_LST[0] == 'Not required':
        try:
            loggs_up_audit()
        except:
            pass
    LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
    if LIVECMD_TEXT.strip().endswith('>'):
        RRSG_enter_command(self, ['quit'], 'chatbuddy')
    clear_all_lists()
    thread = threading.Thread(target=enabled_button, args=(self,))
    thread.start()
    thread.join()
    showMessage_qt('KPI Monitoring completed successfully.', 5000)
    selected_checkboxes.clear()

def RRSG_PMX_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, techeccci, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmd, RBS_RNC):
    def melted_conter(df):
        try:
            DF_config = pd.read_excel(INPATH + 'PM_config.xlsx', sheet_name='WCDMA_Cell_Map', usecols='A,B,C', engine='openpyxl')
            DF_config = DF_config.drop_duplicates(subset=['3G Carrier', 'Sector Mapping'], keep='first')
            DF_config = DF_config.rename(columns={'3G Carrier': 'cellname'})
            DF_sitemap = pd.read_excel(INPATH + 'PM_config.xlsx', sheet_name='SITE_dB', usecols='A,B,E', engine='openpyxl')
            DF_sitemap = DF_sitemap.drop_duplicates(subset=['RBS Node Id', 'RBS Required Node Id', 'cellname'], keep='first')
            DF_sitemap['RBS Required Node Id'] = DF_sitemap['RBS Required Node Id'].astype(str)
            DF_sitemap['RBS Node Id'] = DF_sitemap['RBS Node Id'].astype(str)
            DF_sitemap['cellname'] = DF_sitemap['cellname'].astype(str)
            NEW_SITE_LIST = DF_sitemap['RBS Required Node Id'].unique().tolist()
            UCell_LIST = DF_sitemap['cellname'].unique().tolist()
            CHECKadd_DF_CELLLIST_f2 = UCell_LIST + NEW_SITE_LIST
            try:
                CHECKadd_DF_CELLLIST_f2 = [x for x in CHECKadd_DF_CELLLIST_f2 if not pd.isna(x)]
            except:
                pass
            del DF_sitemap['cellname']
            DF_sitemap = DF_sitemap.drop_duplicates(subset=['RBS Node Id', 'RBS Required Node Id'], keep='first')
            DF_sitemap = DF_sitemap.rename(columns={'RBS Node Id': 'Node Id'})
            df = df.merge(DF_sitemap, on='Node Id', how='left')
            df = df.drop_duplicates(subset=df.columns.tolist(), keep='first')
            del df['Node Id']
            df = df.rename(columns={'RBS Required Node Id': 'Node Id'})
            df.insert(0, 'Node Id', df.pop('Node Id'))
            try:
                del DF_sitemap
            except:
                pass
            try:
                SEPlst = DF_config['Separator used'].tolist()
                SEPlst_USED = SEPlst[0]
            except:
                SEPlst_USED = 'NO'
            try:
                df = df.applymap(lambda x: str(x).strip())
            except:
                pass
            df['Object'] = df['Object'].str.strip()

            def extract_relevant_part(cell_value):
                if 'EUtranCellFDD=' in cell_value:
                    return cell_value.split('EUtranCellFDD=')[(-1)].split(',')[0]
                if 'NodeBLocalCell=' in cell_value:
                    return cell_value.split('NodeBLocalCell=')[(-1)].split(',')[0]
                if 'SectorCarrier=' in cell_value:
                    return cell_value.split('NodeBSectorCarrier=')[(-1)].split(',')[0]
                if 'SectorCarrier=' in cell_value:
                    return cell_value.split('SectorCarrier=')[(-1)].split(',')[0]
                return 'del'
            df['cell'] = df['Object'].apply(extract_relevant_part)

            def postchk(row):
                if 'Iub=' in row['Object']:
                    return row['Node Id']
                return row['cell']
            df['cellname'] = df.apply(lambda row: postchk(row), axis=1)
            del df['cell']

            def postchk33(row):
                if 'ENodeBFunction=1' in row['Object']:
                    return row['Node Id']
                if 'FieldReplaceableUnit=1,BbProcessingResource=1' in row['Object']:
                    return row['Node Id']
                if 'NodeBLocalCellGroup=1' in row['Object']:
                    return row['Node Id']
                return row['cellname']
            df['cellname23'] = df.apply(lambda row: postchk33(row), axis=1)
            condition = df['cellname'] == 'del'
            df.loc[condition, 'cellname'] = df.loc[condition, 'cellname23']
            del df['cellname23']
            new_df = df[df['Object'].str.contains('Iub=')]
            df = df[~df['Object'].str.contains('Iub=')]
            df = df.loc[df['cellname']!= 'del']
            time_column2 = df.columns[3]
            new_df['NEW_TEST'] = new_df['Node Id'].astype(str) + '>' + new_df['Object'].astype(str) + '>' + new_df['Counter'].astype(str) + '>' + new_df[time_column2].astype(str) + '>' + new_df['cellname'].astype(str)
            del new_df['Node Id']
            del new_df['Object']
            del new_df['Counter']
            del new_df[time_column2]
            del new_df['cellname']
            new_df = new_df.drop_duplicates(subset=['NEW_TEST'], keep='first')
            new_df_lst = new_df['NEW_TEST'].tolist()
            forloop_lst = DF_config['Sector Mapping'].tolist()
            try:
                del new_df
            except:
                pass
            Final_lst = []
            for ddd in forloop_lst:
                for tt in new_df_lst:
                    if SEPlst_USED == 'NO':
                        Final_lst.append(tt + str(ddd))
                    else:  # inserted
                        Final_lst.append(tt + str(SEPlst_USED) + str(ddd))
            New_dffinal = pd.DataFrame()
            if Final_lst:
                New_dffinal['final_col'] = Final_lst
                try:
                    Final_lst.clear()
                    del Final_lst
                except:
                    pass
                New_dffinal = New_dffinal.join(New_dffinal['final_col'].str.split('>', expand=True).add_prefix('final_col'))
                del New_dffinal['final_col']
                New_dffinal = New_dffinal.rename(columns={'final_col0': 'Node Id'})
                New_dffinal = New_dffinal.rename(columns={'final_col1': 'Object'})
                New_dffinal = New_dffinal.rename(columns={'final_col2': 'Counter'})
                New_dffinal = New_dffinal.rename(columns={'final_col3': time_column2})
                New_dffinal = New_dffinal.rename(columns={'final_col4': 'cellname'})
                New_dffinal = New_dffinal.drop_duplicates(subset=['Node Id', 'Object', 'Counter', time_column2, 'cellname'], keep='first')
                df = pd.concat([df, New_dffinal], ignore_index=True)
                try:
                    del New_dffinal
                except:
                    pass
            df['cellname'] = df['cellname'].str.strip()
            df['Counter'] = df['Counter'].str.strip()
            df['Object'] = df['Object'].str.strip()
            df = df.merge(DF_config, on='cellname', how='left')
            try:
                del DF_config
            except:
                pass

            def generate_required_column(row):
                if pd.notna(row['Sector Mapping']) and row['Sector Mapping']!= '':
                    if row['Separator used'] == 'NO':
                        return f"{row['Node Id']}{row['Sector Mapping']}"
                    return f"{row['Node Id']}{SEPlst_USED}{row['Sector Mapping']}"
                return row['cellname']
            df['cellname1'] = df.apply(generate_required_column, axis=1)
            del df['cellname']
            del df['Sector Mapping']
            del df['Object']
            df = df.rename(columns={'cellname1': 'cellname'})
            df = df.drop_duplicates(subset=['Node Id', 'Counter', 'cellname'], keep='first')
            time_column = df.columns[2]
            pivot_df = df.pivot(index=['Node Id', 'cellname'], columns='Counter', values=time_column).reset_index()
            try:
                del df
            except:
                pass
            pivot_df.columns.name = None
            pivot_df = pivot_df.rename_axis(None, axis=1)
            pivot_df['Time'] = time_column
            try:
                pivot_df = pivot_df[pivot_df['cellname'].isin(CHECKadd_DF_CELLLIST_f2)]
                try:
                    CHECKadd_DF_CELLLIST_f2.clear()
                    del CHECKadd_DF_CELLLIST_f2
                except:
                    pass
            except:
                pass
            try:
                pivot_df = pivot_df.dropna(axis=1, how='all')
                return pivot_df
            except:
                pass
                return pivot_df
        except:
            return df

    def melted_conter_WCDMA(df):
        try:
            try:
                df['Node Id'] = df['Node Id'].apply(lambda x: x.split('@')[1] if '@' in x else x)
            except:
                pass
            DF_sitemap = pd.read_excel(INPATH + 'PM_config.xlsx', sheet_name='SITE_dB', usecols='C,D', engine='openpyxl')
            DF_sitemap = DF_sitemap.drop_duplicates(subset=['RNC Node Id', 'RNC Required Node Id'], keep='first')
            DF_sitemap['RBS Required Node Id'] = DF_sitemap['RNC Node Id'].astype(str)
            DF_sitemap['RBS Node Id'] = DF_sitemap['RNC Required Node Id'].astype(str)
            DF_sitemap = DF_sitemap.rename(columns={'RNC Node Id': 'Node Id'})
            df = df.merge(DF_sitemap, on='Node Id', how='left')
            df = df.drop_duplicates(subset=df.columns.tolist(), keep='first')
            del df['Node Id']
            df = df.rename(columns={'RNC Required Node Id': 'Node Id'})
            df.insert(0, 'Node Id', df.pop('Node Id'))
            try:
                del DF_sitemap
            except:
                pass
            try:
                df = df.applymap(lambda x: str(x).strip())
            except:
                pass
            df['Object'] = df['Object'].str.strip()

            def extract_relevant_part(cell_value):
                if 'UtranCell=' in cell_value:
                    return cell_value.split('UtranCell=')[(-1)].split(',')[0]
                if 'NodeBLocalCell=' in cell_value:
                    return cell_value.split('NodeBLocalCell=')[(-1)].split(',')[0]
                if 'SectorCarrier=' in cell_value:
                    return cell_value.split('SectorCarrier=')[(-1)].split(',')[0]
                return 'del'
            df['cell'] = df['Object'].apply(extract_relevant_part)

            def postchk(row):
                if 'Iub=' in row['Object']:
                    return row['Node Id']
                return row['cell']
            df['cellname'] = df.apply(lambda row: postchk(row), axis=1)
            del df['cell']

            def postchk33(row):
                if 'IubLink=' in row['Object']:
                    return row['Node Id']
                return row['cellname']
            df['cellname23'] = df.apply(lambda row: postchk33(row), axis=1)
            condition = df['cellname'] == 'del'
            df.loc[condition, 'cellname'] = df.loc[condition, 'cellname23']
            del df['cellname23']
            new_df = df[df['Object'].str.contains('IubLink=')]
            df = df[~df['Object'].str.contains('IubLink=')]
            df = df.loc[df['cellname']!= 'del']
            time_column2 = df.columns[3]
            new_df['NEW_TEST'] = new_df['Node Id'].astype(str) + '>' + new_df['Object'].astype(str) + '>' + new_df['Counter'].astype(str) + '>' + new_df[time_column2].astype(str) + '>' + new_df['cellname'].astype(str)
            del new_df['Node Id']
            del new_df['Object']
            del new_df['Counter']
            del new_df[time_column2]
            del new_df['cellname']
            new_df = new_df.drop_duplicates(subset=['NEW_TEST'], keep='first')
            new_df = new_df.join(new_df['NEW_TEST'].str.split('>', expand=True).add_prefix('final_col'))
            del new_df['NEW_TEST']
            new_df = new_df.rename(columns={'final_col0': 'Node Id'})
            new_df = new_df.rename(columns={'final_col1': 'Object'})
            new_df = new_df.rename(columns={'final_col2': 'Counter'})
            new_df = new_df.rename(columns={'final_col3': time_column2})
            new_df = new_df.rename(columns={'final_col4': 'cellname'})
            new_df = new_df.drop_duplicates(subset=['Node Id', 'Object', 'Counter', time_column2, 'cellname'], keep='first')
            df = pd.concat([df, new_df], ignore_index=True)
            try:
                del new_df
            except:
                pass
            df['cellname'] = df['cellname'].str.strip()
            df['Counter'] = df['Counter'].str.strip()
            df['Object'] = df['Object'].str.strip()
            del df['Object']
            df = df.drop_duplicates(subset=['Node Id', 'Counter', 'cellname'], keep='first')
            time_column = df.columns[2]
            pivot_df = df.pivot(index=['Node Id', 'cellname'], columns='Counter', values=time_column).reset_index()
            pivot_df.columns.name = None
            pivot_df = pivot_df.rename_axis(None, axis=1)
            pivot_df['Time'] = time_column
            try:
                nodellst = pivot_df['Node Id'].unique()
                nodellst = list(nodellst)
                nodellst = [nodellst for nodellst in nodellst if str(nodellst)!= 'nan']
            except:
                nodellst = []
            try:
                CHECKadd_DF_CELLLIST_f = []
                pivot_df_del = pivot_df.copy()
                pivot_df_del['pmAnrRelationAdd'].replace('', np.nan, inplace=True)
                pivot_df_del.dropna(subset=['pmAnrRelationAdd'], inplace=True)
                celllllst = pivot_df_del['cellname'].unique()
                celllllst = list(celllllst)
                celllllst = [celllllst for celllllst in celllllst if str(celllllst)!= 'nan']
                CHECKadd_DF_CELLLIST_f2 = CHECKadd_DF_CELLLIST_f + celllllst + nodellst
            except:
                CHECKadd_DF_CELLLIST_f2 = []
            try:
                del pivot_df_del
            except:
                pass
            try:
                pivot_df = pivot_df[pivot_df['cellname'].isin(CHECKadd_DF_CELLLIST_f2)]
            except:
                pass
            try:
                pivot_df = pivot_df.dropna(axis=1, how='all')
            except:
                pass
            try:
                CHECKadd_DF_CELLLIST_f2.clear()
            except:
                pass
            try:
                celllllst.clear()
            except:
                pass
            try:
                nodellst.clear()
                return pivot_df
            except:
                pass
                return pivot_df
        except:
            return df
    if 'login_error' in LOGINERROR_LST:
        return
    check_lst = []
    CIPRILST = []

    def read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK):
        commnd_sh = []
        commnd_sh.append(cmd_ex)

        def remove_color_codes(input_string):
            ansi_escape = re.compile('\\x1b[^m]*m')
            clean_string = ansi_escape.sub('', input_string)
            html_color = re.compile('<[^>]*>')
            clean_string = html_color.sub('', clean_string)
            return clean_string
        if MOTASK == 'MOBATCH':
            read_received_data_auto.out = ''
            output_B = b''
            combined_cmd = ''
            command_not_found = []
            while not (output_B.endswith(b'> ') or output_B.endswith(b'<') or output_B.endswith(b': ') or output_B.endswith(b'<') or output_B.endswith(b'< ') or output_B.endswith(b':') or output_B.endswith(b'$ ')):
                output_B += remote_conn.recv(99999)
                write_t = output_B.decode(encoding='utf-8')
                write_te_1 = remove_color_codes(write_t)
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('\a', '')
                self.update_textbox(f'{write_te_1}\n')
                if cmd_ex == 'sdir' and 'CPRI links' in write_te_1 and (not CIPRILST):
                    CIPRILST.append(write_te_1)
                try:
                    self.reset_timer(None)
                except:
                    pass
                if 'no such command' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'command not found' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                write_te_1 = '\n'.join([line.replace('coli>/rc/nrat/ue print -admitted', '') for line in write_te_1.split('\n')])
                write_te_1 = '\n'.join([line.replace('NRAT is dormant.', '') for line in write_te_1.split('\n')])
                if 'ue print -admitted'.lower() in command.lower():
                    write_te_1 = '\n'.join([line.replace('coli>', '') for line in write_te_1.split('\n')])
                if 'coli>' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'amos' in command and 'AMOS error' in write_te_1:
                    command_not_found.append('coli>')
                last_line = output_B.decode(encoding='utf-8').split('\n')[(-1)]
                last_line = ' '.join(last_line.split()).strip()
                if check_lst and last_line == check_lst[0]:
                    read_received_data_auto.brek = 'break'
                try:
                    if last_line[(-1)].strip() == '>':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '<':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == ':':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '$':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
            if not command_not_found:
                write_text1 = output_B.decode(encoding='utf-8')
                write_text = remove_color_codes(write_text1)
                write_text = write_text.replace('', '')
                write_text = write_text.replace('', '')
                write_text = write_text.replace('\a', '')
                if Save_logfile == 'YES':
                    if chk_oss!= 'LOGIN_CNM' and chk_oss!= 'no':
                        if techeccci == 'LTE':
                            OUTFOLDERR = os.path.join(OUTFOLDERR_LTE[0]) + '\\'
                        if techeccci == 'WCDMA':
                            OUTFOLDERR = os.path.join(OUTFOLDERR_WCDMA[0]) + '\\'
                        with open(OUTFOLDERR + 'Ropwise_log ' + datetime.now().strftime('%d%m%Y') + '.log', 'a+', encoding='utf-8') as f:
                            f.writelines(f'{write_text}\n\n')
                        with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str(activity_type) + '__' + str(cmd_node) + '__' + str(command) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    if chk_oss == 'OSS_CNM':
                        combined_cmd += f'{write_text}\n'
                crate_txt_df.df_fram = ''
                if chk_oss == 'OSS_CNM':
                    if command!= 'q' and command!= 'quit':
                        def remove_line_containing(text, substring):
                            lines = text.split('\n')
                            filtered_lines = [line for line in lines if substring not in line]
                            return '\n'.join(filtered_lines)
                        combined_cmd = combined_cmd.replace('It is recommended to remove duplicate counter definitions from the PM scanners.', '')
                        combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                        combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo)', 'Date & Time (UTC)   S Specific Problem                    MO (Cause/AdditionalInfo)')
                        combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo) Operator', '')
                        combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                        combined_cmd = remove_line_containing(combined_cmd, 'ACKNOWLEDGED ALARMS')
                        utc_time_pattern_date = 'Start Time: (\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2}) End Time: (\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2})'
                        utc_time_match = re.search(utc_time_pattern_date, combined_cmd)
                        if utc_time_match:
                            utc_time = utc_time_match.group()
                            time_pattern = '\\d{2}:\\d{2}'
                            datat_pattern = '\\d{4}-\\d{2}-\\d{2}'
                            utc_times = re.findall(time_pattern, str(utc_time))
                            date = utc_time_match.group()
                            date_pat = re.findall(datat_pattern, str(date))
                            f_dates = date_pat[0]
                            if PMAX_DATE:
                                break
                            PMAX_DATE.append(f_dates)
                            if PMX_ROP_TIME:
                                break
                            PMX_ROP_TIME.append(utc_times[0])
                            PMX_ROP_TIME.append(utc_times[(-1)])
                        else:  # inserted
                            if PMAX_DATE:
                                break
                            try:
                                date_pattern = '\\d{4}-\\d{2}-\\d{2}'
                                dates = re.findall(date_pattern, combined_cmd)
                                f_dates = dates[(-1)]
                                f_dates = datetime.strptime(f_dates, '%Y-%m-%d').strftime('%m/%d/%Y')
                                PMAX_DATE.append(f_dates)
                            except:
                                f_dates = datetime.today().strftime('%Y-%m-%d')
                                f_dates = datetime.strptime(f_dates, '%Y-%m-%d').strftime('%m/%d/%Y')
                                PMAX_DATE.append(f_dates)
                            if PMX_ROP_TIME:
                                break
                            try:
                                utc_time_pattern = '\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2} UTC to \\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2} UTC'
                                utc_time_match = re.search(utc_time_pattern, combined_cmd)
                                if utc_time_match:
                                    utc_time = utc_time_match.group()
                                    time_pattern = '\\d{2}:\\d{2}'
                                    utc_times = re.findall(time_pattern, str(utc_time))
                                    PMX_ROP_TIME.append(utc_times[0])
                                    PMX_ROP_TIME.append(utc_times[(-1)])
                            except:
                                pass
                        try:
                            STAT_TIM = PMX_ROP_TIME[0]
                            lines = combined_cmd.split('\n')
                            for i, line in enumerate(lines):
                                if 'Object' in line and 'Counter' in line:
                                    lines[i] = 'Object                                                                                                Counter                                        ' + str(STAT_TIM)
                            combined_cmd = '\n'.join(lines)
                        except:
                            pass
                        if 'pmr' in combined_cmd.lower():
                            combined_cmd = remove_line_containing(combined_cmd, 'pmr')
                            combined_cmd = remove_line_containing(combined_cmd, 'PMR')
                        if 'pmx' in combined_cmd.lower():
                            combined_cmd = remove_line_containing(combined_cmd, 'pmx')
                            combined_cmd = remove_line_containing(combined_cmd, 'PMX')
                        combined_cmd = combined_cmd.strip()
                        START_TIME_CHECK = []
                        END_TIME_CHECK = []
                        pattern = 'Start Time: (\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2}) End Time: (\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2})'
                        match = re.search(pattern, combined_cmd)
                        if match:
                            line_with_times = match.group(0)
                            pattern = 'Start Time: \\d{4}-\\d{2}-\\d{2} (\\d{2}):(\\d{2}):(\\d{2}) End Time: \\d{4}-\\d{2}-\\d{2} (\\d{2}):(\\d{2}):(\\d{2})'
                            match = re.search(pattern, line_with_times)
                            if match:
                                start_hour, start_minute = (match.group(1), match.group(2))
                                end_hour, end_minute = (match.group(4), match.group(5))
                                Start_Time = '{}:{}'.format(start_hour, start_minute)
                                End_Time = '{}:{}'.format(end_hour, end_minute)
                                START_TIME_CHECK.append(Start_Time)
                                END_TIME_CHECK.append(End_Time)
                        if '16Qam'.lower() not in command.lower():
                            combined_cmd = log_clening(combined_cmd, cmd_ex, cmd_node)
                        Final_log_lst = final_log_process(combined_cmd)
                        try:
                            combined_cmd = ''
                        except:
                            pass
                        Final_log_lst = [item.strip() for item in Final_log_lst if item.strip()]
                        try:
                            if Final_log_lst[0] == '\n':
                                Final_log_lst = []
                        except:
                            pass
                        if Final_log_lst:
                            i = 1
                            Final_log_lst = [item for item in Final_log_lst if not item.startswith('+')]
                            Final_log_lst = [item for item in Final_log_lst if not item.startswith('\n+')]
                            ALL_DF = []
                            for section in Final_log_lst:
                                lines = section.strip().split('\n')
                                data = [line.split('\t') for line in lines]
                                if len(data) > 1:
                                    crate_df = crate_txt_df(section, cmd_node, cmd_ex)
                                    try:
                                        crate_df.replace('deletestring', '', inplace=True)
                                    except:
                                        pass
                                    if crate_df is not None and (not crate_df.empty):
                                        crate_df = crate_df[~crate_df.apply(lambda row: any(row.astype(str).str.contains('Total:')), axis=1)]
                                        crate_df.columns = crate_df.columns.str.strip()
                                        try:
                                            crate_df = crate_df.applymap(lambda x: x.replace('_', ' '))
                                        except:
                                            pass
                                        try:
                                            crate_df['Node Id'] = cmd_node
                                        except:
                                            pass
                                        try:
                                            crate_df.insert(0, 'Node Id', crate_df.pop('Node Id'))
                                        except:
                                            pass
                                        counter = [col for col in crate_df.columns if 'counter' in col.lower()]
                                        try:
                                            columns_to_check = crate_df.columns.difference(['Node Id', 'Object', 'Time'])
                                            filtered_df = crate_df.loc[:, columns_to_check].dropna(how='all')
                                            crate_df = pd.concat([crate_df.loc[filtered_df.index, ['Node Id', 'Object', 'Time']], filtered_df], axis=1)
                                        except:
                                            pass
                                        if techeccci == 'LTE' and counter:
                                            crate_df = melted_conter(crate_df)
                                        if techeccci == 'WCDMA' and counter:
                                            crate_df = melted_conter_WCDMA(crate_df)
                                        try:
                                            del crate_df['Command']
                                        except:
                                            pass
                                        try:
                                            del crate_df['Baseline Check']
                                        except:
                                            pass
                                        try:
                                            if 'Time' not in crate_df.columns:
                                                crate_df['Time'] = str(START_TIME_CHECK)
                                                crate_df.insert(1, 'Time', crate_df.pop('Time'))
                                        except:
                                            pass
                                        if not crate_df.empty:
                                            Save_logfile = 'YES'
                                            if Save_logfile == 'YES' and chk_oss!= 'LOGIN_CNM' and (chk_oss!= 'no'):
                                                try:
                                                    del crate_df['Command']
                                                except:
                                                    pass
                                                try:
                                                    try:
                                                        crate_df = crate_df.applymap(lambda x: str(x).strip())
                                                    except:
                                                        pass
                                                    ALL_DF.append(crate_df)
                                                except:
                                                    continue
                            try:
                                combined_df = pd.concat(ALL_DF, ignore_index=True)
                                try:
                                    ALL_DF.clear()
                                except:
                                    pass
                                combined_df.replace('', pd.NA, inplace=True)
                                try:
                                    required_output_df = combined_df.groupby(['Node Id', 'Object'], as_index=False).first()
                                except:
                                    try:
                                        required_output_df = combined_df.groupby(['Node Id', 'Time', 'Object'], as_index=False).first()
                                    except:
                                        try:
                                            required_output_df = combined_df.groupby(['Node Id', 'Time', 'cellname'], as_index=False).first()
                                        except:
                                            required_output_df = combined_df.groupby(['Node Id', 'cellname'], as_index=False).first()
                                required_output_df.replace('nan', '', inplace=True)
                                try:
                                    required_output_df['Date'] = str(f_dates)
                                    required_output_df.insert(0, 'Date', required_output_df.pop('Date'))
                                except:
                                    pass
                                try:
                                    if PMX_ROP_TIME:
                                        required_output_df['Time'] = str(PMX_ROP_TIME[0])
                                        required_output_df.insert(1, 'Time', required_output_df.pop('Time'))
                                        required_output_df['Rop Time'] = str(PMX_ROP_TIME[0]) + ' to ' + str(PMX_ROP_TIME[(-1)])
                                        required_output_df.insert(2, 'Rop Time', required_output_df.pop('Rop Time'))
                                        df_cleanedtime = [entry.split(':')[0] for entry in required_output_df['Time']]
                                        required_output_df['Hour'] = df_cleanedtime
                                        required_output_df['Hour'] = required_output_df['Hour'].astype(str)
                                        required_output_df.insert(3, 'Hour', required_output_df.pop('Hour'))
                                except:
                                    pass

                                def split_columns(df):
                                    new_df = df.copy()
                                    for col in df.columns:
                                        if df[col].dtype == object and df[col].str.contains(',').any():
                                            split_col = df[col].str.split(',', expand=True)
                                            split_col.columns = [f'{col}${i}' for i in range(split_col.shape[1])]
                                            new_df = pd.concat([new_df, split_col], axis=1)
                                    return new_df
                                required_output_df = split_columns(required_output_df)
                                ROP_WISE_REPORT.append(required_output_df)
                                try:
                                    del required_output_df
                                except:
                                    pass
                            except:
                                pass
                            try:
                                del Final_log_lst
                            except:
                                pass
                            try:
                                del START_TIME_CHECK
                            except:
                                pass
                            try:
                                del END_TIME_CHECK
                            except:
                                pass
                            try:
                                del command_not_found
                            except:
                                pass
                            combined_cmd = ''

    def send_command(remote_conn, command, chk_oss, iii, cmd_node, Save_logfile, cmd_ex, MOTASK):
        read_received_data_auto.fil = 'None'
        remote_conn.send(command + '\n')
        read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK)
        send_command.fil = read_received_data_auto.fil

    def main(logcommand_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK):
        iii = 1
        paaaawe = None
        try:
            llll = RRSG_auto_login.remote_conn
            paaaawe = 'ok'
        except:
            try:
                con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                del Rdf['indexxx']
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                PID_LST = Rdf['value'].tolist()
                remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
            except:
                RSG_LST_N = []
            if not RSG_LST_N:
                showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 10000)
            else:  # inserted
                paaaawe = OTPPP_tets()
            if paaaawe == None and 'cancle' not in processdone:
                processdone.append('cancle')
            try:
                if VPN_select_METHOD_F[0]!= 'VPN-CAS':
                    if is_non_numeric_string(paaaawe):
                        showMessage_qt('Wrong OTP password please try again.', 10000)
                    else:  # inserted
                        if paaaawe:
                            self.Auto_login_sc(paaaawe, RSG_LST_N)
                else:  # inserted
                    if paaaawe:
                        self.Auto_login_sc(paaaawe, RSG_LST_N)
            except:
                pass
        try:
            chk_rssg = RRSG_auto_login.remote_conn
            chk_rssg_chk = 'OK'
        except:
            chk_rssg_chk = 'NOK'
        if chk_rssg_chk == 'NOK':
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Something wrong in RSG/VPN connection please try again after some time.') + '\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        else:  # inserted
            if not CHECK_ISF_SEQUENCE:
                sig = os.getlogin().lower()
                priority = 'High'
                ISF_API_Based = 'SharePoint'
                KEY = '5b97555705ce40128131d58c8d3596f9'
                try:
                    con = sqlite3.connect('./res/ISFdetails.db')
                    Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                    con.close()
                    del Rdf['indexxx']
                    ppPID_LST = Rdf['value'].tolist()
                    ppPID_LST2 = []
                    for ppp in ppPID_LST:
                        ppPID_LST2.append(decrypt(ppp, 3))
                    del Rdf['value']
                    Rdf['value'] = ppPID_LST2
                    PID_LST = Rdf['value'].tolist()
                    remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                    ISF_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                    try:
                        del Rdf
                    except:
                        pass
                except:
                    ISF_LST = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                chk_line = []
                for llll in ISF_LST:
                    if 'line' in llll.lower():
                        chk_line.append('stop_process')
                        break
                    if 'start_process' not in chk_line:
                        chk_line.append('start_process')
                if 'start_process' in chk_line:
                    projectID = ISF_LST[0]
                    type1 = ISF_LST[1]
                    nodname = ISF_LST[2]
                    wOName = ISF_LST[3]
                    executionPlanName = ISF_LST[4]
                    lastModifiedBy = ISF_LST[5]
                    ISF_STEP_START_stepID = ISF_LST[6]
                    ISF_STEP_START_Task_Id = ISF_LST[7]
                    if not ISF_STEP_START_Task_Id_lst:
                        ISF_STEP_START_Task_Id_lst.append(ISF_STEP_START_Task_Id)
                    ISF_WO_CREATE_woId, STEP_START = ISF_API(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY, ISF_STEP_START_stepID, ISF_STEP_START_Task_Id)
                    WO_LSTT = [ISF_WO_CREATE_woId, STEP_START]
                    ISF_STTUS_MSG = 'sucess'
                    if WO_LSTT[0] == 'Some issue in isf workflow':
                        ISF_STTUS_MSG = 'ISF details not correct. please check and try again.'
                    if ISF_STTUS_MSG == 'sucess' and WO_LSTT[1]!= 'sucess':
                        ISF_STTUS_MSG = 'TaskID and StepID not correct. please provide the valid taskID for the StepID.'
            else:  # inserted
                ISF_STTUS_MSG = 'sucess'
            if ISF_STTUS_MSG == 'sucess' and (not CHECK_ISF_SEQUENCE):
                if 'work_order_raised' not in CHECK_ISF_SEQUENCE:
                    CHECK_ISF_SEQUENCE.append('work_order_raised')

                def sucss_msg(self):
                    self.textbox2.delete(1.0, 'end')
                    self.after(0, self.textbox2.insert, 'end', 'ISF Work Order created.\n\n')
                    self.after(0, self.textbox2.insert, 'end', 'Signum    : ' + str(sig) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'Nodtype   : ' + str(type1) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'Nodname   : ' + str(nodname) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'ProjectID : ' + str(projectID) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'WOID      : ' + str(ISF_WO_CREATE_woId) + '\n')
                    try:
                        self.after(0, self.textbox2.insert, 'end', 'Time      : ' + str(ISF_API.timestr) + '\n\n')
                    except:
                        pass
                    try:
                        self.after(0, self.textbox2.insert, 'end', str(ISF_API.urlcreateWO) + '\n\n')
                    except:
                        return None
                statusAUDITLOGPATH = './res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log'
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Signum__' + str(sig) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Nodtype__' + str(type1) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Nodname__' + str(nodname) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'ProjectID__' + str(projectID) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'WOID__' + str(ISF_WO_CREATE_woId) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Time__' + str(ISF_API.timestr) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                threading.Thread(target=sucss_msg, args=(self,)).start()
            if 'work_order_raised' in CHECK_ISF_SEQUENCE and paaaawe!= None:
                send_command.fil = 'None'
                if 'FAIL' not in establish_ssh_conn_lst and techeccci == 'LTE':
                    for cmd_node in Node_cmd_lst:
                        self.textbox_LIVECMD.delete(1.0, 'end')
                        chk_oss = 'NODE_CNM'
                        cmd_ex_sh = ''
                        send_command(RRSG_auto_login.remote_conn, 'amos ' + str(cmd_node), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                        send_command(RRSG_auto_login.remote_conn, 'lt all', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                        RBS_RBS_switch = self.swich_Sec_rbs.get()
                        if RBS_RBS_switch == 1:
                            RBS_RBS_switch_loe = 'YES'
                        else:  # inserted
                            RBS_RBS_switch_loe = 'NO'
                        if RBS_RBS_switch_loe == 'YES':
                            if RBSPASS_CRED_CRREDD[0] == 'rbs':
                                user_rbs = 'rbs'
                            else:  # inserted
                                try:
                                    user_rbs = RBSPASS_CRED_CRREDD[0]
                                except:
                                    user_rbs = 'rbs'
                            if RBSPASS_CRED_CRREDD[1] == 'rbs':
                                pass_rbs = 'rbs'
                            else:  # inserted
                                try:
                                    pass_rbs = RBSPASS_CRED_CRREDD[1]
                                except:
                                    pass_rbs = 'rbs'
                            send_command(RRSG_auto_login.remote_conn, str(user_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            send_command(RRSG_auto_login.remote_conn, str(pass_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                        for cmd_ex in oss_cmd_lst:
                            cmd_ex_sh = clean_sheet_name(cmd_ex, max_length=28)
                            chk_oss = 'OSS_CNM'
                            send_command(RRSG_auto_login.remote_conn, cmd_ex, chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            iii = iii + 1
                        send_command(RRSG_auto_login.remote_conn, 'quit', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                if techeccci == 'WCDMA':
                    for cmd_node in Node_cmd_lst:
                        self.textbox_LIVECMD.delete(1.0, 'end')
                        chk_oss = 'NODE_CNM'
                        cmd_ex_sh = ''
                        send_command(RRSG_auto_login.remote_conn, 'amos ' + str(cmd_node.split('@')[0].strip()), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                        send_command(RRSG_auto_login.remote_conn, 'lt all', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                        RBS_RBS_switch = self.swich_Sec_rbs.get()
                        if RBS_RBS_switch == 1:
                            RBS_RBS_switch_loe = 'YES'
                        else:  # inserted
                            RBS_RBS_switch_loe = 'NO'
                        if RBS_RBS_switch_loe == 'YES':
                            if RBSPASS_CRED_CRREDD[0] == 'rbs':
                                user_rbs = 'rbs'
                            else:  # inserted
                                try:
                                    user_rbs = RBSPASS_CRED_CRREDD[0]
                                except:
                                    user_rbs = 'rbs'
                            if RBSPASS_CRED_CRREDD[1] == 'rbs':
                                pass_rbs = 'rbs'
                            else:  # inserted
                                try:
                                    pass_rbs = RBSPASS_CRED_CRREDD[1]
                                except:
                                    pass_rbs = 'rbs'
                            send_command(RRSG_auto_login.remote_conn, str(user_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            send_command(RRSG_auto_login.remote_conn, str(pass_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                        for cmd_ex in oss_cmd_lst:
                            cmd_ex_sh = clean_sheet_name(cmd_ex, max_length=28)
                            chk_oss = 'OSS_CNM'
                            cmd_ex = cmd_ex.replace('samplenode', str(cmd_node.split('@')[1].strip()))
                            send_command(RRSG_auto_login.remote_conn, cmd_ex, chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            iii = iii + 1
                        send_command(RRSG_auto_login.remote_conn, 'quit', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
    if SQL_METHOD == 'LOCAL':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check(user_signum, command_to_barred_required, command_to_barred)
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)
    if SQL_METHOD == 'AZURE_CLOUDE':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            if change_User:
                ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check_AZURE(change_User, command_barred, user_signum, command_to_barred_required, command_to_barred)
            else:  # inserted
                if 'amos' in command_to_barred.lower():
                    ucercmd_AUTHENTICATION = 'User_Authenticated'
                else:  # inserted
                    if 'rbs' in command_to_barred.lower():
                        ucercmd_AUTHENTICATION = 'User_Authenticated'
                    else:  # inserted
                        if 'lt all' in command_to_barred.lower():
                            ucercmd_AUTHENTICATION = 'User_Authenticated'
                        else:  # inserted
                            ucercmd_AUTHENTICATION = 'NO'
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)

def merge_file_old(DF, folderr):
    try:
        os.makedirs(folderr + 'temp')
    except:
        pass
    try:
        files = os.listdir(folderr)
        for fname in files:
            try:
                shutil.copy2(os.path.join(folderr, fname), folderr + 'temp')
            except:
                pass
    except:
        pass
    folderr = folderr + 'temp\\'
    Target_File = DF['Target File Name'].unique()
    Target_File = list(Target_File)
    for T_File in Target_File:
        try:
            DF222 = DF.copy()
            DF222 = DF222.loc[DF222['Target File Name'] == T_File]
            S_File = DF222['Source File Name'].unique()
            S_File = list(S_File)
            S_File = [S_File for S_File in S_File if str(S_File)!= 'nan']
            if len(S_File) > 1:
                Master_file = S_File[0]
                S_File.remove(S_File[0])
                try:
                    df_reptt = pd.read_csv(folderr + Master_file, index_col=None, header=0, low_memory=False, encoding='utf-8-sig')
                except:
                    new_file = folderr + pathlib.Path(folderr + Master_file).stem + '.txt'
                    shutil.copy(folderr + Master_file, new_file)
                    df_reptt = pd.read_csv(new_file, sep='\t', index_col=None, header=0, low_memory=False, encoding='utf-16')
                    try:
                        os.remove(new_file)
                    except:
                        pass
                DF333 = DF.copy()
                DF333 = DF333.loc[DF333['Source File Name'] == Master_file]
                DF333 = DF333[['Source Header Name', 'Target Header Name']]
                try:
                    DF333 = DF333[DF333['Target Header Name'].isin(['date', 'cellname', 'nodename'])]
                except:
                    DF333 = DF333[DF333['Target Header Name'].isin(['date', 'nodename'])]
                DF333 = DF333.drop_duplicates(subset=['Target Header Name'], keep='first')
                DF333['datat'] = DF333['Source Header Name'].astype(str) + '$' + DF333['Target Header Name'].astype(str)
                S_datat = DF333['datat'].unique()
                S_datat = list(S_datat)
                for hh in S_datat:
                    try:
                        if hh.split('$')[1].strip() == 'date':
                            DATE_COL = hh.split('$')[0].strip()
                    except:
                        DATE_COL = 'notav'
                    try:
                        if hh.split('$')[1].strip() == 'cellname':
                            CELL_COL = hh.split('$')[0].strip()
                    except:
                        CELL_COL = 'notav'
                    try:
                        if hh.split('$')[1].strip() == 'nodename':
                            NODE_COL = hh.split('$')[0].strip()
                    except:
                        NODE_COL = 'notav'
                try:
                    df_reptt['MERGE_COL'] = df_reptt[DATE_COL].astype(str) + '$' + df_reptt[CELL_COL].astype(str)
                except:
                    df_reptt['MERGE_COL'] = df_reptt[DATE_COL].astype(str) + '$' + df_reptt[NODE_COL].astype(str)
                for mergee in S_File:
                    DF333 = DF.copy()
                    DF333 = DF333.loc[DF333['Source File Name'] == mergee]
                    DF333 = DF333[['Source Header Name', 'Target Header Name']]
                    try:
                        DF333 = DF333[DF333['Target Header Name'].isin(['date', 'cellname', 'nodename'])]
                    except:
                        DF333 = DF333[DF333['Target Header Name'].isin(['date', 'nodename'])]
                    DF333 = DF333.drop_duplicates(subset=['Target Header Name'], keep='first')
                    DF333['datat'] = DF333['Source Header Name'].astype(str) + '$' + DF333['Target Header Name'].astype(str)
                    S_datat = DF333['datat'].unique()
                    S_datat = list(S_datat)
                    for hh in S_datat:
                        try:
                            if hh.split('$')[1].strip() == 'date':
                                DATE_COL = hh.split('$')[0].strip()
                        except:
                            DATE_COL = 'notav'
                        try:
                            if hh.split('$')[1].strip() == 'cellname':
                                CELL_COL = hh.split('$')[0].strip()
                        except:
                            CELL_COL = 'notav'
                        try:
                            if hh.split('$')[1].strip() == 'nodename':
                                NODE_COL = hh.split('$')[0].strip()
                        except:
                            NODE_COL = 'notav'
                    try:
                        df_merg = pd.read_csv(folderr + mergee, index_col=None, header=0, low_memory=False, encoding='utf-8-sig')
                        for i in range(1, 10):
                            if len(df_merg.columns[df_merg.columns.str.contains('unnamed', case=False)]) > 0:
                                df_merg = pd.read_csv(folderr + mergee, skiprows=i, index_col=None, header=0, low_memory=False, encoding='utf-8-sig')
                    except:
                        new_file = folderr + pathlib.Path(folderr + mergee).stem + '.txt'
                        shutil.copy(folderr + mergee, new_file)
                        df_merg = pd.read_csv(new_file, sep='\t', index_col=None, header=0, low_memory=False, encoding='utf-16')
                        try:
                            os.remove(new_file)
                        except:
                            pass
                    try:
                        os.remove(folderr + mergee)
                    except:
                        pass
                    try:
                        df_merg['MERGE_COL'] = df_merg[DATE_COL].astype(str) + '$' + df_merg[CELL_COL].astype(str)
                    except:
                        df_merg['MERGE_COL'] = df_merg[DATE_COL].astype(str) + '$' + df_merg[NODE_COL].astype(str)
                    master_col = df_reptt.columns.tolist()
                    merg_coll = df_merg.columns.tolist()
                    for jjkm in merg_coll:
                        if jjkm!= 'MERGE_COL':
                            if jjkm in master_col:
                                del df_merg[jjkm]
                    df_reptt = df_reptt.merge(df_merg, on='MERGE_COL', how='left')
                    data_mee_lst = df_reptt.columns.tolist()
                    df_reptt = df_reptt.drop_duplicates(subset=data_mee_lst, keep='first')
                    del df_merg
                    DF['Source File Name'] = DF['Source File Name'].astype(str).replace(mergee, Master_file, regex=True).replace(['nan', 'None'], np.nan)
                df_llsstt = df_reptt.columns.tolist()
                if CELL_COL in df_llsstt:
                    try:
                        df_reptt = df_reptt.drop_duplicates(subset=[DATE_COL, CELL_COL, NODE_COL], keep='first')
                    except:
                        pass
                else:  # inserted
                    try:
                        df_reptt = df_reptt.drop_duplicates(subset=[DATE_COL, NODE_COL], keep='first')
                    except:
                        pass
                data_mee_lst = df_reptt.columns.tolist()
                df_reptt = df_reptt.drop_duplicates(subset=data_mee_lst, keep='first')
                df_reptt.to_csv(folderr + Master_file, index=None, header=True, encoding='utf-8-sig')
        except:
            continue
    try:
        merge_file_old.dff = DF.copy()
    except:
        return None

def merge_file_fun(files, output_file, date_column, cell_column, input_path, Master_file, DF, hour_COL):
    try:
        result_df = pd.DataFrame()
        for file in files:
            try:
                file_path = os.path.join(input_path, file)
                try:
                    try:
                        df = pd.read_csv(file_path, index_col=None, header=0, low_memory=False, encoding='utf-8-sig')
                    except:
                        df = pd.read_csv(file_path, sep='\t', index_col=None, header=0, low_memory=False, encoding='utf-16')
                except:
                    new_file = input_path + pathlib.Path(input_path + file).stem + '.txt'
                    df = pd.read_csv(new_file, sep='\t', index_col=None, header=0, low_memory=False, encoding='utf-16')
                try:
                    df[hour_COL] = df[hour_COL].astype(str).replace('\\.00', '', regex=True).replace(['nan', 'None'], np.nan)
                except:
                    pass
                try:
                    os.remove(file_path)
                except:
                    pass
                try:
                    df = df[~df[cell_column].str.contains('DUMMY')]
                except:
                    pass
                try:
                    df = df[~df[cell_column].str.contains('dummy')]
                except:
                    pass
                if date_column not in df.columns or cell_column not in df.columns:
                    continue
                if result_df.empty:
                    result_df = df.copy()
                else:  # inserted
                    duplicate_columns = set(result_df.columns) & set(df.columns)
                    duplicate_columns.discard(date_column)
                    duplicate_columns.discard(cell_column)
                    master_col = result_df.columns.tolist()
                    merg_coll = df.columns.tolist()
                    for jjkm in merg_coll:
                        if jjkm!= date_column and jjkm!= cell_column and (jjkm in master_col):
                            if str(hour_COL).lower() == str(jjkm).lower():
                                continue
                            del df[jjkm]
                    data_mee_lst = df.columns.tolist()
                    df = df.drop_duplicates(subset=data_mee_lst, keep='first')
                    if hour_COL not in df.columns:
                        result_df = result_df.merge(df, left_on=[date_column, cell_column], right_on=[date_column, cell_column], how='left')
                    else:  # inserted
                        try:
                            result_df = result_df.merge(df, left_on=[date_column, hour_COL, cell_column], right_on=[date_column, hour_COL, cell_column], how='left')
                        except:
                            result_df = result_df.merge(df, left_on=[date_column, cell_column], right_on=[date_column, cell_column], how='left')
                    data_mee_lst = result_df.columns.tolist()
                    result_df = result_df.drop_duplicates(subset=data_mee_lst, keep='first')
                    try:
                        DF['Source File Name'] = DF['Source File Name'].astype(str).replace(file, Master_file, regex=True).replace(['nan', 'None'], np.nan)
                    except:
                        pass
            except:
                continue
        output_path = os.path.join(input_path, output_file)
        result_df.to_csv(output_path, index=None, header=True, encoding='utf-8-sig')
        try:
            merge_file_fun.dff = DF.copy()
        except:
            return
    except:
        return None

def merge_file(DF, folderr):
    try:
        folderr = folderr + 'temp\\'
        Target_File = DF['Target File Name'].unique()
        Target_File = list(Target_File)
        for T_File in Target_File:
            try:
                DF222 = DF.copy()
                DF222 = DF222.loc[DF222['Target File Name'] == T_File]
                S_File = DF222['Source File Name'].unique()
                S_File = list(S_File)
                S_File = [S_File for S_File in S_File if str(S_File)!= 'nan']
                if len(S_File) > 1:
                    Master_file = S_File[0]
                    output_file_name = S_File[0]
                    date_column_name = 'date'
                    cell_column_name = 'cellname'
                    hour_COL = 'time'
                    merge_file_fun(S_File, output_file_name, date_column_name, cell_column_name, folderr, Master_file, DF, hour_COL)
            except:
                continue
    except:
        return None

def ERCA_FROMAT(self, INPATH_FOLDER, OUTFOLDERR, BOT_COFIG_PATH, technoo, hOURLY_switch):
    try:
        DF = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_PMX_Map_3G_4G', usecols='A,B,C,D,E,F,G', engine='openpyxl')
        DF = DF.loc[DF['Tech'] == technoo]
        DF = DF[['Source File Name', 'Source Header Name', 'Target Header Name', 'Target File Name']]
        try:
            DF = DF[DF['Target Header Name'].notna()]
        except:
            pass
        mapping_df = DF.copy()
        mapping_df['Source File Name'].replace('', np.nan, inplace=True)
        mapping_df['Source Header Name'].replace('', np.nan, inplace=True)
        mapping_df['Source Header Name'] = mapping_df['Source Header Name'].mask(mapping_df['Source Header Name'].isna(), mapping_df['Target Header Name'])
        mapping_df['New_col'] = mapping_df['Source Header Name'].astype(str) + '#' + mapping_df['Target Header Name'].astype(str)
        uNique_sourcefile = mapping_df['Source File Name'].unique()
        uNique_sourcefile = list(uNique_sourcefile)
        uNique_sourcefile = [uNique_sourcefile for uNique_sourcefile in uNique_sourcefile if str(uNique_sourcefile)!= 'nan']
        try:
            os.makedirs(INPATH_FOLDER + 'temp')
        except:
            pass
        dateee = []
        HOUR_LST = []
        for source_file in uNique_sourcefile:
            mapping_df_v2 = mapping_df.copy()
            mapping_df_v2 = mapping_df_v2.loc[mapping_df_v2['Source File Name'] == source_file]
            uNique_header = mapping_df_v2['New_col'].unique()
            uNique_header = list(uNique_header)
            uNique_header2 = mapping_df_v2['Target Header Name'].unique()
            uNique_header2 = list(uNique_header2)
            if os.path.exists(INPATH_FOLDER + source_file):
                try:
                    df = pd.read_csv(INPATH_FOLDER + source_file.split('>')[0].strip(), index_col=None, header=0, low_memory=False, encoding='utf-8-sig')
                except:
                    new_file = INPATH_FOLDER + pathlib.Path(INPATH_FOLDER + source_file.split('>')[0].strip()).stem + '.txt'
                    shutil.copy(INPATH_FOLDER + source_file, new_file)
                    df = pd.read_csv(new_file, sep='\t', index_col=None, header=0, low_memory=False, encoding='utf-16')
                    try:
                        os.remove(new_file)
                    except:
                        pass
                s_list = df.columns.tolist()
                column_mapping = {}
                for item in uNique_header:
                    original, new = item.split('#')
                    column_mapping[original] = new
                for i, col_name in enumerate(s_list):
                    if col_name in column_mapping:
                        s_list[i] = column_mapping[col_name]
                df.rename(columns=dict(zip(df.columns, s_list)), inplace=True)
                for hedd in uNique_header2:
                    if hedd not in df.columns:
                        df[hedd] = np.nan
                try:
                    if df['cellname'].isnull().values.any() or (df['cellname'] == '').any():
                        df['cellname'] = df['nodename']
                except:
                    pass
                try:
                    if df['nodename'].isnull().values.any() or (df['nodename'] == '').any():
                        df['nodename'] = df['cellname']
                except:
                    pass
                try:
                    if 'cellname' not in df.columns:
                        df['cellname'] = df['nodename']
                except:
                    pass
                try:
                    if 'nodename' not in df.columns:
                        df['nodename'] = df['cellname']
                except:
                    pass
                df['date'] = pd.to_datetime(df['date'])
                df['date'] = pd.to_datetime(df['date'], format='%m/%d/%y').dt.strftime('%m/%d/%Y')
                so_datatt = df['date'].unique()
                try:
                    so_datatt = list(so_datatt)
                except:
                    so_datatt = []
                for dd in so_datatt:
                    if dd not in dateee:
                        dateee.append(dd)
                try:
                    so_HOUR = df['time'].unique()
                    so_HOUR = list(so_HOUR)
                except:
                    so_HOUR = []
                for dd in so_HOUR:
                    if dd not in HOUR_LST:
                        HOUR_LST.append(dd)
                try:
                    if technoo == 'GSM':
                        SITE_C = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_2G', usecols='I', engine='openpyxl')
                        Auto_SITE = SITE_C.iloc[0, 0]
                        site_dounter = SITE_C.iloc[1, 0]
                        site_elenemt = int(SITE_C.iloc[2, 0])
                        try:
                            if Auto_SITE == 'YES':
                                df['nodename'] = df[site_dounter].str[:-site_elenemt]
                        except:
                            pass
                    if technoo == 'WCDMA':
                        SITE_C = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_3G', usecols='I', engine='openpyxl')
                        Auto_SITE = SITE_C.iloc[0, 0]
                        site_dounter = SITE_C.iloc[1, 0]
                        site_elenemt = int(SITE_C.iloc[2, 0])
                        try:
                            if Auto_SITE == 'YES':
                                df['nodename'] = df[site_dounter].str[:-site_elenemt]
                        except:
                            pass
                    if technoo == 'LTE':
                        SITE_C = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_4G', usecols='I', engine='openpyxl')
                        Auto_SITE = SITE_C.iloc[0, 0]
                        site_dounter = SITE_C.iloc[1, 0]
                        site_elenemt = int(SITE_C.iloc[2, 0])
                        try:
                            if Auto_SITE == 'YES':
                                df['nodename'] = df[site_dounter].str[:-site_elenemt]
                        except:
                            pass
                    if technoo == '5G':
                        SITE_C = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_5G', usecols='I', engine='openpyxl')
                        Auto_SITE = SITE_C.iloc[0, 0]
                        site_dounter = SITE_C.iloc[1, 0]
                        site_elenemt = int(SITE_C.iloc[2, 0])
                        try:
                            if Auto_SITE == 'YES':
                                df['nodename'] = df[site_dounter].str[:-site_elenemt]
                        except:
                            pass
                except:
                    pass
                df.to_csv(INPATH_FOLDER + 'temp\\' + source_file, index=None, header=True, encoding='utf-8-sig')
                try:
                    del df
                except:
                    continue
        merge_file(DF, INPATH_FOLDER)
        try:
            DF = merge_file_fun.dff.copy()
        except:
            pass
        try:
            INPATH_FOLDER = INPATH_FOLDER + 'temp\\'
        except:
            pass
        date_col = 'date'
        mapping_df_target = DF['Target File Name'].unique()
        mapping_df_target = list(mapping_df_target)
        ma_targetLST = []
        for ttgt in mapping_df_target:
            mapping_df2 = DF.copy()
            mapping_df2 = mapping_df2.loc[mapping_df2['Target File Name'] == ttgt]
            elemment = mapping_df2.iloc[0, 0]
            mapping_df2['Source File Name'] = str(elemment)
            ma_targetLST.append(mapping_df2)
        DF = pd.concat(ma_targetLST, axis=0, ignore_index=True)
        mapping_df = DF.copy()
        mapping_df['comb'] = mapping_df['Source File Name'].astype(str) + '>' + mapping_df['Target File Name'].astype(str)
        filename_lst = mapping_df['comb'].unique()
        filename_lst = list(filename_lst)
        source_filename = []
        target_filename = []
        for fileee in filename_lst:
            try:
                source_filename.append(fileee.split('>')[0].strip())
                target_filename.append(fileee.split('>')[1].strip())
            except:
                continue
        target_filename = [target_filename for target_filename in target_filename if str(target_filename)!= 'nan']
        kk = 0
        for file in filename_lst:
            try:
                filtered_DF = DF.copy()
                filtered_DF = filtered_DF.loc[filtered_DF['Target File Name'] == file.split('>')[1].strip()]
                target_hR1 = filtered_DF['Target Header Name'].tolist()
                target_hR = list(dict.fromkeys(target_hR1))
                try:
                    df_report = pd.read_csv(INPATH_FOLDER + file.split('>')[0].strip(), index_col=None, header=0, low_memory=False, encoding='utf-8-sig')
                    df_report = df_report[df_report[date_col].notna()]
                    for i in range(1, 10):
                        if len(df_report.columns[df_report.columns.str.contains('unnamed', case=False)]) > 0:
                            df_report = pd.read_csv(INPATH_FOLDER + file, skiprows=i, index_col=None, header=0, low_memory=False, encoding='utf-8-sig')
                except:
                    new_file = INPATH_FOLDER + pathlib.Path(INPATH_FOLDER + file.split('>')[0].strip()).stem + '.txt'
                    shutil.copy(INPATH_FOLDER + file, new_file)
                    df_report = pd.read_csv(new_file, sep='\t', index_col=None, header=0, low_memory=False, encoding='utf-16')
                    try:
                        os.remove(new_file)
                    except:
                        pass
                except:
                    df_report = pd.DataFrame(columns=target_hR)
                if 'ON' in switch_PMX_REQ_CHECK:
                    if technoo == 'LTE':
                        try:
                            df_report['pmA3InterFBestCellEvalReport'].replace('', np.nan, inplace=True)
                            df_report.dropna(subset=['pmA3InterFBestCellEvalReport'], inplace=True)
                        except:
                            pass
                    if technoo == 'WCDMA':
                        try:
                            df_report['pmDchFramesReceived'].replace('', np.nan, inplace=True)
                            df_report.dropna(subset=['pmDchFramesReceived'], inplace=True)
                        except:
                            pass
                try:
                    if technoo == 'GSM':
                        add_DF = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_2G', usecols='A,B,C,D,E,F', engine='openpyxl')
                    if technoo == 'WCDMA':
                        add_DF = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_3G', usecols='A,B,C,D,E,F', engine='openpyxl')
                    if technoo == 'LTE':
                        add_DF = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_4G', usecols='A,B,C,D,E,F', engine='openpyxl')
                    if technoo == '5G':
                        add_DF = pd.read_excel(BOT_COFIG_PATH, sheet_name='ERCA_KPI_Map_5G', usecols='A,B,C,D,E,F', engine='openpyxl')
                    add_DF['Additional_Column'].replace('', np.nan, inplace=True)
                    add_DF.dropna(subset=['Additional_Column'], inplace=True)
                    add_DF = add_DF.loc[add_DF['Source File Name'] == file]
                    add_DF['Add_Col'] = add_DF['Target Header Name'].astype(str) + '$' + add_DF['Additional_Column'].astype(str)
                    filen_lst = add_DF['Add_Col'].unique()
                    filen_lst = list(filen_lst)
                    for coll in filen_lst:
                        try:
                            source_col = coll.split('$')[0].strip()
                            add_col = coll.split('$')[1].strip()
                            add_col1 = str(add_col.split('_')[0].strip())
                            add_col2 = str(add_col.split('_')[1].strip())
                            if add_col2 == '1':
                                df_report[source_col] = df_report[add_col1].astype(str) + '_' + str(1)
                            else:  # inserted
                                df_report[source_col] = df_report[add_col1].astype(str) + '_' + df_report[add_col2].astype(str)
                        except:
                            pass
                except:
                    pass
                df_report = df_report[target_hR]
                out_filename = str(target_filename[kk])
                excll = out_filename.split('_')[(-1)].strip()
                out_filename = ' '.join(out_filename.replace(excll, '').split())
                try:
                    cpoll = df_report.columns.tolist()
                    if 'managedElementId' in cpoll:
                        df_report['managedElementId'] = df_report['nodename']
                except:
                    pass
                try:
                    cpoll = df_report.columns.tolist()
                    for dfgw in cpoll:
                        df_report[dfgw].replace('', np.nan, inplace=True)
                        try:
                            if len(df_report[dfgw].unique().tolist()) == 1:
                                if 'ENodeBFunctionId'.lower() in dfgw.lower():
                                    df_report[dfgw].fillna('1', inplace=True)
                                if 'gNBDUFunctionId'.lower() in dfgw.lower():
                                    df_report[dfgw].fillna('1', inplace=True)
                                if 'gNBCUUPFunctionId'.lower() in dfgw.lower():
                                    df_report[dfgw].fillna('1', inplace=True)
                                if 'gNBCUCPFunctionId'.lower() in dfgw.lower():
                                    df_report[dfgw].fillna('1', inplace=True)
                        except:
                            pass
                except:
                    pass
                try:
                    cpoll = df_report.columns.tolist()
                    for dfgw in cpoll:
                        df_report[dfgw].replace('', np.nan, inplace=True)
                        try:
                            if len(df_report[dfgw].unique().tolist()) == 1 and 'time' == dfgw.lower():
                                df_report[dfgw].fillna('0', inplace=True)
                        except:
                            pass
                except:
                    pass
                try:
                    data_mee = pd.read_excel(BOT_COFIG_PATH, sheet_name='CGI_FGI', usecols='A,B,C,D,E', engine='openpyxl')
                    data_mee = data_mee.loc[data_mee['Technology'] == technoo]
                    data_mee_lst = data_mee.columns.tolist()
                    data_mee = data_mee.drop_duplicates(subset=data_mee_lst, keep='first')
                    try:
                        del data_mee['Technology']
                    except:
                        pass
                    cpoll = df_report.columns.tolist()
                    if 'cellname' in cpoll:
                        try:
                            del data_mee['nodename']
                        except:
                            pass
                    else:  # inserted
                        try:
                            del data_mee['cellname']
                        except:
                            pass
                cgic = 'no'
                fgif = 'no'
                if 'cgi' in cpoll:
                    cgic = 'yes'
                if 'fgi' in cpoll:
                    fgif = 'yes'
                if fgif == 'yes':
                    try:
                        del data_mee['cgi']
                    except:
                        pass
                    try:
                        del df_report['fgi']
                    except:
                        pass
                    data_mee_lst = data_mee.columns.tolist()
                    data_mee = data_mee.drop_duplicates(subset=data_mee_lst, keep='first')
                    try:
                        df_report = df_report.merge(data_mee, on='cellname', how='left')
                        data_mee_lst = df_report.columns.tolist()
                        df_report = df_report.drop_duplicates(subset=data_mee_lst, keep='first')
                    except:
                        df_report = df_report.merge(data_mee, on='nodename', how='left')
                        data_mee_lst = df_report.columns.tolist()
                        df_report = df_report.drop_duplicates(subset=data_mee_lst, keep='first')
                if cgic == 'yes':
                    try:
                        del data_mee['fgi']
                    except:
                        pass
                    try:
                        del df_report['cgi']
                    except:
                        pass
                    data_mee_lst = data_mee.columns.tolist()
                    data_mee = data_mee.drop_duplicates(subset=data_mee_lst, keep='first')
                    try:
                        df_report = df_report.merge(data_mee, on='cellname', how='left')
                        data_mee_lst = df_report.columns.tolist()
                        df_report = df_report.drop_duplicates(subset=data_mee_lst, keep='first')
                    except:
                        df_report = df_report.merge(data_mee, on='nodename', how='left')
                        data_mee_lst = df_report.columns.tolist()
                        df_report = df_report.drop_duplicates(subset=data_mee_lst, keep='first')
                except:
                    pass
                if len(dateee) == 0:
                    df_rep = df_report.copy()
                    df_r_date = datetime.now().strftime('%Y%m%d')
                    out_fil = out_filename + str(df_r_date) + '.csv'
                    data_mee_lst = df_rep.columns.tolist()
                    df_rep = df_rep.drop_duplicates(subset=data_mee_lst, keep='first')
                    try:
                        if df_rep.empty:
                            coldf = df_rep.columns.tolist()
                            for dfg in coldf:
                                try:
                                    if 'id' in dfg.lower():
                                        df_rep[dfg] = [1, 1, 1]
                                    else:  # inserted
                                        df_rep[dfg] = ['None', 'None', 'None']
                                    if dfg == 'date':
                                        df_rep[dfg] = [dateee[0], dateee[0], dateee[0]]
                                    if 'time' in dfg.lower():
                                        df_rep[dfg] = ['0', '0', '0']
                                except:
                                    pass
                    except:
                        pass
                    if hOURLY_switch == 'HOURLY':
                        try:
                            for Hattt in HOUR_LST:
                                try:
                                    DF2 = df_rep.copy()
                                    DF2['time'] = DF2['time'].astype(str)
                                    if 'time1' not in DF2['time'].unique():
                                        try:
                                            DF2 = DF2.loc[DF2['time'] == str(Hattt)]
                                        except:
                                            pass
                                    DF2['time'] = str(Hattt)
                                    filename = out_fil.replace(str('.csv'), '').strip()
                                    if len(str(Hattt)) == 1:
                                        Hattt2 = '0' + str(Hattt)
                                    else:  # inserted
                                        Hattt2 = str(Hattt)
                                    filename = filename + str(Hattt2) + '.csv'
                                    DF2.to_csv(OUTFOLDERR + str(filename), index=None, header=True, encoding='utf-8-sig')
                                except:
                                    pass
                        except:
                            pass
                    else:  # inserted
                        df_rep.to_csv(OUTFOLDERR + str(out_fil), index=None, header=True, encoding='utf-8-sig')
                    try:
                        del df_rep
                    except:
                        pass
                else:  # inserted
                    tt = 0
                    for new_date in dateee:
                        if df_report.empty:
                            df_rep = df_report.copy()
                            try:
                                df_r_date = datetime.strptime(new_date, '%m/%d/%Y')
                            except:
                                df_r_date = datetime.strptime(new_date, '%d/%m/%Y')
                            df_r_date = df_r_date.strftime('%Y%m%d')
                            out_fil = out_filename + str(df_r_date) + '.csv'
                            coldf22 = df_rep.columns.tolist()
                            data_mee_lst = df_rep.columns.tolist()
                            df_rep = df_rep.drop_duplicates(subset=data_mee_lst, keep='first')
                            try:
                                if df_rep.empty:
                                    coldf = df_rep.columns.tolist()
                                    for dfg in coldf:
                                        try:
                                            if 'id' in dfg.lower():
                                                df_rep[dfg] = [1, 1, 1]
                                            else:  # inserted
                                                df_rep[dfg] = ['None', 'None', 'None']
                                            if dfg == 'date':
                                                df_rep[dfg] = [new_date, new_date, new_date]
                                            if hOURLY_switch == 'HOURLY':
                                                if 'time' in dfg.lower():
                                                    df_rep[dfg] = ['time', 'time1', 'time2']
                                            else:  # inserted
                                                if 'time' in dfg.lower():
                                                    df_rep[dfg] = ['0', '0', '0']
                                        except:
                                            pass
                            except:
                                pass
                            if hOURLY_switch == 'HOURLY':
                                try:
                                    for Hattt in HOUR_LST:
                                        try:
                                            DF2 = df_rep.copy()
                                            try:
                                                DF2['time'] = DF2['time'].astype(str)
                                            except:
                                                pass
                                            if 'time1' not in DF2['time'].unique():
                                                try:
                                                    DF2 = DF2.loc[DF2['time'] == str(Hattt)]
                                                except:
                                                    pass
                                            DF2['time'] = str(Hattt)
                                            try:
                                                filename = out_fil.replace(str('.csv'), '').strip()
                                            except:
                                                pass
                                            if len(str(Hattt)) == 1:
                                                Hattt2 = '0' + str(Hattt)
                                            else:  # inserted
                                                Hattt2 = str(Hattt)
                                            filename = filename + str(Hattt2) + '.csv'
                                            DF2.to_csv(OUTFOLDERR + str(filename), index=None, header=True, encoding='utf-8-sig')
                                        except:
                                            pass
                                except:
                                    pass
                            else:  # inserted
                                df_rep.to_csv(OUTFOLDERR + str(out_fil), index=None, header=True, encoding='utf-8-sig')
                            try:
                                del df_rep
                            except:
                                pass
                            tt = tt + 1
                        else:  # inserted
                            df_rep = df_report.copy()
                            try:
                                df_r_date = datetime.strptime(new_date, '%m/%d/%Y')
                            except:
                                df_r_date = datetime.strptime(new_date, '%d/%m/%Y')
                            df_r_date = df_r_date.strftime('%Y%m%d')
                            out_fil = out_filename + str(df_r_date) + '.csv'
                            df_rep = df_rep.loc[df_rep['date'] == new_date]
                            data_mee_lst = df_rep.columns.tolist()
                            df_rep = df_rep.drop_duplicates(subset=data_mee_lst, keep='first')
                            try:
                                if df_rep.empty:
                                    coldf = df_rep.columns.tolist()
                                    for dfg in coldf:
                                        try:
                                            if 'id' in dfg.lower():
                                                df_rep[dfg] = [1, 1, 1]
                                            else:  # inserted
                                                df_rep[dfg] = ['None', 'None', 'None']
                                            if dfg == 'date':
                                                df_rep[dfg] = [new_date, new_date, new_date]
                                            if 'time' in dfg.lower():
                                                df_rep[dfg] = ['0', '0', '0']
                                        except:
                                            pass
                            except:
                                pass
                            if hOURLY_switch == 'HOURLY':
                                try:
                                    for Hattt in HOUR_LST:
                                        try:
                                            DF2 = df_rep.copy()
                                            DF2['time'] = DF2['time'].astype(str)
                                            if 'time1' not in DF2['time'].unique():
                                                try:
                                                    DF2 = DF2.loc[DF2['time'] == str(Hattt)]
                                                except:
                                                    pass
                                            DF2['time'] = str(Hattt)
                                            filename = out_fil.replace(str('.csv'), '').strip()
                                            if len(str(Hattt)) == 1:
                                                Hattt2 = '0' + str(Hattt)
                                            else:  # inserted
                                                Hattt2 = str(Hattt)
                                            filename = filename + str(Hattt2) + '.csv'
                                            DF2.to_csv(OUTFOLDERR + str(filename), index=None, header=True, encoding='utf-8-sig')
                                        except:
                                            pass
                                except:
                                    pass
                            else:  # inserted
                                df_rep.to_csv(OUTFOLDERR + str(out_fil), index=None, header=True, encoding='utf-8-sig')
                            try:
                                del df_rep
                            except:
                                pass
                            tt = tt + 1
            kk = kk + 1
            except:
                continue
        if hOURLY_switch == 'HOURLY':
            try:
                csv_files = [f for f in os.listdir(OUTFOLDERR) if f.endswith('.csv')]
                for file in csv_files:
                    try:
                        file_path = os.path.join(OUTFOLDERR, file)
                        df = pd.read_csv(file_path)
                        df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
                        output_file_path = os.path.join(OUTFOLDERR, file)
                        df.to_csv(output_file_path, index=False)
                    except:
                        pass
            except:
                pass
        thread = threading.Thread(target=enabled_button, args=(self,))
        thread.start()
        thread.join()
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ERCA report conversion finished.\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    except Exception as e:
        thread = threading.Thread(target=enabled_button, args=(self,))
        thread.start()
        thread.join()
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : There seems to be an issue with the ERCA report conversion.\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        error_log = ('Error found on Line {}'.format(sys.exc_info()[(-1)].tb_lineno), type(e).__name__)
        error_log4 = list(error_log)
        Final_error = str(error_log4)[1:(-1)]
        with open(OUTPATH + 'Error log ' + datetime.now().strftime('%d%m%Y') + '.txt', 'w') as f:
            f.writelines(Final_error + '\n' + str(e))
        f.close()

def RRSG_PMX_Module(self, lte_pmx_value, node_list_value, hourly_combo_value, start_hour_value, end_hour_value, start_date_value, end_date_value, selected_radio_button, mfolder_create, BOTTYPE):
    def remove_empty_folders(folder_path):
        for root, dirs, files in os.walk(folder_path, topdown=False):
            try:
                for dir_name in dirs:
                    try:
                        dir_path = os.path.join(root, dir_name)
                        if not os.listdir(dir_path):
                            os.rmdir(dir_path)
                    except:
                        pass
            except:
                continue
    try:
        remove_empty_folders(OUTPATH + 'OSS_logs_' + str(mfolder_create))
    except:
        pass
    Start_date = start_date_value
    End_date = end_date_value
    Start_date = datetime.strptime(str(Start_date), '%Y-%m-%d')
    Start_date = Start_date.strftime('%m/%d/%Y')
    End_date = datetime.strptime(str(End_date), '%Y-%m-%d')
    End_date = End_date.strftime('%m/%d/%Y')
    S_HOUR = start_hour_value
    E_HOUR = end_hour_value
    try:
        con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
        Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
        con.close()
        del Rdf['indexxx']
        ppPID_LST = Rdf['value'].tolist()
        ppPID_LST2 = []
        for ppp in ppPID_LST:
            ppPID_LST2.append(decrypt(ppp, 3))
        del Rdf['value']
        Rdf['value'] = ppPID_LST2
        PID_LST = Rdf['value'].tolist()
        remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
        Login_comm_lst = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
    except:
        Login_comm_lst = []
    try:
        db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db')
        con = sqlite3.connect(db_path)
        cursor = con.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
        con.commit()
        Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
        con.close()
        Rdf = Rdf.drop_duplicates(keep='first')
        PID_LST = Rdf['value'].tolist()
        HOSTIP = PID_LST[0]
    except:
        HOSTIP = '148.135.15.71'
    try:
        self.LIVECMD_TEXT.delete(1.0, 'end')
    except:
        pass
    OUTFOLDERR_LTE.clear()
    router = HOSTIP
    username = os.getlogin().lower()
    password = 'OK'
    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait while the PM report is being extracted!\n')
    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
    start_index = '1.0'
    while True:
        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
        if not start_index:
            break
        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
        start_index = end_index
    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    Node_cmd_lst = node_list_value
    Save_logfile = 'YES'
    MOTASK = 'MOBATCH'
    Login_comm_lst = ['']
    activity_type = 'KPI MONITORING'
    PATHHHH_TIME_AUDIT = time.strftime('%d%m%Y%H%M%S')
    LTE_NODE_LST = [item for item in Node_cmd_lst if '@' not in item]
    WCDMA_NODE_LST = [item for item in Node_cmd_lst if '@' in item]
    if LTE_NODE_LST:
        check_tec = 'LTE'
        try:
            os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + 'LTE' + str(' PM_Report_') + str(time.strftime('%d%m%Y%H%M%S')))
        except:
            pass
        try:
            OUTFOLDERR_LTE.append(os.path.join(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + 'LTE' + str(' PM_Report_') + str(time.strftime('%d%m%Y%H%M%S')) + '\\'))
        except:
            pass
        if hourly_combo_value == 'SPECIFIC ROP':
            date_obj = datetime.strptime(Start_date, '%m/%d/%Y')
            formatted_date_s = date_obj.strftime('%Y%m%d')
            date_obje = datetime.strptime(End_date, '%m/%d/%Y')
            formatted_date_e = date_obje.strftime('%Y%m%d')
            time_list = [f'{hour}-{hour + 1}' for hour in range(int(S_HOUR), int(E_HOUR))]
            for timme in time_list:
                S_HOUR_1 = timme.split('-')[0]
                E_HOUR_1 = timme.split('-')[(-1)]
                if len(str(S_HOUR_1)) == 1:
                    S_HOUR22 = '0' + str(S_HOUR_1)
                if len(str(E_HOUR_1)) == 1:
                    E_HOUR22 = '0' + str(E_HOUR_1)
                try:
                    if len(str(S_HOUR22)) == 2:
                        S_HOUR22 = str(S_HOUR22) + '00'
                except:
                    if len(str(S_HOUR_1)) > 2:
                        S_HOUR22 = str(S_HOUR_1)
                    else:  # inserted
                        S_HOUR22 = str(S_HOUR_1) + '00'
                try:
                    if len(str(E_HOUR22)) == 2:
                        E_HOUR22 = str(E_HOUR22) + '00'
                except:
                    if len(str(E_HOUR_1)) > 2:
                        E_HOUR22 = str(E_HOUR_1)
                    else:  # inserted
                        E_HOUR22 = str(E_HOUR_1) + '00'
                commandk = 'pmx . pm .* -s ' + str(formatted_date_s) + '.' + str(S_HOUR22) + ' -e ' + str(formatted_date_e) + '.' + str(E_HOUR22) + ' -a'
                LTE_LST = [commandk]
                LSTTT_cmdd = []
                RRSG_PMX_process(self, Login_comm_lst, LTE_NODE_LST, LTE_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                PMX_ROP_TIME.clear()
                PMAX_DATE.clear()
        else:  # inserted
            LTE_LST = ['pmx . pm.']
            LSTTT_cmdd = []
            RRSG_PMX_process(self, Login_comm_lst, LTE_NODE_LST, LTE_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
            PMX_ROP_TIME.clear()
            PMAX_DATE.clear()
        try:
            try:
                final_df = pd.concat(ROP_WISE_REPORT, ignore_index=True)
            except:
                pass
            try:
                ROP_WISE_REPORT.clear()
            except:
                pass
            try:
                final_df = final_df.dropna(axis=1, how='all')
            except:
                pass
            final_df_22 = final_df.copy()
            final_df_22['cellname_del'] = final_df_22['Node Id'].astype(str)

            def compare_columns(row):
                if row['cellname'] == row['cellname_del']:
                    return 'OK'
                return 'NOK'
            final_df_22['status'] = final_df_22.apply(compare_columns, axis=1)
            final_df_22_node = final_df_22.copy()
            final_df_22_cell = final_df_22.copy()
            final_df_22_node = final_df_22_node.loc[final_df_22_node['status'] == 'OK']
            del final_df_22_node['status']
            del final_df_22_node['cellname_del']
            final_df_22_node.replace('', np.nan, inplace=True)
            final_df_22_node = final_df_22_node.dropna(axis=1, how='all')
            final_df_22_cell = final_df_22_cell.loc[final_df_22_cell['status'] == 'NOK']
            del final_df_22_cell['status']
            del final_df_22_cell['cellname_del']
            final_df_22_cell.replace('', np.nan, inplace=True)
            final_df_22_cell = final_df_22_cell.dropna(axis=1, how='all')
            final_df_22_cell_LTE = final_df_22_cell.copy()
            try:
                final_df_22_cell_LTE['pmA3InterFBestCellEvalReport'].replace('', np.nan, inplace=True)
                final_df_22_cell_LTE.dropna(subset=['pmA3InterFBestCellEvalReport'], inplace=True)
            except:
                pass
            final_df_22_cell_wcdma = final_df_22_cell.copy()
            try:
                final_df_22_cell_wcdma['pmDchFramesReceived'].replace('', np.nan, inplace=True)
                final_df_22_cell_wcdma.dropna(subset=['pmDchFramesReceived'], inplace=True)
            except:
                pass
            if lte_pmx_value == 'WCDMA-PMX' and (not final_df_22_cell_wcdma.empty):
                final_df_22_cell_wcdma = final_df_22_cell_wcdma.dropna(axis=1, how='all')
                try:
                    final_df_22_cell_wcdma.to_csv(os.path.join(OUTFOLDERR_LTE[0]) + 'WCDMA PMX_report_cell_v1.csv', index=None, header=True, encoding='utf-8-sig')
                except:
                    pass
            try:
                final_df_22_node.to_csv(os.path.join(OUTFOLDERR_LTE[0]) + str(check_tec) + ' PMX_report_node.csv', index=None, header=True, encoding='utf-8-sig')
            except:
                pass
            try:
                final_df_22_cell_LTE.to_csv(os.path.join(OUTFOLDERR_LTE[0]) + str(check_tec) + ' PMX_report_cell.csv', index=None, header=True, encoding='utf-8-sig')
            except:
                pass
        except:
            pass
        try:
            del final_df_22
        except:
            pass
        try:
            del final_df
        except:
            pass
        try:
            ROP_WISE_REPORT.clear()
        except:
            pass
        try:
            del final_df_22_cell_wcdma
        except:
            pass
        try:
            del final_df_22_cell_LTE
        except:
            pass
    if WCDMA_NODE_LST:
        check_tec = 'WCDMA'
        try:
            os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + check_tec + str(' PM_Report_') + str(time.strftime('%d%m%Y%H%M%S')))
        except:
            pass
        try:
            OUTFOLDERR_WCDMA.append(os.path.join(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + check_tec + str(' PM_Report_') + str(time.strftime('%d%m%Y%H%M%S')) + '\\'))
        except:
            pass
        try:
            try:
                shutil.copy(os.path.join(OUTFOLDERR_LTE[0]) + 'WCDMA PMX_report_cell_v1.csv', os.path.join(OUTFOLDERR_WCDMA[0]) + 'WCDMA PMX_report_cell_v1.csv')
            except:
                pass
            try:
                shutil.copy(os.path.join(OUTFOLDERR_LTE[0]) + 'Ropwise_log ' + datetime.now().strftime('%d%m%Y') + '.log', os.path.join(OUTFOLDERR_WCDMA[0]) + 'Ropwise_logv1 ' + datetime.now().strftime('%d%m%Y') + '.log')
            except:
                pass
            shutil.rmtree(os.path.join(OUTFOLDERR_LTE[0]))
        except:
            pass
        if hourly_combo_value == 'SPECIFIC ROP':
            date_obj = datetime.strptime(Start_date, '%m/%d/%Y')
            formatted_date_s = date_obj.strftime('%Y%m%d')
            date_obje = datetime.strptime(End_date, '%m/%d/%Y')
            formatted_date_e = date_obje.strftime('%Y%m%d')
            time_list = [f'{hour}-{hour + 1}' for hour in range(int(S_HOUR), int(E_HOUR))]
            for timme in time_list:
                S_HOUR_1 = timme.split('-')[0]
                E_HOUR_1 = timme.split('-')[(-1)]
                if len(str(S_HOUR_1)) == 1:
                    S_HOUR22 = '0' + str(S_HOUR_1)
                if len(str(E_HOUR_1)) == 1:
                    E_HOUR22 = '0' + str(E_HOUR_1)
                try:
                    if len(str(S_HOUR22)) == 2:
                        S_HOUR22 = str(S_HOUR22) + '00'
                except:
                    if len(str(S_HOUR_1)) > 2:
                        S_HOUR22 = str(S_HOUR_1)
                    else:  # inserted
                        S_HOUR22 = str(S_HOUR_1) + '00'
                try:
                    if len(str(E_HOUR22)) == 2:
                        E_HOUR22 = str(E_HOUR22) + '00'
                except:
                    if len(str(E_HOUR_1)) > 2:
                        E_HOUR22 = str(E_HOUR_1)
                    else:  # inserted
                        E_HOUR22 = str(E_HOUR_1) + '00'
                samplenode = 'samplenode'
                commandk = 'pmx ' + str(samplenode) + ' pm .* -s ' + str(formatted_date_s) + '.' + str(S_HOUR22) + ' -e ' + str(formatted_date_e) + '.' + str(E_HOUR22) + ' -a'
                LTE_LST = [commandk]
                LSTTT_cmdd = []
                RNC_LST = []
                for ee in WCDMA_NODE_LST:
                    if ee.split('@')[0].strip() not in RNC_LST:
                        RNC_LST.append(ee.split('@')[0].strip())
                RNC_NODELST = []
                for rrn in RNC_LST:
                    for rncnod in WCDMA_NODE_LST:
                        if rrn in rncnod:
                            RNC_NODELST.append(rncnod)
                    RRSG_PMX_process(self, Login_comm_lst, RNC_NODELST, LTE_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                    PMX_ROP_TIME.clear()
                    PMAX_DATE.clear()
                    RNC_NODELST.clear()
        else:  # inserted
            LTE_LST = ['pmx . pm.']
            LSTTT_cmdd = []
            RNC_LST = []
            for ee in WCDMA_NODE_LST:
                if ee.split('@')[0].strip() not in RNC_LST:
                    RNC_LST.append(ee.split('@')[0].strip())
            RNC_NODELST = []
            for rrn in RNC_LST:
                for rncnod in WCDMA_NODE_LST:
                    if rrn in rncnod:
                        RNC_NODELST.append(rncnod)
                RRSG_PMX_process(self, Login_comm_lst, RNC_NODELST, LTE_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
            PMX_ROP_TIME.clear()
            PMAX_DATE.clear()
        try:
            final_df = pd.concat(ROP_WISE_REPORT, ignore_index=True)
        except:
            pass
        try:
            ROP_WISE_REPORT.clear()
        except:
            pass
        try:
            final_df['status'] = final_df.apply(lambda row: 'Node' if row['Node Id'] == row['cellname'] else 'Cell', axis=1)
        except:
            pass
        final_df_22 = final_df.copy()
        try:
            final_df_22 = final_df_22[final_df_22['status'].str.contains('Node')]
            del final_df_22['status']
        except:
            pass
        final_df_23 = final_df.copy()
        try:
            final_df_23 = final_df_23[final_df_23['status'].str.contains('Cell')]
            del final_df_23['status']
        except:
            pass
        try:
            final_df_22.to_csv(os.path.join(OUTFOLDERR_WCDMA[0]) + str(check_tec) + ' PMX_report_node.csv', index=None, header=True, encoding='utf-8-sig')
        except:
            pass
        try:
            final_df_23.to_csv(os.path.join(OUTFOLDERR_WCDMA[0]) + str(check_tec) + ' PMX_report_cell.csv', index=None, header=True, encoding='utf-8-sig')
        except:
            pass
        try:
            del final_df_22
        except:
            pass
        try:
            del final_df_23
        except:
            pass
        try:
            del final_df
        except:
            pass
        try:
            ROP_WISE_REPORT.clear()
        except:
            pass
        try:
            ROP_WISE_REPORT.clear()
        except:
            pass
        try:
            PMAX_DATE.clear()
        except:
            pass
        try:
            PMX_ROP_TIME.clear()
        except:
            pass
    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : PM Report extraction finished.\n')
    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
    start_index = '1.0'
    while True:
        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
        if not start_index:
            break
        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
        start_index = end_index
    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    if selected_radio_button == 'ERCA FORMATE':
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please hold on while the ERCA report is being converted.\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        if lte_pmx_value == 'LTE-PMX':
            technoo = 'LTE'
            INPATH_new = os.path.join(OUTFOLDERR_LTE[0])
        else:  # inserted
            if lte_pmx_value == 'WCDMA-PMX':
                technoo = 'WCDMA'
                INPATH_new = os.path.join(OUTFOLDERR_WCDMA[0])
            else:  # inserted
                if lte_pmx_value == '5G-PMX':
                    technoo = '5G'
                    INPATH_new = os.path.join(OUTFOLDERR_WCDMA[0])
        PMBOT_config_FILE = 'PM_config.xlsx'
        hOURLY_switch = 'HOURLY'
        try:
            os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + technoo + '_ERCA_Report_' + str(mfolder_create))
        except:
            pass
        OUTFOLDERR_k = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + technoo + '_ERCA_Report_' + str(mfolder_create) + '\\'
        threading.Thread(target=disable_button, args=(self,)).start()
        thread_erca_xl = threading.Thread(target=ERCA_FROMAT, args=(self, INPATH_new, OUTFOLDERR_k, INPATH + PMBOT_config_FILE, technoo, hOURLY_switch))
        thread_erca_xl.start()
        thread_erca_xl.join()
    if 'cancle' not in processdone:
        try:
            ISF_STEP_CLOSE(ISF_API.API_Based, ISF_STEP_START_Task_Id_lst[0], ISF_API.woid, ISF_API.headers)
        except:
            pass
        try:
            ISF_WO_CLOSE(ISF_API.API_Based, ISF_API.headers, ISF_API.woid)
            try:
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + str(activity_type) + str(' WOID') + '__' + str(ISF_API.woid) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            except:
                pass
        except:
            pass
    try:
        shutil.rmtree(os.path.join(OUTFOLDERR_LTE[0]) + 'temp')
    except:
        pass
    try:
        shutil.rmtree(os.path.join(OUTFOLDERR_WCDMA[0]) + 'temp')
    except:
        pass
    try:
        OUTFOLDERR_LTE.clear()
    except:
        pass
    try:
        OUTFOLDERR_WCDMA.clear()
    except:
        pass
    clear_all_lists()
    try:
        cmmmddd.clear()
    except:
        pass
    try:
        siteddd.clear()
    except:
        pass
    try:
        checknxtsit.clear()
    except:
        pass
    try:
        processdone.clear()
    except:
        return None

def RRSG_Audit_Module(self, selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST, activity_type, BOTTYPE):
    PATHHHH_TIME_AUDIT = time.strftime('%d%m%Y%H%M%S')
    try:
        con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
        Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
        con.close()
        del Rdf['indexxx']
        ppPID_LST = Rdf['value'].tolist()
        ppPID_LST2 = []
        for ppp in ppPID_LST:
            ppPID_LST2.append(decrypt(ppp, 3))
        del Rdf['value']
        Rdf['value'] = ppPID_LST2
        PID_LST = Rdf['value'].tolist()
        remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
        Login_comm_lst = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
    except:
        Login_comm_lst = []
    try:
        db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db')
        con = sqlite3.connect(db_path)
        cursor = con.cursor()
        cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
        con.commit()
        Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
        con.close()
        Rdf = Rdf.drop_duplicates(keep='first')
        PID_LST = Rdf['value'].tolist()
        HOSTIP = PID_LST[0]
    except:
        HOSTIP = '148.135.15.71'
    try:
        self.LIVECMD_TEXT.delete(1.0, 'end')
    except:
        pass
    AZURE_CREAD_FAIL.clear()
    Azue_OTP_lst.clear()
    establish_ssh_conn_lst.clear()
    KEYWORD_ADD.clear()
    excel_df_LST.clear()
    amos_node.clear()
    Buddy_request_lst.clear()
    buddy_node.clear()
    buddy_node_curr.clear()
    amos_nod2.clear()
    ltall_nod2.clear()
    check_yesno.clear()
    ASKSITID.clear()
    cmmmddd.clear()
    siteddd.clear()
    checknxtsit.clear()
    processdone.clear()
    combined_cmd_window.clear()
    combined_cmd_window_chk.clear()
    checkamos.clear()
    command.clear()
    LSTTT_cmdd = []
    for check_tec in selected_checkboxes:
        if check_tec == 'GSM':
            if activity_type == 'Health Check':
                Node_cmd_lst = GSM_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, GSM_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                try:
                    filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                    df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                    try:
                        if not df_summary.empty:
                            with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                                try:
                                    df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                    wb = writer.book
                                    summary_sheet = wb['Summary']
                                    wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                    workbook = writer.book
                                    headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                                except:
                                    pass
                            try:
                                workbook = openpyxl.load_workbook(filelocation)
                                sheet = workbook['Summary']
                                red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                white_font = Font(color='FFFFFF', bold=True)
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value == 'NOK':
                                            cell.fill = red_fill
                                            cell.font = white_font
                                workbook.save(filelocation)
                            except:
                                pass
                    except:
                        pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
            if activity_type == 'Parameter Audit':
                Node_cmd_lst = GSM_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'KGET'
                oss_cmd_lst = ['kget all']
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'KGET')
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
            if activity_type == 'Health Check & Parameter':
                Node_cmd_lst = GSM_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, GSM_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                try:
                    filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                    df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                    try:
                        if not df_summary.empty:
                            with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                                try:
                                    df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                    wb = writer.book
                                    summary_sheet = wb['Summary']
                                    wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                    workbook = writer.book
                                    headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                                except:
                                    pass
                            try:
                                workbook = openpyxl.load_workbook(filelocation)
                                sheet = workbook['Summary']
                                red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                white_font = Font(color='FFFFFF', bold=True)
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value == 'NOK':
                                            cell.fill = red_fill
                                            cell.font = white_font
                                workbook.save(filelocation)
                            except:
                                pass
                    except:
                        pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
                Node_cmd_lst = GSM_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'KGET'
                oss_cmd_lst = ['kget all']
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
        if check_tec == 'WCDMA':
            if activity_type == 'Health Check':
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                WCDMA_LST_RBS = []
                WCDMA_LST_RNC = []
                for item in WCDMA_LST:
                    try:
                        prefix, suffix = item.split('@')
                        if prefix == 'RBS':
                            WCDMA_LST_RBS.append(suffix)
                        else:  # inserted
                            if prefix == 'RNC':
                                WCDMA_LST_RNC.append(suffix)
                    except:
                        WCDMA_LST_RBS.append(item)
                Node_cmd_RBS = []
                Node_cmd_RNC = []
                for item in WCDMA_NODE_LST:
                    if '@' not in item:
                        if item not in Node_cmd_RBS:
                            Node_cmd_RBS.append(item)
                    else:  # inserted
                        if item not in Node_cmd_RNC:
                            Node_cmd_RNC.append(item)
                FINAL_RNC_COMMAND_LST = []
                FINAL_RNC_NODE_LST = []
                for item in Node_cmd_RNC:
                    for item2 in WCDMA_LST_RNC:
                        item2 = item2.replace('CELL_NAME', str(item.split('@')[(-1)]).strip()).strip()
                        if str(item2) not in FINAL_RNC_COMMAND_LST:
                            FINAL_RNC_COMMAND_LST.append(str(item2))
                        if str(item.split('@')[0]).strip() not in FINAL_RNC_NODE_LST:
                            FINAL_RNC_NODE_LST.append(str(item.split('@')[0]).strip())
                if Node_cmd_RBS:
                    RRSG_Audit_process(self, Login_comm_lst, Node_cmd_RBS, WCDMA_LST_RBS, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                if FINAL_RNC_NODE_LST:
                    RRSG_Audit_process(self, Login_comm_lst, FINAL_RNC_NODE_LST, FINAL_RNC_COMMAND_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RNC')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                try:
                    filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                    df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                    if not df_summary.empty:
                        with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                            try:
                                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                wb = writer.book
                                summary_sheet = wb['Summary']
                                wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                workbook = writer.book
                                headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                            except:
                                pass
                        try:
                            workbook = openpyxl.load_workbook(filelocation)
                            sheet = workbook['Summary']
                            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                            white_font = Font(color='FFFFFF', bold=True)
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value == 'NOK':
                                        cell.fill = red_fill
                                        cell.font = white_font
                            workbook.save(filelocation)
                        except:
                            pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
            if activity_type == 'Parameter Audit':
                WCDMA_NODE_LST_F = []
                WCDMA_RNC_LST_F = []
                for ccc in WCDMA_NODE_LST:
                    try:
                        WCDMA_NODE_LST_F.append(ccc.split('@')[0])
                    except:
                        pass
                    try:
                        WCDMA_RNC_LST_F.append(ccc.split('@')[(-1)])
                    except:
                        continue
                if WCDMA_RNC_LST_F:
                    Node_cmd_lst = WCDMA_RNC_LST_F
                    router = HOSTIP
                    username = os.getlogin().lower()
                    password = 'OK'
                    Save_logfile = 'YES'
                    MOTASK = 'KGET'
                    oss_cmd_lst = ['kget all']
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                    RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'KGET')
                    try:
                        with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    except:
                        pass
                else:  # inserted
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : RNC List is empty.\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            if activity_type == 'Health Check & Parameter':
                Node_cmd_lst = WCDMA_NODE_LST_F
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, WCDMA_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                try:
                    filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                    df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                    try:
                        if not df_summary.empty:
                            with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                                try:
                                    df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                    wb = writer.book
                                    summary_sheet = wb['Summary']
                                    wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                    workbook = writer.book
                                    headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                                except:
                                    pass
                            try:
                                workbook = openpyxl.load_workbook(filelocation)
                                sheet = workbook['Summary']
                                red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                white_font = Font(color='FFFFFF', bold=True)
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value == 'NOK':
                                            cell.fill = red_fill
                                            cell.font = white_font
                                workbook.save(filelocation)
                            except:
                                pass
                    except:
                        pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
                if WCDMA_RNC_LST_F:
                    Node_cmd_lst = WCDMA_RNC_LST_F
                    router = HOSTIP
                    username = os.getlogin().lower()
                    password = 'OK'
                    Save_logfile = 'YES'
                    MOTASK = 'KGET'
                    oss_cmd_lst = ['kget all']
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                    RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                    try:
                        with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    except:
                        pass
                else:  # inserted
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : RNC List is empty.\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        if check_tec == 'LTE':
            if activity_type == 'Health Check':
                Node_cmd_lst = LTE_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, LTE_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                try:
                    filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                    df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                    try:
                        if not df_summary.empty:
                            with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                                try:
                                    df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                    wb = writer.book
                                    summary_sheet = wb['Summary']
                                    wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                    workbook = writer.book
                                    headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                                except:
                                    pass
                            try:
                                workbook = openpyxl.load_workbook(filelocation)
                                sheet = workbook['Summary']
                                red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                white_font = Font(color='FFFFFF', bold=True)
                                for row in sheet.iter_rows():
                                    for cell in row:
                                        if cell.value == 'NOK':
                                            cell.fill = red_fill
                                            cell.font = white_font
                                workbook.save(filelocation)
                            except:
                                pass
                    except:
                        pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
            if activity_type == 'Parameter Audit':
                Node_cmd_lst = LTE_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'KGET'
                oss_cmd_lst = ['kget all']
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'KGET')
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
            if activity_type == 'Health Check & Parameter':
                Node_cmd_lst = LTE_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, LTE_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                try:
                    if not df_summary.empty:
                        with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                            try:
                                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                wb = writer.book
                                summary_sheet = wb['Summary']
                                wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                workbook = writer.book
                                headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                            except:
                                pass
                        try:
                            workbook = openpyxl.load_workbook(filelocation)
                            sheet = workbook['Summary']
                            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                            white_font = Font(color='FFFFFF', bold=True)
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value == 'NOK':
                                        cell.fill = red_fill
                                        cell.font = white_font
                            workbook.save(filelocation)
                        except:
                            pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
                Node_cmd_lst = LTE_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'KGET'
                oss_cmd_lst = ['kget all']
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
        if check_tec == 'NR':
            if activity_type == 'Health Check':
                Node_cmd_lst = NR_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, NR_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                try:
                    if not df_summary.empty:
                        with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                            try:
                                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                wb = writer.book
                                summary_sheet = wb['Summary']
                                wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                workbook = writer.book
                                headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                            except:
                                pass
                        try:
                            workbook = openpyxl.load_workbook(filelocation)
                            sheet = workbook['Summary']
                            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                            white_font = Font(color='FFFFFF', bold=True)
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value == 'NOK':
                                        cell.fill = red_fill
                                        cell.font = white_font
                            workbook.save(filelocation)
                        except:
                            pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
            if activity_type == 'Parameter Audit':
                Node_cmd_lst = NR_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'KGET'
                oss_cmd_lst = ['kget all']
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'KGET')
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
            if activity_type == 'Health Check & Parameter':
                Node_cmd_lst = NR_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'MOBATCH'
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Health check audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, NR_LST, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    eyye = RRSG_auto_login.remote_conn
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Health check audit completed !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                except:
                    pass
                filelocation = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(check_tec) + '\\' + str(check_tec) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                df_summary = SUMMARY_SH(filelocation, LSTTT_cmdd)
                try:
                    if not df_summary.empty:
                        with pd.ExcelWriter(filelocation, engine='openpyxl', mode='a') as writer:
                            try:
                                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                                wb = writer.book
                                summary_sheet = wb['Summary']
                                wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))
                                workbook = writer.book
                                headr_colour_summary(writer, df_summary, 'Summary', '0070C0', ['KPI'], 'red')
                            except:
                                pass
                        try:
                            workbook = openpyxl.load_workbook(filelocation)
                            sheet = workbook['Summary']
                            red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                            white_font = Font(color='FFFFFF', bold=True)
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value == 'NOK':
                                        cell.fill = red_fill
                                        cell.font = white_font
                            workbook.save(filelocation)
                        except:
                            pass
                    initial_kpi_format(INPATH + 'Baseline.xlsx', filelocation, 'Initial KPI Threshold', 'Initial KPI', check_tec)
                except:
                    pass
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
                Node_cmd_lst = NR_NODE_LST
                router = HOSTIP
                username = os.getlogin().lower()
                password = 'OK'
                Save_logfile = 'YES'
                MOTASK = 'KGET'
                oss_cmd_lst = ['kget all']
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Parameter audit ongoing !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, check_tec, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmdd, 'RBS')
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(check_tec) + '__' + str(activity_type) + '__' + str('Close') + '__NA' + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    continue
    try:
        chk_rssg = RRSG_auto_login.remote_conn
        chk_rssg_chk = 'OK'
    except:
        chk_rssg_chk = 'NOK'
    if chk_rssg_chk == 'OK':
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Done! data has been successfully fetched. Feel free to ask for more assistance or information.') + '\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    thread = threading.Thread(target=enabled_button, args=(self,))
    thread.start()
    thread.join()
    DATABASE = 'CAUDIT_CLOSE_WINDOW'
    SQLITE_DB_CLEAR(DATABASE)
    DATABASE = 'AUDDIT_START'
    SQLITE_DB_CLEAR(DATABASE)
    DATABASE = 'ACTIVITYY'
    SQLITE_DB_CLEAR(DATABASE)
    if 'cancle' not in processdone:
        try:
            ISF_STEP_CLOSE(ISF_API.API_Based, ISF_STEP_START_Task_Id_lst[0], ISF_API.woid, ISF_API.headers)
        except:
            pass
        try:
            ISF_WO_CLOSE(ISF_API.API_Based, ISF_API.headers, ISF_API.woid)
            try:
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + str(activity_type) + str(' WOID') + '__' + str(ISF_API.woid) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            except:
                pass
        except:
            pass
    try:
        loggs_up_audit()
    except:
        pass
    selected_checkboxes.clear()
    if chk_rssg_chk == 'OK':
        showMessage_qt('Audit completed successfully.', 5000)
    clear_all_lists()

def headr_colour_summary(writer2, dfff, sh_name, colour, hed_lst, tab):
    try:
        worksheet = writer2.sheets[sh_name]
        worksheet.sheet_properties.tabColor = colour
        header_format = Font(bold=True, color='FFFFFF')
        fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        for col_num, value in enumerate(dfff.columns.values, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = value
            cell.font = header_format
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        num_rows = worksheet.max_row
        for row_num in range(1, num_rows + 1):
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if row_num == 1:
                    cell.font = Font(bold=True, color='FFFFFF')
                else:  # inserted
                    cell.font = Font(name='Calibri', size=11, color='000000')
                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=False)
        for column in dfff:
            column_letter = get_column_letter(dfff.columns.get_loc(column) + 1)
            max_length = max(dfff[column].astype(str).apply(len).max(), len(column))
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width
        return dfff
    except:
        return None

def ask_yes_or_no_qt(message, timeout=600000):
    app = QApplication(sys.argv)
    msgBox = QMessageBox()
    msgBox.setIcon(QMessageBox.Question)
    msgBox.setText(message)
    msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
    msgBox.setWindowFlags(msgBox.windowFlags() | Qt.WindowStaysOnTopHint)
    msgBox.setWindowTitle('Question')
    msgBox.setStyleSheet('QMessageBox { background-color: #f0f0f0; }QPushButton { color: #ffffff; background-color: #fe6f5e; }QPushButton:hover { background-color: #fe6f5e; }')
    timer = QTimer()
    timer.timeout.connect(lambda: msgBox.close() if msgBox.isVisible() else None)
    timer.start(timeout)
    button_clicked = msgBox.exec_()
    timer.stop()
    if button_clicked == QMessageBox.Yes:
        return True
    return False

def showMessage_qt(message, timeout=5000):
    app = QApplication(sys.argv)
    msgBox = QMessageBox()
    msgBox.setIcon(QMessageBox.Information)
    msgBox.setText(message)
    msgBox.setWindowFlags(msgBox.windowFlags() | Qt.WindowStaysOnTopHint)
    msgBox.setWindowTitle('Information')
    msgBox.setStyleSheet('QMessageBox { background-color: #f0f0f0; }QPushButton { color: #ffffff; background-color: #fe6f5e; }QPushButton:hover { background-color: #fe6f5e; }')
    QTimer.singleShot(timeout, msgBox.close)
    msgBox.exec_()

def SUMMARY_SH(FILEPATH, LSTTT):
    WS_Df = pd.ExcelFile(FILEPATH)
    sheet_lst = WS_Df.sheet_names
    ALLDF = []
    for sh in sheet_lst:
        try:
            M_DF1 = WS_Df.parse(sh)
            M_DF1 = M_DF1[['Node Id', 'Baseline Check']]
            M_DF1 = M_DF1.drop_duplicates(subset=['Node Id', 'Baseline Check'], keep='first')
            M_DF1['Baseline Check'].replace('', np.nan, inplace=True)
            M_DF1['Baseline Check'].fillna('Pass', inplace=True)
            M_DF1['countif'] = M_DF1.groupby('Node Id')['Node Id'].transform('count')
            M_DF1 = M_DF1[~((M_DF1['Baseline Check'] == 'Pass') & (M_DF1['Baseline Check'].shift((-1)) == 'Fail'))]
            M_DF1['Status'] = M_DF1['countif'].apply(lambda x: 'Pass' if x == 1 else 'Fail')
            M_DF1 = M_DF1.drop_duplicates(subset=['Node Id', 'Status'], keep='first')
            M_DF1['Status'] = M_DF1['Baseline Check'].apply(lambda x: 'OK' if x == 'Pass' else 'NOK')
            del M_DF1['countif']
            del M_DF1['Baseline Check']
            M_DF1 = M_DF1.drop_duplicates(subset=['Node Id', 'Status'], keep='first')
            M_DF1['Activity'] = str(sh)
            ALLDF.append(M_DF1)
        except:
            continue
    try:
        combined_df = pd.concat(ALLDF, ignore_index=True)
        table = pd.pivot_table(combined_df, values='Status', index=['Node Id'], columns=['Activity'], aggfunc='sum')
        df_summary = table.rename_axis(None, axis=1).reset_index(drop=False)
        for col in df_summary.columns:
            df_summary = df_summary.rename(columns={col: col.replace('_', ' ')})
        for col in df_summary.columns:
            for lcol in LSTTT:
                try:
                    cmdd = lcol.split('_')[0]
                    cmdd2 = lcol.split('_')[1]
                    if col.lower() == cmdd.lower():
                        df_summary = df_summary.rename(columns={col: cmdd2})
                except:
                    pass
        try:
            df_summary = df_summary.fillna('OK')
            return
            return df_summary
        except:
            return df_summary
    except:
        df_summary = pd.DataFrame()

def RRSG_monitoring_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, techeccci, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmd):
    try:
        os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT))
    except:
        pass
    try:
        os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci))
    except:
        pass
    if 'login_error' in LOGINERROR_LST:
        return
    check_lst = []
    CIPRILST = []

    def read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK):
        commnd_sh = []
        commnd_sh.append(cmd_ex)

        def remove_color_codes(input_string):
            ansi_escape = re.compile('\\x1b[^m]*m')
            clean_string = ansi_escape.sub('', input_string)
            html_color = re.compile('<[^>]*>')
            clean_string = html_color.sub('', clean_string)
            return clean_string
        if MOTASK == 'MOBATCH':
            read_received_data_auto.out = ''
            output_B = b''
            combined_cmd = ''
            command_not_found = []
            while not (output_B.endswith(b'> ') or output_B.endswith(b'<') or output_B.endswith(b': ') or output_B.endswith(b'<') or output_B.endswith(b'< ') or output_B.endswith(b':') or output_B.endswith(b'$ ')):
                output_B += remote_conn.recv(99999)
                write_t = output_B.decode(encoding='utf-8')
                write_te_1 = remove_color_codes(write_t)
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('\a', '')
                self.update_textbox(f'{write_te_1}\n')
                if cmd_ex == 'sdir' and 'CPRI links' in write_te_1 and (not CIPRILST):
                    CIPRILST.append(write_te_1)
                try:
                    self.reset_timer(None)
                except:
                    pass
                if 'no such command' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'command not found' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'coli>' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'amos' in command and 'AMOS error' in write_te_1 and ('no such command' not in command_not_found):
                    command_not_found.append('no such command')
                last_line = output_B.decode(encoding='utf-8').split('\n')[(-1)]
                last_line = ' '.join(last_line.split()).strip()
                if check_lst and last_line == check_lst[0]:
                    read_received_data_auto.brek = 'break'
                try:
                    if last_line[(-1)].strip() == '>':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '<':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == ':':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '$':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
            if not command_not_found:
                write_text1 = output_B.decode(encoding='utf-8')
                write_text = remove_color_codes(write_text1)
                write_text = write_text.replace('', '')
                write_text = write_text.replace('', '')
                write_text = write_text.replace('\a', '')
                if Save_logfile == 'YES':
                    if chk_oss!= 'LOGIN_CNM' and chk_oss!= 'no':
                        with open(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\' + techeccci + '_' + str(cmd_node) + '.Log', 'a+', encoding='utf-8') as f:
                            f.writelines(f'{write_text}\n\n')
                        with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str(activity_type) + '__' + str(cmd_node) + '__' + str(command) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    if chk_oss == 'OSS_CNM':
                        combined_cmd += f'{write_text}\n'
                crate_txt_df.df_fram = ''
                if chk_oss == 'OSS_CNM':
                    if command!= 'q' and command!= 'quit':
                        def remove_line_containing(text, substring):
                            lines = text.split('\n')
                            filtered_lines = [line for line in lines if substring not in line]
                            return '\n'.join(filtered_lines)
                        combined_cmd = combined_cmd.replace('It is recommended to remove duplicate counter definitions from the PM scanners.', '')
                        combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                        combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo)', 'Date & Time (UTC)   S Specific Problem                    MO (Cause/AdditionalInfo)')
                        combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo) Operator', '')
                        combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                        combined_cmd = remove_line_containing(combined_cmd, 'ACKNOWLEDGED ALARMS')
                        if 'pmr' in combined_cmd.lower():
                            combined_cmd = remove_line_containing(combined_cmd, 'pmr')
                            combined_cmd = remove_line_containing(combined_cmd, 'PMR')
                        combined_cmd = combined_cmd.strip()
                        START_TIME_CHECK = []
                        END_TIME_CHECK = []
                        pattern = 'Start Time: (\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2}) End Time: (\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2})'
                        match = re.search(pattern, combined_cmd)
                        if match:
                            line_with_times = match.group(0)
                            pattern = 'Start Time: \\d{4}-\\d{2}-\\d{2} (\\d{2}):(\\d{2}):(\\d{2}) End Time: \\d{4}-\\d{2}-\\d{2} (\\d{2}):(\\d{2}):(\\d{2})'
                            match = re.search(pattern, line_with_times)
                            if match:
                                start_hour, start_minute = (match.group(1), match.group(2))
                                end_hour, end_minute = (match.group(4), match.group(5))
                                Start_Time = '{}:{}'.format(start_hour, start_minute)
                                End_Time = '{}:{}'.format(end_hour, end_minute)
                                START_TIME_CHECK.append(Start_Time)
                                END_TIME_CHECK.append(End_Time)
                        if '16Qam'.lower() not in command.lower():
                            combined_cmd = log_clening(combined_cmd, cmd_ex, cmd_node)
                        Final_log_lst = final_log_process(combined_cmd)
                        try:
                            combined_cmd = ''
                        except:
                            pass
                        Final_log_lst = [item.strip() for item in Final_log_lst if item.strip()]
                        try:
                            if Final_log_lst[0] == '\n':
                                Final_log_lst = []
                        except:
                            pass
                        if Final_log_lst:
                            i = 1
                            Final_log_lst = [item for item in Final_log_lst if not item.startswith('+')]
                            Final_log_lst = [item for item in Final_log_lst if not item.startswith('\n+')]
                            ALL_DF = []
                            for section in Final_log_lst:
                                lines = section.strip().split('\n')
                                data = [line.split('\t') for line in lines]
                                if len(data) > 1:
                                    crate_df = crate_txt_df(section, cmd_node, cmd_ex)
                                    try:
                                        crate_df.replace('deletestring', '', inplace=True)
                                    except:
                                        pass
                                    if crate_df is not None and (not crate_df.empty):
                                        crate_df = crate_df[~crate_df.apply(lambda row: any(row.astype(str).str.contains('Total:')), axis=1)]
                                        crate_df.columns = crate_df.columns.str.strip()
                                        if not crate_df.empty:
                                            try:
                                                try:
                                                    crate_df['VSWR (RL)'] = crate_df['VSWR (RL)'].str.strip()
                                                except:
                                                    pass
                                                VSWR = [col for col in crate_df.columns if 'VSWR' in col]
                                                if VSWR:
                                                    try:
                                                        crate_df[['VSWR', 'RL']] = crate_df['VSWR_(RL)'].str.split('_', expand=True)
                                                    except:
                                                        crate_df[['VSWR', 'RL']] = crate_df['VSWR (RL)'].str.split(' ', expand=True)
                                                    crate_df['RL'] = crate_df['RL'].str.strip('()')
                                                    crate_df['VSWR'] = crate_df['VSWR'].astype(float)
                                                    crate_df['RL'] = crate_df['RL'].astype(float)
                                                    try:
                                                        del crate_df['VSWR_(RL)']
                                                    except:
                                                        try:
                                                            del crate_df['VSWR (RL)']
                                                        except:
                                                            pass
                                            except:
                                                pass
                                            try:
                                                crate_df = crate_df.applymap(lambda x: x.replace('_', ' '))
                                            except:
                                                pass
                                            try:
                                                crate_df['Node Id'] = cmd_node
                                            except:
                                                pass
                                            try:
                                                crate_df.insert(0, 'Node Id', crate_df.pop('Node Id'))
                                            except:
                                                pass
                                            counter = [col for col in crate_df.columns if 'counter' in col.lower()]
                                            if counter:
                                                crate_df = melted_conter(crate_df)
                                            try:
                                                del crate_df['Command']
                                            except:
                                                pass
                                            try:
                                                del crate_df['Baseline Check']
                                            except:
                                                pass
                                            try:
                                                if ROP_ACT[0] == 'ROP by ROP' and 'Time' not in crate_df.columns:
                                                    crate_df['Time'] = 'Single-Rop (' + str(START_TIME_CHECK[0]) + '-' + str(END_TIME_CHECK[0]) + ')'
                                                    crate_df.insert(1, 'Time', crate_df.pop('Time'))
                                            except:
                                                pass
                                            if crate_txt_df.df_fram!= 'error':
                                                tabulate.PRESERVE_WHITESPACE = True
                                                data_list = crate_df.to_dict(orient='records')
                                                table = tabulate(data_list, headers='keys', tablefmt='presto')
                                                table = cleeend_text(table)
                                                self.update_textbox_big_window(table, iii)
                                                try:
                                                    self.reset_timer(None)
                                                except:
                                                    pass
                                                if 'done' not in processdone:
                                                    processdone.append('done')
                                                if not crate_df.empty:
                                                    Save_log_switch = self.swich_Sav_loss.get()
                                                    if Save_log_switch == 1:
                                                        Save_logfile = 'YES'
                                                    else:  # inserted
                                                        Save_logfile = 'NO'
                                                    if Save_logfile == 'YES' and chk_oss!= 'LOGIN_CNM' and (chk_oss!= 'no'):
                                                        try:
                                                            del crate_df['Command']
                                                        except:
                                                            pass
                                                        try:
                                                            try:
                                                                crate_df = crate_df.applymap(lambda x: str(x).strip())
                                                            except:
                                                                pass
                                                            ALL_DF.append(crate_df)
                                                        except:
                                                            continue
                            try:
                                combined_df = pd.concat(ALL_DF, ignore_index=True)
                                combined_df.replace('', pd.NA, inplace=True)
                                try:
                                    required_output_df = combined_df.groupby(['Node Id', 'Object'], as_index=False).first()
                                except:
                                    try:
                                        required_output_df = combined_df.groupby(['Node Id', 'Time', 'Object'], as_index=False).first()
                                    except:
                                        try:
                                            required_output_df = combined_df.groupby(['Node Id', 'Time', 'Cell Name'], as_index=False).first()
                                        except:
                                            required_output_df = combined_df.groupby(['Node Id', 'Cell Name'], as_index=False).first()
                                write_excel_new_monitoring(required_output_df, str(techeccci) + '_' + 'KPI Monitoring ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'Initial KPI', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KPI Monitoring_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\')
                                try:
                                    del required_output_df
                                except:
                                    pass
                            except:
                                pass
                            try:
                                del Final_log_lst
                            except:
                                pass
                            try:
                                del START_TIME_CHECK
                            except:
                                pass
                            try:
                                del END_TIME_CHECK
                            except:
                                pass
                            try:
                                del command_not_found
                            except:
                                pass
                            combined_cmd = ''

    def send_command(remote_conn, command, chk_oss, iii, cmd_node, Save_logfile, cmd_ex, MOTASK):
        read_received_data_auto.fil = 'None'
        remote_conn.send(command + '\n')
        read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK)
        send_command.fil = read_received_data_auto.fil

    def main(logcommand_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK):
        iii = 1
        paaaawe = None
        try:
            llll = RRSG_auto_login.remote_conn
            paaaawe = 'ok'
        except:
            try:
                con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                del Rdf['indexxx']
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                PID_LST = Rdf['value'].tolist()
                remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
            except:
                RSG_LST_N = []
            if not RSG_LST_N:
                showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 10000)
            else:  # inserted
                paaaawe = OTPPP_tets()
            if paaaawe == None and 'cancle' not in processdone:
                processdone.append('cancle')
            try:
                if VPN_select_METHOD_F[0]!= 'VPN-CAS':
                    if is_non_numeric_string(paaaawe):
                        showMessage_qt('Wrong OTP password please try again.', 10000)
                    else:  # inserted
                        if paaaawe:
                            self.Auto_login_sc(paaaawe, RSG_LST_N)
                else:  # inserted
                    if paaaawe:
                        self.Auto_login_sc(paaaawe, RSG_LST_N)
            except:
                pass
        try:
            chk_rssg = RRSG_auto_login.remote_conn
            chk_rssg_chk = 'OK'
        except:
            chk_rssg_chk = 'NOK'
        if chk_rssg_chk == 'NOK':
            clear_all_lists()
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Something wrong in RSG/VPN connection please try again after some time.') + '\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        else:  # inserted
            if not CHECK_ISF_SEQUENCE:
                sig = os.getlogin().lower()
                priority = 'High'
                ISF_API_Based = 'SharePoint'
                KEY = '5b97555705ce40128131d58c8d3596f9'
                try:
                    con = sqlite3.connect('./res/ISFdetails.db')
                    Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                    con.close()
                    del Rdf['indexxx']
                    ppPID_LST = Rdf['value'].tolist()
                    ppPID_LST2 = []
                    for ppp in ppPID_LST:
                        ppPID_LST2.append(decrypt(ppp, 3))
                    del Rdf['value']
                    Rdf['value'] = ppPID_LST2
                    PID_LST = Rdf['value'].tolist()
                    remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                    ISF_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                    try:
                        del Rdf
                    except:
                        pass
                except:
                    ISF_LST = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                chk_line = []
                for llll in ISF_LST:
                    if 'line' in llll.lower():
                        chk_line.append('stop_process')
                        break
                    if 'start_process' not in chk_line:
                        chk_line.append('start_process')
                if 'start_process' in chk_line:
                    projectID = ISF_LST[0]
                    type1 = ISF_LST[1]
                    nodname = ISF_LST[2]
                    wOName = ISF_LST[3]
                    executionPlanName = ISF_LST[4]
                    lastModifiedBy = ISF_LST[5]
                    ISF_STEP_START_stepID = ISF_LST[6]
                    ISF_STEP_START_Task_Id = ISF_LST[7]
                    if not ISF_STEP_START_Task_Id_lst:
                        ISF_STEP_START_Task_Id_lst.append(ISF_STEP_START_Task_Id)
                    ISF_WO_CREATE_woId, STEP_START = ISF_API(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY, ISF_STEP_START_stepID, ISF_STEP_START_Task_Id)
                    WO_LSTT = [ISF_WO_CREATE_woId, STEP_START]
                    ISF_STTUS_MSG = 'sucess'
                    if WO_LSTT[0] == 'Some issue in isf workflow':
                        ISF_STTUS_MSG = 'ISF details not correct. please check and try again.'
                    if ISF_STTUS_MSG == 'sucess' and WO_LSTT[1]!= 'sucess':
                        ISF_STTUS_MSG = 'TaskID and StepID not correct. please provide the valid taskID for the StepID.'
            else:  # inserted
                ISF_STTUS_MSG = 'sucess'
            if ISF_STTUS_MSG == 'sucess' and (not CHECK_ISF_SEQUENCE):
                if 'work_order_raised' not in CHECK_ISF_SEQUENCE:
                    CHECK_ISF_SEQUENCE.append('work_order_raised')

                def sucss_msg(self):
                    self.textbox2.delete(1.0, 'end')
                    self.after(0, self.textbox2.insert, 'end', 'ISF Work Order created.\n\n')
                    self.after(0, self.textbox2.insert, 'end', 'Signum    : ' + str(sig) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'Nodtype   : ' + str(type1) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'Nodname   : ' + str(nodname) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'ProjectID : ' + str(projectID) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'WOID      : ' + str(ISF_WO_CREATE_woId) + '\n')
                    try:
                        self.after(0, self.textbox2.insert, 'end', 'Time      : ' + str(ISF_API.timestr) + '\n\n')
                    except:
                        pass
                    try:
                        self.after(0, self.textbox2.insert, 'end', str(ISF_API.urlcreateWO) + '\n\n')
                    except:
                        return None
                statusAUDITLOGPATH = './res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log'
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Signum__' + str(sig) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Nodtype__' + str(type1) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Nodname__' + str(nodname) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'ProjectID__' + str(projectID) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'WOID__' + str(ISF_WO_CREATE_woId) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Time__' + str(ISF_API.timestr) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                threading.Thread(target=sucss_msg, args=(self,)).start()
            if 'work_order_raised' in CHECK_ISF_SEQUENCE:
                KGETFILE = []
                if paaaawe!= None:
                    send_command.fil = 'None'
                    if 'FAIL' not in establish_ssh_conn_lst:
                        for cmd_node in Node_cmd_lst:
                            self.textbox_LIVECMD.delete(1.0, 'end')
                            chk_oss = 'NODE_CNM'
                            cmd_ex_sh = ''
                            send_command(RRSG_auto_login.remote_conn, 'amos ' + str(cmd_node), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            send_command(RRSG_auto_login.remote_conn, 'lt all', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            RBS_RBS_switch = self.swich_Sec_rbs.get()
                            if RBS_RBS_switch == 1:
                                RBS_RBS_switch_loe = 'YES'
                            else:  # inserted
                                RBS_RBS_switch_loe = 'NO'
                            if RBS_RBS_switch_loe == 'YES':
                                if RBSPASS_CRED_CRREDD[0] == 'rbs':
                                    user_rbs = 'rbs'
                                else:  # inserted
                                    try:
                                        user_rbs = RBSPASS_CRED_CRREDD[0]
                                    except:
                                        user_rbs = 'rbs'
                                if RBSPASS_CRED_CRREDD[1] == 'rbs':
                                    pass_rbs = 'rbs'
                                else:  # inserted
                                    try:
                                        pass_rbs = RBSPASS_CRED_CRREDD[1]
                                    except:
                                        pass_rbs = 'rbs'
                                send_command(RRSG_auto_login.remote_conn, str(user_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                                send_command(RRSG_auto_login.remote_conn, str(pass_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            for cmd_ex in oss_cmd_lst:
                                cmd_ex_sh = clean_sheet_name(cmd_ex, max_length=28)
                                chk_oss = 'OSS_CNM'
                                send_command(RRSG_auto_login.remote_conn, cmd_ex, chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                                iii = iii + 1
                            send_command(RRSG_auto_login.remote_conn, 'quit', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            if MOTASK == 'KGET':
                                KGETFILE.append(str(cmd_node) + '$' + send_command.fil)
                        if MOTASK == 'KGET':
                            if KGETFILE:
                                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Kget processing ongoing !\n')
                                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                start_index = '1.0'
                                while True:
                                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                    if not start_index:
                                        break
                                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                    start_index = end_index
                                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                threads = []
                                for kk in KGETFILE:
                                    cmd_node = str(kk).split('$')[0].strip()
                                    node_path = str(kk).split('$')[(-1)].strip()
                                    self.update_textbox(f'Please wait {cmd_node} Kget underprocessing. \n\n')
                                    try:
                                        self.reset_timer(None)
                                    except:
                                        pass
                                    t = threading.Thread(target=KGET_PARSER_AUDIT, args=(self, node_path, cmd_node, techeccci))
                                    t.start()
                                    threads.append(t)
                                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str(cmd_node) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                                for t in threads:
                                    t.join()
                                dump_path = os.path.dirname(node_path)
                                TECHNO = techeccci
                                BASE_INPATH = 'C:\\00_OSS_CHAT_BOT\\'
                                thread_combined_xl = threading.Thread(target=combined_xl_audit, args=(dump_path, TECHNO))
                                thread_combined_xl.start()
                                thread_combined_xl.join()
                                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Parameters baseline check ongoing!\n')
                                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                start_index = '1.0'
                                while True:
                                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                    if not start_index:
                                        break
                                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                    start_index = end_index
                                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                Dump1_Name = str(TECHNO) + '_' + str('Combined_KGET_Dump') + '_' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                                thread_ENM_BASELINE = threading.Thread(target=ENM_BASELINE_AUDIT, args=('Baseline.xlsx', Dump1_Name, TECHNO, dump_path + '\\', dump_path + '\\', BASE_INPATH))
                                thread_ENM_BASELINE.start()
                                thread_ENM_BASELINE.join()
                                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str('Parameter baseline audit') + '__' + str('Multi') + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                                if KGETFILE:
                                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Parameter Audit completed !\n')
                                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                    start_index = '1.0'
                                    while True:
                                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                        if not start_index:
                                            break
                                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                        start_index = end_index
                                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                    time.sleep(2)
                                self.deiconify()
                            self.update_textbox('Kget logs converted into excel check in output folder. \n\n')
                            try:
                                self.reset_timer(None)
                            except:
                                return None
                        else:  # inserted
                            pass
            else:  # inserted
                self.textbox2.delete(1.0, 'end')
                self.after(0, self.textbox2.insert, 'end', ISF_STTUS_MSG + '\n\n')
    if SQL_METHOD == 'LOCAL':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check(user_signum, command_to_barred_required, command_to_barred)
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)
    if SQL_METHOD == 'AZURE_CLOUDE':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            if change_User:
                ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check_AZURE(change_User, command_barred, user_signum, command_to_barred_required, command_to_barred)
            else:  # inserted
                if 'amos' in command_to_barred.lower():
                    ucercmd_AUTHENTICATION = 'User_Authenticated'
                else:  # inserted
                    if 'rbs' in command_to_barred.lower():
                        ucercmd_AUTHENTICATION = 'User_Authenticated'
                    else:  # inserted
                        if 'lt all' in command_to_barred.lower():
                            ucercmd_AUTHENTICATION = 'User_Authenticated'
                        else:  # inserted
                            ucercmd_AUTHENTICATION = 'NO'
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)

def RRSG_Audit_process(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, techeccci, activity_type, PATHHHH_TIME_AUDIT, LSTTT_cmd, RBS_RNC):
    try:
        os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT))
    except:
        pass
    try:
        os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci))
    except:
        pass
    if 'login_error' in LOGINERROR_LST:
        return
    check_lst = []
    CIPRILST = []

    def read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK):
        commnd_sh = []
        commnd_sh.append(cmd_ex)

        def remove_color_codes(input_string):
            ansi_escape = re.compile('\\x1b[^m]*m')
            clean_string = ansi_escape.sub('', input_string)
            html_color = re.compile('<[^>]*>')
            clean_string = html_color.sub('', clean_string)
            return clean_string
        if MOTASK == 'KGET':
            read_received_data_auto.out = ''
            output_B = b''
            while not output_B.endswith(b'$'):
                output_B = remote_conn.recv(999999)
                write_t = output_B.decode(encoding='utf-8')
                write_te_1 = remove_color_codes(write_t)
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('\a', '')
                self.update_textbox(f'{write_te_1}\n')
                try:
                    self.reset_timer(None)
                except:
                    pass
                if Save_logfile == 'YES':
                    with open(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\' + techeccci + '_' + str(cmd_node) + '_KGET.Log', 'a+', encoding='utf-8') as f:
                        f.writelines(f'{write_te_1}\n\n')
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str(cmd_node) + '__' + str(command) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                last_line = str(write_te_1.strip().split('\n')[(-1)]).split(' ')[0].strip()
                try:
                    last_line2 = str(write_te_1.strip().split('\n')[(-1)])[(-1)]
                except:
                    last_line2 = '=============='
                    pass
                if last_line == 'Total:':
                    break
                if last_line == (str(cmd_node).strip() + '>').strip():
                    break
                if last_line2 == '$':
                    break
                if last_line == 'e:':
                    break
                if last_line2 == ':':
                    break
            read_received_data_auto.fil = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\' + str(techeccci) + '_' + str(cmd_node) + '_KGET.Log'
        if MOTASK == 'MOBATCH':
            read_received_data_auto.out = ''
            output_B = b''
            combined_cmd = ''
            command_not_found = []
            while not (output_B.endswith(b'> ') or output_B.endswith(b'<') or output_B.endswith(b': ') or output_B.endswith(b'<') or output_B.endswith(b'< ') or output_B.endswith(b':') or output_B.endswith(b'$ ')):
                output_B += remote_conn.recv(99999)
                write_t = output_B.decode(encoding='utf-8')
                write_te_1 = remove_color_codes(write_t)
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('\a', '')
                self.update_textbox(f'{write_te_1}\n')
                if cmd_ex == 'sdir' and 'CPRI links' in write_te_1 and (not CIPRILST):
                    CIPRILST.append(write_te_1)
                try:
                    self.reset_timer(None)
                except:
                    pass
                if 'no such command' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'command not found' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                write_te_1 = '\n'.join([line.replace('coli>/rc/nrat/ue print -admitted', '') for line in write_te_1.split('\n')])
                write_te_1 = '\n'.join([line.replace('NRAT is dormant.', '') for line in write_te_1.split('\n')])
                if 'ue print -admitted'.lower() in command.lower():
                    write_te_1 = '\n'.join([line.replace('coli>', '') for line in write_te_1.split('\n')])
                if 'coli>' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'amos' in command and 'AMOS error' in write_te_1 and ('no such command' not in command_not_found):
                    command_not_found.append('no such command')
                last_line = output_B.decode(encoding='utf-8').split('\n')[(-1)]
                last_line = ' '.join(last_line.split()).strip()
                if check_lst and last_line == check_lst[0]:
                    read_received_data_auto.brek = 'break'
                try:
                    if last_line[(-1)].strip() == '>':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '<':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == ':':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '$':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
            if not command_not_found:
                write_text1 = output_B.decode(encoding='utf-8')
                write_text = remove_color_codes(write_text1)
                write_text = write_text.replace('', '')
                write_text = write_text.replace('', '')
                write_text = write_text.replace('\a', '')
                if Save_logfile == 'YES':
                    if chk_oss!= 'LOGIN_CNM' and chk_oss!= 'no':
                        with open(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\' + techeccci + '_' + str(cmd_node) + '.Log', 'a+', encoding='utf-8') as f:
                            f.writelines(f'{write_text}\n\n')
                        with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str(activity_type) + '__' + str(cmd_node) + '__' + str(command) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    if chk_oss == 'OSS_CNM':
                        combined_cmd += f'{write_text}\n'
                crate_txt_df.df_fram = ''
                if chk_oss == 'OSS_CNM':
                    if command!= 'q':
                        if command!= 'quit':
                            def remove_line_containing(text, substring):
                                lines = text.split('\n')
                                filtered_lines = [line for line in lines if substring not in line]
                                return '\n'.join(filtered_lines)
                            combined_cmd = combined_cmd.replace('It is recommended to remove duplicate counter definitions from the PM scanners.', '')
                            combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                            combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo)', 'Date & Time (UTC)   S Specific Problem                    MO (Cause/AdditionalInfo)')
                            combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo) Operator', '')
                            combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                            combined_cmd = remove_line_containing(combined_cmd, 'ACKNOWLEDGED ALARMS')
                            if 'pmr' in combined_cmd.lower():
                                combined_cmd = remove_line_containing(combined_cmd, 'pmr')
                                combined_cmd = remove_line_containing(combined_cmd, 'PMR')
                            combined_cmd = combined_cmd.strip()
                            if '16Qam'.lower() not in command.lower():
                                combined_cmd = log_clening(combined_cmd, cmd_ex, cmd_node)
                            Final_log_lst = final_log_process(combined_cmd)
                            try:
                                combined_cmd = ''
                            except:
                                pass
                            Final_log_lst = [item.strip() for item in Final_log_lst if item.strip()]
                            try:
                                if Final_log_lst[0] == '\n':
                                    Final_log_lst = []
                            except:
                                pass
                            if Final_log_lst:
                                i = 1
                                Final_log_lst = [item for item in Final_log_lst if not item.startswith('+')]
                                Final_log_lst = [item for item in Final_log_lst if not item.startswith('\n+')]
                                for section in Final_log_lst:
                                    lines = section.strip().split('\n')
                                    data = [line.split('\t') for line in lines]
                                    if len(data) > 1:
                                        crate_df = crate_txt_df(section, cmd_node, cmd_ex)
                                        try:
                                            crate_df.replace('deletestring', '', inplace=True)
                                        except:
                                            pass
                                        if crate_df is not None and (not crate_df.empty):
                                            crate_df = crate_df[~crate_df.apply(lambda row: any(row.astype(str).str.contains('Total:')), axis=1)]
                                            crate_df.columns = crate_df.columns.str.strip()
                                            if not crate_df.empty:
                                                try:
                                                    try:
                                                        crate_df['VSWR (RL)'] = crate_df['VSWR (RL)'].str.strip()
                                                    except:
                                                        pass
                                                    VSWR = [col for col in crate_df.columns if 'VSWR' in col]
                                                    if VSWR:
                                                        try:
                                                            crate_df[['VSWR', 'RL']] = crate_df['VSWR_(RL)'].str.split('_', expand=True)
                                                        except:
                                                            crate_df[['VSWR', 'RL']] = crate_df['VSWR (RL)'].str.split(' ', expand=True)
                                                        crate_df['RL'] = crate_df['RL'].str.strip('()')
                                                        crate_df['VSWR'] = crate_df['VSWR'].astype(float)
                                                        crate_df['RL'] = crate_df['RL'].astype(float)
                                                        try:
                                                            del crate_df['VSWR_(RL)']
                                                        except:
                                                            try:
                                                                del crate_df['VSWR (RL)']
                                                            except:
                                                                pass
                                                except:
                                                    pass
                                                try:
                                                    crate_df = crate_df.applymap(lambda x: x.replace('_', ' '))
                                                except:
                                                    pass
                                                try:
                                                    crate_df['Node Id'] = cmd_node
                                                except:
                                                    pass
                                                try:
                                                    crate_df.insert(0, 'Node Id', crate_df.pop('Node Id'))
                                                except:
                                                    pass
                                                excel_df = crate_df.copy()
                                                counter = [col for col in excel_df.columns if 'counter' in col.lower()]
                                                if counter:
                                                    excel_df = melted_conter(excel_df)
                                                try:
                                                    del crate_df['Command']
                                                except:
                                                    pass
                                                try:
                                                    del crate_df['Baseline Check']
                                                except:
                                                    pass
                                                if crate_txt_df.df_fram!= 'error':
                                                    tabulate.PRESERVE_WHITESPACE = True
                                                    data_list = crate_df.to_dict(orient='records')
                                                    table = tabulate(data_list, headers='keys', tablefmt='presto')
                                                    table = cleeend_text(table)
                                                    self.update_textbox_big_window(table, iii)
                                                    try:
                                                        self.reset_timer(None)
                                                    except:
                                                        pass
                                                    if 'done' not in processdone:
                                                        processdone.append('done')
                                                    if not excel_df.empty:
                                                        Save_log_switch = self.swich_Sav_loss.get()
                                                        if Save_log_switch == 1:
                                                            Save_logfile = 'YES'
                                                        else:  # inserted
                                                            Save_logfile = 'NO'
                                                        if Save_logfile == 'YES' and chk_oss!= 'LOGIN_CNM' and (chk_oss!= 'no'):
                                                            try:
                                                                del excel_df['Command']
                                                            except:
                                                                pass
                                                            try:
                                                                if len(Final_log_lst) > 1:
                                                                    cmd_ex = commnd_sh[0] + str(i)
                                                                else:  # inserted
                                                                    pass
                                                            except:
                                                                pass
                                                                try:
                                                                    if commnd_sh[0] == 'sdir':
                                                                        cmd_ex = 'sdir'
                                                                except:
                                                                    pass
                                                                try:
                                                                    if commnd_sh[0] == 'cabex':
                                                                        cmd_ex = 'cabex'
                                                                except:
                                                                    pass
                                                                try:
                                                                    if commnd_sh[0] == 'sdi':
                                                                        cmd_ex = 'sdi'
                                                                except:
                                                                    pass
                                                                try:
                                                                    excel_df = excel_df.applymap(lambda x: str(x).strip())
                                                                except:
                                                                    pass
                                                                try:
                                                                    if techeccci == 'WCDMA':
                                                                        if RBS_RNC!= 'KGET':
                                                                            write_excel_new(excel_df, str(techeccci) + '_' + 'Audit_Output ' + str(RBS_RNC) + ' ' + datetime.now().strftime('%d%m%Y') + '.xlsx', cmd_ex, OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\')
                                                                        if RBS_RNC == 'KGET':
                                                                            write_excel_new(excel_df, str(techeccci) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', cmd_ex, OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\')
                                                                    else:  # inserted
                                                                        write_excel_new(excel_df, str(techeccci) + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', cmd_ex, OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\')
                                                                    try:
                                                                        del excel_df
                                                                    except:
                                                                        pass
                                                                except:
                                                                    pass
                                                            except:
                                                                continue
                                try:
                                    if command.strip().lower() == 'sdir':
                                        CIPRI_DFF = CIPRIDFFF(CIPRILST[0])
                                        CIPRILST.clear()
                                        if CIPRI_DFF is not None and (not CIPRI_DFF.empty):
                                            try:
                                                CIPRI_DFF['Node Id'] = cmd_node
                                            except:
                                                pass
                                            try:
                                                CIPRI_DFF.insert(0, 'Node Id', CIPRI_DFF.pop('Node Id'))
                                            except:
                                                pass
                                            if not CIPRI_DFF.empty:
                                                try:
                                                    CIPRI_DFF = CIPRI_DFF.applymap(lambda x: str(x).strip())
                                                except:
                                                    pass
                                                try:
                                                    def conditione(row):
                                                        if int(row['CPRI Link NOK']) > 0:
                                                            return 'Fail'
                                                        if int(row['CPRI Link OKW']) > 0:
                                                            return 'Fail'
                                                        return 'Pass'
                                                    CIPRI_DFF['Baseline Check'] = CIPRI_DFF.apply(conditione, axis=1)
                                                except:
                                                    pass
                                                if techeccci == 'WCDMA':
                                                    if RBS_RNC!= 'KGET':
                                                        write_excel_new(CIPRI_DFF, techeccci + '_' + 'Audit_Output ' + str(RBS_RNC) + ' ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'CPRI_link', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\')
                                                    if RBS_RNC == 'KGET':
                                                        write_excel_new(CIPRI_DFF, techeccci + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'CPRI_link', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\')
                                                else:  # inserted
                                                    write_excel_new(CIPRI_DFF, techeccci + '_' + 'Audit_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'CPRI_link', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('Health check and audit_') + str(PATHHHH_TIME_AUDIT) + '\\' + str(techeccci) + '\\')
                                                tabulate.PRESERVE_WHITESPACE = True
                                                data_list = CIPRI_DFF.to_dict(orient='records')
                                                table = tabulate(data_list, headers='keys', tablefmt='presto')
                                                table = cleeend_text(table)
                                                CIPRI_PRINT(self, table)
                                            try:
                                                del CIPRI_DFF
                                            except:
                                                pass
                                except:
                                    pass
                                try:
                                    del Final_log_lst
                                except:
                                    pass
                                try:
                                    del command_not_found
                                except:
                                    pass
                                combined_cmd = ''

    def send_command(remote_conn, command, chk_oss, iii, cmd_node, Save_logfile, cmd_ex, MOTASK):
        read_received_data_auto.fil = 'None'
        remote_conn.send(command + '\n')
        read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK)
        send_command.fil = read_received_data_auto.fil

    def main(logcommand_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK):
        iii = 1
        paaaawe = None
        try:
            llll = RRSG_auto_login.remote_conn
            paaaawe = 'ok'
        except:
            try:
                con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                del Rdf['indexxx']
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                PID_LST = Rdf['value'].tolist()
                remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
            except:
                RSG_LST_N = []
            if not RSG_LST_N:
                showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 10000)
            else:  # inserted
                paaaawe = OTPPP_tets()
            if paaaawe == None and 'cancle' not in processdone:
                processdone.append('cancle')
            try:
                if VPN_select_METHOD_F[0]!= 'VPN-CAS':
                    if is_non_numeric_string(paaaawe):
                        showMessage_qt('Wrong OTP password please try again.', 10000)
                    else:  # inserted
                        if paaaawe:
                            self.Auto_login_sc(paaaawe, RSG_LST_N)
                else:  # inserted
                    if paaaawe:
                        self.Auto_login_sc(paaaawe, RSG_LST_N)
            except:
                pass
        try:
            chk_rssg = RRSG_auto_login.remote_conn
            chk_rssg_chk = 'OK'
        except:
            chk_rssg_chk = 'NOK'
        if chk_rssg_chk == 'NOK':
            clear_all_lists()
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Something wrong in RSG/VPN connection please try again after some time.') + '\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        else:  # inserted
            if not CHECK_ISF_SEQUENCE:
                sig = os.getlogin().lower()
                priority = 'High'
                ISF_API_Based = 'SharePoint'
                KEY = '5b97555705ce40128131d58c8d3596f9'
                try:
                    con = sqlite3.connect('./res/ISFdetails.db')
                    Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                    con.close()
                    del Rdf['indexxx']
                    ppPID_LST = Rdf['value'].tolist()
                    ppPID_LST2 = []
                    for ppp in ppPID_LST:
                        ppPID_LST2.append(decrypt(ppp, 3))
                    del Rdf['value']
                    Rdf['value'] = ppPID_LST2
                    PID_LST = Rdf['value'].tolist()
                    remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                    ISF_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                    try:
                        del Rdf
                    except:
                        pass
                except:
                    ISF_LST = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                chk_line = []
                for llll in ISF_LST:
                    if 'line' in llll.lower():
                        chk_line.append('stop_process')
                        break
                    if 'start_process' not in chk_line:
                        chk_line.append('start_process')
                if 'start_process' in chk_line:
                    projectID = ISF_LST[0]
                    type1 = ISF_LST[1]
                    nodname = ISF_LST[2]
                    wOName = ISF_LST[3]
                    executionPlanName = ISF_LST[4]
                    lastModifiedBy = ISF_LST[5]
                    ISF_STEP_START_stepID = ISF_LST[6]
                    ISF_STEP_START_Task_Id = ISF_LST[7]
                    if not ISF_STEP_START_Task_Id_lst:
                        ISF_STEP_START_Task_Id_lst.append(ISF_STEP_START_Task_Id)
                    ISF_WO_CREATE_woId, STEP_START = ISF_API(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY, ISF_STEP_START_stepID, ISF_STEP_START_Task_Id)
                    WO_LSTT = [ISF_WO_CREATE_woId, STEP_START]
                    ISF_STTUS_MSG = 'sucess'
                    if WO_LSTT[0] == 'Some issue in isf workflow':
                        ISF_STTUS_MSG = 'ISF details not correct. please check and try again.'
                    if ISF_STTUS_MSG == 'sucess' and WO_LSTT[1]!= 'sucess':
                        ISF_STTUS_MSG = 'TaskID and StepID not correct. please provide the valid taskID for the StepID.'
            else:  # inserted
                ISF_STTUS_MSG = 'sucess'
            if ISF_STTUS_MSG == 'sucess' and (not CHECK_ISF_SEQUENCE):
                if 'work_order_raised' not in CHECK_ISF_SEQUENCE:
                    CHECK_ISF_SEQUENCE.append('work_order_raised')

                def sucss_msg(self):
                    self.textbox2.delete(1.0, 'end')
                    self.after(0, self.textbox2.insert, 'end', 'ISF Work Order created.\n\n')
                    self.after(0, self.textbox2.insert, 'end', 'Signum    : ' + str(sig) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'Nodtype   : ' + str(type1) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'Nodname   : ' + str(nodname) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'ProjectID : ' + str(projectID) + '\n')
                    self.after(0, self.textbox2.insert, 'end', 'WOID      : ' + str(ISF_WO_CREATE_woId) + '\n')
                    try:
                        self.after(0, self.textbox2.insert, 'end', 'Time      : ' + str(ISF_API.timestr) + '\n\n')
                    except:
                        pass
                    try:
                        self.after(0, self.textbox2.insert, 'end', str(ISF_API.urlcreateWO) + '\n\n')
                    except:
                        return None
                statusAUDITLOGPATH = './res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log'
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Signum__' + str(sig) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Nodtype__' + str(type1) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Nodname__' + str(nodname) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'ProjectID__' + str(projectID) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'WOID__' + str(ISF_WO_CREATE_woId) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('Health check and audit') + '__' + 'Time__' + str(ISF_API.timestr) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                threading.Thread(target=sucss_msg, args=(self,)).start()
            if 'work_order_raised' in CHECK_ISF_SEQUENCE:
                KGETFILE = []
                if paaaawe!= None:
                    send_command.fil = 'None'
                    if 'FAIL' not in establish_ssh_conn_lst:
                        for cmd_node in Node_cmd_lst:
                            self.textbox_LIVECMD.delete(1.0, 'end')
                            chk_oss = 'NODE_CNM'
                            cmd_ex_sh = ''
                            send_command(RRSG_auto_login.remote_conn, 'amos ' + str(cmd_node), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            send_command(RRSG_auto_login.remote_conn, 'lt all', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            RBS_RBS_switch = self.swich_Sec_rbs.get()
                            if RBS_RBS_switch == 1:
                                RBS_RBS_switch_loe = 'YES'
                            else:  # inserted
                                RBS_RBS_switch_loe = 'NO'
                            if RBS_RBS_switch_loe == 'YES':
                                if RBSPASS_CRED_CRREDD[0] == 'rbs':
                                    user_rbs = 'rbs'
                                else:  # inserted
                                    try:
                                        user_rbs = RBSPASS_CRED_CRREDD[0]
                                    except:
                                        user_rbs = 'rbs'
                                if RBSPASS_CRED_CRREDD[1] == 'rbs':
                                    pass_rbs = 'rbs'
                                else:  # inserted
                                    try:
                                        pass_rbs = RBSPASS_CRED_CRREDD[1]
                                    except:
                                        pass_rbs = 'rbs'
                                send_command(RRSG_auto_login.remote_conn, str(user_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                                send_command(RRSG_auto_login.remote_conn, str(pass_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            for cmd_ex in oss_cmd_lst:
                                cmd_ex_sh = clean_sheet_name(cmd_ex, max_length=28)
                                chk_oss = 'OSS_CNM'
                                send_command(RRSG_auto_login.remote_conn, cmd_ex, chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                                iii = iii + 1
                            send_command(RRSG_auto_login.remote_conn, 'quit', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            if MOTASK == 'KGET':
                                KGETFILE.append(str(cmd_node) + '$' + send_command.fil)
                        if MOTASK == 'KGET':
                            if KGETFILE:
                                threads = []
                                for kk in KGETFILE:
                                    cmd_node = str(kk).split('$')[0].strip()
                                    node_path = str(kk).split('$')[(-1)].strip()
                                    self.update_textbox(f'Please wait {cmd_node} Kget underprocessing. \n\n')
                                    try:
                                        self.reset_timer(None)
                                    except:
                                        pass
                                    t = threading.Thread(target=KGET_PARSER_AUDIT, args=(self, node_path, cmd_node, techeccci))
                                    t.start()
                                    threads.append(t)
                                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str(cmd_node) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                                for t in threads:
                                    t.join()
                                dump_path = os.path.dirname(node_path)
                                TECHNO = techeccci
                                BASE_INPATH = 'C:\\00_OSS_CHAT_BOT\\'
                                thread_combined_xl = threading.Thread(target=combined_xl_audit, args=(dump_path, TECHNO))
                                thread_combined_xl.start()
                                thread_combined_xl.join()
                                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Parameters baseline check ongoing!\n')
                                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                start_index = '1.0'
                                while True:
                                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                    if not start_index:
                                        break
                                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                    start_index = end_index
                                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                Dump1_Name = str(TECHNO) + '_' + str('Combined_KGET_Dump') + '_' + datetime.now().strftime('%d%m%Y') + '.xlsx'
                                thread_ENM_BASELINE = threading.Thread(target=ENM_BASELINE_AUDIT, args=('Baseline.xlsx', Dump1_Name, TECHNO, dump_path + '\\', dump_path + '\\', BASE_INPATH))
                                thread_ENM_BASELINE.start()
                                thread_ENM_BASELINE.join()
                                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(techeccci) + '__' + str('Parameter baseline audit') + '__' + str('Multi') + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                                if KGETFILE:
                                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Parameter Audit completed !\n')
                                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                    start_index = '1.0'
                                    while True:
                                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                        if not start_index:
                                            break
                                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                        start_index = end_index
                                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                            try:
                                self.reset_timer(None)
                            except:
                                return None
                        else:  # inserted
                            pass
            else:  # inserted
                self.textbox2.delete(1.0, 'end')
                self.after(0, self.textbox2.insert, 'end', ISF_STTUS_MSG + '\n\n')
    if SQL_METHOD == 'LOCAL':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check(user_signum, command_to_barred_required, command_to_barred)
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)
    if SQL_METHOD == 'AZURE_CLOUDE':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            if change_User:
                ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check_AZURE(change_User, command_barred, user_signum, command_to_barred_required, command_to_barred)
            else:  # inserted
                if 'amos' in command_to_barred.lower():
                    ucercmd_AUTHENTICATION = 'User_Authenticated'
                else:  # inserted
                    if 'rbs' in command_to_barred.lower():
                        ucercmd_AUTHENTICATION = 'User_Authenticated'
                    else:  # inserted
                        if 'lt all' in command_to_barred.lower():
                            ucercmd_AUTHENTICATION = 'User_Authenticated'
                        else:  # inserted
                            ucercmd_AUTHENTICATION = 'NO'
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)

def ERCA_PM_Tool():
    check_close = []

    class TemplateFreePMReport(QDialog):
        def __init__(self):
            super().__init__()
            self.user_closed = False
            self.initUI()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def initUI(self):
            self.setFixedSize(500, 600)
            self.center_window()
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowCloseButtonHint)
            main_layout = QHBoxLayout()
            left_layout = QVBoxLayout()
            self.lte_pmx_combo = QComboBox()
            self.lte_pmx_combo.addItem('LTE-PMX')
            self.lte_pmx_combo.addItem('WCDMA-PMX')
            self.lte_pmx_combo.addItem('5G-PMX')
            self.lte_pmx_combo.setStyleSheet('\n                QComboBox {\n                    background-color: #3498db;  /* Dropdown button background color */\n                    color: white;  /* Text color */\n                    border-radius: 5px;  /* Rounded corners */\n                    padding: 5px;  /* Padding inside the dropdown */\n                    font-size: 16px;  /* Font size */\n                    border: 1px solid #2980b9;  /* Border color */\n                }\n                QComboBox::drop-down {\n                    subcontrol-origin: padding;\n                    subcontrol-position: top right;\n                    width: 20px;\n                    border-left-width: 1px;\n                    border-left-color: #2980b9;  /* Border color for the drop-down button */\n                    border-left-style: solid;\n                    border-top-right-radius: 3px;  /* Rounded corner for the top-right */\n                    border-bottom-right-radius: 3px;  /* Rounded corner for the bottom-right */\n                    background-color: #2980b9;  /* Background color for the drop-down button */\n                }\n                QComboBox::down-arrow {\n                    image: url(down-arrow-icon.png);  /* Custom down arrow icon, replace with your own */\n                    width: 12px;\n                    height: 12px;\n                }\n                QComboBox QAbstractItemView {\n                    background-color: #ecf0f1;  /* Dropdown list background color */\n                    color: #2c3e50;  /* Dropdown list text color */\n                    selection-background-color: #2980b9;  /* Background color of selected item */\n                    selection-color: white;  /* Text color of selected item */\n                }\n            ')
            self.node_list_textbox = QTextEdit()
            self.node_list_textbox.setPlaceholderText('Enter Node List')
            self.node_list_textbox.setStyleSheet('font-size: 16px;')
            self.radio_button_group = QButtonGroup(self)
            self.radio_button_1 = QRadioButton('ONLY PM REPORT')
            self.radio_button_2 = QRadioButton('ERCA FORMATE')
            self.radio_button_1.setChecked(True)
            self.radio_button_group.addButton(self.radio_button_1)
            self.radio_button_group.addButton(self.radio_button_2)
            left_layout.addWidget(self.lte_pmx_combo)
            left_layout.addWidget(self.node_list_textbox)
            left_layout.addWidget(self.radio_button_1)
            left_layout.addWidget(self.radio_button_2)
            right_layout = QVBoxLayout()
            self.hourly_combo = QComboBox()
            self.hourly_combo.addItem('15 Minute ROP')
            self.hourly_combo.addItem('SPECIFIC ROP')
            self.hourly_combo.setStyleSheet('\n                            QComboBox {\n                                background-color: #3498db;  /* Dropdown button background color */\n                                color: white;  /* Text color */\n                                border-radius: 5px;  /* Rounded corners */\n                                padding: 5px;  /* Padding inside the dropdown */\n                                font-size: 16px;  /* Font size */\n                                border: 1px solid #2980b9;  /* Border color */\n                            }\n                            QComboBox::drop-down {\n                                subcontrol-origin: padding;\n                                subcontrol-position: top right;\n                                width: 20px;\n                                border-left-width: 1px;\n                                border-left-color: #2980b9;  /* Border color for the drop-down button */\n                                border-left-style: solid;\n                                border-top-right-radius: 3px;  /* Rounded corner for the top-right */\n                                border-bottom-right-radius: 3px;  /* Rounded corner for the bottom-right */\n                                background-color: #2980b9;  /* Background color for the drop-down button */\n                            }\n                            QComboBox::down-arrow {\n                                image: url(down-arrow-icon.png);  /* Custom down arrow icon, replace with your own */\n                                width: 12px;\n                                height: 12px;\n                            }\n                            QComboBox QAbstractItemView {\n                                background-color: #ecf0f1;  /* Dropdown list background color */\n                                color: #2c3e50;  /* Dropdown list text color */\n                                selection-background-color: #2980b9;  /* Background color of selected item */\n                                selection-color: white;  /* Text color of selected item */\n                            }\n                        ')
            start_hour_label = QLabel('Start Hour')
            start_hour_label.setStyleSheet('font-size: 16px;')
            self.start_hour_spinbox = QSpinBox()
            self.start_hour_spinbox.setRange(0, 23)
            self.start_hour_spinbox.setStyleSheet('font-size: 16px;')
            end_hour_label = QLabel('End Hour')
            end_hour_label.setStyleSheet('font-size: 16px;')
            self.end_hour_spinbox = QSpinBox()
            self.end_hour_spinbox.setRange(0, 23)
            self.end_hour_spinbox.setValue(1)
            self.end_hour_spinbox.setStyleSheet('font-size: 16px;')
            date_layout = QHBoxLayout()
            start_date_layout = QVBoxLayout()
            start_date_label = QLabel('Start Date :')
            start_date_label.setStyleSheet('font-size: 16px;')
            self.start_date_edit = QDateEdit()
            self.start_date_edit.setDate(QDate.currentDate())
            self.start_date_edit.setStyleSheet('font-size: 16px;')
            start_date_layout.addWidget(start_date_label)
            start_date_layout.addWidget(self.start_date_edit)
            end_date_layout = QVBoxLayout()
            end_date_label = QLabel('End Date :')
            end_date_label.setStyleSheet('font-size: 16px;')
            self.end_date_edit = QDateEdit()
            self.end_date_edit.setDate(QDate.currentDate())
            self.end_date_edit.setStyleSheet('font-size: 16px;')
            end_date_layout.addWidget(end_date_label)
            end_date_layout.addWidget(self.end_date_edit)
            date_layout.addLayout(start_date_layout)
            date_layout.addLayout(end_date_layout)
            start_button = QPushButton('START')
            start_button.setStyleSheet('font-size: 14px; padding: 10px 20px;')
            start_button.clicked.connect(self.get_all_field_values)
            CLOSE_button = QPushButton('CLOSE')
            CLOSE_button.setStyleSheet('font-size: 14px; padding: 10px 20px;')
            CLOSE_button.clicked.connect(self.CLOSE_button_click)
            right_layout.addWidget(self.hourly_combo)
            right_layout.addStretch()
            right_layout.addWidget(start_hour_label)
            right_layout.addWidget(self.start_hour_spinbox)
            right_layout.addWidget(end_hour_label)
            right_layout.addWidget(self.end_hour_spinbox)
            right_layout.addStretch()
            right_layout.addLayout(date_layout)
            right_layout.addStretch()
            right_layout.addWidget(CLOSE_button)
            right_layout.addWidget(start_button)
            main_layout.addLayout(left_layout)
            main_layout.addLayout(right_layout)
            self.setLayout(main_layout)
            self.setWindowTitle('Template-Free PM Report')
            style = 'QLineEdit { background-color: #dcf5e3; border: 2px solid #4CAF50; border-radius: 8px; padding: 2px; font-size: 19px; font-family: Segoe UI Semibold;}QPushButton { background-color: #4CAF51; color: white; border: none; border-radius: 8px; padding: 8px; font-size: 18px; }'
            self.setStyleSheet(style)

        def CLOSE_button_click(self):
            self.close()
            check_close.append('cloose')

        def get_all_field_values(self):
            if 'cloose' not in check_close:
                node_list_textbox_value = self.node_list_textbox.toPlainText().strip()
                if not node_list_textbox_value:
                    QMessageBox.warning(self, 'Input Error', 'Node List is empty! Please enter a node list before continuing.')
                    return
                lte_pmx_value = self.lte_pmx_combo.currentText()
                node_list_textbox_value = self.node_list_textbox.toPlainText().splitlines()
                hourly_combo_value = self.hourly_combo.currentText()
                start_hour_value = self.start_hour_spinbox.value()
                end_hour_value = self.end_hour_spinbox.value()
                start_date_value = self.start_date_edit.date().toString('yyyy-MM-dd')
                end_date_value = self.end_date_edit.date().toString('yyyy-MM-dd')
                selected_radio_button = self.radio_button_group.checkedButton().text() if self.radio_button_group.checkedButton() else None
                self.close()
                return (lte_pmx_value, node_list_textbox_value, hourly_combo_value, start_hour_value, end_hour_value, start_date_value, end_date_value, selected_radio_button)
    app = QApplication(sys.argv)
    dialog = TemplateFreePMReport()
    dialog.exec_()
    return dialog.get_all_field_values()

def module_setting(self):
    clossd_buton = []

    def center_window(window):
        self.thread_running = False
        window.update_idletasks()
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100
        if scaleFactor == 1.5:
            width = 180
            height = 200
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2.7
            window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
            return
        width = 240
        height = 230
        x = (screen_width - width + 20) // 2
        y = (screen_height - height - 300) // 2
        window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

    def on_close():
        if clossd_buton or not self.thread_running:
            self.thread_running = True
            threading.Thread(target=enabled_button, args=(self,)).start()
    customtkinter.set_appearance_mode(MODEEE)
    customtkinter.set_default_color_theme('green')
    app = customtkinter.CTk()
    tool_setting.app = app
    app.geometry('350x200')
    app.title('Tool Setting')
    center_window(app)
    app.bind('<Destroy>', lambda e: on_close())

    def button_function():
        if 'ididi' not in clossd_buton:
            clossd_buton.append('ididi')

        def delayed_destr22():
            if self.winfo_exists():
                app.destroy()
            try:
                clossd_buton.clear()
                del clossd_buton
            except:
                return None
        self.after(100, delayed_destr22)
        try:
            if amos_node:
                try:
                    eyye = RRSG_auto_login.remote_conn
                    threading.Thread(target=RRSG_enter_command, args=(self, ['quit'], 'chatbuddy')).start()
                except:
                    pass
        except:
            pass

        def START_FUNNNN(self):
            COMMAND_DEFALT_AUTO = []
            selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST = AUDIT_main_SELECT(COMMAND_DEFALT_AUTO)
            try:
                GSM_NODE_LST = [item.upper() for item in GSM_NODE_LST]
            except:
                pass
            try:
                WCDMA_NODE_LST = [item.upper() for item in WCDMA_NODE_LST]
            except:
                pass
            try:
                LTE_NODE_LST = [item.upper() for item in LTE_NODE_LST]
            except:
                pass
            try:
                NR_NODE_LST = [item.upper() for item in NR_NODE_LST]
            except:
                pass
            DATABASE = 'CAUDIT_CLOSE_WINDOW'
            DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
            if 'close' not in DATBASE_VALUE_READ:
                try:
                    DATBASE_VALUE_READ.clear()
                    del DATBASE_VALUE_READ
                except:
                    pass
                DATABASE = 'ACTIVITYY'
                DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                RRSG_Audit_Module(self, selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST, DATBASE_VALUE_READ[0], 'UIBOOT')
                try:
                    DATBASE_VALUE_READ.clear()
                    del DATBASE_VALUE_READ
                except:
                    return None
            else:  # inserted
                thread = threading.Thread(target=enabled_button, args=(self,))
                thread.start()
                thread.join()
                DATABASE = 'CAUDIT_CLOSE_WINDOW'
                SQLITE_DB_CLEAR(DATABASE)
                DATABASE = 'AUDDIT_START'
                SQLITE_DB_CLEAR(DATABASE)
                DATABASE = 'ACTIVITYY'
                SQLITE_DB_CLEAR(DATABASE)
                GSM_LST.clear()
                LTE_LST.clear()
                WCDMA_LST.clear()
                NR_LST.clear()
                selected_checkboxes.clear()
                GSM_NODE_LST.clear()
                LTE_NODE_LST.clear()
                WCDMA_NODE_LST.clear()
                NR_NODE_LST.clear()
                COMMAND_DEFALT_AUTO.clear()
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Thanks! how Chat Buddy will assist you !') + '\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        threading.Thread(target=START_FUNNNN, args=(self,)).start()

    def button_function2():
        if 'ididi' not in clossd_buton:
            clossd_buton.append('ididi')

        def delayed_destr22():
            if self.winfo_exists():
                app.destroy()
            try:
                clossd_buton.clear()
                del clossd_buton
            except:
                return None
        self.after(100, delayed_destr22)

        def ERCA_PM_Tool_new():
            try:
                lte_pmx_value, node_list_value, hourly_combo_value, start_hour_value, end_hour_value, start_date_value, end_date_value, selected_radio_button = ERCA_PM_Tool()
                RRSG_PMX_Module(self, lte_pmx_value, node_list_value, hourly_combo_value, start_hour_value, end_hour_value, start_date_value, end_date_value, selected_radio_button, mfolder_create, 'UIBOOT')
            except:
                pass
            thread = threading.Thread(target=enabled_button, args=(self,))
            thread.start()
            thread.join()
        threading.Thread(target=ERCA_PM_Tool_new).start()

    def button_function3():
        app.withdraw()
        ISFLoginApp_fun()
        app.deiconify()
    execution_count = [0]

    def button_function4():
        try:
            if amos_node:
                try:
                    eyye = RRSG_auto_login.remote_conn
                    threading.Thread(target=RRSG_enter_command, args=(self, ['quit'], 'chatbuddy')).start()
                except:
                    pass
        except:
            pass
        DATABASE = 'CAUDIT_CLOSE_WINDOW'
        SQLITE_DB_CLEAR(DATABASE)
        DATABASE = 'AUDDIT_START'
        SQLITE_DB_CLEAR(DATABASE)
        DATABASE = 'ACTIVITYY'
        SQLITE_DB_CLEAR(DATABASE)
        GSM_LST.clear()
        LTE_LST.clear()
        WCDMA_LST.clear()
        NR_LST.clear()
        selected_checkboxes.clear()
        GSM_NODE_LST.clear()
        LTE_NODE_LST.clear()
        WCDMA_NODE_LST.clear()
        NR_NODE_LST.clear()
        COMMAND_DEFALT_AUTO.clear()
        if 'ididi' not in clossd_buton:
            clossd_buton.append('ididi')

        def delayed_destr22():
            if self.winfo_exists():
                app.destroy()
            try:
                clossd_buton.clear()
                del clossd_buton
            except:
                return None
        self.after(100, delayed_destr22)

        def START_FUNNNN():
            COMMAND_DEFALT_AUTO = []
            FREQUENCY_count_LST.clear()
            FREQUENCY_LST.clear()
            EMAIL_METHOOD_LST.clear()
            selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST = Monitoring_main_SELECT(COMMAND_DEFALT_AUTO)
            try:
                GSM_NODE_LST = [item.upper() for item in GSM_NODE_LST]
            except:
                pass
            try:
                WCDMA_NODE_LST = [item.upper() for item in WCDMA_NODE_LST]
            except:
                pass
            try:
                LTE_NODE_LST = [item.upper() for item in LTE_NODE_LST]
            except:
                pass
            try:
                NR_NODE_LST = [item.upper() for item in NR_NODE_LST]
            except:
                pass
            DATABASE = 'CAUDIT_CLOSE_WINDOW'
            DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
            if 'close' not in DATBASE_VALUE_READ:
                try:
                    DATBASE_VALUE_READ.clear()
                    del DATBASE_VALUE_READ
                except:
                    pass
                execution_count[0] = 0

                def START_FUNNNN_22():
                    NEW_WCDMA_LST = []
                    if WCDMA_LST:
                        for nod in WCDMA_NODE_LST:
                            NEW_WCDMA_LST.append(WCDMA_LST[0] + nod.split('@')[(-1)])
                    NEW_WCDMA_NODE_LST = []
                    if WCDMA_NODE_LST:
                        for nod in WCDMA_NODE_LST:
                            NEW_WCDMA_NODE_LST.append(nod.split('@')[0])
                    RRSG_KPI_MONITORING_Module(self, selected_checkboxes, GSM_LST, NEW_WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, NEW_WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST, 'KPI MONITORING', 'UIBOOT')
                    if len(FREQUENCY_LST) > 1 and FREQUENCY_LST[0]!= 'Not required':
                        execution_count[0] += 1
                        if execution_count[0] < int(FREQUENCY_count_LST[0]):
                            current_time = datetime.now()
                            new_time = current_time + timedelta(minutes=int(FREQUENCY_LST[0]))
                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Please wait next execution scheduled on ' + new_time.strftime('%H:%M:%S') + ' Hour') + '\n')
                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                            start_index = '1.0'
                            while True:
                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                if not start_index:
                                    break
                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                start_index = end_index
                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                if len(FREQUENCY_LST) > 1 and FREQUENCY_LST[0] == 'Not required':
                    START_FUNNNN_22()
                    FREQUENCY_count_LST.clear()
                    FREQUENCY_LST.clear()
                    EMAIL_METHOOD_LST.clear()
                    DATABASE = 'CAUDIT_CLOSE_WINDOW'
                    SQLITE_DB_CLEAR(DATABASE)
                    DATABASE = 'AUDDIT_START'
                    SQLITE_DB_CLEAR(DATABASE)
                    DATABASE = 'ACTIVITYY'
                    SQLITE_DB_CLEAR(DATABASE)
                    GSM_LST.clear()
                    LTE_LST.clear()
                    WCDMA_LST.clear()
                    NR_LST.clear()
                    selected_checkboxes.clear()
                    GSM_NODE_LST.clear()
                    LTE_NODE_LST.clear()
                    WCDMA_NODE_LST.clear()
                    NR_NODE_LST.clear()
                    COMMAND_DEFALT_AUTO.clear()
                    return
                START_FUNNNN_22()
                if len(FREQUENCY_LST) > 1:
                    schedule.every(int(FREQUENCY_LST[0])).minutes.do(START_FUNNNN_22)
                    while execution_count[0] < int(FREQUENCY_count_LST[0]):
                        schedule.run_pending()
                        time.sleep(1)
                FREQUENCY_count_LST.clear()
                FREQUENCY_LST.clear()
                EMAIL_METHOOD_LST.clear()
                DATABASE = 'CAUDIT_CLOSE_WINDOW'
                SQLITE_DB_CLEAR(DATABASE)
                DATABASE = 'AUDDIT_START'
                SQLITE_DB_CLEAR(DATABASE)
                DATABASE = 'ACTIVITYY'
                SQLITE_DB_CLEAR(DATABASE)
                GSM_LST.clear()
                LTE_LST.clear()
                WCDMA_LST.clear()
                NR_LST.clear()
                selected_checkboxes.clear()
                GSM_NODE_LST.clear()
                LTE_NODE_LST.clear()
                WCDMA_NODE_LST.clear()
                NR_NODE_LST.clear()
                COMMAND_DEFALT_AUTO.clear()
                self.radio_var.set((-1))
                self.activity_BU.set(0)
            else:  # inserted
                thread = threading.Thread(target=enabled_button, args=(self,))
                thread.start()
                thread.join()
                DATABASE = 'CAUDIT_CLOSE_WINDOW'
                SQLITE_DB_CLEAR(DATABASE)
                DATABASE = 'AUDDIT_START'
                SQLITE_DB_CLEAR(DATABASE)
                DATABASE = 'ACTIVITYY'
                SQLITE_DB_CLEAR(DATABASE)
                GSM_LST.clear()
                LTE_LST.clear()
                WCDMA_LST.clear()
                NR_LST.clear()
                selected_checkboxes.clear()
                GSM_NODE_LST.clear()
                LTE_NODE_LST.clear()
                WCDMA_NODE_LST.clear()
                NR_NODE_LST.clear()
                COMMAND_DEFALT_AUTO.clear()
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Thanks! how Chat Buddy will assist you !') + '\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        threading.Thread(target=START_FUNNNN).start()

    def on_closing_ii():
        app.destroy()
    if MODEEE == 'Dark':
        butoncolour = '#7b6d48'
    if MODEEE == 'light':
        butoncolour = '#A8A970'
    scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100
    health_button = customtkinter.CTkButton(master=app, text='Health check & audit', fg_color=butoncolour, font=customtkinter.CTkFont(size=12, weight='bold'), width=160, command=button_function)
    if scaleFactor == 1.5:
        health_button.grid(row=1, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        health_button.grid(row=1, column=0, padx=30, pady=(18, 0), sticky='w')
    pmx_button = customtkinter.CTkButton(master=app, text='Template-Free PM Report', fg_color=butoncolour, font=customtkinter.CTkFont(size=12, weight='bold'), state='normal', width=160, command=button_function2)
    if scaleFactor == 1.5:
        pmx_button.grid(row=2, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        pmx_button.grid(row=2, column=0, padx=30, pady=(18, 0), sticky='w')
    db_button = customtkinter.CTkButton(master=app, text='MRR Recording', font=customtkinter.CTkFont(size=12, weight='bold'), state='disabled', width=160, command=button_function3)
    if scaleFactor == 1.5:
        db_button.grid(row=3, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        db_button.grid(row=3, column=0, padx=30, pady=(18, 0), sticky='w')
    db_button = customtkinter.CTkButton(master=app, text='KPI Monitoring', fg_color='#CD5C5C', font=customtkinter.CTkFont(size=12, weight='bold'), state='normal', width=160, command=button_function4)
    if scaleFactor == 1.5:
        db_button.grid(row=4, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        db_button.grid(row=4, column=0, padx=30, pady=(18, 0), sticky='w')
    app.protocol('WM_DELETE_WINDOW', on_closing_ii)
    app.attributes('-topmost', True)
    app.mainloop()

def RBSPASSApp_fun():
    class RBSAppLoginApp(QWidget):
        def __init__(self):
            super().__init__()
            self.load_credentials()
            self.setGeometry(300, 300, 200, 300)
            self.setWindowTitle('RBS PASSWORD DETAILS')
            self.setFixedSize(420, 300)
            self.center_window()
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.create_gui()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def load_credentials(self):
            try:
                db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_RBSPASS_CRED.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                con.close()
                self.PID_LST = Rdf['value'].tolist()
                while len(self.PID_LST) < 4:
                    self.PID_LST.append(f'LINE{len(self.PID_LST) + 1}')
            except:
                self.PID_LST = [f'LINE{i + 1}' for i in range(4)]
            if self.PID_LST[0] == 'LINE1':
                self.PID_LST[0] = 'rbs'
            if self.PID_LST[1] == 'LINE2':
                self.PID_LST[1] = 'rbs'

        def create_gui(self):
            layout = QVBoxLayout(self)
            label = QLabel('RBS PASSWORD DETAILS', self)
            label.setStyleSheet('font-size: 24px; color: #fe6f5e; font-weight: bold;')
            layout.addWidget(label)
            self.entries = []
            GATEWAY_label = QLabel('RBS User Nmae:', self)
            GATEWAY_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(GATEWAY_label)
            GATEWAY_label_entry = QLineEdit(self)
            GATEWAY_label_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            GATEWAY_label_entry.setFont(font)
            GATEWAY_label_entry.setFixedSize(390, 35)
            GATEWAY_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(GATEWAY_label_entry)
            GATEWAY_label_entry.setText(self.PID_LST[0])
            self.entries.append(GATEWAY_label_entry)
            Username_label = QLabel('RBS Password:', self)
            Username_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(Username_label)
            Username_label_entry = QLineEdit(self)
            Username_label_entry.setEchoMode(QLineEdit.Password)
            Username_label_entry.setFont(font)
            Username_label_entry.setFixedSize(390, 35)
            Username_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(Username_label_entry)
            Username_label_entry.setText(self.PID_LST[1])
            self.entries.append(Username_label_entry)
            update_button = QPushButton('UPDATE', self)
            update_button.clicked.connect(self.update_credentials)
            update_button.setStyleSheet('padding: 8px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;')
            layout.addWidget(update_button)
            self.setLayout(layout)

        def update_credentials(self):
            values = [entry.toPlainText() if isinstance(entry, QTextEdit) else entry.text() for entry in self.entries]
            DF = pd.DataFrame({'indexxx': [f'Line{i + 1}' for i in range(len(values))], 'value': values})
            ppPID_LST = DF['value'].tolist()
            ppPID_LST2 = []
            for ppp in ppPID_LST:
                ppPID_LST2.append(encrypt(ppp, 3))
            del DF['value']
            DF['value'] = ppPID_LST2
            try:
                os.remove(os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_RBSPASS_CRED.db'))
            except:
                pass
            con = sqlite3.connect(os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_RBSPASS_CRED.db'))
            try:
                DF.to_sql('Credd', con, index=False, if_exists='replace')
            except:
                pass
            finally:  # inserted
                con.close()
            RBSPASS_CRED_CRREDD.clear()
            try:
                for ppp in ppPID_LST2:
                    RBSPASS_CRED_CRREDD.append(decrypt(ppp, 3))
            except:
                pass
            self.close()
    if __name__ == '__main__':
        app = QApplication([])
        rsg_app = RBSAppLoginApp()
        rsg_app.show()
        app.exec_()

def VPNApp_fun():
    class VPNAppLoginApp(QWidget):
        def __init__(self):
            super().__init__()
            self.load_credentials()
            self.setGeometry(300, 300, 200, 300)
            self.setWindowTitle('VPN & CAS DETAILS')
            self.setFixedSize(420, 300)
            self.center_window()
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.create_gui()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def load_credentials(self):
            try:
                db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_VPN_CRED.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                con.close()
                self.PID_LST = Rdf['value'].tolist()
                while len(self.PID_LST) < 4:
                    self.PID_LST.append(f'LINE{len(self.PID_LST) + 1}')
            except:
                self.PID_LST = [f'LINE{i + 1}' for i in range(4)]
            if self.PID_LST[0] == 'LINE1':
                self.PID_LST[0] = 'VPN $ CAS Username'
            if self.PID_LST[1] == 'LINE2':
                self.PID_LST[1] = 'VPN $ CAS Password'

        def create_gui(self):
            layout = QVBoxLayout(self)
            label = QLabel('VPN & CAS DETAILS', self)
            label.setStyleSheet('font-size: 24px; color: #fe6f5e; font-weight: bold;')
            layout.addWidget(label)
            self.entries = []
            GATEWAY_label = QLabel('User Name:', self)
            GATEWAY_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(GATEWAY_label)
            GATEWAY_label_entry = QLineEdit(self)
            GATEWAY_label_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            GATEWAY_label_entry.setFont(font)
            GATEWAY_label_entry.setFixedSize(390, 35)
            GATEWAY_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(GATEWAY_label_entry)
            GATEWAY_label_entry.setText(self.PID_LST[0])
            self.entries.append(GATEWAY_label_entry)
            Username_label = QLabel('Password:', self)
            Username_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(Username_label)
            Username_label_entry = QLineEdit(self)
            Username_label_entry.setEchoMode(QLineEdit.Password)
            Username_label_entry.setFont(font)
            Username_label_entry.setFixedSize(390, 35)
            Username_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(Username_label_entry)
            Username_label_entry.setText(self.PID_LST[1])
            self.entries.append(Username_label_entry)
            update_button = QPushButton('UPDATE', self)
            update_button.clicked.connect(self.update_credentials)
            update_button.setStyleSheet('padding: 8px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;')
            layout.addWidget(update_button)
            self.setLayout(layout)

        def update_credentials(self):
            values = [entry.toPlainText() if isinstance(entry, QTextEdit) else entry.text() for entry in self.entries]
            DF = pd.DataFrame({'indexxx': [f'Line{i + 1}' for i in range(len(values))], 'value': values})
            ppPID_LST = DF['value'].tolist()
            ppPID_LST2 = []
            for ppp in ppPID_LST:
                ppPID_LST2.append(encrypt(ppp, 3))
            del DF['value']
            DF['value'] = ppPID_LST2
            try:
                os.remove(os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_VPN_CRED.db'))
            except:
                pass
            con = sqlite3.connect(os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_VPN_CRED.db'))
            try:
                DF.to_sql('Credd', con, index=False, if_exists='replace')
            except:
                pass
            finally:  # inserted
                con.close()
            self.close()
    if __name__ == '__main__':
        app = QApplication([])
        rsg_app = VPNAppLoginApp()
        rsg_app.show()
        app.exec_()

def EMAILApp_fun():
    class EMAILAppLoginApp(QWidget):
        def __init__(self):
            super().__init__()
            self.load_credentials()
            self.setGeometry(300, 300, 420, 526)
            self.setWindowTitle('EMAIL DETAILS')
            self.setFixedSize(420, 680)
            self.center_window()
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.create_gui()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def load_credentials(self):
            try:
                db_path = os.path.join('./res/EMAIL_CRED.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                con.close()
                self.PID_LST = Rdf['value'].tolist()
                while len(self.PID_LST) < 4:
                    self.PID_LST.append(f'LINE{len(self.PID_LST) + 1}')
            except:
                self.PID_LST = [f'LINE{i + 1}' for i in range(4)]
            if self.PID_LST[0] == 'LINE1':
                self.PID_LST[0] = 'Email Id'
            if self.PID_LST[1] == 'LINE2':
                self.PID_LST[1] = 'Len password'

        def create_gui(self):
            layout = QVBoxLayout(self)
            label = QLabel('EMAIL DETAILS', self)
            label.setStyleSheet('font-size: 24px; color: #fe6f5e; font-weight: bold;')
            layout.addWidget(label)
            self.entries = []
            GATEWAY_label = QLabel('Email Id:', self)
            GATEWAY_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(GATEWAY_label)
            GATEWAY_label_entry = QLineEdit(self)
            font = QFont('Calibri', 11)
            GATEWAY_label_entry.setFont(font)
            GATEWAY_label_entry.setFixedSize(390, 35)
            GATEWAY_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(GATEWAY_label_entry)
            GATEWAY_label_entry.setText(self.PID_LST[0])
            self.entries.append(GATEWAY_label_entry)
            Username_label = QLabel('Len password:', self)
            Username_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(Username_label)
            Username_label_entry = QLineEdit(self)
            Username_label_entry.setEchoMode(QLineEdit.Password)
            Username_label_entry.setFont(font)
            Username_label_entry.setFixedSize(390, 35)
            Username_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(Username_label_entry)
            Username_label_entry.setText(self.PID_LST[1])
            self.entries.append(Username_label_entry)
            recipient_label_to = QLabel('Email To:', self)
            recipient_label_to.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(recipient_label_to)
            recipient_entry_to = QTextEdit(self)
            recipient_entry_to.setFont(font)
            recipient_entry_to.setFixedSize(390, 180)
            recipient_entry_to.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(recipient_entry_to)
            recipient_entry_to.setText(self.PID_LST[2])
            self.entries.append(recipient_entry_to)
            recipient_label_cc = QLabel('Email CC:', self)
            recipient_label_cc.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(recipient_label_cc)
            recipient_entry_cc = QTextEdit(self)
            recipient_entry_cc.setFont(font)
            recipient_entry_cc.setFixedSize(390, 180)
            recipient_entry_cc.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(recipient_entry_cc)
            recipient_entry_cc.setText(self.PID_LST[3])
            self.entries.append(recipient_entry_cc)
            update_button = QPushButton('UPDATE', self)
            update_button.clicked.connect(self.update_credentials)
            update_button.setStyleSheet('padding: 8px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;')
            layout.addWidget(update_button)
            self.setLayout(layout)

        def update_credentials(self):
            values = [entry.toPlainText() if isinstance(entry, QTextEdit) else entry.text() for entry in self.entries]
            DF = pd.DataFrame({'indexxx': [f'Line{i + 1}' for i in range(len(values))], 'value': values})
            ppPID_LST = DF['value'].tolist()
            ppPID_LST2 = []
            for ppp in ppPID_LST:
                ppPID_LST2.append(encrypt(ppp, 3))
            del DF['value']
            DF['value'] = ppPID_LST2
            try:
                os.remove(os.path.join('./res/EMAIL_CRED.db'))
            except:
                pass
            con = sqlite3.connect(os.path.join('./res/EMAIL_CRED.db'))
            try:
                DF.to_sql('Credd', con, index=False, if_exists='replace')
            except:
                pass
            finally:  # inserted
                con.close()
            self.close()
    if __name__ == '__main__':
        app = QApplication([])
        rsg_app = EMAILAppLoginApp()
        rsg_app.show()
        app.exec_()

def HoatnameLoginApp_fun():
    class AZURELoginApp(QWidget):
        def __init__(self):
            super().__init__()
            self.load_credentials()
            self.setGeometry(300, 300, 420, 526)
            self.setWindowTitle('Hostname Details')
            self.setFixedSize(420, 180)
            self.center_window()
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.create_gui()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def load_credentials(self):
            try:
                db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                con.close()
                Rdf = Rdf.drop_duplicates(keep='first')
                self.PID_LST = Rdf['value'].tolist()
                while len(self.PID_LST) < 1:
                    self.PID_LST.append(f'LINE{len(self.PID_LST) + 1}')
            except:
                self.PID_LST = [f'LINE{i + 1}' for i in range(1)]
            if self.PID_LST[0] == 'LINE1':
                self.PID_LST[0] = '148.135.15.71'

        def create_gui(self):
            layout = QVBoxLayout(self)
            label = QLabel('Hostname', self)
            label.setStyleSheet('font-size: 24px; color: #fe6f5e; font-weight: bold;')
            layout.addWidget(label)
            self.entries = []
            GATEWAY_label = QLabel('Hostname:', self)
            GATEWAY_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(GATEWAY_label)
            GATEWAY_label_entry = QLineEdit(self)
            GATEWAY_label_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            GATEWAY_label_entry.setFont(font)
            GATEWAY_label_entry.setFont(font)
            GATEWAY_label_entry.setFixedSize(390, 35)
            GATEWAY_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(GATEWAY_label_entry)
            GATEWAY_label_entry.setText(self.PID_LST[0])
            self.entries.append(GATEWAY_label_entry)
            update_button = QPushButton('UPDATE', self)
            update_button.clicked.connect(self.login)
            update_button.setStyleSheet('padding: 8px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;')
            layout.addWidget(update_button)
            self.setLayout(layout)

        def login(self):
            values = [entry.text() if entry.text() else f'LINE{i + 1}' for i, entry in enumerate(self.entries)]
            DF = pd.DataFrame({'indexxx': [f'Line{i + 1}' for i in range(1)], 'value': values})
            try:
                os.remove(os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db'))
            except:
                pass
            con = sqlite3.connect(os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db'))
            try:
                DF.to_sql('Credd', con, index=False, if_exists='replace')
            except:
                pass
            finally:  # inserted
                con.close()
            self.close()
    if __name__ == '__main__':
        app = QApplication([])
        rsg_app = AZURELoginApp()
        rsg_app.show()
        app.exec_()

def AZURELoginApp_fun():
    class AZURELoginApp(QWidget):
        def __init__(self):
            super().__init__()
            self.load_credentials()
            self.setGeometry(300, 300, 420, 526)
            self.setWindowTitle('AZURE DETAILS')
            self.setFixedSize(420, 540)
            self.center_window()
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
            self.create_gui()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def load_credentials(self):
            try:
                db_path = os.path.join('./res/AZURE_CRED.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                con.close()
                Rdf = Rdf.drop_duplicates(keep='first')
                self.PID_LST = Rdf['value'].tolist()
                while len(self.PID_LST) < 5:
                    self.PID_LST.append(f'LINE{len(self.PID_LST) + 1}')
            except:
                self.PID_LST = [f'LINE{i + 1}' for i in range(6)]
            if self.PID_LST[0] == 'LINE1':
                self.PID_LST[0] = 'Gateway host name'
            if self.PID_LST[1] == 'LINE2':
                self.PID_LST[1] = 'Len password'
            if self.PID_LST[2] == 'LINE3':
                self.PID_LST[2] = 'Azure host'
            if self.PID_LST[3] == 'LINE4':
                self.PID_LST[3] = 'Azure login'
            if self.PID_LST[4] == 'LINE5':
                self.PID_LST[4] = 'Azure password'

        def create_gui(self):
            layout = QVBoxLayout(self)
            label = QLabel('AZURE DETAILS', self)
            label.setStyleSheet('font-size: 24px; color: #fe6f5e; font-weight: bold;')
            layout.addWidget(label)
            self.entries = []
            GATEWAY_label = QLabel('Gateway Host:', self)
            GATEWAY_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(GATEWAY_label)
            GATEWAY_label_entry = QLineEdit(self)
            font = QFont('Calibri', 11)
            GATEWAY_label_entry.setFont(font)
            GATEWAY_label_entry.setFont(font)
            GATEWAY_label_entry.setFixedSize(390, 35)
            GATEWAY_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(GATEWAY_label_entry)
            GATEWAY_label_entry.setText(self.PID_LST[0])
            self.entries.append(GATEWAY_label_entry)
            password_label = QLabel('Len Password:', self)
            password_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(password_label)
            password_entry = QLineEdit(self)
            password_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            password_entry.setFont(font)
            password_entry.setFont(font)
            password_entry.setFixedSize(390, 35)
            password_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(password_entry)
            password_entry.setText(self.PID_LST[1])
            self.entries.append(password_entry)
            AZURE_HOST = QLabel('Azure Host Name:', self)
            AZURE_HOST.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(AZURE_HOST)
            AZURE_HOST_entry = QLineEdit(self)
            AZURE_HOST_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            AZURE_HOST_entry.setFont(font)
            AZURE_HOST_entry.setFont(font)
            AZURE_HOST_entry.setFixedSize(390, 35)
            AZURE_HOST_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(AZURE_HOST_entry)
            AZURE_HOST_entry.setText(self.PID_LST[2])
            self.entries.append(AZURE_HOST_entry)
            AUsername_label = QLabel('Azure Login:', self)
            AUsername_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(AUsername_label)
            AUsername_label_entry = QLineEdit(self)
            AUsername_label_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            AUsername_label_entry.setFont(font)
            AUsername_label_entry.setFont(font)
            AUsername_label_entry.setFixedSize(390, 35)
            AUsername_label_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(AUsername_label_entry)
            AUsername_label_entry.setText(self.PID_LST[3])
            self.entries.append(AUsername_label_entry)
            Apassword_label = QLabel('Azure Password:', self)
            Apassword_label.setStyleSheet('font-size: 17px; color: #333; font-weight: bold;')
            layout.addWidget(Apassword_label)
            Apassword_entry = QLineEdit(self)
            Apassword_entry.setEchoMode(QLineEdit.Password)
            font = QFont('Calibri', 11)
            Apassword_entry.setFont(font)
            Apassword_entry.setFont(font)
            Apassword_entry.setFixedSize(390, 35)
            Apassword_entry.setStyleSheet('background-color: #eff4f9; border: 1px solid #ccc;')
            layout.addWidget(Apassword_entry)
            Apassword_entry.setText(self.PID_LST[4])
            self.entries.append(Apassword_entry)
            update_button = QPushButton('UPDATE', self)
            update_button.clicked.connect(self.login)
            update_button.setStyleSheet('padding: 8px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;')
            layout.addWidget(update_button)
            self.setLayout(layout)

        def login(self):
            values = [encrypt(entry.text(), 3) if encrypt(entry.text(), 3) else f'LINE{i + 1}' for i, entry in enumerate(self.entries)]
            DF = pd.DataFrame({'indexxx': [f'Line{i + 1}' for i in range(5)], 'value': values})
            try:
                os.remove(os.path.join('./res/AZURE_CRED.db'))
            except:
                pass
            con = sqlite3.connect(os.path.join('./res/AZURE_CRED.db'))
            try:
                DF.to_sql('Credd', con, index=False, if_exists='replace')
            except:
                pass
            finally:  # inserted
                con.close()
            self.close()
    if __name__ == '__main__':
        app = QApplication([])
        rsg_app = AZURELoginApp()
        rsg_app.show()
        app.exec_()

def tool_setting(self):
    self.thread_running = False

    def center_window(window):
        window.update_idletasks()
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100
        if scaleFactor == 1.5:
            width = 180
            height = 385
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2.0
            window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
        else:  # inserted
            width = 230
            height = 430
            x = (screen_width - width + 20) // 2
            y = (screen_height - height - 200) // 2
            window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

    def on_close():
        if not self.thread_running:
            threading.Thread(target=enabled_button, args=(self,)).start()
            self.thread_running = True
    customtkinter.set_appearance_mode(MODEEE)
    customtkinter.set_default_color_theme('green')
    app = customtkinter.CTk()
    tool_setting.app = app
    app.geometry('350x200')
    app.title('Tool Setting')
    center_window(app)
    app.bind('<Destroy>', lambda e: on_close())

    def delayaqd_destr22():
        if app.winfo_exists():
            app.destroy()

    def button_function():
        app.withdraw()
        RSGLoginApp_fun()
        app.after(100, delayaqd_destr22)

    def button_function2():
        app.withdraw()
        AZURELoginApp_fun()
        app.after(100, delayaqd_destr22)

    def button_function3():
        app.withdraw()
        ISFLoginApp_fun()
        app.after(100, delayaqd_destr22)

    def button_function5():
        app.withdraw()
        EMAILApp_fun()
        app.after(100, delayaqd_destr22)

    def button_function6():
        app.withdraw()
        HoatnameLoginApp_fun()
        app.after(100, delayaqd_destr22)

    def button_function_RBS():
        app.withdraw()
        RBSPASSApp_fun()
        app.after(100, delayaqd_destr22)

    def button_function_VPN():
        app.withdraw()
        VPNApp_fun()
        app.after(100, delayaqd_destr22)

    def button_function4():
        def delayed_destr22():
            if self.winfo_exists():
                app.destroy()
        self.after(100, delayed_destr22)
        try:
            eyye = RRSG_auto_login.remote_conn
            RRSG_enter_command(self, ['quit'], 'chatbuddy')
            clear_all_lists()
        except:
            pass
        try:
            ssh.close()
            RRSG_auto_login.remote_conn.close()
            del RRSG_auto_login.remote_conn
        except:
            pass
        try:
            combined_cmd = ''
        except:
            pass
        clear_all_lists()
        try:
            RRSG_auto_login.remote_conn
        except:
            pass
        self.radio_var.set((-1))
        self.activity_BU.set(0)
        threading.Thread(target=enabled_button, args=(self,)).start()
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Reset Done.\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        try:
            self.textbox_LIVECMD.delete(1.0, 'end')
        except:
            return None
    if MODEEE == 'Dark':
        butoncolour = '#7b6d48'
    if MODEEE == 'light':
        butoncolour = '#A8A970'
    scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100
    rsg_button = customtkinter.CTkButton(master=app, text='RSG Setting', fg_color=butoncolour, font=customtkinter.CTkFont(size=12, weight='bold'), width=150, command=button_function)
    if scaleFactor == 1.5:
        rsg_button.grid(row=1, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        rsg_button.grid(row=1, column=0, padx=30, pady=(18, 0), sticky='w')
    VPN_button = customtkinter.CTkButton(master=app, text='VPN Setting', fg_color=butoncolour, font=customtkinter.CTkFont(size=12, weight='bold'), width=150, command=button_function_VPN)
    if scaleFactor == 1.5:
        VPN_button.grid(row=2, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        VPN_button.grid(row=2, column=0, padx=30, pady=(18, 0), sticky='w')
    if 'SME' in USER_MAPPED_ROLE:
        SME_ROLE = 'normal'
    else:  # inserted
        SME_ROLE = 'disabled'
    if os.getlogin().lower() in SME_LST:
        AZ_state = 'normal'
    else:  # inserted
        AZ_state = 'disabled'
    azure_button = customtkinter.CTkButton(master=app, text='Azure Setting', fg_color=butoncolour, state=AZ_state, font=customtkinter.CTkFont(size=12, weight='bold'), width=150, command=button_function2)
    if scaleFactor == 1.5:
        azure_button.grid(row=3, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        azure_button.grid(row=3, column=0, padx=30, pady=(18, 0), sticky='w')
    db_button = customtkinter.CTkButton(master=app, text='ISF Setting', font=customtkinter.CTkFont(size=12, weight='bold'), state=SME_ROLE, width=150, command=button_function3)
    if scaleFactor == 1.5:
        db_button.grid(row=4, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        db_button.grid(row=4, column=0, padx=30, pady=(18, 0), sticky='w')
    email_button = customtkinter.CTkButton(master=app, text='Email', fg_color=butoncolour, font=customtkinter.CTkFont(size=12, weight='bold'), width=150, command=button_function5)
    if scaleFactor == 1.5:
        email_button.grid(row=5, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        email_button.grid(row=5, column=0, padx=30, pady=(18, 0), sticky='w')
    db_button = customtkinter.CTkButton(master=app, text='Hostname', fg_color=butoncolour, font=customtkinter.CTkFont(size=12, weight='bold'), width=150, command=button_function6)
    if scaleFactor == 1.5:
        db_button.grid(row=6, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        db_button.grid(row=6, column=0, padx=30, pady=(18, 0), sticky='w')
    RBSPASS_button = customtkinter.CTkButton(master=app, text='RBS Password', fg_color=butoncolour, font=customtkinter.CTkFont(size=12, weight='bold'), width=150, command=button_function_RBS)
    if scaleFactor == 1.5:
        RBSPASS_button.grid(row=7, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        RBSPASS_button.grid(row=7, column=0, padx=30, pady=(18, 0), sticky='w')
    db_button = customtkinter.CTkButton(master=app, text='Reset', fg_color='#CD5C5C', font=customtkinter.CTkFont(size=12, weight='bold'), width=150, command=button_function4)
    if scaleFactor == 1.5:
        db_button.grid(row=8, column=0, padx=14, pady=(18, 0), sticky='w')
    else:  # inserted
        db_button.grid(row=8, column=0, padx=30, pady=(18, 0), sticky='w')
    app.attributes('-topmost', True)
    app.mainloop()

def Azure_OTP():
    class OTPDialog(QDialog):
        def __init__(self):
            super().__init__()
            self.init_ui()

        def init_ui(self):
            self.setWindowTitle('AZURE OTP')
            self.setGeometry(800, 400, 190, 100)
            layout = QVBoxLayout()
            self.label = QLabel('AZURE One Time OTP', self)
            self.label.setStyleSheet('padding: 0px; font: 9pt \'Segoe UI Semibold\';')
            self.edit = QLineEdit(self)
            self.edit.setEchoMode(QLineEdit.Password)
            self.edit.returnPressed.connect(self.process_otp)
            self.edit.setFixedWidth(190)
            self.edit.setAlignment(Qt.AlignCenter)
            layout.addWidget(self.label)
            layout.addWidget(self.edit)
            self.setLayout(layout)
            self.setStyleSheet('\n                QDialog {\n                    background-color: #e6e6e6;\n                }\n                QLabel {\n                    font: 14pt \"Calibri\";\n                }\n                QLineEdit {\n                    font: 12pt \"Calibri\";\n                    padding: 4px;\n                    border: 2px solid #afa04c;\n                    border-radius: 4px;\n                }\n            ')
            self.setWindowFlag(Qt.WindowStaysOnTopHint)

        def process_otp(self):
            user_input = self.edit.text()
            if user_input:
                self.accept()
            else:  # inserted
                QMessageBox.warning(self, 'Warning', 'Please enter a valid OTP.')

    def get_otp():
        app = QApplication([])
        dialog = OTPDialog()
        result = dialog.exec_()
        if result == QDialog.Accepted:
            user_input = dialog.edit.text()
            return user_input
        cmmmddd.clear()
        siteddd.clear()
        checknxtsit.clear()
        processdone.clear()
    otp = get_otp()
    return otp

def autenti_fail(self):
    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : RSG Authentication failed. Please check your credentials and try again !\n')
    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
    start_index = '1.0'
    while True:
        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
        if not start_index:
            break
        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
        start_index = end_index
    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    thread = threading.Thread(target=enabled_button, args=(self,))
    thread.start()
    thread.join()

def program_expired(path):
    con = sqlite3.connect(path)
    expire = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
    con.close()
    del expire['indexxx']
    year = int(expire.iloc[0, 0])
    mon = int(expire.iloc[1, 0])
    day = int(expire.iloc[2, 0])
    app_date = datetime(year=year, month=mon, day=day)
    now = datetime.now()
    program_expired.exp_on = (app_date - now).days + 1
    exp_d = now + timedelta(days=program_expired.exp_on)
    program_expired.exp_date = exp_d.strftime('%a %b %d %Y %H:%M:%S')
    if program_expired.exp_on <= 10:
        try:
            os.mkdir(OUTPATH + 'warning')
        except:
            pass
        if program_expired.exp_on > 0:
            program_expired.expire_on = 'tool_expire_on'
            program_expired.expire_msg = 'Warning ,Your tool expire after ' + str(program_expired.exp_on) + ' days on ' + program_expired.exp_date + ' please contact with developer..'
            with open(OUTPATH + 'warning' + '\\' + 'Important Note.txt', 'w') as file:
                file.write('Warning ,Your tool expire after ' + str(program_expired.exp_on) + ' days on ' + program_expired.exp_date + ' please contact with developer..')
            showMessage('Warning ,Your tool expire after ' + str(program_expired.exp_on) + ' days on ' + program_expired.exp_date + ' please contact with developer..')
            return 'continue'
    if program_expired.exp_on <= 0:
        try:
            os.mkdir(OUTPATH + 'warning')
        except:
            pass
        with open(OUTPATH + 'warning' + '\\' + 'Tool Expire.txt', 'w') as file:
            file.write('Illegal Copy ,Your tool has expired, please contact the developer..')
        program_expired.expire = 'tool_expire'
        del expire
        showMessage('Illegal Copy ,Your tool has expired, please contact the developer..')
        try:
            sys.exit()
            return 'Tool_expired'
        except:
            return 'Tool_expired'

def checkuser_expired(path):
    con = sqlite3.connect(path)
    expire_use = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
    expire_user = expire_use['value'].unique()
    expire_user = list(expire_user)
    expire_user = [item.lower() for item in expire_user]
    con.close()
    if os.getlogin().lower() in expire_user:
        try:
            os.mkdir(OUTPATH + 'warning')
        except:
            pass
        with open(OUTPATH + 'warning' + '\\' + 'Unauthorized Access.txt', 'w') as file:
            file.write('Unauthorized Access, you are not authorized to access this app,please contact the developer..')
        program_expired.expire = 'Unauthorized_Access'
        del expire_user
        showMessage('Unauthorized Access, you are not authorized to access this app,please contact the developer..')
        try:
            sys.exit()
            return 'Unauthorized_Access'
        except:
            return 'Unauthorized_Access'

def conn_ssh_sqlwrite(gateway_pass, gateway_Token, user, gateway_host, remote_host, remote_username, remote_password, remote_file_path, SQL_TABLE, keyword, command):
    global my_transport  # inserted
    global remote_ssh_client  # inserted

    def inter_handler(title, instructions, prompt_list):
        resp = []
        for pr in prompt_list:
            if str(pr[0]).strip() == 'EGAD-Password:':
                resp.append(gateway_pass)
            if 'Enter Your Microsoft verification code' in str(pr[0]).strip():
                resp.append(gateway_Token)
        return tuple(resp)
    if my_transport is None or not my_transport.is_active():
        gateway_Token = Azure_OTP()
        gateway_port = 22
        my_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        my_socket.connect((gateway_host, gateway_port))
        my_transport = paramiko.Transport(my_socket)
        my_transport.start_client(timeout=5020)
        my_transport.auth_interactive(user, inter_handler)
        channel = my_transport.open_session()
        transport = channel.get_transport()
        dest_addr = (remote_host, 22)
        local_addr = ('', 0)
        tunnel = transport.open_channel('direct-tcpip', dest_addr, local_addr)
        remote_ssh_client = paramiko.SSHClient()
        remote_ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        remote_ssh_client.connect(remote_host, username=remote_username, password=remote_password, sock=tunnel)
    if remote_ssh_client:
        command_to_execute = f'sqlite3 {remote_file_path} \"INSERT INTO conversations (text) VALUES (\'{keyword}_{command}\');\"'
        stdin, stdout, stderr = remote_ssh_client.exec_command(command_to_execute)
        remote_file_path = '/home/inputdata_TIGO-TZ/Input/Ericsson/4G/RCA/Test_data/conversations.db'
        CONVERSION_TABLE = 'SELECT text FROM conversations'
        connection_strin22 = f'sqlite3 {remote_file_path} \"{CONVERSION_TABLE};\"'
        stdin, stdout, stderr = remote_ssh_client.exec_command(connection_strin22)
        conversations11 = stdout.readlines()
        conversations11 = [string.replace('\n', '') for string in conversations11]
        conversations11 = [element.lower() for element in conversations11]
        DF = pd.DataFrame()
        DF['text'] = conversations11
        try:
            os.remove(REPORT_INPATH22 + 'conv.db')
        except:
            pass
        con = sqlite3.connect(REPORT_INPATH22 + 'conv.db')
        try:
            DF.to_sql('conversations', con, index=False, if_exists='replace')
        except:
            pass
        try:
            del DF
        except:
            pass
        try:
            conversations11.clear()
        except:
            pass
        CHK_TO_CONTinue2.append('connected')

def USER_AUTHENTICATION_check_AZURE(change_User, command_barred, user_signum, command_to_barred_required, command_to_barred):
    if SQL_METHOD == 'AZURE_CLOUDE':
        def generate_key():
            key = secrets.token_hex(16)
            return key

        def authenticate_key(input_key, generated_key):
            return input_key == generated_key

        def authenticate_user(command_to_barred, user_signum, command_to_barred_required):
            con_change_writes = change_User
            if user_signum.lower() in con_change_writes:
                user_autenticate = 'YES'
            else:  # inserted
                user_autenticate = 'NO'
            if user_autenticate == 'YES':
                if command_to_barred_required == 'YES':
                    conversations_barred = command_barred
                    conversations_barred = [element.lower() for element in conversations_barred]
                    if command_to_barred.lower() in conversations_barred:
                        return False
                return True
            command_to_barred_required = 'YES'
            if command_to_barred_required == 'YES':
                conversations_barred = command_barred
                conversations_barred = [element.lower() for element in conversations_barred]
                for uuin in conversations_barred:
                    if command_to_barred.lower() == uuin.lower():
                        return False
                else:  # inserted
                    if command_to_barred.lower()[:3] == 'set':
                        return False
                    return True
            else:  # inserted
                return True
        if authenticate_user(command_to_barred.lower(), user_signum, command_to_barred_required):
            user_key = generate_key()
            validate_key = user_key
            if authenticate_key(validate_key, user_key):
                USER_AUTHENTICATION = 'User_Authenticated'
                return USER_AUTHENTICATION
            USER_AUTHENTICATION = 'Not_Authenticated'
            return USER_AUTHENTICATION
        USER_AUTHENTICATION = 'Not_Authenticated'
        return USER_AUTHENTICATION

def USER_AUTHENTICATION_check(user_signum, command_to_barred_required, command_to_barred):
    def generate_key():
        key = secrets.token_hex(16)
        return key

    def authenticate_key(input_key, generated_key):
        return input_key == generated_key

    def authenticate_user(command_to_barred, user_signum, command_to_barred_required):
        conn = sqlite3.connect(REPORT_INPATH22 + 'User_access.db')
        cursor = conn.cursor()
        cursor.execute('SELECT signum FROM change_writes')
        rows = cursor.fetchall()
        con_change_writes = [row[0] for row in rows]
        if user_signum.lower() in con_change_writes:
            user_autenticate = 'YES'
        else:  # inserted
            user_autenticate = 'NO'
        if user_autenticate == 'YES':
            if command_to_barred_required == 'YES':
                cursor.execute('SELECT text FROM conversations_barred')
                rows = cursor.fetchall()
                conversations_barred = [row[0] for row in rows]
                conversations_barred = [element.lower() for element in conversations_barred]
                if command_to_barred.lower() in conversations_barred:
                    conn.close()
                    return False
            conn.close()
            return True
        command_to_barred_required = 'YES'
        if command_to_barred_required == 'YES':
            cursor.execute('SELECT text FROM conversations_barred')
            rows = cursor.fetchall()
            conversations_barred = [row[0] for row in rows]
            conversations_barred = [element.lower() for element in conversations_barred]
            for uuin in conversations_barred:
                if command_to_barred.lower() == uuin.lower():
                    conn.close()
                    return False
            else:  # inserted
                if command_to_barred.lower()[:3] == 'set':
                    conn.close()
                    return False
                return True
        else:  # inserted
            conn.close()
            return True
    if authenticate_user(command_to_barred.lower(), user_signum, command_to_barred_required):
        user_key = generate_key()
        validate_key = user_key
        if authenticate_key(validate_key, user_key):
            USER_AUTHENTICATION = 'User_Authenticated'
            return USER_AUTHENTICATION
        USER_AUTHENTICATION = 'Not_Authenticated'
        return USER_AUTHENTICATION
    USER_AUTHENTICATION = 'Not_Authenticated'
    return USER_AUTHENTICATION

def check_GSM_CMD(self):
    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : GSM Commands module will available soon..\n')
    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
    start_index = '1.0'
    while True:
        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
        if not start_index:
            break
        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
        start_index = end_index
    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')

def PDF_COMMAND_GSM(REQUIRED_CMD, cmdfile_path):
    REQUIRED_CMDlst = REQUIRED_CMD.split(' ')
    for ttxt in REQUIRED_CMDlst:
        command_data = []

        def extract_data_until_command(pdf_path, keyword):
            doc = fitz.open(pdf_path)
            breafpoint = 'NO'
            for page_num in range(doc.page_count):
                page = doc[page_num]
                text = page.get_text()
                lines = text.split('\n')
                for i, line in enumerate(lines):
                    if keyword.lower() in line.lower():
                        start_line = i
                        end_line = find_command_line(lines, i)
                        if end_line is not None:
                            context = '\n'.join(lines[start_line:end_line])
                            result_list = context.lower().split()
                            for jjj in result_list:
                                if 'rl' in jjj and jjj[(-1)] == 'p':
                                    for tt in context.splitlines():
                                        try:
                                            command.append(tt.strip())
                                            break
                                        except:
                                            continue
                                    command_data.append(context)
                                    breafpoint = 'YES'
                                    break
                if breafpoint == 'YES':
                    break

        def find_command_line(lines, start_line):
            for i in range(start_line, len(lines)):
                if 'Command:' in lines[i]:
                    return i + 2
            else:  # inserted
                return None

        def extract_additional_details(context):
            additional_details = {}
            for line in context.split('\n'):
                parts = line.split(':')
                if len(parts) == 2:
                    key, value = (parts[0].strip(), parts[1].strip())
                    additional_details[key] = value
            return additional_details
        pdf_path = cmdfile_path
        keyword_to_extract = REQUIRED_CMD
        extracted_data = extract_data_until_command(pdf_path, ttxt)
    chkcmd = []
    REQ_CMD = []
    for data in command_data:
        if 'Command'.lower() in data.lower():
            chkcmd.append('cccmmmdd')
        if 'cccmmmdd' in chkcmd:
            REQ_CMD.append(data.split('\n')[(-1)])
    return REQ_CMD

class SplashScreen_conncting(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setGeometry(710, 490, 400, 200)
        self.setWindowTitle('Loading...')
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.close)
        self.timer.start(4000)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), QColor(0, 0, 0, 0))
        font = QFont('Consolas', 9)
        painter.setFont(font)
        painter.drawText(self.rect(), Qt.AlignCenter, 'Connecting RSG please wait...')

class splasch_connecting:
    def __init__(self):
        self.app = QApplication(sys.argv)
        splash_screen = SplashScreen_conncting()
        splash_screen.show()
        self.load_resources()
        self.main_window_content()
        QTimer.singleShot(4000, SplashScreen_conncting.close)
        sys.exit(self.app.exec_())

    def load_resources(self):
        self.app.processEvents()

    def main_window_content(self):
        return

class SplashScreencon(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowStaysOnTopHint | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setGeometry(710, 490, 400, 200)
        self.setWindowTitle('Loading...')
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.close)
        self.timer.start(1500)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), QColor(0, 0, 0, 0))
        font = QFont('Consolas', 9)
        painter.setFont(font)
        painter.drawText(self.rect(), Qt.AlignCenter, 'RSG connected...')

class SplashConnectingApp:
    def __init__(self):
        self.app = QApplication(sys.argv)
        splash_screen = SplashScreencon()
        splash_screen.show()
        self.load_resources()
        QTimer.singleShot(1500, splash_screen.close)
        sys.exit(self.app.exec_())

    def load_resources(self):
        self.app.processEvents()

class TextUpdater:
    def __init__(self, app_instance):
        self.app_instance = app_instance
        self.root = tk.Tk()
        self.root.title('OSS Terminal')
        self.text_widget = tk.Text(self.root, wrap='none', width=1500, height=600, bg='#e6f7f6')
        self.text_widget.grid(row=0, column=0, sticky='nsew')
        font_size = 8
        new_font = ('Consolas', font_size)
        self.text_widget.configure(font=new_font)
        self.scrollbar_y = tk.Scrollbar(self.root, command=self.text_widget.yview)
        self.scrollbar_y.grid(row=0, column=1, sticky='ns')
        self.text_widget.config(yscrollcommand=self.scrollbar_y.set)
        self.scrollbar_x = tk.Scrollbar(self.root, orient=tk.HORIZONTAL, command=self.text_widget.xview)
        self.scrollbar_x.grid(row=1, column=0, sticky='ew')
        self.text_widget.config(xscrollcommand=self.scrollbar_x.set)
        self.text_widget.bind('<Return>', self.on_enter_pressed)
        self.update_counter = 0
        self.after_id = None
        self.center_window()
        self.start_text()

    def center_window(self):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = 1725
        window_height = 750
        scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100
        if scaleFactor == 1.5:
            x_position = screen_width + 540 - window_width
            y_position = screen_height + 150 - window_height
        else:  # inserted
            x_position = screen_width + 300 - window_width
            y_position = screen_height + 20 - window_height
        self.root.geometry(f'{window_width}x{window_height}+{x_position}+{y_position}')

    def update_text_widget(self):
        if len(combined_cmd_window) >= 20:
            combined_cmd_window2 = combined_cmd_window[20:]
            table_msg = '\n'.join(combined_cmd_window2)
            self.text_widget.insert(tk.END, table_msg)
            self.text_widget.see(tk.END)
        else:  # inserted
            table_msg = '\n'.join(combined_cmd_window)
            self.text_widget.insert(tk.END, table_msg)
            self.text_widget.see(tk.END)

    def periodic_update(self):
        self.update_text_widget()
        if combined_cmd_window_chk:
            self.after_id = self.root.after(1000, self.periodic_update)
        else:  # inserted
            self.after_id = None

    def start_text(self):
        entered_text = self.text_widget.get('1.0', 'end-1c')
        if not entered_text.strip():
            self.text_widget.insert('end-1c', 'Enter command> ')
        self.update_text_widget()
        self.after_id = self.root.after(1000, self.periodic_update)
        self.root.protocol('WM_DELETE_WINDOW', self.on_closing)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.root.mainloop()

    def on_closing(self):
        if self.after_id:
            self.root.after_cancel(self.after_id)
        self.root.destroy()
        self.app_instance.deiconify()

    def on_enter_pressed(self, event):
        entered_text = self.text_widget.get('end-1l linestart', 'end-1l lineend')
        if '>' in entered_text:
            entered_command = entered_text.split('>')[(-1)].strip()
            self.update_text_widget()
            self.after_id = self.root.after(1000, self.periodic_update)
            if 'stoppp' not in combined_cmd_window_chk:
                combined_cmd_window_chk.append('stoppp')
            threading.Thread(target=RRSG_enter_command, args=(self.app_instance, [str(entered_command)], 'nochatbuddy')).start()

class TextUpdater2:
    def __init__(self, app_instance):
        self.app_instance = app_instance
        self.root = tk.Tk()
        self.root.title('OSS Terminal')
        self.text_widget = tk.Text(self.root, wrap='none', width=1500, height=600, bg='#e6f7f6')
        self.text_widget.grid(row=0, column=0, sticky='nsew')
        font_size = 8
        new_font = ('Consolas', font_size)
        self.text_widget.configure(font=new_font)
        self.scrollbar_y = tk.Scrollbar(self.root, command=self.text_widget.yview)
        self.scrollbar_y.grid(row=0, column=1, sticky='ns')
        self.text_widget.config(yscrollcommand=self.scrollbar_y.set)
        self.scrollbar_x = tk.Scrollbar(self.root, orient=tk.HORIZONTAL, command=self.text_widget.xview)
        self.scrollbar_x.grid(row=1, column=0, sticky='ew')
        self.text_widget.config(xscrollcommand=self.scrollbar_x.set)
        self.update_counter = 0
        self.after_id = None
        self.center_window()
        self.start_text()

    def center_window(self):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = 1725
        window_height = 750
        scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100
        if scaleFactor == 1.5:
            x_position = screen_width + 540 - window_width
            y_position = screen_height + 150 - window_height
        else:  # inserted
            x_position = screen_width + 300 - window_width
            y_position = screen_height + 20 - window_height
        self.root.geometry(f'{window_width}x{window_height}+{x_position}+{y_position}')

    def update_text_widget(self):
        return

    def start_text(self):
        CHATBUDDY_textbox2 = self.app_instance.textbox2.get('1.0', 'end')
        self.text_widget.insert('end-1c', CHATBUDDY_textbox2)
        self.update_text_widget()
        self.root.protocol('WM_DELETE_WINDOW', self.on_closing)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.root.mainloop()

    def on_closing(self):
        if self.after_id:
            self.root.after_cancel(self.after_id)
        self.root.destroy()
        self.app_instance.deiconify()

def BASELINE_CHK(DFF22, baselin_DF, cmd_exxx):
    if cmd_exxx.lower() == 'altk':
        try:
            baselllin_DFhh = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='Service impacting alarm list', usecols='A,D', engine='openpyxl')
            ALTK_REQUIRED = baselllin_DFhh.columns[1]
        except:
            ALTK_REQUIRED = 'Not Required'
        if ALTK_REQUIRED == 'Required':
            Alarm_required_list = baselllin_DFhh['Specific Problem'].unique()
            Alarm_required_list = list(Alarm_required_list)
            try:
                def conditione(row):
                    if row['Severty'] == 'C' or row['Severty'] == 'M':
                        return 'Fail'
                    return 'Pass'
                DFF22['Baseline Check'] = DFF22.apply(conditione, axis=1)
                DFF22.loc[DFF22['Specific Problem'].isin(Alarm_required_list), 'Baseline Check'] = 'Fail'
                return DFF22
            except:
                return None
        if ALTK_REQUIRED == 'Only Service impacting alarm':
            Alarm_required_list = baselllin_DFhh['Specific Problem'].unique()
            Alarm_required_list = list(Alarm_required_list)
            try:
                def conditione(row):
                    if row['Severty'] == 'C' or row['Severty'] == 'M':
                        return 'Fail'
                    return 'Pass'
                DFF22['Baseline Check'] = DFF22.apply(conditione, axis=1)
                DFF22['Baseline Check'] = DFF22['Specific Problem'].apply(lambda x: 'Fail' if x in Alarm_required_list else 'Pass')
                return DFF22
            except:
                return None
        try:
            def conditione(row):
                if row['Severty'] == 'C' or row['Severty'] == 'M':
                    return 'Fail'
                return 'Pass'
            DFF22['Baseline Check'] = DFF22.apply(conditione, axis=1)
            return DFF22
        except:
            return None
    if cmd_exxx.lower() == 'alarm':
        try:
            baselllin_DFhh = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='Service impacting alarm list', usecols='A,D', engine='openpyxl')
            ALTK_REQUIRED = baselllin_DFhh.columns[1]
        except:
            ALTK_REQUIRED = 'Not Required'
        if ALTK_REQUIRED == 'Required':
            Alarm_required_list = baselllin_DFhh['Specific Problem'].unique()
            Alarm_required_list = list(Alarm_required_list)
            try:
                def conditione(row):
                    if row['Severty'] == 'C' or row['Severty'] == 'M':
                        return 'Fail'
                    return 'Pass'
                DFF22['Baseline Check'] = DFF22.apply(conditione, axis=1)
                DFF22.loc[DFF22['Specific Problem'].isin(Alarm_required_list), 'Baseline Check'] = 'Fail'
                return DFF22
            except:
                return None
        if ALTK_REQUIRED == 'Only Service impacting alarm':
            Alarm_required_list = baselllin_DFhh['Specific Problem'].unique()
            Alarm_required_list = list(Alarm_required_list)
            try:
                def conditione(row):
                    if row['Severty'] == 'C' or row['Severty'] == 'M':
                        return 'Fail'
                    return 'Pass'
                DFF22['Baseline Check'] = DFF22.apply(conditione, axis=1)
                DFF22['Baseline Check'] = DFF22['Specific Problem'].apply(lambda x: 'Fail' if x in Alarm_required_list else 'Pass')
                return DFF22
            except:
                return None
        try:
            def conditione(row):
                if row['Severty'] == 'C' or row['Severty'] == 'M':
                    return 'Fail'
                return 'Pass'
            DFF22['Baseline Check'] = DFF22.apply(conditione, axis=1)
            return DFF22
        except:
            return None
    baselin_DF_pro = baselin_DF.copy()
    try:
        baselin_DF_pro = baselin_DF_pro.loc[baselin_DF_pro['Tab Name'] == cmd_exxx]
    except:
        baselin_DF_pro = pd.DataFrame()
    if not baselin_DF_pro.empty:
        try:
            del baselin_DF_pro['Tab Name']
        except:
            pass
        try:
            if cmd_exxx.lower() == 'xyzzz':
                try:
                    def conditione(row):
                        if row['Severty'] == 'C' or row['Severty'] == 'M':
                            return 'Fail'
                        return 'Pass'
                    DFF22['Baseline Check'] = DFF22.apply(conditione, axis=1)
                except:
                    pass
            else:  # inserted
                DFF = DFF22.copy()
                baselllin_lst = baselin_DF_pro['Index'].unique()
                baselllin_lst = list(baselllin_lst)
                crate_df_chk = DFF.columns.tolist()
                crate_df_chk = [col.strip().lower() for col in crate_df_chk]
                baselllin_lst2 = baselin_DF_pro['MO1'].unique()
                baselllin_lst2 = list(baselllin_lst2)
                baselllin_lst2 = [baselllin_lst2 for baselllin_lst2 in baselllin_lst2 if str(baselllin_lst2)!= 'nan']
                for ccl in baselllin_lst:
                    if baselllin_lst2:
                        try:
                            if ccl.lower() in crate_df_chk:
                                DFF['MO1'] = DFF['MO'].astype(str)
                                for pattern in baselllin_lst2:
                                    DFF['MO1'] = DFF['MO1'].apply(lambda x: re.sub(pattern, pattern, x))
                                for string in baselllin_lst2:
                                    DFF['MO1'] = DFF['MO1'].apply(lambda x: string if string in x else x)
                                DFF['MO1'] = DFF['MO1'].apply(lambda x: '' if '.*' not in x else x)
                                baselin_DF_pro = baselin_DF_pro.loc[baselin_DF_pro['Index'].str.lower() == ccl.lower()]
                                DFF = DFF.merge(baselin_DF_pro, on='MO1', how='left')
                                try:
                                    del DFF['check']
                                except:
                                    pass
                                try:
                                    del DFF['Index']
                                except:
                                    pass
                                for ffg in DFF.columns:
                                    try:
                                        DFF[ffg] = DFF[ffg].astype(float)
                                    except:
                                        continue
                                if DFF['Threshold'].dtype == 'float':
                                    DFF['Value2'] = DFF[ccl]
                                    DFF[ccl] = pd.to_numeric(DFF[ccl], errors='coerce')
                                    DFF['Condition'].fillna('==', inplace=True)
                                DFF['Baseline Check'] = DFF.apply(lambda row, ccl=ccl: 'Fail' if eval(f"row[ccl] {row['Condition']} row[\'Threshold\']") else 'Pass', axis=1)
                                DFF['MO1'].replace('', np.nan, inplace=True)
                                DFF['MO1'].fillna('XXXXXXXX', inplace=True)
                                DFF.loc[DFF['MO1'] == 'XXXXXXXX', 'Baseline Check'] = np.nan
                                DFF.loc[DFF['Attribute']!= DFF['Attribute '], 'Baseline Check'] = ''
                                try:
                                    del DFF['MO1']
                                except:
                                    pass
                                try:
                                    del DFF['Value']
                                except:
                                    pass
                                try:
                                    del DFF['Attribute ']
                                except:
                                    pass
                                DFF = DFF.rename(columns={'Value2': 'Value'})
                                break
                        except:
                            continue
                    else:  # inserted
                        try:
                            if ccl.lower() in crate_df_chk:
                                baselin_DF_pro['check'] = 'check'
                                DFF['check'] = 'check'
                                baselin_DF_pro = baselin_DF_pro.loc[baselin_DF_pro['Index'].str.lower() == ccl.lower()]
                                DFF = DFF.merge(baselin_DF_pro, on='check', how='left')
                                try:
                                    del DFF['check']
                                except:
                                    pass
                                try:
                                    del DFF['Index']
                                except:
                                    pass
                                for ffg in DFF.columns:
                                    try:
                                        DFF[ffg] = DFF[ffg].astype(float)
                                    except:
                                        continue
                                DFF['Baseline Check'] = DFF.apply(lambda row, ccl=ccl: 'Fail' if eval(f"row[ccl] {row['Condition']} row[\'Threshold\']") else 'Pass', axis=1)
                                break
                        except:
                            continue
            try:
                del DFF['Condition']
            except:
                pass
            try:
                del DFF['Threshold']
                return DFF
            except:
                pass
                return DFF
        except:
            return DFF22

def check_empty_textbox_thread(self):
    try:
        time.sleep(60)
        check_empty_textbox(self)
    except:
        return None

def check_empty_textbox(self):
    try:
        text_content = self.CHATBUDDY_textbox4.get('1.0', 'end').strip()
        if not text_content:
            self.CHATBUDDY_textbox4.insert('0.10', 'Type your message..')
    except:
        return None

def check_empty_textbox_live(self):
    try:
        time.sleep(60)
        check__textbox_live(self)
    except:
        return None

def check__textbox_live(self):
    try:
        text_content = self.textbox8.get('1.0', 'end').strip()
        if not text_content:
            self.after(0, self.textbox8.delete, '1.0', 'end')
            self.after(0, self.textbox8.insert, 'end', 'Press enter to excute command...')
            self.after(0, self.textbox8.see, 'end')
    except:
        return None

def CIPRI_PRINT(self, table):
    self.after(0, self.textbox2.insert, 'end', table + '\n\n')
    self.after(0, self.textbox2.see, 'end')
    self.textbox2.tag_config('red', foreground='red')

def BASE_LINE_COLOUR(self, table, iii):
    Thrshlod_baseline = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='Baseline', usecols='a,b,c,d,e,f')
    threshold_value = None
    threshold_SYMBOL = '='
    for ind in Thrshlod_baseline['Index'].unique():
        if ind in table:
            threshold_value = Thrshlod_baseline[Thrshlod_baseline['Index'] == ind]['Threshold'].iloc[0]
    for ind in Thrshlod_baseline['Index'].unique():
        if ind in table:
            threshold_SYMBOL = Thrshlod_baseline[Thrshlod_baseline['Index'] == ind]['Condition'].iloc[0]
    lines = table.split('\n')
    table22 = []
    for ll in lines:
        if '---' not in ll:
            if ll.strip():
                table22.append(str(ll) + ' |')
        else:  # inserted
            if ll.strip():
                table22.append(str(ll))
    table = '\n'.join(table22)
    terminal_TEXT = self.textbox2.get('1.0', 'end').replace('\n', '').strip()
    self.after(0, self.textbox2.insert, 'end', table + '\n\n')
    self.after(0, self.textbox2.see, 'end')
    self.textbox2.tag_config('red', foreground='red')

    def find_separator(start, end):
        index = start
        pass
        index = self.textbox2.search('|', index, end)
        if not index or self.textbox2.compare(index, '>', end):
            return None
        return index
    start_index = self.textbox2.search('|', '1.0', 'end')
    while start_index:
        end_index = find_separator(f'{start_index} + 1 chars', 'end')
        if not end_index:
            return
        try:
            value = float(self.textbox2.get(f'{start_index} + 1 chars', end_index).strip())
            if threshold_SYMBOL == '>=':
                if threshold_value is not None:
                    if value >= float(threshold_value):
                        self.textbox2.tag_add('red', f'{start_index} + 1 chars', end_index)
            else:  # inserted
                if threshold_SYMBOL == '!=':
                    if threshold_value is not None:
                        if value!= float(threshold_value):
                            self.textbox2.tag_add('red', f'{start_index} + 1 chars', end_index)
                else:  # inserted
                    if threshold_SYMBOL == '<=':
                        if threshold_value is not None:
                            if value <= float(threshold_value):
                                self.textbox2.tag_add('red', f'{start_index} + 1 chars', end_index)
                    else:  # inserted
                        if threshold_SYMBOL == '>':
                            if threshold_value is not None and value > float(threshold_value):
                                self.textbox2.tag_add('red', f'{start_index} + 1 chars', end_index)
                        else:  # inserted
                            if threshold_SYMBOL == '<':
                                if threshold_value is not None and value < float(threshold_value):
                                    self.textbox2.tag_add('red', f'{start_index} + 1 chars', end_index)
                            else:  # inserted
                                if threshold_SYMBOL == '=':
                                    if threshold_value is not None and value == float(threshold_value):
                                        self.textbox2.tag_add('red', f'{start_index} + 1 chars', end_index)
                                else:  # inserted
                                    if threshold_value is not None:
                                        if value == float(threshold_value):
                                            self.textbox2.tag_add('red', f'{start_index} + 1 chars', end_index)
        except ValueError:
            pass
        start_index = find_separator(end_index, 'end')

def initial_kpi_format(KPIBASELINE_excel_file, existing_excel_file, KPI_ThresholdSH, INITIAL_KPI_TAB, check_tec):
    try:
        thresholds_df = pd.read_excel(KPIBASELINE_excel_file, sheet_name=KPI_ThresholdSH)
        thresholds_df = thresholds_df.loc[thresholds_df['TECH'] == check_tec]
        thresholds_df = thresholds_df[['KPI', 'Condition', 'Threshold', 'Status']]
        try:
            thresholds_df = thresholds_df.loc[thresholds_df['Status'] == 'Required']
            required_kpi = thresholds_df['KPI'].unique()
            required_kpi = list(required_kpi)
            colindexx = 4
        except:
            colindexx = 1
        df = pd.read_excel(existing_excel_file, sheet_name=INITIAL_KPI_TAB)
        DF_LST = df.columns
        allheaderlast = ['Node Id', 'Time', 'Cell Name', 'Object'] + required_kpi
        DF_LST = list(DF_LST)
        for col in DF_LST:
            if col not in allheaderlast:
                del df[col]
        FINAL_DF_LST = df.columns
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        wb = load_workbook(existing_excel_file)
        ws = wb[INITIAL_KPI_TAB]
        for index, row in thresholds_df.iterrows():
            column_name = row['KPI']
            condition = row['Condition']
            threshold = row['Threshold']
            col_idx = df.columns.get_loc(column_name) + colindexx
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for row_idx, cell_value in enumerate(df[column_name], start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if condition == '<' and float(cell_value) < float(threshold):
                        cell.fill = red_fill
                except (TypeError, ValueError):
                    pass
        try:
            columns_to_remove = []
            for col in ws.iter_cols():
                col_name = col[0].column_letter
                if ws[col_name + '1'].value not in FINAL_DF_LST:
                    columns_to_remove.append(col)
            for col in columns_to_remove:
                ws.delete_cols(col[0].column)
        except:
            pass
        wb.save(existing_excel_file)
    except:
        return None

def write_excel_new(dfff, excel_file, sheet_name, OUTPATH):
    try:
        sh_name = []
        dflst = []
        sheet_name1 = clean_sheet_name(sheet_name, max_length=28)
        sheet_name = str(sheet_name1).strip()
        sheet_name = sheet_name.replace('-', '').strip()
        if sh_map_lst:
            for ssh in sh_map_lst:
                try:
                    cmd_sh = str(ssh).split('@')[0].strip()
                    if cmd_sh.lower().strip() == sheet_name.lower().strip():
                        req_sh = str(ssh).split('@')[(-1)].strip()
                        sheet_name = req_sh
                except:
                    continue
        try:
            crate_df_base = BASELINE_CHK(dfff, baselllin_DF, sheet_name)
            if crate_df_base is not None:
                dfff = crate_df_base.copy()
        except:
            pass
        try:
            del crate_df_base
        except:
            pass
        try:
            with pd.ExcelFile(OUTPATH + excel_file) as xls:
                if sheet_name in xls.sheet_names:
                    all_sh = xls.sheet_names
                else:  # inserted
                    all_sh = xls.sheet_names + [sheet_name]
                for sh in all_sh:
                    if sh not in sh_name:
                        sh_name.append(sh)
                    if sh in xls.sheet_names:
                        existing_df = pd.read_excel(xls, sh)
                        if sheet_name == sh:
                            combined_df = pd.concat([existing_df, dfff], ignore_index=True)
                            duplst = combined_df.columns.tolist()
                            combined_df = combined_df.drop_duplicates(subset=duplst, keep='first')
                            dflst.append(combined_df)
                        else:  # inserted
                            duplst = existing_df.columns.tolist()
                            existing_df = existing_df.drop_duplicates(subset=duplst, keep='first')
                            dflst.append(existing_df)
                    else:  # inserted
                        duplst = dfff.columns.tolist()
                        dfff = dfff.drop_duplicates(subset=duplst, keep='first')
                        dflst.append(dfff)
                with pd.ExcelWriter(OUTPATH + excel_file, engine='xlsxwriter') as writer:
                    i = 0
                    for sh in sh_name:
                        h_lst = dflst[i].columns.tolist()
                        for header in h_lst:
                            for dfdf in ['VSWR']:
                                if dfdf in header:
                                    try:
                                        dflst[i] = dflst[i].replace('', pd.NA).dropna(subset=[header])
                                    except:
                                        pass
                        dflst[i] = dflst[i].dropna(axis=1, how='all')
                        dflst[i] = dflst[i].rename(columns=lambda x: x.strip())
                        dflst[i].to_excel(writer, sheet_name=sh, index=False)
                        try:
                            column = chr(ord('A') + dflst[i].columns.get_loc('Baseline Check'))
                            num_rows, num_cols = dflst[i].shape
                            last_cell = f"{chr(ord('A') + num_cols - 1)}{num_rows}"
                            fail_format = writer.book.add_format({'bg_color': 'red', 'font_color': 'white'})
                            writer.sheets[sh].conditional_format(f'{column + str(1)}:{last_cell + str(1)}', {'type': 'text', 'criteria': 'containing', 'value': 'Fail', 'format': fail_format})
                        except:
                            pass
                        headr_colour_multi(writer, dflst[i], sh, '#facf' + str('91'), ['KPI'], '#facf' + str('91'))
                        i = i + 1
        except FileNotFoundError:
            duplst = dfff.columns.tolist()
            dfff = dfff.drop_duplicates(subset=duplst, keep='first')
            with pd.ExcelWriter(OUTPATH + excel_file, engine='xlsxwriter') as writer:
                h_lst = dfff.columns.tolist()
                for header in h_lst:
                    for dfdf in ['VSWR']:
                        if dfdf in header:
                            try:
                                dfff = dfff.replace('', pd.NA).dropna(subset=[header])
                            except:
                                pass
                dfff = dfff.dropna(axis=1, how='all')
                dfff = dfff.rename(columns=lambda x: x.strip())
                dfff.to_excel(writer, sheet_name=sheet_name, index=False)
                headr_colour_multi(writer, dfff, sheet_name, '#facf' + str('91'), ['KPI'], '#facf' + str('91'))
        except:
            return
    except:
        return None

def write_excel_new_monitoring(dfff, excel_file, sheet_name, OUTPATH):
    try:
        sh_name = []
        dflst = []
        if sh_map_lst:
            for ssh in sh_map_lst:
                try:
                    cmd_sh = str(ssh).split('_')[0].strip()
                    if cmd_sh.lower().strip() == sheet_name.lower().strip():
                        req_sh = str(ssh).split('_')[(-1)].strip()
                        sheet_name = req_sh
                except:
                    continue
        try:
            crate_df_base = BASELINE_CHK(dfff, baselllin_DF, sheet_name)
            if crate_df_base is not None:
                dfff = crate_df_base.copy()
        except:
            pass
        try:
            del crate_df_base
        except:
            pass
        try:
            with pd.ExcelFile(OUTPATH + excel_file) as xls:
                if sheet_name in xls.sheet_names:
                    all_sh = xls.sheet_names
                else:  # inserted
                    all_sh = xls.sheet_names + [sheet_name]
                for sh in all_sh:
                    if sh not in sh_name:
                        sh_name.append(sh)
                    if sh in xls.sheet_names:
                        existing_df = pd.read_excel(xls, sh)
                        if sheet_name == sh:
                            combined_df = pd.concat([existing_df, dfff], ignore_index=True)
                            duplst = combined_df.columns.tolist()
                            combined_df = combined_df.drop_duplicates(subset=duplst, keep='first')
                            dflst.append(combined_df)
                        else:  # inserted
                            duplst = existing_df.columns.tolist()
                            existing_df = existing_df.drop_duplicates(subset=duplst, keep='first')
                            dflst.append(existing_df)
                    else:  # inserted
                        duplst = dfff.columns.tolist()
                        dfff = dfff.drop_duplicates(subset=duplst, keep='first')
                        dflst.append(dfff)
                with pd.ExcelWriter(OUTPATH + excel_file, engine='xlsxwriter') as writer:
                    i = 0
                    for sh in sh_name:
                        h_lst = dflst[i].columns.tolist()
                        for header in h_lst:
                            for dfdf in ['VSWR']:
                                if dfdf in header:
                                    try:
                                        dflst[i] = dflst[i].replace('', pd.NA).dropna(subset=[header])
                                    except:
                                        pass
                        dflst[i] = dflst[i].dropna(axis=1, how='all')
                        dflst[i] = dflst[i].rename(columns=lambda x: x.strip())
                        dflst[i].to_excel(writer, sheet_name=sh, index=False)
                        try:
                            column = chr(ord('A') + dflst[i].columns.get_loc('Baseline Check'))
                            num_rows, num_cols = dflst[i].shape
                            last_cell = f"{chr(ord('A') + num_cols - 1)}{num_rows}"
                            fail_format = writer.book.add_format({'bg_color': 'red', 'font_color': 'white'})
                            writer.sheets[sh].conditional_format(f'{column + str(1)}:{last_cell + str(1)}', {'type': 'text', 'criteria': 'containing', 'value': 'Fail', 'format': fail_format})
                        except:
                            pass
                        headr_colour_multi(writer, dflst[i], sh, '#facf' + str('91'), ['KPI'], '#facf' + str('91'))
                        i = i + 1
        except FileNotFoundError:
            duplst = dfff.columns.tolist()
            dfff = dfff.drop_duplicates(subset=duplst, keep='first')
            with pd.ExcelWriter(OUTPATH + excel_file, engine='xlsxwriter') as writer:
                h_lst = dfff.columns.tolist()
                for header in h_lst:
                    for dfdf in ['VSWR']:
                        if dfdf in header:
                            try:
                                dfff = dfff.replace('', pd.NA).dropna(subset=[header])
                            except:
                                pass
                dfff = dfff.dropna(axis=1, how='all')
                dfff = dfff.rename(columns=lambda x: x.strip())
                dfff.to_excel(writer, sheet_name=sheet_name, index=False)
                headr_colour_multi(writer, dfff, sheet_name, '#facf' + str('91'), ['KPI'], '#facf' + str('91'))
        except:
            return
    except:
        return None

def melted_conter(df):
    try:
        try:
            df = df.applymap(lambda x: str(x).strip())
        except:
            pass
        numeric_columns = df.columns.tolist()
        try:
            numeric_columns.remove('Node Id')
        except:
            pass
        try:
            numeric_columns.remove('Time')
        except:
            pass
        try:
            numeric_columns.remove('Counter')
        except:
            pass
        try:
            numeric_columns.remove('Object')
        except:
            pass
        df[numeric_columns] = df[numeric_columns].apply(pd.to_numeric, errors='coerce')
        try:
            df_transformed = df.melt(id_vars=['Node Id', 'Time', 'Counter'], var_name='Cell Name', value_name='Value')
            df_transformed = df_transformed.pivot_table(index=['Node Id', 'Time', 'Cell Name'], columns='Counter', values='Value').reset_index()
        except:
            df_transformed = df.melt(id_vars=['Node Id', 'Object', 'Counter'], var_name='Time', value_name='Value')
            df_transformed = df_transformed.pivot_table(index=['Node Id', 'Object', 'Time'], columns='Counter', values='Value').reset_index()
        try:
            df_transformed = df_transformed.dropna(axis=1, how='all')
            return df_transformed
        except:
            pass
            return df_transformed
    except:
        return df

def amoooss(self):
    threading.Thread(target=disable_button, args=(self,)).start()
    LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
    if not LIVECMD_TEXT.strip().endswith('>'):
        NODEID = get_nodeid()
        if NODEID:
            if NODEID!= None:
                try:
                    NODEID = NODEID.lower().replace('amos', '').strip()
                except:
                    pass

                def eneter():
                    self.after(0, self.textbox8.delete, '1.0', 'end')
                    self.after(0, self.textbox8.insert, 'end', str('amos ' + str(NODEID).upper()))
                    self.after(0, self.textbox8.see, 'end')
                try:
                    nodedb = pd.read_excel(INPATH + 'Node_db.xlsx', sheet_name='Nodedb', usecols='A')
                    nodedb_lst = nodedb['Node list'].unique()
                    nodedb_lst = list(nodedb_lst)
                    for nnd in nodedb_lst:
                        if NODEID.lower() in nnd.lower():
                            NODEID = str(nnd).strip().upper()
                            break
                except:
                    pass
                eneter()
                RRSG_enter_command(self, ['amos ' + str(NODEID).upper()], 'nochatbuddy')
                RRSG_enter_command(self, ['lt all'], 'nochatbuddy')
                RBS_RBS_switch = self.swich_Sec_rbs.get()
                if RBS_RBS_switch == 1:
                    RBS_RBS_switch_loe = 'YES'
                else:  # inserted
                    RBS_RBS_switch_loe = 'NO'
                if RBS_RBS_switch_loe == 'YES':
                    if RBSPASS_CRED_CRREDD[0] == 'rbs':
                        user_rbs = 'rbs'
                    else:  # inserted
                        try:
                            user_rbs = RBSPASS_CRED_CRREDD[0]
                        except:
                            user_rbs = 'rbs'
                    if RBSPASS_CRED_CRREDD[1] == 'rbs':
                        pass_rbs = 'rbs'
                    else:  # inserted
                        try:
                            pass_rbs = RBSPASS_CRED_CRREDD[1]
                        except:
                            pass_rbs = 'rbs'
                    RRSG_enter_command(self, [str(user_rbs)], 'nochatbuddy')
                    RRSG_enter_command(self, [str(pass_rbs)], 'nochatbuddy')
                thread = threading.Thread(target=enabled_button, args=(self,))
                thread.start()
                thread.join()
                if 'cancle' not in processdone:
                    cmmmddd.clear()
                    siteddd.clear()
                    checknxtsit.clear()
                    processdone.clear()
                else:  # inserted
                    thread = threading.Thread(target=enabled_button, args=(self,))
                    thread.start()
                    thread.join()
    else:  # inserted
        thread = threading.Thread(target=enabled_button, args=(self,))
        thread.start()
        thread.join()

def OTPPP_tets():
    if VPN_select_METHOD_F[0] == 'VPN-CAS':
        try:
            db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_VPN_CRED.db')
            con = sqlite3.connect(db_path)
            cursor = con.cursor()
            cursor.execute('CREATE TABLE IF NOT EXISTS Credd (value TEXT)')
            con.commit()
            Rdf = pd.read_sql_query('SELECT * FROM Credd', con)
            ppPID_LST = Rdf['value'].tolist()
            for ppp in ppPID_LST:
                VPN_RSSG_CRREDD.append(decrypt(ppp, 3))
            try:
                ppPID_LST.clear()
            except:
                pass
            return VPN_RSSG_CRREDD[1]
        except:
            return None
    else:  # inserted
        class OTPDialog(QDialog):
            def __init__(self):
                super().__init__()
                self.init_ui()

            def init_ui(self):
                self.setWindowTitle('OTP')
                self.setGeometry(800, 400, 190, 100)
                layout = QVBoxLayout()
                self.label = QLabel('   RSG One Time OTP', self)
                self.label.setStyleSheet('padding: 0px; font: 9pt \'Segoe UI Semibold\';')
                self.edit = QLineEdit(self)
                self.edit.setEchoMode(QLineEdit.Password)
                self.edit.returnPressed.connect(self.process_otp)
                self.edit.setFixedWidth(190)
                self.edit.setAlignment(Qt.AlignCenter)
                layout.addWidget(self.label)
                layout.addWidget(self.edit)
                self.setLayout(layout)
                self.setStyleSheet('\n                    QDialog {\n                        background-color: #e6e6e6;\n                    }\n                    QLabel {\n                        font: 14pt \"Calibri\";\n                    }\n                    QLineEdit {\n                        font: 12pt \"Calibri\";\n                        padding: 4px;\n                        border: 2px solid #4CAF50;\n                        border-radius: 4px;\n                    }\n                ')
                self.setWindowFlag(Qt.WindowStaysOnTopHint)

            def process_otp(self):
                user_input = self.edit.text()
                if user_input:
                    self.accept()
                else:  # inserted
                    QMessageBox.warning(self, 'Warning', 'Please enter a valid OTP.')

        def get_otp():
            app = QApplication([])
            dialog = OTPDialog()
            result = dialog.exec_()
            if result == QDialog.Accepted:
                user_input = dialog.edit.text()
                return user_input
            return None
        otp = get_otp()
        return otp

def ENM_BASELINE_AUDIT(para_file, Dump1_Name, TECHNO, INPATH134, OUTPATH134, BASE_INPATH):
    def headr_colour_multi(writer2, dfff, sh_name, colour, hed_lst, tab):
        workbook = writer2.book
        worksheet = writer2.sheets[sh_name]
        worksheet.set_tab_color(tab)
        workbook.formats[0].set_font_size(9)
        workbook.formats[0].set_align('center')
        workbook.formats[0].set_font('Calibri')
        header_format = workbook.add_format({'bold': True, 'font_name': 'Calibri', 'font_size': 10, 'font_color': '#FFFFFF', 'bg_color': '#262626', 'align': 'left', 'valign': 'top', 'text_wrap': False})
        header_format_row = workbook.add_format({'font_name': 'Calibri', 'font_size': 9, 'align': 'left', 'valign': 'top', 'text_wrap': False})
        for col_num, value in enumerate(dfff.columns.values):
            worksheet.write(0, col_num, value, header_format)
        num_rows = worksheet.dim_rowmax
        for row_num in range(num_rows + 1):
            worksheet.set_row(row_num, None, header_format_row)
        for column in dfff:
            if column!= 'MO' and column!= 'NodeId':
                column_length = max(dfff[column].astype(str).apply(len).max(), len(column))
                col_idx = dfff.columns.get_loc(column)
                writer2.sheets[sh_name].set_column(col_idx, col_idx, column_length)
        return dfff

    def FEATURE_AUDIT(TECHNO, MO_NAME, Feature_state_df):
        try:
            if MO_NAME == 'FeatureState':
                Bas_line2 = pd.read_excel(BASE_INPATH + para_file, sheet_name='Parameter_Baseline', usecols='a,b,c,d,e', engine='openpyxl')
                Bas_line2 = Bas_line2.loc[Bas_line2['Tech'] == TECHNO]
                Bas_line2 = Bas_line2[['MO Class', 'Parameters', 'Value', 'MO1']]
                Bas_line2 = Bas_line2.loc[Bas_line2['MO Class'] == 'FeatureState']
                Bas_line2 = Bas_line2.rename(columns={'Parameters': 'featureStateId'})
                Bas_line2 = Bas_line2[['featureStateId', 'Value']]
                Feature_state_df = Feature_state_df[['MO', 'featureStateId', 'featureState']]
                Feature_state_df = Feature_state_df.merge(Bas_line2, on='featureStateId', how='left')
                Feature_state_df['Value'].replace('', np.nan, inplace=True)
                Feature_state_df.dropna(subset=['Value'], inplace=True)
                Feature_state_df['check'] = Feature_state_df.apply(lambda row: 'True' if row['licenseState'] == row['Value'] else 'False', axis=1)
                Feature_state_df = Feature_state_df.loc[Feature_state_df['check'] == 'False']
                del Feature_state_df['check']
                Feature_state_df = Feature_state_df.rename(columns={'featureStateId': 'Parameter'})
                Feature_state_df = Feature_state_df.rename(columns={'licenseState': 'Current Value'})
                Feature_state_df = Feature_state_df.rename(columns={'Value': 'Required Value'})
                MO_LST = Feature_state_df['MO'].tolist()
                MO_LST = [MO_LST for MO_LST in MO_LST if str(MO_LST)!= 'nan']
                if TECHNO == 'WCDMA':
                    mecotext_CELL(MO_LST, 'WCDMA')
                if TECHNO == 'LTE':
                    mecotext_CELL(MO_LST, 'LTE')
                if TECHNO == 'GSM':
                    mecotext_CELL(MO_LST, 'LTE')
                if TECHNO == 'VOLTE':
                    mecotext_CELL(MO_LST, 'LTE')
                if TECHNO == 'NR':
                    mecotext_CELL(MO_LST, 'LTE')
                if TECHNO == 'WCDMA':
                    Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                    Feature_state_df['RNC'] = mecotext_CELL.F_RNC_BSC
                if TECHNO == 'GSM':
                    Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                    Feature_state_df['BSC'] = mecotext_CELL.F_RNC_BSC
                if TECHNO == 'LTE':
                    Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                    Feature_state_df['RNC'] = ''
                if TECHNO == 'VOLTE':
                    Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                    Feature_state_df['RNC'] = ''
                if TECHNO == 'NR':
                    Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                    Feature_state_df['RNC'] = ''
                Feature_state_df['MO class'] = MO_NAME
                Feature_state_df = Feature_state_df[['MO', 'NodeId', 'MO class', 'Parameter', 'Current Value', 'Required Value']]
                try:
                    lssst = Feature_state_df.columns
                    Feature_state_df = Feature_state_df.drop_duplicates(subset=list(lssst), keep='first')
                except:
                    pass
                Feature_state_df = Feature_state_df.loc[Feature_state_df['Current Value']!= 'nan']
                try:
                    Feature_state_df['Current Value'] = pd.to_numeric(Feature_state_df['Current Value'])
                except:
                    Feature_state_df['Current Value'] = Feature_state_df['Current Value'].astype(str)
                DF_STORE.append(Feature_state_df)
                del Feature_state_df
        except:
            pass
        if MO_NAME == 'RncFeature':
            Bas_line2 = pd.read_excel(BASE_INPATH + para_file, sheet_name='Parameter_Baseline', usecols='a,b,c,d,e', engine='openpyxl')
            Bas_line2 = Bas_line2.loc[Bas_line2['Tech'] == TECHNO]
            Bas_line2 = Bas_line2[['MO Class', 'Parameters', 'Value', 'MO1']]
            Bas_line2 = Bas_line2.loc[Bas_line2['MO Class'] == 'RncFeature']
            Bas_line2 = Bas_line2.rename(columns={'Parameters': 'keyId'})
            Bas_line2 = Bas_line2[['keyId', 'Value']]
            Feature_state_df = Feature_state_df[['MO', 'keyId', 'featureState']]
            Feature_state_df = Feature_state_df.merge(Bas_line2, on='keyId', how='left')
            Feature_state_df['Value'].replace('', np.nan, inplace=True)
            Feature_state_df.dropna(subset=['Value'], inplace=True)
            Feature_state_df['check'] = Feature_state_df.apply(lambda row: 'True' if row['featureState'] == row['Value'] else 'False', axis=1)
            Feature_state_df = Feature_state_df.loc[Feature_state_df['check'] == 'False']
            del Feature_state_df['check']
            Feature_state_df = Feature_state_df.rename(columns={'keyId': 'Parameter'})
            Feature_state_df = Feature_state_df.rename(columns={'featureState': 'Current Value'})
            Feature_state_df = Feature_state_df.rename(columns={'Value': 'Required Value'})
            MO_LST = Feature_state_df['MO'].tolist()
            MO_LST = [MO_LST for MO_LST in MO_LST if str(MO_LST)!= 'nan']
            if TECHNO == 'WCDMA':
                mecotext_CELL(MO_LST, 'WCDMA')
            if TECHNO == 'LTE':
                mecotext_CELL(MO_LST, 'LTE')
            if TECHNO == 'GSM':
                mecotext_CELL(MO_LST, 'LTE')
            if TECHNO == 'VOLTE':
                mecotext_CELL(MO_LST, 'LTE')
            if TECHNO == 'NR':
                mecotext_CELL(MO_LST, 'LTE')
            if TECHNO == 'WCDMA':
                Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                Feature_state_df['RNC'] = mecotext_CELL.F_RNC_BSC
            if TECHNO == 'GSM':
                Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                Feature_state_df['BSC'] = mecotext_CELL.F_RNC_BSC
            if TECHNO == 'LTE':
                Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                Feature_state_df['RNC'] = ''
            if TECHNO == 'VOLTE':
                Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                Feature_state_df['RNC'] = ''
            if TECHNO == 'NR':
                Feature_state_df['NodeId'] = mecotext_CELL.CELL_MO
                Feature_state_df['RNC'] = ''
            Feature_state_df['MO class'] = MO_NAME
            Feature_state_df = Feature_state_df[['MO', 'NodeId', 'RNC', 'MO class', 'Parameter', 'Current Value', 'Required Value']]
            try:
                lssst = Feature_state_df.columns
                Feature_state_df = Feature_state_df.drop_duplicates(subset=list(lssst), keep='first')
            except:
                pass
            Feature_state_df = Feature_state_df.loc[Feature_state_df['Current Value']!= 'nan']
            try:
                Feature_state_df['Current Value'] = pd.to_numeric(Feature_state_df['Current Value'])
            except:
                Feature_state_df['Current Value'] = Feature_state_df['Current Value'].astype(str)
            DF_STORE.append(Feature_state_df)
            del Feature_state_df

    def write_chukes_large_df(DF, Filename, TECHNO, sh_name):
        if DF.empty:
            with open(OUTPATH134 + 'Note ' + datetime.now().strftime('%d%m%Y') + '.txt', 'w') as f:
                f.writelines('No discrepancy found\n')
            f.close()
            return
        if len(sh_name) > 31:
            sh_name = sh_name[:len(sh_name) - 31]
        yesee = datetime.now()
        dateee = yesee.strftime('%d%m%Y')
        fileee = ' ' + Filename + ' '
        ddump_FILE = TECHNO + fileee + dateee

        def headr_colour():
            workbook = writer.book
            workbook.formats[0].set_font_size(10)
            workbook.formats[0].set_align('center')
            workbook.formats[0].set_font('Calibri')
        chunk_size = 800000
        k = 0
        for i in range(0, len(DF), chunk_size):
            if k == 0:
                writer = pd.ExcelWriter(OUTPATH134 + ddump_FILE + '.xlsx'.format(k), engine='xlsxwriter')
            if k > 0:
                writer = pd.ExcelWriter(OUTPATH134 + ddump_FILE + '_{:02d}.xlsx'.format(k), engine='xlsxwriter')
            df_chunk = DF.iloc[i:i + chunk_size]
            df_chunk.T.reset_index().T.to_excel(writer, sheet_name=sh_name, engine='xlsxwriter', index=False, header=None)
            writer.save()
            k += 1

    def mecotext_CELL(MO_LS, techh):
        mecotext_CELL.CELL_MO = []
        mecotext_CELL.F_RNC_BSC = []
        if techh == 'GSM':
            CELLLNAMEE = 'GeranCell'
        if techh == 'WCDMA':
            CELLLNAMEE = 'UtranCell'
        if techh == 'LTE':
            CELLLNAMEE = 'EUtranCellFDD'
        for kk in MO_LS:
            xx = 'MeContext' in kk
            if xx == True:
                try:
                    lst = str(kk).split(',')
                    new_ls = []
                    for zz in lst:
                        lstring = str(zz).split('=')[0]
                        new_ls.append(lstring)
                    inde = [i for i, x in enumerate(new_ls) if x == 'MeContext']
                    try:
                        indxxx = inde[0]
                        new_ls.clear()
                        cell_site = str(kk).split(',')[indxxx]
                        Final_cell = str(cell_site).split('=')[1]
                        mecotext_CELL.F_RNC_BSC.append(Final_cell)
                    except:
                        new_ls.clear()
                        mecotext_CELL.F_RNC_BSC.append('NA')
                except:
                    pass
            else:  # inserted
                mecotext_CELL.F_RNC_BSC.append('NA')
            yy = CELLLNAMEE in kk
            if yy == True:
                try:
                    lst = str(kk).split(',')
                    new_ls = []
                    for zz in lst:
                        lstring = str(zz).split('=')[0]
                        new_ls.append(lstring)
                    inde = [i for i, x in enumerate(new_ls) if x == CELLLNAMEE]
                    try:
                        indxxx = inde[0]
                        new_ls.clear()
                        cell_site = str(kk).split(',')[indxxx]
                        Final_cell = str(cell_site).split('=')[1]
                        mecotext_CELL.CELL_MO.append(Final_cell)
                    except:
                        new_ls.clear()
                        mecotext_CELL.CELL_MO.append(kk)
                except:
                    continue
            else:  # inserted
                mecotext_CELL.CELL_MO.append(kk)
    try:
        para_SELECT_PARA_req = 'ALL'
        CELLS_LIT = []
        para_SELECT_para = 'ALL'
        mult_excel = 'Combined Dump'
        CELL_M0 = 'MO'
        cell_list_req = 'NO'
        Bas_line = pd.read_excel(BASE_INPATH + para_file, sheet_name='Parameter_Baseline', usecols='a,b,c,d,e', engine='openpyxl')
        Bas_line = Bas_line.loc[Bas_line['Tech'] == TECHNO]
        Bas_line_new = Bas_line.copy()
        Bas_line = Bas_line[['MO Class', 'Parameters', 'Value', 'MO1']]
        Bas_MOLST = Bas_line['MO Class'].unique()
        Bas_MOLST = list(Bas_MOLST)
        Pararr_LST = Bas_line['Parameters'].unique()
        Pararr_LST = list(Pararr_LST)
        try:
            Bas_line['Parameters2'] = Bas_line['Parameters']
        except:
            pass
        try:
            Bas_line['Parameters'] = Bas_line['Parameters'].astype(str).str.upper()
        except:
            pass
        try:
            Bas_line['MO Class'].replace('', np.nan, inplace=True)
        except:
            pass
        try:
            Bas_line.dropna(subset=['MO Class'], inplace=True)
        except:
            pass
        Bas_line['MO_PARA'] = Bas_line['MO Class'].astype(str) + '$' + Bas_line['Parameters'].astype(str)
        MO_PARA_LST = list(Bas_line['MO_PARA'])
        F_mo_lst = []
        [F_mo_lst.append(x) for x in MO_PARA_LST if x not in F_mo_lst]
        F_mo_lst = [F_mo_lst for F_mo_lst in F_mo_lst if str(F_mo_lst)!= 'nan']
        Filter_Bas_line = Bas_line.loc[Bas_line['Value']!= 'NR']
        Filter_Bas_line['MO1'] = Filter_Bas_line['MO1'].astype(str).replace('NA', np.nan, regex=True).replace(['nan', 'None'], np.nan)
        Filter_Bas_line['MO1'] = Filter_Bas_line['MO1'].astype(str).replace('NR', np.nan, regex=True).replace(['nan', 'None'], np.nan)
        Filter_Bas_line['MO1'] = Filter_Bas_line['MO1'].astype(str).replace('', np.nan, regex=True).replace(['nan', 'None'], np.nan)
        Filter_Bas_line['concol'] = Filter_Bas_line.apply(lambda row: row['Parameters'] + '_' + row['MO1'] if pd.notnull(row['MO1']) else row['Parameters'], axis=1)
        if mult_excel == 'Combined Dump':
            WS_Df = pd.ExcelFile(INPATH134 + Dump1_Name)
            sheet_lst = WS_Df.sheet_names
            sheet_lst = [x for x in sheet_lst if x in Bas_MOLST]
        try:
            CELL_M0 = CELL_M0.upper()
        except:
            pass
        sheet_lst2 = sheet_lst.copy()
        DF_STORE = []
        i = 0
        for sh in sheet_lst2:
            if sh.lower() == 'FeatureState'.lower():
                M_DF1 = WS_Df.parse(sh)
                FEATURE_AUDIT(TECHNO, 'FeatureState', M_DF1)
            else:  # inserted
                if sh.lower() == 'RncFeature'.lower():
                    M_DF1 = WS_Df.parse(sh)
                    FEATURE_AUDIT(TECHNO, 'RncFeature', M_DF1)
                    continue
                try:
                    M_DF1 = WS_Df.parse(sh)
                    M_DF1_collst = M_DF1.columns
                    M_DF1_collst = [x for x in M_DF1_collst if x in Pararr_LST]
                    M_DF1 = M_DF1[['MO'] + M_DF1_collst]
                    M_DF1['MO1'] = M_DF1['MO'].astype(str)
                    Bas_line_new2 = Bas_line_new.copy()
                    Bas_line_new2 = Bas_line_new2.loc[Bas_line_new2['MO Class'] == str(sh)]
                    Bas_line_new2['MO1'].replace('', np.nan, inplace=True)
                    Bas_line_new2.dropna(subset=['MO1'], inplace=True)
                    base_lstt = Bas_line_new2['MO1'].unique()
                    del Bas_line_new2
                    for pattern in base_lstt:
                        M_DF1['MO1'] = M_DF1['MO1'].apply(lambda x: re.sub(pattern, pattern, x))
                    for string in base_lstt:
                        M_DF1['MO1'] = M_DF1['MO1'].apply(lambda x: string if string in x else x)
                    M_DF1['MO1'] = M_DF1['MO1'].apply(lambda x: '' if '.*' not in x else x)
                    try:
                        M_DF1.columns = [x.upper() for x in M_DF1.columns]
                    except:
                        pass
                    if CELL_M0 in M_DF1.columns:
                        try:
                            M_DF1[CELL_M0].fillna('NA22', inplace=True)
                        except:
                            pass
                        try:
                            M_DF1 = M_DF1.loc[M_DF1[CELL_M0]!= 'NA22']
                        except:
                            pass
                        try:
                            MO_LST = list(M_DF1[CELL_M0])
                        except:
                            pass
                        MO_LST = [MO_LST for MO_LST in MO_LST if str(MO_LST)!= 'nan']
                        if TECHNO == 'WCDMA':
                            mecotext_CELL(MO_LST, 'WCDMA')
                        if TECHNO == 'LTE':
                            mecotext_CELL(MO_LST, 'LTE')
                        if TECHNO == 'GSM':
                            mecotext_CELL(MO_LST, 'GSM')
                        if TECHNO == 'VOLTE':
                            mecotext_CELL(MO_LST, 'VOLTE')
                        if TECHNO == 'NR':
                            mecotext_CELL(MO_LST, 'NR')
                        col_net = list(M_DF1.columns)
                        col_newl = []
                        [col_newl.append(x) for x in col_net if x not in col_newl]
                        col_newl = [col_newl for col_newl in col_newl if str(col_newl)!= 'nan']
                        New_new = []
                        if mult_excel == 'Combined Dump':
                            for n_para in col_newl:
                                par_con = str(sh) + '$' + str(n_para)
                                New_new.append(par_con)
                        if mult_excel == 'Mult Excel Dump':
                            for n_para in col_newl:
                                par_con = str(sh) + '$' + str(n_para)
                                New_new.append(par_con)
                        if CELL_M0 in col_net:
                            chk_lst = list(map(lambda st: str.replace(st, str(sh) + '$', ''), F_mo_lst))
                            chk_lst = [item.upper() for item in chk_lst]
                            if TECHNO!= 'GSM' and chk_lst:
                                try:
                                    M_DF1[CELL_M0].fillna('NA22', inplace=True)
                                except:
                                    pass
                                try:
                                    M_DF1 = M_DF1.loc[M_DF1[CELL_M0]!= 'NA22']
                                except:
                                    pass
                                for para in chk_lst:
                                    if para_SELECT_PARA_req == 'ALL':
                                        try:
                                            if para!= CELL_M0:
                                                ndf = pd.DataFrame()
                                                try:
                                                    ndf[CELL_M0] = M_DF1[CELL_M0].astype(str)
                                                except:
                                                    pass
                                                try:
                                                    ndf['MO2'] = M_DF1['MO1'].astype(str)
                                                except:
                                                    pass
                                                if TECHNO == 'WCDMA':
                                                    ndf['NodeId'] = mecotext_CELL.CELL_MO
                                                    ndf['RNC'] = mecotext_CELL.F_RNC_BSC
                                                if TECHNO == 'GSM':
                                                    ndf['NodeId'] = mecotext_CELL.CELL_MO
                                                    ndf['RNC'] = mecotext_CELL.F_RNC_BSC
                                                if TECHNO == 'LTE':
                                                    ndf['NodeId'] = mecotext_CELL.CELL_MO
                                                    ndf['RNC'] = ''
                                                if TECHNO == 'VOLTE':
                                                    ndf['NodeId'] = mecotext_CELL.CELL_MO
                                                    ndf['RNC'] = ''
                                                if TECHNO == 'NR':
                                                    ndf['NodeId'] = mecotext_CELL.CELL_MO
                                                    ndf['RNC'] = ''
                                                if mult_excel == 'Combined Dump':
                                                    ndf['MO class'] = str(sh)
                                                if mult_excel == 'Mult Excel Dump':
                                                    ndf['MO class'] = str(sh)
                                                Filter_Bas_line2 = Filter_Bas_line.copy()
                                                Filter_Bas_line2 = Filter_Bas_line2.loc[Filter_Bas_line2['MO Class'] == str(sh)]
                                                Filter_Bas_line2['Parameters_2'] = Filter_Bas_line2['concol'].astype(str)
                                                da1 = ndf.copy()
                                                da1['Parameters'] = str(para)
                                                da1['pre_val'] = M_DF1[str(para.split('_')[0])].astype(str)

                                                def check_blank(row):
                                                    if row['MO2'] == '':
                                                        return row['Parameters']
                                                    return row['Parameters'] + '_' + row['MO2']
                                                da1['Parameters_2'] = da1.apply(lambda row: check_blank(row), axis=1)
                                                da1 = da1.merge(Filter_Bas_line2, on='Parameters_2', how='left')
                                                da1['Value'].fillna('NO Activity', inplace=True)
                                                da1 = da1.loc[da1['Value']!= 'NO Activity']
                                                try:
                                                    da1['pre_val'] = da1['pre_val'].astype(str)
                                                except:
                                                    pass
                                                try:
                                                    da1['pre_val'].fillna('NA22', inplace=True)
                                                except:
                                                    pass
                                                try:
                                                    da1['pre_val'] = pd.to_numeric(da1['pre_val'], errors='coerce')
                                                except:
                                                    da1['pre_val'] = da1['pre_val'].astype(str)
                                                da1['pre_val_N2'] = da1['pre_val']
                                                da1['Value_N2'] = da1['Value']
                                                try:
                                                    da1['pre_val'] = da1['pre_val'].astype(str).str.lower()
                                                except:
                                                    pass
                                                try:
                                                    da1['Value'] = da1['Value'].astype(str).str.lower()
                                                except:
                                                    pass
                                                try:
                                                    da1['Value'].replace('', np.nan, inplace=True)
                                                    da1.dropna(subset=['Value'], inplace=True)
                                                    da1 = da1.loc[da1['Value']!= 'nan']
                                                except:
                                                    pass
                                                try:
                                                    da1['pre_val'].replace('', np.nan, inplace=True)
                                                    da1.dropna(subset=['pre_val'], inplace=True)
                                                    da1 = da1.loc[da1['pre_val']!= 'nan']
                                                except:
                                                    pass
                                                try:
                                                    da1['Value'] = pd.to_numeric(da1['Value'], errors='coerce')
                                                except:
                                                    da1['Value'] = da1['Value'].astype(str)
                                                try:
                                                    da1['pre_val'] = pd.to_numeric(da1['pre_val'], errors='coerce')
                                                except:
                                                    da1['pre_val'] = da1['pre_val'].astype(str)
                                                da1['checkk'] = da1['pre_val'] == da1['Value']
                                                extract_df = da1.loc[da1['checkk'] == False]
                                                try:
                                                    del extract_df['pre_val']
                                                except:
                                                    pass
                                                try:
                                                    del extract_df['Value']
                                                except:
                                                    pass
                                                extract_df = extract_df.rename(columns={'pre_val_N2': 'pre_val'})
                                                extract_df = extract_df.rename(columns={'Value_N2': 'Value'})
                                                extract_df = extract_df[[CELL_M0, 'NodeId', 'RNC', 'MO class', 'Parameters2', 'pre_val', 'Value', 'MO2']]
                                                extract_df = extract_df.rename(columns={'Parameters2': 'Parameter'})
                                                extract_df = extract_df.rename(columns={'pre_val': 'Current Value'})
                                                extract_df = extract_df.rename(columns={'Value': 'Required Value'})
                                                extract_df = extract_df.rename(columns={'MO2': 'MO1'})
                                                if TECHNO == 'GSM':
                                                    extract_df = extract_df.rename(columns={'RNC': 'BSC'})
                                                if TECHNO == 'LTE':
                                                    del extract_df['RNC']
                                                if TECHNO == 'VOLTE':
                                                    del extract_df['RNC']
                                                if TECHNO == 'NR':
                                                    del extract_df['RNC']
                                                try:
                                                    lssst = extract_df.columns
                                                    extract_df = extract_df.drop_duplicates(subset=list(lssst), keep='first')
                                                except:
                                                    pass
                                                extract_df = extract_df.loc[extract_df['Current Value']!= 'nan']
                                                try:
                                                    extract_df['Current Value'] = pd.to_numeric(extract_df['Current Value'])
                                                except:
                                                    extract_df['Current Value'] = extract_df['Current Value'].astype(str)
                                                DF_STORE.append(extract_df)
                                                del extract_df
                                        except:
                                            continue
                    i = i + 1
                except Exception as e:
                    if os.path.exists(OUTPATH134 + 'Error log ' + datetime.now().strftime('%d%m%Y') + '.txt'):
                        append_write = 'a'
                    else:  # inserted
                        append_write = 'w'
                    error_log = ('Error found on Line {}'.format(sys.exc_info()[(-1)].tb_lineno), type(e).__name__)
                    error_log4 = list(error_log)
                    Final_error = str(error_log4)[1:(-1)]
                    if str(e)!= 'No objects to concatenate':
                        with open(OUTPATH134 + 'Error log ' + datetime.now().strftime('%d%m%Y') + '.txt', append_write) as f:
                            f.writelines(Final_error + '\n' + str(e) + '\n')
                        f.close()
                    else:  # inserted
                        with open(OUTPATH134 + 'Note ' + datetime.now().strftime('%d%m%Y') + '.txt', 'w') as f:
                            f.writelines('No discrepancy found\n')
                        f.close()
        Final_baseline = pd.concat(DF_STORE, axis=0, ignore_index=True)
        Final_baseline['Current Value'].fillna('NA22', inplace=True)
        Final_baseline = Final_baseline.loc[Final_baseline['Current Value']!= 'NA22']
        Final_baseline.insert(0, CELL_M0, Final_baseline.pop(CELL_M0))
        Final_baseline.insert(1, 'MO class', Final_baseline.pop('MO class'))
        Final_baseline[CELL_M0].replace('', np.nan, inplace=True)
        Final_baseline.dropna(subset=[CELL_M0], inplace=True)
        Final_baseline['MO class'].replace('', np.nan, inplace=True)
        Final_baseline.dropna(subset=['MO class'], inplace=True)
        Final_baseline['NodeId'].replace('', np.nan, inplace=True)
        Final_baseline.dropna(subset=['NodeId'], inplace=True)
        Final_baseline['Current Value'].replace('', np.nan, inplace=True)
        Final_baseline.dropna(subset=['Current Value'], inplace=True)
        Final_baseline['Required Value'].replace('', np.nan, inplace=True)
        Final_baseline.dropna(subset=['Required Value'], inplace=True)
        try:
            Final_baseline['MO1'].replace('', np.nan, inplace=True)
            if Final_baseline['MO1'].isnull().all():
                Final_baseline = Final_baseline.drop('MO1', axis=1)
        except:
            pass
        try:
            lssst = Final_baseline.columns
            Final_baseline = Final_baseline.drop_duplicates(subset=list(lssst), keep='first')
        except:
            pass
        if Final_baseline.empty:
            with open(OUTPATH134 + 'Note ' + datetime.now().strftime('%d%m%Y') + '.txt', 'w') as f:
                f.writelines('No discrepancy found\n')
            f.close()
        else:  # inserted
            with pd.ExcelWriter(OUTPATH134 + TECHNO + str('_Baseline_Audit_') + datetime.now().strftime('%d%m%Y') + '.xlsx') as writer:
                Final_baseline.T.reset_index().T.to_excel(writer, sheet_name='Baseline Audit', engine='xlsxwriter', index=False, header=None)
                headr_colour_multi(writer, Final_baseline, 'Baseline Audit', '#facf' + str('91'), ['KPI'], '#facf' + str('91'))
        try:
            del Final_baseline
        except:
            return
    except Exception as e:
        if os.path.exists(OUTPATH134 + 'Error log ' + datetime.now().strftime('%d%m%Y') + '.txt'):
            append_write = 'a'
        else:  # inserted
            append_write = 'w'
        error_log = ('Error found on Line {}'.format(sys.exc_info()[(-1)].tb_lineno), type(e).__name__)
        error_log4 = list(error_log)
        Final_error = str(error_log4)[1:(-1)]
        if str(e)!= 'No objects to concatenate':
            with open(OUTPATH134 + 'Error log ' + datetime.now().strftime('%d%m%Y') + '.txt', append_write) as f:
                f.writelines(Final_error + '\n' + str(e) + '\n')
            f.close()
        else:  # inserted
            with open(OUTPATH134 + 'Note ' + datetime.now().strftime('%d%m%Y') + '.txt', 'w') as f:
                f.writelines('No discrepancy found\n')
            f.close()
        return None

def combined_xl(directory11):
    try:
        files = os.listdir(directory11)
        merged_data = {}
        for file in files:
            if file.endswith('.xlsx'):
                sheets = pd.read_excel(os.path.join(directory11, file), sheet_name=None)
                for sheet_name, sheet_data in sheets.items():
                    if sheet_name in merged_data:
                        merged_data[sheet_name] = pd.concat([merged_data[sheet_name], sheet_data], ignore_index=True)
                    else:  # inserted
                        merged_data[sheet_name] = sheet_data.copy()
        with pd.ExcelWriter(directory11 + '\\' + str('Combined_KGET_Dump') + '_' + datetime.now().strftime('%d%m%Y') + '.xlsx') as writer:
            for sheet_name, sheet_data in merged_data.items():
                sheet_data.T.reset_index().T.to_excel(writer, sheet_name=sheet_name, engine='xlsxwriter', index=False, header=None)
        try:
            writer.close()
        except:
            return
    except:
        return None

def KGET_PARSER_AUDIT(self, XMLFILE, cmd_node, TECHNOY):
    try:
        OUTPUT_MO_LST = 'ALL'
        CELLMOLIST = []
        MO_NAME = []
        Final_DATA1 = []
        Final_DATA2 = []
        moclass_LS = []
        file1 = open(XMLFILE, 'rt', encoding='utf-8-sig')
        Lines = file1.readlines()
        Mew_mo = 'NOMO'
        count = 0
        LST_EX = ['Temporary', 'Proxy Id', 'dnPrefix', 'locationName', 'logicalName', 'managedElementId', 'managedElementType', 'managedElementTypeList[9]', 'networkManagedElementId', 'release', 'siteLocation', 'swVersion', 'userDefinedState', 'userLabel', 'vendorName', '$ssh_pid', 'reservedBy', 'description', 'condition', 'publicKey', 'recommendedAction', 'legalNotice', 'additionalInfo', 'extensionContent', 'Struct[0]', 'Struct[1]', 'Struct[2]', 'Struct[3]', 'Struct[4]', 'Struct[5]', 'Struct[6]', 'Struct[7]', 'Struct[8]', 'Struct[9]', 'Struct[10]', 'Struct[11]', 'Struct[12]', 'Struct[13]', 'Struct[14]', 'Struct[15]', 'Struct[16]']
        for line in Lines:
            count += 1
            if len(line) > 1 and '========' not in line:
                line = line.strip()
                line1 = line.replace(' >>> ', '')
                line1 = line1.replace('=', ' ')
                line2 = line.replace(' >>> ', '')
                line2 = line2.replace('>>>', '')
                line2 = line2.replace('>>> ', '')
                line2 = line2.replace('>', '')
                lin1 = re.split('\\s{2,}', line1)
                lin2 = re.split('\\s{2,}', line2)
                if lin1[0][:3]!= 'zzz' and lin1[0] not in LST_EX:
                    try:
                        if lin2[0] == 'MO':
                            MO_nammm = lin2[1]
                            Mew_mo = MO_nammm
                        else:  # inserted
                            MO_nammm = Mew_mo
                    except:
                        if lin2[0] == 'mo':
                            MO_nammm = lin2[1]
                            Mew_mo = MO_nammm
                        else:  # inserted
                            MO_nammm = Mew_mo
                    moclass1 = MO_nammm.split(',')[(-1)]
                    moclass = moclass1.split('=')[0]
                    para = lin1[0].split('.')[(-1)]
                    try:
                        if len(lin1[1]) < 200:
                            appandd = 'YES'
                    except:
                        appandd = 'NO'
                    para_SELECT = 'NO'
                    if appandd == 'YES':
                        if para_SELECT == 'Selected':
                            if para.upper() in para_SELECT:
                                try:
                                    Final_DATA1.append(para.strip())
                                except:
                                    Final_DATA1.append('')
                                try:
                                    Final_DATA2.append(lin1[1].strip())
                                except:
                                    Final_DATA2.append('')
                                MO_NAME.append(MO_nammm.strip())
                                moclass_LS.append(moclass.strip())
                        else:  # inserted
                            try:
                                Final_DATA1.append(para.strip())
                            except:
                                Final_DATA1.append('')
                            try:
                                Final_DATA2.append(lin1[1].strip())
                            except:
                                Final_DATA2.append('')
                            MO_NAME.append(MO_nammm.strip())
                            moclass_LS.append(moclass.strip())
                        try:
                            del MO_nammm
                        except:
                            pass
                        try:
                            del para
                        except:
                            continue
        df = pd.DataFrame()
        df['MO CLASS'] = moclass_LS
        df['MO'] = MO_NAME
        df['Tag'] = Final_DATA1
        df['Value'] = Final_DATA2
        try:
            del MO_NAME
        except:
            pass
        try:
            del Final_DATA1
        except:
            pass
        try:
            del Final_DATA2
        except:
            pass
        df = df.loc[df['Tag']!= 'MO']
        df = df.loc[df['MO']!= 'NOMO']
        df.replace({'MO': np.nan}, inplace=True)
        df.dropna(subset=['MO'], inplace=True)
        df.replace({'Tag': np.nan}, inplace=True)
        df.dropna(subset=['Tag'], inplace=True)
        df.replace({'Value': np.nan}, inplace=True)
        df.dropna(subset=['Value'], inplace=True)
        if OUTPUT_MO_LST == 'ALL':
            MOCLASS_LST = df['MO CLASS'].unique()
            MOCLASS_LST = [i for i in MOCLASS_LST if i]
        try:
            MOCLASS_LST.remove('None')
        except:
            pass
        try:
            MOCLASS_LST.remove('NR')
        except:
            pass
        try:
            MOCLASS_LST.remove('NA')
        except:
            pass
        try:
            MOCLASS_LST = [x for x in MOCLASS_LST if x]
        except:
            pass
        OUTPUT_TYPE_excel = 'Combined_excel'
        if OUTPUT_TYPE_excel == 'Combined_excel':
            XML_OptiON = 'CM/KGET Dump export'
            with pd.ExcelWriter(os.path.dirname(XMLFILE) + '\\' + str(TECHNOY) + '_' + str(cmd_node) + '_' + datetime.now().strftime('%d%m%Y') + '.xlsx', engine='xlsxwriter') as writer:
                for moclss in MOCLASS_LST:
                    try:
                        df_newout = df.copy()
                        df = df.loc[df['MO CLASS']!= str(moclss)]
                        df_newout = df_newout.loc[df_newout['MO CLASS'] == moclss]
                        df_newout = df_newout[['MO', 'Tag', 'Value']]
                        if moclss not in CELLMOLIST:
                            def fix_my_stuff(x):
                                x = x.tolist()
                                x = ' '.join([str(y) for y in x])
                                return x
                            df_newout = df_newout.groupby(['MO', 'Tag']).agg(lambda x: fix_my_stuff(x)).reset_index()
                            df_newout = df_newout.groupby(['MO', 'Tag'])[['Value']].max().unstack(['Tag']).droplevel(0, axis='columns').rename_axis(columns=None).reset_index()
                            df_newout = df_newout.replace('nan', '')
                            try:
                                if df_newout.empty:
                                    df_newout = pd.DataFrame()
                                if XML_OptiON == 'CM/KGET Dump export':
                                    if df_newout.empty:
                                        df_newout = pd.DataFrame()
                                    df_newout.T.reset_index().T.to_excel(writer, sheet_name=str(moclss), engine='xlsxwriter', index=False, header=None)
                            except:
                                pass
                        if moclss in CELLMOLIST:
                            df_newout = df_newout.groupby(['MO', 'Tag'])[['Value']].max().unstack(['Tag']).droplevel(0, axis='columns').rename_axis(columns=None).reset_index()
                            df_newout = df_newout.replace('nan', '')
                            try:
                                if XML_OptiON == 'CM/KGET Dump export':
                                    if df_newout.empty:
                                        df_newout = pd.DataFrame()
                                    df_newout.T.reset_index().T.to_excel(writer, sheet_name=str(moclss), engine='xlsxwriter', index=False, header=None)
                            except:
                                pass
                    except:
                        continue
    except:
        return None

def KGET_PARSER(self, XMLFILE, cmd_node):
    try:
        OUTPUT_MO_LST = 'ALL'
        CELLMOLIST = []
        MO_NAME = []
        Final_DATA1 = []
        Final_DATA2 = []
        moclass_LS = []
        file1 = open(XMLFILE, 'rt', encoding='utf-8-sig')
        Lines = file1.readlines()
        Mew_mo = 'NOMO'
        count = 0
        LST_EX = ['Temporary', 'Proxy Id', 'dnPrefix', 'locationName', 'logicalName', 'managedElementId', 'managedElementType', 'managedElementTypeList[9]', 'networkManagedElementId', 'release', 'siteLocation', 'swVersion', 'userDefinedState', 'userLabel', 'vendorName', '$ssh_pid', 'reservedBy', 'description', 'condition', 'publicKey', 'recommendedAction', 'legalNotice', 'additionalInfo', 'extensionContent', 'Struct[0]', 'Struct[1]', 'Struct[2]', 'Struct[3]', 'Struct[4]', 'Struct[5]', 'Struct[6]', 'Struct[7]', 'Struct[8]', 'Struct[9]', 'Struct[10]', 'Struct[11]', 'Struct[12]', 'Struct[13]', 'Struct[14]', 'Struct[15]', 'Struct[16]']
        for line in Lines:
            count += 1
            if len(line) > 1:
                if '========' not in line:
                    line = line.strip()
                    line1 = line.replace(' >>> ', '')
                    line1 = line1.replace('=', ' ')
                    line2 = line.replace(' >>> ', '')
                    line2 = line2.replace('>>>', '')
                    line2 = line2.replace('>>> ', '')
                    line2 = line2.replace('>', '')
                    lin1 = re.split('\\s{2,}', line1)
                    lin2 = re.split('\\s{2,}', line2)
                if lin1[0][:3]!= 'zzz' and lin1[0] not in LST_EX:
                    try:
                        if lin2[0] == 'MO':
                            MO_nammm = lin2[1]
                            Mew_mo = MO_nammm
                        else:  # inserted
                            MO_nammm = Mew_mo
                    except:
                        if lin2[0] == 'mo':
                            MO_nammm = lin2[1]
                            Mew_mo = MO_nammm
                        else:  # inserted
                            MO_nammm = Mew_mo
                    moclass1 = MO_nammm.split(',')[(-1)]
                    moclass = moclass1.split('=')[0]
                    para = lin1[0].split('.')[(-1)]
                    try:
                        if len(lin1[1]) < 200:
                            appandd = 'YES'
                    except:
                        appandd = 'NO'
                    para_SELECT = 'NO'
                    if appandd == 'YES':
                        if para_SELECT == 'Selected':
                            if para.upper() in para_SELECT:
                                try:
                                    Final_DATA1.append(para.strip())
                                except:
                                    Final_DATA1.append('')
                                try:
                                    Final_DATA2.append(lin1[1].strip())
                                except:
                                    Final_DATA2.append('')
                                MO_NAME.append(MO_nammm.strip())
                                moclass_LS.append(moclass.strip())
                        else:  # inserted
                            try:
                                Final_DATA1.append(para.strip())
                            except:
                                Final_DATA1.append('')
                            try:
                                Final_DATA2.append(lin1[1].strip())
                            except:
                                Final_DATA2.append('')
                            MO_NAME.append(MO_nammm.strip())
                            moclass_LS.append(moclass.strip())
                        try:
                            del MO_nammm
                        except:
                            pass
                        try:
                            del para
                        except:
                            continue
        df = pd.DataFrame()
        df['MO CLASS'] = moclass_LS
        df['MO'] = MO_NAME
        df['Tag'] = Final_DATA1
        df['Value'] = Final_DATA2
        try:
            del MO_NAME
        except:
            pass
        try:
            del Final_DATA1
        except:
            pass
        try:
            del Final_DATA2
        except:
            pass
        df = df.loc[df['Tag']!= 'MO']
        df = df.loc[df['MO']!= 'NOMO']
        df.replace({'MO': np.nan}, inplace=True)
        df.dropna(subset=['MO'], inplace=True)
        df.replace({'Tag': np.nan}, inplace=True)
        df.dropna(subset=['Tag'], inplace=True)
        df.replace({'Value': np.nan}, inplace=True)
        df.dropna(subset=['Value'], inplace=True)
        if OUTPUT_MO_LST == 'ALL':
            MOCLASS_LST = df['MO CLASS'].unique()
            MOCLASS_LST = [i for i in MOCLASS_LST if i]
        try:
            MOCLASS_LST.remove('None')
        except:
            pass
        try:
            MOCLASS_LST.remove('NR')
        except:
            pass
        try:
            MOCLASS_LST.remove('NA')
        except:
            pass
        try:
            MOCLASS_LST = [x for x in MOCLASS_LST if x]
        except:
            pass
        OUTPUT_TYPE_excel = 'Combined_excel'
        if OUTPUT_TYPE_excel == 'Combined_excel':
            XML_OptiON = 'CM/KGET Dump export'
            with pd.ExcelWriter(os.path.dirname(XMLFILE) + '\\' + str(cmd_node) + '_' + datetime.now().strftime('%d%m%Y') + '.xlsx', engine='xlsxwriter') as writer:
                for moclss in MOCLASS_LST:
                    try:
                        df_newout = df.copy()
                        df = df.loc[df['MO CLASS']!= str(moclss)]
                        df_newout = df_newout.loc[df_newout['MO CLASS'] == moclss]
                        df_newout = df_newout[['MO', 'Tag', 'Value']]
                        if moclss not in CELLMOLIST:
                            def fix_my_stuff(x):
                                x = x.tolist()
                                x = ' '.join([str(y) for y in x])
                                return x
                            df_newout = df_newout.groupby(['MO', 'Tag']).agg(lambda x: fix_my_stuff(x)).reset_index()
                            df_newout = df_newout.groupby(['MO', 'Tag'])[['Value']].max().unstack(['Tag']).droplevel(0, axis='columns').rename_axis(columns=None).reset_index()
                            df_newout = df_newout.replace('nan', '')
                            try:
                                if df_newout.empty:
                                    df_newout = pd.DataFrame()
                                if XML_OptiON == 'CM/KGET Dump export':
                                    if df_newout.empty:
                                        df_newout = pd.DataFrame()
                                    df_newout.T.reset_index().T.to_excel(writer, sheet_name=str(moclss), engine='xlsxwriter', index=False, header=None)
                            except:
                                pass
                        if moclss in CELLMOLIST:
                            df_newout = df_newout.groupby(['MO', 'Tag'])[['Value']].max().unstack(['Tag']).droplevel(0, axis='columns').rename_axis(columns=None).reset_index()
                            df_newout = df_newout.replace('nan', '')
                            try:
                                if XML_OptiON == 'CM/KGET Dump export':
                                    if df_newout.empty:
                                        df_newout = pd.DataFrame()
                                    df_newout.T.reset_index().T.to_excel(writer, sheet_name=str(moclss), engine='xlsxwriter', index=False, header=None)
                            except:
                                pass
                    except:
                        continue
    except:
        return None

def clean_sheet_name_old(sheet_name, max_length=30):
    sheet_name = sheet_name.replace('hget', '').strip()
    sheet_name = sheet_name.replace('get', '').strip()
    sheet_name = sheet_name.replace('HGET', '').strip()
    sheet_name = sheet_name.replace('GET', '').strip()
    cleaned_name = re.sub('[^\\w\\s]', '', sheet_name)
    cleaned_name = cleaned_name[:max_length].strip()
    return cleaned_name

def clean_sheet_name(sheet_name, max_length=30):
    try:
        sheet_name = sheet_name.replace('.', '')
    except:
        pass
    try:
        sheet_name = sheet_name.split(' ', 1)[1].strip()
    except:
        pass
    try:
        vectorizer = CountVectorizer()
        X = vectorizer.fit_transform([sheet_name])
        transformer = TfidfTransformer()
        tfidf_matrix = transformer.fit_transform(X)
        feature_names = vectorizer.get_feature_names_out()
        tfidf_values = tfidf_matrix.toarray()[0]
        keywords = [(feature_names[i], tfidf_values[i]) for i in range(len(feature_names))]
        keywords = sorted(keywords, key=lambda x: x[1], reverse=True)
        Final_klst = []
        EXCLUDE_KEY = ['hget', 'label', 'state', 'get', 'calibrationstatus', 'grep', 'pmr', 'PMR', 'pget', 'pmx']
        for key in keywords:
            fkey = str(key).replace('.', '')
            fkey = str(fkey).replace('(', '')
            fkey = str(fkey).replace(')', '')
            fkey = str(fkey).replace('\'', '')
            fkey = fkey.split(',')[0].strip()
            try:
                intt = int(fkey)
            except:
                if fkey not in EXCLUDE_KEY and 'status' not in str(fkey):
                    Final_klst.append(fkey)
                    break
    cleaned_name = Final_klst[0].strip()
    return cleaned_name
    except:
        return sheet_name

def is_non_numeric_string(input_string):
    return not input_string.isdigit()

def cleeend_text22(text):
    lines = text.split('\n')
    max_line_length = max((len(line) for line in lines))
    cleaned_text = '=' * max_line_length + '\n' + re.sub('^[=]+$', '', text, flags=re.MULTILINE) + '\n' + '=' * max_line_length
    cleaned_text = re.sub('\\n\\s*\\n', '\n', cleaned_text)
    return cleaned_text

def cleeend_text(text):
    lines = text.split('\n')
    max_line_length = max((len(line) for line in lines))
    cleaned_text = '\n' + '-' * max_line_length + '\n' + re.sub('^[=]+$', ' ' * max_line_length, text, flags=re.MULTILINE) + '\n' + '-' * max_line_length + '\n'
    cleaned_text = re.sub('\\n\\s*\\n', '\n', cleaned_text)
    return cleaned_text

def PROJECT_USRE_DETAILS_ADD_SME_ROLE():
    class TextBoxDialog(QDialog):
        def __init__(self, parent=None):
            super(TextBoxDialog, self).__init__(parent)
            style = 'QLineEdit { background-color: #dcf5e3; border: 2px solid #4CAF50; border-radius: 8px; padding: 2px; font-size: 19px; font-family: Segoe UI Semibold;}QPushButton { background-color: #4CAF51; color: white; border: none; border-radius: 8px; padding: 8px; font-size: 18px; }'
            self.setStyleSheet(style)
            self.setWindowTitle('Add User In Project')
            self.setGeometry(800, 300, 500, 500)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            try:
                con = sqlite3.connect(REPORT_INPATH22 + SELECTED_PROJECT[0] + '_PROJECT_MAPPING.db')
                PROJECT_MAPPING_usww = pd.read_sql_query(f"SELECT * FROM {'MyTable'}", con)
                PROJECT_LST_Signum = PROJECT_MAPPING_usww['Signum'].unique()
                PROJECT_LST_Signum = list(PROJECT_LST_Signum)
                seen = set()
                PROJECT_LST_Signum = [item for item in PROJECT_LST_Signum if not (item in seen or seen.add(item))]
            except:
                PROJECT_LST_Signum = []
            content = '\n'.join([item.lower() for item in PROJECT_LST_Signum])
            if not str(content).strip():
                content = 'Enter user signum here...'
            layout = QVBoxLayout(self)
            self.text_box = QTextEdit(self)
            font = self.text_box.font()
            font.setPointSize(11)
            self.text_box.setFont(font)
            if content == 'Enter user signum here...':
                self.text_box.setPlaceholderText(content)
            else:  # inserted
                self.text_box.setText(content)
            layout.addWidget(self.text_box)
            self.submit_button = QPushButton('Submit', self)
            self.submit_button.clicked.connect(self.submit)
            layout.addWidget(self.submit_button)

        def submit(self):
            try:
                input_text = self.text_box.toPlainText()
                input_text_list = input_text.split()
                seen = set()
                input_text_list = [item.lower() for item in input_text_list if not (item in seen or seen.add(item))]
                df = pd.DataFrame()
                df['Signum'] = input_text_list
                df['PROJECT ID'] = SELECTED_PROJECT[0]
                df = df.drop_duplicates(subset=['Signum'], keep='first')
                try:
                    os.remove(os.path.join(REPORT_INPATH22 + SELECTED_PROJECT[0] + '_PROJECT_MAPPING.db'))
                except:
                    pass
                con = sqlite3.connect(os.path.join(REPORT_INPATH22 + SELECTED_PROJECT[0] + '_PROJECT_MAPPING.db'))
                try:
                    df.to_sql('MyTable', con, index=False)
                except:
                    pass
                finally:  # inserted
                    con.close()

                def uplode_datata():
                    FILENNNAME = SELECTED_PROJECT[0] + '_PROJECT_MAPPING.db'
                    sit_url = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal'
                    sh_ul_path = '//ericsson.sharepoint.com/sites/MMEAPerformanceReportingandmonitoringautomationPortal/Shared%20Documents/Project%20wise%20Reposiory/ZAuthentication/CHTBOT/PROJECT_MAPPING/PID-' + str(SELECTED_PROJECT[0])
                    client_id = 'd8c4d3d7-e0e4-41e4-ba43-63cbb6405da1'
                    client_secret = 'Dy18Z0hudKcOHsQw0AIBF5JVw888DJDAYSeHlgB1Feg='
                    sp = da_tran_SP365(site_url='https:' + sit_url, client_id=client_id, client_secret=client_secret)
                    up_path = sp.create_link('https:' + sh_ul_path + '/' + FILENNNAME)
                    sp.upload(sharepoint_location=up_path, local_location=REPORT_INPATH22 + SELECTED_PROJECT[0] + '_PROJECT_MAPPING.db')
                threading.Thread(target=uplode_datata).start()
            except:
                pass
            self.close()
    dialog = TextBoxDialog()
    dialog.exec_()

def get_nodeid():
    class YourApp(QWidget):
        def __init__(self):
            super(YourApp, self).__init__()
            self.init_ui()
            self.node_id = None
            self.setWindowTitle('Node Id')
            self.setGeometry(770, 480, 450, 110)

        def init_ui(self):
            style = 'QLineEdit { background-color: #dcf5e3; border: 2px solid #4CAF50; border-radius: 8px; padding: 2px; font-size: 19px; font-family: Segoe UI Semibold;}QPushButton { background-color: #4CAF51; color: white; border: none; border-radius: 8px; padding: 8px; font-size: 18px; }'
            self.setStyleSheet(style)
            self.node_id_label = QLabel('Enter Node Id:', self)
            self.node_id_label.setFont(QFont('Calibri', 11))
            self.node_id_input = QLineEdit(self)
            self.node_id_input.setFont(QFont('Calibri', 11))
            submit_button = QPushButton('Submit', self)
            submit_button.setFont(QFont('Segoe UI Semibold', 17))
            submit_button.clicked.connect(self.submit_values)
            self.node_id_input.returnPressed.connect(self.submit_values)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            layout = QVBoxLayout(self)
            layout.addWidget(self.node_id_label)
            layout.addWidget(self.node_id_input)
            layout.addWidget(submit_button)

        def submit_values(self):
            self.node_id = self.node_id_input.text()
            if not self.node_id:
                return
            self.close()

    def enter_cmd():
        app = QApplication(sys.argv)
        ex = YourApp()
        ex.show()
        app.exec_()
        return ex.node_id
    node_id22 = enter_cmd()
    if node_id22:
        try:
            node_id22 = node_id22.lower().replace('amos', '').strip()
            node_id = str(node_id22).strip()
        except:
            pass
        try:
            nodedb = pd.read_excel(INPATH + 'Node_db.xlsx', sheet_name='Nodedb', usecols='A')
            nodedb_lst = nodedb['Node list'].unique()
            nodedb_lst = list(nodedb_lst)
            for nnd in nodedb_lst:
                if node_id.lower() in nnd.lower():
                    node_id = str(nnd).strip().upper()
                    return
                    return node_id
        except:
            node_id = node_id22

def submit_node22(self):
    class YourApp(QWidget):
        def __init__(self):
            super(YourApp, self).__init__()
            self.init_ui()
            self.keyword = None
            self.command = None
            self.setWindowTitle('Command Input')
            self.setGeometry(750, 400, 550, 210)

        def init_ui(self):
            style = 'QLineEdit { background-color: #dcf5e3;border: 2px solid #4CAF50; border-radius: 8px; padding: 2px; font-size: 19px; font-family: Segoe UI Semibold;}QPushButton { background-color: #4CAF51; color: white; border: none; border-radius: 8px; padding: 8px; font-size: 18px; }'
            self.setStyleSheet(style)
            self.command_keyword_label = QLabel('Command short description(e.g.,alarm/vswr):', self)
            self.command_keyword_label.setFont(QFont('Calibri', 11))
            self.command_input_label = QLabel('Specific command (e.g., altk/cabex) :', self)
            self.command_input_label.setFont(QFont('Calibri', 11))
            self.command_keyword_display = QLineEdit(self)
            self.command_keyword_display.setFont(QFont('Calibri', 11))
            self.command_input_display = QLineEdit(self)
            self.command_input_display.setFont(QFont('Calibri', 11))
            submit_button = QPushButton('Submit', self)
            submit_button.setFont(QFont('Segoe UI Semibold', 17))
            submit_button.clicked.connect(self.submit_values)
            self.command_input_display.returnPressed.connect(self.submit_values)
            layout = QVBoxLayout(self)
            layout.addWidget(self.command_keyword_label)
            layout.addWidget(self.command_keyword_display)
            layout.addWidget(self.command_input_label)
            layout.addWidget(self.command_input_display)
            layout.addWidget(submit_button)

        def submit_values(self):
            self.keyword = self.command_keyword_display.text()
            self.command = self.command_input_display.text()
            QApplication.instance().quit()

    def enter_cmd():
        app = QApplication(sys.argv)
        ex = YourApp()
        ex.show()
        app.exec_()
        return (ex.keyword, ex.command)
    keyword, command = enter_cmd()
    check_yesno.clear()
    KEYWORD_ADD.clear()
    KEYWORD_ADD.append(str(keyword).strip() + '_' + str(command).strip())
    return str(str(keyword).strip() + '_' + str(command).strip())

def RSGLoginApp_fun():
    class RSGLoginApp(QWidget):
        def __init__(self):
            super().__init__()
            self.load_credentials()
            self.setGeometry(300, 300, 420, 526)
            self.setWindowTitle('RSG DETAILS')
            self.setFixedSize(420, 550)
            self.center_window()
            self.create_gui()

        def center_window(self):
            frame_geometry = self.frameGeometry()
            center_point = QDesktopWidget().availableGeometry().center()
            frame_geometry.moveCenter(center_point)
            self.move(frame_geometry.topLeft())

        def load_credentials(self):
            try:
                db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                con.close()
                self.PID_LST = Rdf['value'].tolist()
                while len(self.PID_LST) < 10:
                    self.PID_LST.append(f'LINE{len(self.PID_LST) + 1}')
            except:
                self.PID_LST = [f'LINE{i + 1}' for i in range(10)]

        def create_gui(self):
            layout = QVBoxLayout(self)
            label = QLabel('RSG DETAILS', self)
            label.setStyleSheet('font-size: 24px; color: #fe6f5e; font-weight: bold;')
            layout.addWidget(label)
            self.entries = []
            for i in range(10):
                entry = QLineEdit(self)
                entry.setEchoMode(QLineEdit.Password)
                font = QFont('Calibri', 12)
                entry.setFont(font)
                entry.setFixedSize(390, 35)
                entry.setText(self.PID_LST[i])
                entry.setStyleSheet('padding: 4px; background-color: #eff4f9; border: 1px solid #ccc;')
                layout.addWidget(entry)
                self.entries.append(entry)
                self.setWindowFlag(Qt.WindowStaysOnTopHint)
            update_button = QPushButton('UPDATE', self)
            update_button.clicked.connect(self.login)
            update_button.setStyleSheet('padding: 8px; background-color: #4CAF50; color: white; border: none; border-radius: 4px; font-weight: bold;')
            layout.addWidget(update_button)
            self.setLayout(layout)

        def login(self):
            values = [encrypt(entry.text(), 3) if encrypt(entry.text(), 3) else f'LINE{i + 1}' for i, entry in enumerate(self.entries)]
            DF = pd.DataFrame({'indexxx': [f'Line{i + 1}' for i in range(10)], 'value': values})
            try:
                os.remove('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
            except:
                pass
            con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
            try:
                DF.to_sql('Credd', con, index=False, if_exists='replace')
            except:
                pass
            finally:  # inserted
                con.close()
            self.close()
    if __name__ == '__main__':
        app = QApplication([])
        rsg_app = RSGLoginApp()
        rsg_app.show()
        app.exec_()

def Monitoring_main_SELECT(COMMAND_DEFALT_AUTO):
    AZURE_CREAD_FAIL.clear()
    Azue_OTP_lst.clear()
    establish_ssh_conn_lst.clear()
    KEYWORD_ADD.clear()
    excel_df_LST.clear()
    amos_node.clear()
    Buddy_request_lst.clear()
    buddy_node.clear()
    buddy_node_curr.clear()
    amos_nod2.clear()
    ltall_nod2.clear()
    check_yesno.clear()
    ASKSITID.clear()
    cmmmddd.clear()
    siteddd.clear()
    checknxtsit.clear()
    processdone.clear()
    combined_cmd_window.clear()
    combined_cmd_window_chk.clear()
    checkamos.clear()
    command.clear()

    class defalcmdtWindow(QDialog):
        def __init__(self, technologies, parent=None):
            super(defalcmdtWindow, self).__init__(parent)

    class NodeSelectWindow(QDialog):
        def __init__(self, technologies, parent=None):
            super(NodeSelectWindow, self).__init__(parent)
            self.setWindowTitle('Node List')
            self.setGeometry(650, 250, 650, 420)
            layout = QVBoxLayout(self)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            self.tab_widget = QTabWidget()
            layout.addWidget(self.tab_widget)
            self.tab_technology_mapping = {tech: idx for idx, tech in enumerate(technologies)}
            for technology in technologies:
                tab = QWidget()
                self.tab_widget.addTab(tab, technology)
                tab_layout = QVBoxLayout(tab)
                command_text_edit = QTextEdit()
                command_text_edit.setWordWrapMode(QTextOption.NoWrap)
                tab_layout.addWidget(command_text_edit)
                submit_button = QPushButton('Submit')
                submit_button.clicked.connect(self.submit_command_final)
                self.apply_bold_font(submit_button)
                tab_layout.addWidget(submit_button)
            self.tab_widget.tabBar().setStyleSheet('background-color: lightblue;')

        def apply_bold_font(self, widget):
            font = widget.font()
            font.setBold(True)
            widget.setFont(font)

        def closeEvent(self, event):
            self.window_closing = True
            event.accept()
            INSERT_COM = ['close']
            DATABASE = 'CAUDIT_CLOSE_WINDOW'
            DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)

        def submit_command_final(self):
            try:
                for i in range(self.tab_widget.count()):
                    current_tab = self.tab_widget.widget(i)
                    command_text_edit = current_tab.findChild(QTextEdit)
                    command_list = command_text_edit.toPlainText().splitlines()
                    command_list = [command for command in command_list if command.strip()]
                    technology = list(self.tab_technology_mapping.keys())[i]
                    if technology == 'Volte':
                        GSM_NODE_LST.extend(command_list)
                    else:  # inserted
                        if technology == 'WCDMA':
                            WCDMA_NODE_LST.extend(command_list)
                        else:  # inserted
                            if technology == 'LTE':
                                LTE_NODE_LST.extend(command_list)
                            else:  # inserted
                                if technology == 'NR':
                                    NR_NODE_LST.extend(command_list)
            except:
                pass
            missing_tech = []
            for tttt in selected_checkboxes:
                if tttt == 'Volte' and (not GSM_NODE_LST):
                    missing_tech.append('Volte')
                else:  # inserted
                    if tttt == 'WCDMA' and (not WCDMA_NODE_LST):
                        missing_tech.append('WCDMA')
                    else:  # inserted
                        if tttt == 'LTE' and (not LTE_NODE_LST):
                            missing_tech.append('LTE')
                        else:  # inserted
                            if tttt == 'NR' and (not NR_NODE_LST):
                                missing_tech.append('NR')
            if missing_tech:
                missing_tech_names = ', '.join(missing_tech)
                QMessageBox.warning(self, 'Warning', f'Please enter Node list for the following technology before continuing: {missing_tech_names}')
            else:  # inserted
                self.accept()

    class MainWindow(QDialog):
        def __init__(self, parent=None):
            super(MainWindow, self).__init__(parent)
            ROP_ACT.clear()
            self.setWindowTitle('Initial KPI Monitoring')
            self.setGeometry(650, 250, 650, 480)
            layout = QVBoxLayout(self)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            self.command_text_edit = QTextEdit()
            font = QFont('Consolas', 12)
            self.command_text_edit.setFont(font)
            self.command_text_edit.setPlainText('\n\n                    Initial KPI Monitoring\n\n   Please select all required info like technology and Period')
            layout.addWidget(self.command_text_edit)
            combo_layout = QVBoxLayout()
            self.activity_combo = QComboBox()
            self.activity_combo.setFont(QFont('Calibri', 12))
            self.activity_combo.addItem('ROP by ROP')
            self.activity_combo.addItem('Hour by Hour')
            self.activity_combo.addItem('Whole Period')
            combo_layout.addWidget(self.activity_combo)
            layout.addLayout(combo_layout)
            spacer_item = QSpacerItem(30, 20, QSizePolicy.Minimum, QSizePolicy.Fixed)
            layout.addItem(spacer_item)
            freq_layout = QHBoxLayout()
            freq_label = QLabel('Frequency in minute:')
            font = QFont()
            font.setPointSize(11)
            freq_label.setFont(font)
            freq_layout.addWidget(freq_label)
            self.freq_spin_box = QComboBox()
            self.freq_spin_box.addItems(['Not required', '1', '5', '10', '15', '30', '60', '120', '180'])
            font = QFont()
            font.setPointSize(11)
            self.freq_spin_box.setFont(font)
            freq_layout.addWidget(self.freq_spin_box)
            layout.addLayout(freq_layout)
            spacer_item = QSpacerItem(5, 5, QSizePolicy.Minimum, QSizePolicy.Fixed)
            layout.addItem(spacer_item)

            def onFreqComboBoxIndexChanged(index):
                if index == 0:
                    self.freq_loop_spin_box.setEnabled(False)
                else:  # inserted
                    self.freq_loop_spin_box.setEnabled(True)
            self.freq_spin_box.currentIndexChanged.connect(onFreqComboBoxIndexChanged)
            freq_loop_layout = QHBoxLayout()
            freq_loop_label = QLabel('Set Frequency count:')
            font = QFont()
            font.setPointSize(11)
            freq_loop_label.setFont(font)
            freq_loop_layout.addWidget(freq_loop_label)
            self.freq_loop_spin_box = QSpinBox()
            self.freq_loop_spin_box.setEnabled(False)
            self.freq_loop_spin_box.setRange(2, 30)
            self.freq_loop_spin_box.setValue(2)
            font = QFont()
            font.setPointSize(11)
            self.freq_loop_spin_box.setFont(font)
            freq_loop_layout.addWidget(self.freq_loop_spin_box)
            layout.addLayout(freq_loop_layout)
            spacer_item = QSpacerItem(5, 5, QSizePolicy.Minimum, QSizePolicy.Fixed)
            layout.addItem(spacer_item)
            Email_layout = QHBoxLayout()
            Email_label = QLabel('Email method:')
            font = QFont()
            font.setPointSize(11)
            Email_label.setFont(font)
            Email_layout.addWidget(Email_label)
            self.Email_spin_box = QComboBox()
            self.Email_spin_box.addItems(['Multi Nodes', 'Node Wise'])
            font = QFont()
            font.setPointSize(11)
            self.Email_spin_box.setFont(font)
            Email_layout.addWidget(self.Email_spin_box)
            layout.addLayout(Email_layout)
            spacer_item = QSpacerItem(20, 20, QSizePolicy.Minimum, QSizePolicy.Fixed)
            layout.addItem(spacer_item)
            checkbox_layout = QHBoxLayout()
            self.GSM_checkbox = QCheckBox('Volte')
            self.apply_bold_font(self.GSM_checkbox)
            checkbox_layout.addWidget(self.GSM_checkbox)
            checkbox_layout.addSpacing(100)
            self.lte_checkbox = QCheckBox('LTE')
            self.apply_bold_font(self.lte_checkbox)
            checkbox_layout.addWidget(self.lte_checkbox)
            checkbox_layout.addSpacing(100)
            self.wcdma_checkbox = QCheckBox('WCDMA')
            self.apply_bold_font(self.wcdma_checkbox)
            checkbox_layout.addWidget(self.wcdma_checkbox)
            checkbox_layout.addSpacing(100)
            self.nr_checkbox = QCheckBox('NR')
            self.apply_bold_font(self.nr_checkbox)
            checkbox_layout.addWidget(self.nr_checkbox)
            layout.addLayout(checkbox_layout)
            submit_button = QPushButton('NEXT')
            submit_button.clicked.connect(self.submit_command)
            self.apply_bold_font(submit_button)
            layout.addWidget(submit_button)

        def apply_bold_font(self, widget):
            font = widget.font()
            font.setPointSize(11)
            widget.setFont(font)

        def radio_button_toggled(self):
            sender = self.sender()
            if sender.isChecked():
                DATABASE = 'ACTIVITYY'
                INSERT_COM = [sender.text()]
                DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)

        def closeEvent(self, event):
            self.window_closing = True
            event.accept()
            INSERT_COM = ['close']
            DATABASE = 'CAUDIT_CLOSE_WINDOW'
            DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)

        def submit_command(self):
            activitw = self.activity_combo.currentText()
            COMMAND_DEFALT_AUTO.clear()
            selected_checkboxes.clear()
            FREQUENCY_box = self.freq_spin_box.currentText()
            FREQUENCY_LST.append(FREQUENCY_box)
            FREQUENCY_count = self.freq_loop_spin_box.value()
            FREQUENCY_count_LST.append(FREQUENCY_count)
            EMAIL_METHOOD = self.Email_spin_box.currentText()
            EMAIL_METHOOD_LST.append(EMAIL_METHOOD)
            if self.GSM_checkbox.isChecked():
                selected_checkboxes.append('Volte')
            if self.lte_checkbox.isChecked():
                selected_checkboxes.append('LTE')
            if self.wcdma_checkbox.isChecked():
                selected_checkboxes.append('WCDMA')
            if self.nr_checkbox.isChecked():
                selected_checkboxes.append('NR')
            if selected_checkboxes:
                self.accept()
                ROP_ACT.append(activitw)
                if activitw == 'ROP by ROP':
                    if 'Volte' in selected_checkboxes:
                        GSM_LST.extend(['pmr -r 205'])
                    else:  # inserted
                        GSM_LST.extend([])
                    if 'WCDMA' in selected_checkboxes:
                        WCDMA_PMR = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='Initial KPI_WCDMA', usecols='a,b,c,d', engine='openpyxl')
                        WCDMA_PMR['ROP Data'].replace('Single Rop 15 Minute', 'ROP by ROP', inplace=True)
                        WCDMA_PMR = WCDMA_PMR.loc[WCDMA_PMR['ROP Data'] == 'ROP by ROP']
                        WCDMA_PMR = WCDMA_PMR.loc[WCDMA_PMR['Status'] == 'Required']
                        WCDMA_PMR_LST = WCDMA_PMR['PMR No'].unique()
                        Final_PMR = 'pmr -m 1 -a -r ' + ','.join(map(str, WCDMA_PMR_LST)) + ' | grep '
                        WCDMA_LST.extend([Final_PMR])
                    else:  # inserted
                        WCDMA_LST.extend([])
                    if 'LTE' in selected_checkboxes:
                        LTE_LST.extend(['pmr -r 206'])
                    else:  # inserted
                        LTE_LST.extend([])
                    if 'NR' in selected_checkboxes:
                        NR_LST.extend(['pmr -r 406', 'pmr -r 409'])
                    else:  # inserted
                        NR_LST.extend([])
                    node_window = NodeSelectWindow(selected_checkboxes)
                    node_window.exec_()
                if activitw == 'Hour by Hour':
                    if 'Volte' in selected_checkboxes:
                        GSM_LST.extend(['pmr -r 211'])
                    else:  # inserted
                        GSM_LST.extend([])
                    if 'WCDMA' in selected_checkboxes:
                        WCDMA_PMR = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='Initial KPI_WCDMA', usecols='a,b,c,d', engine='openpyxl')
                        WCDMA_PMR['ROP Data'].replace('Single Rop 15 Minute', 'ROP by ROP', inplace=True)
                        WCDMA_PMR = WCDMA_PMR.loc[WCDMA_PMR['ROP Data'] == 'Hour by Hour']
                        WCDMA_PMR = WCDMA_PMR.loc[WCDMA_PMR['Status'] == 'Required']
                        WCDMA_PMR_LST = WCDMA_PMR['PMR No'].unique()
                        Final_PMR = 'pmr -m 1 -a -r ' + ','.join(map(str, WCDMA_PMR_LST)) + ' | grep '
                        WCDMA_LST.extend([Final_PMR])
                    else:  # inserted
                        WCDMA_LST.extend([])
                    if 'LTE' in selected_checkboxes:
                        LTE_LST.extend(['pmr -r 205'])
                    else:  # inserted
                        LTE_LST.extend([])
                    if 'NR' in selected_checkboxes:
                        NR_LST.extend(['pmr -r 405', 'pmr -r 408'])
                    else:  # inserted
                        NR_LST.extend([])
                    node_window = NodeSelectWindow(selected_checkboxes)
                    node_window.exec_()
                if activitw == 'Whole Period':
                    if 'Volte' in selected_checkboxes:
                        GSM_LST.extend(['pmr -r 210'])
                    else:  # inserted
                        GSM_LST.extend([])
                    if 'WCDMA' in selected_checkboxes:
                        WCDMA_PMR = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='Initial KPI_WCDMA', usecols='a,b,c,d', engine='openpyxl')
                        WCDMA_PMR['ROP Data'].replace('Single Rop 15 Minute', 'ROP by ROP', inplace=True)
                        WCDMA_PMR = WCDMA_PMR.loc[WCDMA_PMR['ROP Data'] == 'Whole Period']
                        WCDMA_PMR = WCDMA_PMR.loc[WCDMA_PMR['Status'] == 'Required']
                        WCDMA_PMR_LST = WCDMA_PMR['PMR No'].unique()
                        Final_PMR = 'pmr -m 1 -a -r ' + ','.join(map(str, WCDMA_PMR_LST)) + ' | grep '
                        WCDMA_LST.extend([Final_PMR])
                    else:  # inserted
                        WCDMA_LST.extend([])
                    if 'LTE' in selected_checkboxes:
                        LTE_LST.extend(['pmr -r 204'])
                    else:  # inserted
                        LTE_LST.extend([])
                    if 'NR' in selected_checkboxes:
                        NR_LST.extend(['pmr -r 404', 'pmr -r 407'])
                    else:  # inserted
                        NR_LST.extend([])
                    node_window = NodeSelectWindow(selected_checkboxes)
                    node_window.exec_()
            else:  # inserted
                QMessageBox.warning(self, 'Warning', 'Please select technology before continue.')
    app = QApplication(sys.argv)
    app.setStyleSheet('\n                    QPushButton {\n                        background-color: #4CAF50;\n                        border: none;\n                        color: white;\n                        padding: 10px 20px;\n                        text-align: center;\n                        text-decoration: none;\n                        display: inline-block;\n                        font-size: 16px;\n                        margin: 4px 2px;\n                        cursor: pointer;\n                        border-radius: 4px;\n                    }\n\n                    QTextEdit {\n                        font-size: 16px;\n                        border: 1px solid white;\n                        border-radius: 4px;\n                        padding: 8px;\n                    }\n\n                    CommandSelectWindow {\n                        background-color: #f2f2f2;\n                    }\n                ')
    window = MainWindow()
    window.show()
    app.exec_()
    return (selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST)

def AUDIT_main_SELECT(COMMAND_DEFALT_AUTO):
    def create_db_and_table():
        conn = sqlite3.connect(REPORT_INPATH22 + 'Audit_command_list.db')
        cursor = conn.cursor()
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS GSM (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS WCDMA (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS LTE (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS NR (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        conn.commit()
        conn.close()

    def insert_commands_into_dbcm(command_list, technology):
        try:
            command_list = list(dict.fromkeys(command_list))
        except:
            pass
        conn = sqlite3.connect(REPORT_INPATH22 + 'Audit_command_list.db')
        cursor = conn.cursor()
        table_name = technology.replace(' ', '_')
        cursor.execute(f'DELETE FROM {table_name}')
        if technology == 'GSM':
            cursor.execute('\n                               CREATE TABLE IF NOT EXISTS GSM (\n                                   id INTEGER PRIMARY KEY AUTOINCREMENT,\n                                   command TEXT\n                               )\n                           ')
        if technology == 'WCDMA':
            cursor.execute('\n                           CREATE TABLE IF NOT EXISTS WCDMA (\n                               id INTEGER PRIMARY KEY AUTOINCREMENT,\n                               command TEXT\n                           )\n                       ')
        if technology == 'LTE':
            cursor.execute('\n                           CREATE TABLE IF NOT EXISTS LTE (\n                               id INTEGER PRIMARY KEY AUTOINCREMENT,\n                               command TEXT\n                           )\n                       ')
        if technology == 'NR':
            cursor.execute('\n                           CREATE TABLE IF NOT EXISTS NR (\n                               id INTEGER PRIMARY KEY AUTOINCREMENT,\n                               command TEXT\n                           )\n                       ')
        for command in command_list:
            cursor.execute(f'INSERT INTO {table_name} (command) VALUES (?)', (command,))
        conn.commit()
        conn.close()

    class CommandSelectWindow(QDialog):
        def __init__(self, technologies, parent=None):
            super(CommandSelectWindow, self).__init__(parent)
            self.setWindowTitle('Command')
            self.setGeometry(650, 250, 650, 420)
            layout = QVBoxLayout(self)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            self.tab_widget = QTabWidget()
            layout.addWidget(self.tab_widget)
            self.tab_technology_mapping = {tech: idx for idx, tech in enumerate(technologies)}
            for technology in technologies:
                tab = QWidget()
                self.tab_widget.addTab(tab, technology)
                tab_layout = QVBoxLayout(tab)
                command_text_edit = QTextEdit()
                command_text_edit.setPlainText(self.load_data_from_dbcm(technology))
                command_text_edit.setWordWrapMode(QTextOption.NoWrap)
                tab_layout.addWidget(command_text_edit)
                submit_button = QPushButton('Next')
                submit_button.clicked.connect(self.submit_command)
                self.apply_bold_font(submit_button)
                tab_layout.addWidget(submit_button)
            self.tab_widget.tabBar().setStyleSheet('background-color: lightblue;')

        def apply_bold_font(self, widget):
            font = widget.font()
            font.setBold(True)
            widget.setFont(font)

        def closeEvent(self, event):
            self.window_closing = True
            event.accept()
            INSERT_COM = ['close']
            DATABASE = 'CAUDIT_CLOSE_WINDOW'
            DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)

        def submit_command(self):
            technology_lists = {'GSM': [], 'WCDMA': [], 'LTE': [], 'NR': []}
            for i in range(self.tab_widget.count()):
                current_tab = self.tab_widget.widget(i)
                command_text_edit = current_tab.findChild(QTextEdit)
                command_list = command_text_edit.toPlainText().splitlines()
                command_list = [command for command in command_list if command.strip()]
                technology = list(self.tab_technology_mapping.keys())[i]
                technology_lists[technology].extend(command_list)
                if 'Enter command list' in technology_lists[technology]:
                    technology_lists[technology].remove('Enter command list')
            for technology, commands in technology_lists.items():
                if technology == 'GSM':
                    GSM_LST.extend(commands)
                else:  # inserted
                    if technology == 'WCDMA':
                        WCDMA_LST.extend(commands)
                    else:  # inserted
                        if technology == 'LTE':
                            LTE_LST.extend(commands)
                        else:  # inserted
                            if technology == 'NR':
                                NR_LST.extend(commands)
            for teec in list(self.tab_technology_mapping.keys()):
                if teec == 'GSM':
                    command_list = GSM_LST
                if teec == 'WCDMA':
                    command_list = WCDMA_LST
                if teec == 'LTE':
                    command_list = LTE_LST
                if teec == 'NR':
                    command_list = NR_LST
                insert_commands_into_dbcm(command_list, teec)
            CHHKLST = []
            for tttt in selected_checkboxes:
                if tttt == 'GSM':
                    if not GSM_LST:
                        QMessageBox.warning(self, 'Warning', 'Please enter ' + str(tttt) + ' command list before continue.')
                        CHHKLST.append('stop')
                        break
                    CHHKLST.clear()
                if tttt == 'WCDMA':
                    if not WCDMA_LST:
                        QMessageBox.warning(self, 'Warning', 'Please enter ' + str(tttt) + ' command list before continue.')
                        CHHKLST.append('stop')
                        break
                    CHHKLST.clear()
                if tttt == 'LTE':
                    if not LTE_LST:
                        QMessageBox.warning(self, 'Warning', 'Please enter ' + str(tttt) + ' command list before continue.')
                        CHHKLST.append('stop')
                        break
                    CHHKLST.clear()
                if tttt == 'NR':
                    if not NR_LST:
                        QMessageBox.warning(self, 'Warning', 'Please enter ' + str(tttt) + ' command list before continue.')
                        CHHKLST.append('stop')
                        break
                    CHHKLST.clear()
            if not CHHKLST:
                self.accept()

        def load_data_from_dbcm(self, technology):
            conn = sqlite3.connect(REPORT_INPATH22 + 'Audit_command_list.db')
            cursor = conn.cursor()
            try:
                cursor.execute('SELECT command FROM ' + str(technology))
            except sqlite3.OperationalError:
                create_db_and_table()
                cursor.execute('SELECT command FROM ' + str(technology))
            data = cursor.fetchall()
            conn.close()
            content = '\n'.join([item[0] for item in data])
            if not str(content).strip():
                content = 'Enter command list'
            return content

    class defalcmdtWindow(QDialog):
        def __init__(self, technologies, parent=None):
            super(defalcmdtWindow, self).__init__(parent)

    class NodeSelectWindow(QDialog):
        def __init__(self, technologies, parent=None):
            super(NodeSelectWindow, self).__init__(parent)
            self.setWindowTitle('Node List')
            self.setGeometry(650, 250, 650, 420)
            layout = QVBoxLayout(self)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            self.tab_widget = QTabWidget()
            layout.addWidget(self.tab_widget)
            self.tab_technology_mapping = {tech: idx for idx, tech in enumerate(technologies)}
            for technology in technologies:
                tab = QWidget()
                self.tab_widget.addTab(tab, technology)
                tab_layout = QVBoxLayout(tab)
                command_text_edit = QTextEdit()
                command_text_edit.setWordWrapMode(QTextOption.NoWrap)
                tab_layout.addWidget(command_text_edit)
                submit_button = QPushButton('Submit')
                submit_button.clicked.connect(self.submit_command_final)
                self.apply_bold_font(submit_button)
                tab_layout.addWidget(submit_button)
            self.tab_widget.tabBar().setStyleSheet('background-color: lightblue;')

        def apply_bold_font(self, widget):
            font = widget.font()
            font.setBold(True)
            widget.setFont(font)

        def closeEvent(self, event):
            self.window_closing = True
            event.accept()
            INSERT_COM = ['close']
            DATABASE = 'CAUDIT_CLOSE_WINDOW'
            DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)

        def submit_command_final(self):
            try:
                for i in range(self.tab_widget.count()):
                    current_tab = self.tab_widget.widget(i)
                    command_text_edit = current_tab.findChild(QTextEdit)
                    command_list = command_text_edit.toPlainText().splitlines()
                    command_list = [command for command in command_list if command.strip()]
                    technology = list(self.tab_technology_mapping.keys())[i]
                    if technology == 'GSM':
                        GSM_NODE_LST.extend(command_list)
                    else:  # inserted
                        if technology == 'WCDMA':
                            WCDMA_NODE_LST.extend(command_list)
                        else:  # inserted
                            if technology == 'LTE':
                                LTE_NODE_LST.extend(command_list)
                            else:  # inserted
                                if technology == 'NR':
                                    NR_NODE_LST.extend(command_list)
            except:
                pass
            missing_tech = []
            for tttt in selected_checkboxes:
                if tttt == 'GSM' and (not GSM_NODE_LST):
                    missing_tech.append('GSM')
                else:  # inserted
                    if tttt == 'WCDMA' and (not WCDMA_NODE_LST):
                        missing_tech.append('WCDMA')
                    else:  # inserted
                        if tttt == 'LTE' and (not LTE_NODE_LST):
                            missing_tech.append('LTE')
                        else:  # inserted
                            if tttt == 'NR' and (not NR_NODE_LST):
                                missing_tech.append('NR')
            if missing_tech:
                missing_tech_names = ', '.join(missing_tech)
                QMessageBox.warning(self, 'Warning', f'Please enter Node list for the following technology before continuing: {missing_tech_names}')
            else:  # inserted
                self.accept()

    class MainWindow(QDialog):
        def __init__(self, parent=None):
            super(MainWindow, self).__init__(parent)
            self.setWindowTitle('Health check and audit')
            self.setGeometry(650, 250, 650, 420)
            layout = QVBoxLayout(self)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            self.command_text_edit = QTextEdit()
            font = QFont('Consolas', 12)
            self.command_text_edit.setFont(font)
            self.command_text_edit.setPlainText('\n\n                HEALTH CHECK AND PARAMETER AUDIT\n\nPlease select all required info like technology and activity type')
            layout.addWidget(self.command_text_edit)
            combo_layout = QVBoxLayout()
            self.activity_combo = QComboBox()
            self.activity_combo.setFont(QFont('Calibri', 12))
            self.activity_combo.addItem('Default Command')
            self.activity_combo.addItem('Customized Command')
            combo_layout.addWidget(self.activity_combo)
            layout.addLayout(combo_layout)
            spacer_item = QSpacerItem(30, 20, QSizePolicy.Minimum, QSizePolicy.Fixed)
            layout.addItem(spacer_item)
            radio_layout = QHBoxLayout()
            self.radio_button1 = QRadioButton('Health Check')
            self.apply_bold_font(self.radio_button1)
            radio_layout.addWidget(self.radio_button1)
            self.radio_button2 = QRadioButton('Parameter Audit')
            self.apply_bold_font(self.radio_button2)
            radio_layout.addWidget(self.radio_button2)
            radio_layout.addSpacing(55)
            self.radio_button3 = QRadioButton('Health Check & Parameter')
            self.apply_bold_font(self.radio_button3)
            radio_layout.addWidget(self.radio_button3)
            layout.addLayout(radio_layout)
            spacer_item = QSpacerItem(30, 20, QSizePolicy.Minimum, QSizePolicy.Fixed)
            layout.addItem(spacer_item)
            checkbox_layout = QHBoxLayout()
            self.GSM_checkbox = QCheckBox('GSM')
            self.apply_bold_font(self.GSM_checkbox)
            checkbox_layout.addWidget(self.GSM_checkbox)
            checkbox_layout.addSpacing(100)
            self.lte_checkbox = QCheckBox('LTE')
            self.apply_bold_font(self.lte_checkbox)
            checkbox_layout.addWidget(self.lte_checkbox)
            checkbox_layout.addSpacing(100)
            self.wcdma_checkbox = QCheckBox('WCDMA')
            self.apply_bold_font(self.wcdma_checkbox)
            checkbox_layout.addWidget(self.wcdma_checkbox)
            checkbox_layout.addSpacing(100)
            self.nr_checkbox = QCheckBox('NR')
            self.apply_bold_font(self.nr_checkbox)
            checkbox_layout.addWidget(self.nr_checkbox)
            layout.addLayout(checkbox_layout)
            submit_button = QPushButton('NEXT')
            submit_button.clicked.connect(self.submit_command)
            self.apply_bold_font(submit_button)
            layout.addWidget(submit_button)
            self.radio_button1.toggled.connect(self.radio_button_toggled)
            self.radio_button2.toggled.connect(self.radio_button_toggled)
            self.radio_button3.toggled.connect(self.radio_button_toggled)

        def apply_bold_font(self, widget):
            font = widget.font()
            font.setPointSize(11)
            widget.setFont(font)

        def radio_button_toggled(self):
            sender = self.sender()
            if sender.isChecked():
                DATABASE = 'ACTIVITYY'
                INSERT_COM = [sender.text()]
                DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)

        def closeEvent(self, event):
            self.window_closing = True
            event.accept()
            INSERT_COM = ['close']
            DATABASE = 'CAUDIT_CLOSE_WINDOW'
            DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)

        def submit_command(self):
            activitw = self.activity_combo.currentText()
            COMMAND_DEFALT_AUTO.clear()
            COMMAND_DEFALT_AUTO.append(activitw)
            selected_checkboxes.clear()
            if self.GSM_checkbox.isChecked():
                selected_checkboxes.append('GSM')
            if self.lte_checkbox.isChecked():
                selected_checkboxes.append('LTE')
            if self.wcdma_checkbox.isChecked():
                selected_checkboxes.append('WCDMA')
            if self.nr_checkbox.isChecked():
                selected_checkboxes.append('NR')
            if selected_checkboxes:
                DATABASE = 'ACTIVITYY'
                DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                if DATBASE_VALUE_READ:
                    self.accept()
                    if DATBASE_VALUE_READ[0] == 'Parameter Audit':
                        command_window = 'default'
                    else:  # inserted
                        if COMMAND_DEFALT_AUTO[0] == 'Default Command':
                            command_window = 'default'
                        else:  # inserted
                            command_window = CommandSelectWindow(selected_checkboxes)
                    if command_window == 'default':
                        AUDITALL_TECH = []

                        def defalttt_comm_fun():
                            for ttec in ['GSM', 'LTE', 'NR', 'WCDMA']:
                                try:
                                    conn = sqlite3.connect(REPORT_INPATH22 + 'Audit_command_list_default.db')
                                    cursor = conn.cursor()
                                    try:
                                        cursor.execute('SELECT command FROM ' + str(ttec))
                                    except sqlite3.OperationalError:
                                        pass
                                    data = cursor.fetchall()
                                    conn.close()
                                    content = [item[0] for item in data]
                                    AUDITALL_TECH.append(content)
                                except:
                                    if ttec == 'GSM':
                                        AUDITALL_TECH.append(['altk', 'hget RetSubUnit=1 userlabel|operationalState|electricalAntennaTilt', 'st cell', 'st ret', 'sdir'])
                                    if ttec == 'WCDMA':
                                        AUDITALL_TECH.append(['altk', 'hget RetSubUnit=1 userlabel|operationalState|electricalAntennaTilt', 'st cell', 'st ret', 'get . cellRange', 'get . maxdl', 'get . noofradio', 'pmr -m 1 -r 103', 'sdir'])
                                    if ttec == 'LTE':
                                        AUDITALL_TECH.append(['altk', 'hget RetSubUnit=1 userlabel|operationalState|electricalAntennaTilt', 'st cell', 'st ret', 'get . cellRange', 'get . maxdl', 'get . noofradio', 'pmr -r 205', 'pmr -m 1 -r 103', 'sdir'])
                                    if ttec == 'NR':
                                        AUDITALL_TECH.append(['altk', 'hget RetSubUnit=1 userlabel|operationalState|electricalAntennaTilt', 'st cell', 'st ret', 'get . cellRange', 'get . maxdl', 'get . noofradio', 'sdir'])
                        defalttt_comm_fun()
                        if 'GSM' in selected_checkboxes:
                            GSM_LST.extend(AUDITALL_TECH[0])
                        else:  # inserted
                            GSM_LST.extend([])
                        if 'WCDMA' in selected_checkboxes:
                            WCDMA_LST.extend(AUDITALL_TECH[3])
                        else:  # inserted
                            WCDMA_LST.extend([])
                        if 'LTE' in selected_checkboxes:
                            LTE_LST.extend(AUDITALL_TECH[1])
                        else:  # inserted
                            LTE_LST.extend([])
                        if 'NR' in selected_checkboxes:
                            NR_LST.extend(AUDITALL_TECH[2])
                        else:  # inserted
                            NR_LST.extend([])
                        node_window = NodeSelectWindow(selected_checkboxes)
                        node_window.exec_()
                        COMMAND_DEFALT_AUTO.clear()
                    else:  # inserted
                        if command_window.exec_():
                            node_window = NodeSelectWindow(selected_checkboxes)
                            node_window.exec_()
                            COMMAND_DEFALT_AUTO.clear()
                else:  # inserted
                    QMessageBox.warning(self, 'Warning', 'Please select Activity type before continue.')
                try:
                    DATBASE_VALUE_READ.claer()
                    del DATBASE_VALUE_READ
                except:
                    return None
            else:  # inserted
                QMessageBox.warning(self, 'Warning', 'Please select technology before continue.')
    app = QApplication(sys.argv)
    app.setStyleSheet('\n                    QPushButton {\n                        background-color: #4CAF50;\n                        border: none;\n                        color: white;\n                        padding: 10px 20px;\n                        text-align: center;\n                        text-decoration: none;\n                        display: inline-block;\n                        font-size: 16px;\n                        margin: 4px 2px;\n                        cursor: pointer;\n                        border-radius: 4px;\n                    }\n\n                    QTextEdit {\n                        font-size: 16px;\n                        border: 1px solid white;\n                        border-radius: 4px;\n                        padding: 8px;\n                    }\n\n                    CommandSelectWindow {\n                        background-color: #f2f2f2;\n                    }\n                ')
    window = MainWindow()
    window.show()
    app.exec_()
    return (selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST)

def site_audit(querry):
    responses = {'audit_audit': ['Welcome! Ready for the health check and audit? Reply \'yes\' or \'no\'.'], 'yesy': ['Please provide required details to continue.'], 'noooo': ['Thanks! how Chat Buddy will assist you !'], 'default': ['I\'m not quite clear what you meant, can you confirm with a \'yes\' or \'no\'? To exit, simply reply with \'quit\'.'], 'quiitt': ['Thanks! how Chat Buddy will assist you !']}

    def get_response(input_text):
        if any((word in input_text.lower() for word in ['audit_audit'])):
            return random.choice(responses['audit_audit'])
        if any((word in input_text.lower() for word in ['yes'])):
            return random.choice(responses['yesy'])
        if any((word in input_text.lower() for word in ['no'])):
            return random.choice(responses['noooo'])
        if any((word in input_text.lower() for word in ['quit', 'exit'])):
            return random.choice(responses['quiitt'])
        return random.choice(responses['default'])
    return get_response(querry)

def COMMAND_SELECT():
    GSM_LST = []
    LTE_LST = []
    WCDMA_LST = []
    NR_LST = []

    def create_db_and_table():
        conn = sqlite3.connect('./res/command_list.db')
        cursor = conn.cursor()
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS GSM (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS WCDMA (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS LTE (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        cursor.execute('\n                       CREATE TABLE IF NOT EXISTS NR (\n                           id INTEGER PRIMARY KEY AUTOINCREMENT,\n                           command TEXT\n                       )\n                   ')
        conn.commit()
        conn.close()

    def insert_commands_into_dbcm(command_list, technology):
        try:
            command_list = list(dict.fromkeys(command_list))
        except:
            pass
        conn = sqlite3.connect('./res/command_list.db')
        cursor = conn.cursor()
        table_name = technology.replace(' ', '_')
        cursor.execute(f'DELETE FROM {table_name}')
        if technology == 'GSM':
            cursor.execute('\n                               CREATE TABLE IF NOT EXISTS GSM (\n                                   id INTEGER PRIMARY KEY AUTOINCREMENT,\n                                   command TEXT\n                               )\n                           ')
        if technology == 'WCDMA':
            cursor.execute('\n                           CREATE TABLE IF NOT EXISTS WCDMA (\n                               id INTEGER PRIMARY KEY AUTOINCREMENT,\n                               command TEXT\n                           )\n                       ')
        if technology == 'LTE':
            cursor.execute('\n                           CREATE TABLE IF NOT EXISTS LTE (\n                               id INTEGER PRIMARY KEY AUTOINCREMENT,\n                               command TEXT\n                           )\n                       ')
        if technology == 'NR':
            cursor.execute('\n                           CREATE TABLE IF NOT EXISTS NR (\n                               id INTEGER PRIMARY KEY AUTOINCREMENT,\n                               command TEXT\n                           )\n                       ')
        for command in command_list:
            cursor.execute(f'INSERT INTO {table_name} (command) VALUES (?)', (command,))
        conn.commit()
        conn.close()

    class CommandSelectWindow(QDialog):
        def __init__(self, parent=None):
            super(CommandSelectWindow, self).__init__(parent)
            self.setWindowTitle('Command')
            self.setGeometry(700, 250, 500, 500)
            layout = QVBoxLayout(self)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            self.tab_widget = QTabWidget()
            layout.addWidget(self.tab_widget)
            self.tab_technology_mapping = {0: 'GSM', 1: 'WCDMA', 2: 'LTE', 3: 'NR'}
            for i in range(4):
                tab = QWidget()
                self.tab_widget.addTab(tab, self.tab_technology_mapping[i])
                tab_layout = QVBoxLayout(tab)
                command_text_edit = QTextEdit()
                command_text_edit.setPlainText(self.load_data_from_dbcm(self.tab_technology_mapping[i]))
                command_text_edit.setWordWrapMode(QTextOption.NoWrap)
                tab_layout.addWidget(command_text_edit)
                browse_button = QPushButton('Browse')
                browse_button.clicked.connect(self.browse_file)
                self.apply_bold_font(browse_button)
                tab_layout.addWidget(browse_button)
                submit_button = QPushButton('Submit')
                submit_button.clicked.connect(self.submit_command)
                self.apply_bold_font(submit_button)
                tab_layout.addWidget(submit_button)
            self.tab_widget.tabBar().setStyleSheet('background-color: lightblue;')

        def apply_bold_font(self, widget):
            font = widget.font()
            font.setBold(True)
            widget.setFont(font)

        def browse_file(self):
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getOpenFileName(self, 'Select File')
            if file_path:
                with open(file_path, 'r') as file:
                    data = file.read()
                    current_tab = self.tab_widget.currentWidget()
                    command_text_edit = current_tab.findChild(QTextEdit)
                    command_text_edit.setPlainText(data)

        def submit_command(self):
            for i in range(self.tab_widget.count()):
                current_tab = self.tab_widget.widget(i)
                command_text_edit = current_tab.findChild(QTextEdit)
                command_list = command_text_edit.toPlainText().splitlines()
                command_list = [command for command in command_list if command.strip()]
                technology = self.tab_technology_mapping[i]
                create_db_and_table()
                insert_commands_into_dbcm(command_list, technology)
                command_text_edit.setPlainText(self.load_data_from_dbcm(technology))
                if technology == 'GSM':
                    for lll in command_list:
                        GSM_LST.append(lll)
                if technology == 'WCDMA':
                    for lll in command_list:
                        WCDMA_LST.append(lll)
                if technology == 'LTE':
                    for lll in command_list:
                        LTE_LST.append(lll)
                if technology == 'NR':
                    for lll in command_list:
                        NR_LST.append(lll)
            self.accept()

        def load_data_from_dbcm(self, technology):
            conn = sqlite3.connect('./res/command_list.db')
            cursor = conn.cursor()
            try:
                cursor.execute('SELECT command FROM ' + str(technology))
            except sqlite3.OperationalError:
                create_db_and_table()
                cursor.execute('SELECT command FROM ' + str(technology))
            data = cursor.fetchall()
            conn.close()
            content = '\n'.join([item[0] for item in data])
            if not str(content).strip():
                content = 'Enter command list'
            return content
    app = QApplication(sys.argv)
    window = CommandSelectWindow()
    app.setStyleSheet('\n                QPushButton {\n                    background-color: #4CAF50;\n                    border: none;\n                    color: white;\n                    padding: 10px 20px;\n                    text-align: center;\n                    text-decoration: none;\n                    display: inline-block;\n                    font-size: 16px;\n                    margin: 4px 2px;\n                    cursor: pointer;\n                    border-radius: 4px;\n                }\n\n                QTextEdit {\n                    font-size: 15px;\n                    border: 2px solid #4CAF50;\n                    border-radius: 4px;\n                    padding: 8px;\n                }\n\n                CommandSelectWindow {\n                    background-color: #f2f2f2;\n                }\n            ')
    window.exec_()
    return (GSM_LST, WCDMA_LST, LTE_LST, NR_LST)

def NODE_SELECT():
    def create_db_and_table():
        conn = sqlite3.connect('./res/Node_list.db')
        cursor = conn.cursor()
        cursor.execute('\n            CREATE TABLE IF NOT EXISTS command_list (\n                id INTEGER PRIMARY KEY AUTOINCREMENT,\n                command TEXT\n            )\n        ')
        conn.commit()
        conn.close()

    def insert_commands_into_db(command_list):
        conn = sqlite3.connect('./res/Node_list.db')
        cursor = conn.cursor()
        for command in command_list:
            cursor.execute('INSERT INTO command_list (command) VALUES (?)', (command,))
        conn.commit()
        conn.close()

    class CommandSelectWindow(QDialog):
        def __init__(self, parent=None):
            super(CommandSelectWindow, self).__init__(parent)
            self.setWindowTitle('Node')
            self.setGeometry(800, 250, 350, 500)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            layout = QVBoxLayout(self)
            self.browse_button = QPushButton('Browse')
            self.browse_button.clicked.connect(self.browse_file)
            self.apply_bold_font(self.browse_button)
            layout.addWidget(self.browse_button)
            self.submit_button = QPushButton('Submit')
            self.submit_button.clicked.connect(self.NODE_SELECT22221)
            self.apply_bold_font(self.submit_button)
            layout.addWidget(self.submit_button)
            self.command_text_edit = QTextEdit()
            self.command_text_edit.setPlainText(self.load_data_from_db())
            self.command_text_edit.setWordWrapMode(QTextOption.NoWrap)
            layout.addWidget(self.command_text_edit)

        def apply_bold_font(self, widget):
            font = widget.font()
            font.setBold(True)
            widget.setFont(font)

        def browse_file(self):
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getOpenFileName(self, 'Select File')
            if file_path:
                with open(file_path, 'r') as file:
                    data = file.read()
                    self.command_text_edit.setPlainText(data)

        def load_data_from_db(self):
            conn = sqlite3.connect('./res/Node_list.db')
            cursor = conn.cursor()
            try:
                cursor.execute('SELECT command FROM command_list')
            except sqlite3.OperationalError:
                create_db_and_table()
                cursor.execute('SELECT command FROM command_list')
            data = cursor.fetchall()
            conn.close()
            content = '\n'.join([item[0] for item in data])
            if not str(content).strip():
                content = 'Enter NodeId'
            return content

        def NODE_SELECT22221(self):
            try:
                os.remove('./res/Node_list.db')
            except:
                pass
            command_list = self.command_text_edit.toPlainText().splitlines()
            command_list = [command.upper() for command in command_list if command.strip()]
            create_db_and_table()
            insert_commands_into_db(command_list)
            self.command_text_edit.setPlainText(self.load_data_from_db())
            self.accept()
            return command_list

    def NODE_SELECT22221():
        app = QApplication(sys.argv)
        window = CommandSelectWindow()
        app.setStyleSheet('\n            QPushButton {\n                background-color: #4CAF50;\n                border: none;\n                color: white;\n                padding: 10px 20px;\n                text-align: center;\n                text-decoration: none;\n                display: inline-block;\n                font-size: 16px;\n                margin: 4px 2px;\n                cursor: pointer;\n                border-radius: 4px;\n            }\n\n            QTextEdit {\n                font-size: 17px;\n                border: 2px solid #4CAF50;\n                border-radius: 4px;\n                padding: 8px;\n            }\n\n            CommandSelectWindow {\n                background-color: #f2f2f2;\n            }\n        ')
        result = window.exec_()
        lines_list = window.command_text_edit.toPlainText().splitlines()
        nodelst = []
        for line in lines_list:
            nodelst.append(line.upper().strip())
        if nodelst == ['ENTER NODEID']:
            nodelst = []
        return nodelst
    return NODE_SELECT22221()

def ui_commands():
    class UICommands(QWidget):
        def __init__(self):
            super().__init__()
            self.setWindowTitle('command')
            self.init_ui()

        def init_ui(self):
            con = sqlite3.connect(aql_path23)
            Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
            con.close()
            Rdf['New_lst'] = Rdf['indexxx'].astype(str) + '_' + Rdf['value'].astype(str)
            COMMAND_LST = Rdf['New_lst'].unique()
            COMMAND_LST = list(COMMAND_LST)
            COMMAND_LST2 = []
            for cmm in COMMAND_LST:
                COMMAND_LST2.append(cmm.split('_')[0])
            try:
                del Rdf
            except:
                pass
            self.search_box = QLineEdit(self)
            self.search_box.setPlaceholderText('Search command...')
            self.search_box.textChanged.connect(self.update_button_visibility)
            scroll_area = QScrollArea()
            scroll_widget = QWidget()
            self.layout = QVBoxLayout(scroll_widget)
            self.buttons = COMMAND_LST2
            self.create_buttons()
            self.layout.setSpacing(5)
            self.setWindowFlag(Qt.WindowStaysOnTopHint)
            scroll_area.setWidgetResizable(True)
            scroll_area.setWidget(scroll_widget)
            main_layout = QVBoxLayout(self)
            main_layout.addWidget(self.search_box)
            main_layout.addWidget(scroll_area)
            self.center_on_screen()
            self.adjustSize()

        def create_buttons(self):
            for button_label in self.buttons:
                button = QPushButton(button_label)
                button.clicked.connect(self.button_clicked)
                button.setStyleSheet('\n                    QPushButton {\n                         background-color: #2ecc71;\n                        color: white;\n                        border: 2px solid #27ae60;\n                        padding: 10px 20px;\n                        font-size: 18px;\n                        border-radius: 5px;\n                        margin-bottom: 10px;\n                    }\n\n                    QPushButton:hover {\n                        background-color: #2980b9;\n                        border: 2px solid #3498db;\n                    }\n                ')
                self.layout.addWidget(button)

        def update_button_visibility(self):
            search_text = self.search_box.text().lower()
            for index in range(self.layout.count()):
                item = self.layout.itemAt(index)
                if isinstance(item.widget(), QPushButton):
                    button = item.widget()
                    button.setVisible(search_text in button.text().lower())

        def center_on_screen(self):
            qr = self.frameGeometry()
            cp = QDesktopWidget().availableGeometry().center()
            qr.moveCenter(cp)
            self.move(qr.topLeft())

        def button_clicked(self):
            sender = self.sender()
            self.selected_option = sender.text()
            self.close()

    def ui_commands_qt():
        app = QApplication(sys.argv)
        window = UICommands()
        window.setGeometry(850, 300, 300, window.height())
        window.show()
        app.exec_()
        try:
            return window.selected_option
        except AttributeError:
            return None
    result = ui_commands_qt()
    return result

def log_clening(combined_cmd, cmd_ex_sh, cmd_node):
    filter_listl = ['Start Time', 'End Time', 'Your Choice:', 'Enter the report number', 'Eg \"1\"', 'Date:', 'Node SW:', 'Report from', 'ropfiles)', 'gzip', 'warning:', '/tmp/', 'Collecting', 'Added', '.....', 'quit', 'amos', 'get', 'INFO:', 'Checking', 'Connected', 'Getting', 'BASIC', 'OTHER', 'HELP', 'Bye', 'Framework', 'Ericsson', 'moshell', 'COMMANDS', 'lt all', 'rbs', '$ssh', '|', '.txt', '.gz', 'Node:', str(cmd_node.upper()) + '>']

    def contains_only_numeric_and_parenthesis(input_str):
        pattern = '^\\d+\\)$'
        return bool(re.match(pattern, input_str))

    def clentext_b2(text):
        lines = text.split('\n')
        cleane_text = []
        for ll in lines:
            try:
                cleane_text.append(ll.strip())
                if '------' in str(cleane_text[(-1)].strip()) and str(cleane_text[(-2)].strip())!= '' and (str(cleane_text[(-3)].strip()) == ''):
                    del cleane_text[(-2)]
            except:
                cleane_text.append(ll.strip())
        cleaned_input_string = '\n'.join(cleane_text)
        return cleaned_input_string
    lines = combined_cmd.splitlines()
    filtered_lines = [line for line in lines if not any((element in line for element in filter_listl))]
    filtered_lines = '\n'.join(filtered_lines)
    tetstst2 = clentext_b2(filtered_lines)
    NEW_TXT = []
    for lln in tetstst2.splitlines():
        if 'start time' in str(lln).lower().strip():
            continue
        NEW_TXT.append(str(lln).strip())
    final_lines = '\n'.join(NEW_TXT)
    final_lines = '\n'.join([line.replace('Adm State', 'AdmState') for line in final_lines.split('\n')])
    final_lines = '\n'.join([line.replace('Op. State', 'Op.State') for line in final_lines.split('\n')])
    final_lines = '\n'.join([line.replace(' (', '_(') for line in final_lines.split('\n')])
    NEW_TXT = []
    for lln in final_lines.splitlines():
        if '..' not in str(lln).strip():
            numberoption = contains_only_numeric_and_parenthesis(str(lln.split(' ')[0].strip()))
            if numberoption == True:
                continue
            NEW_TXT.append(str(lln).strip())
    final_lines = '\n'.join(NEW_TXT)

    def add_separator(text):
        lines = text.split('\n')
        checkdash = []
        for lllllll in lines:
            if '====' in lllllll.strip():
                checkdash.append(lllllll)
            if '----' in lllllll.strip():
                checkdash.append(lllllll)
        if len(checkdash) < 1:
            while lines and (not lines[0].strip()):
                lines.pop(0)
            while lines and (not lines[(-1)].strip()):
                lines.pop()
            lines.insert(0, '=================================================================================================================')
            lines.append('=================================================================================================================')
            return '\n'.join(lines)
        else:  # inserted
            return text
    final_lines = add_separator(final_lines)
    return final_lines

def final_log_process(input_string):
    def replace_multiple_blank_lines(text):
        pattern = re.compile('\\n\\s*\\n+', re.MULTILINE)
        result_text = re.sub(pattern, '\n++++++++++++++++++++++++++++++++++++++++++++++++++\n', text)
        return result_text

    def process_log(text, filter_list, remove_string_lst, command_llst, text_to_remove):
        pattern = re.compile('={3,}|-{3,}|^>>> Total:', re.MULTILINE)
        matches = pattern.finditer(text)
        start_index = None
        end_index = None
        for match in matches:
            if start_index is None:
                start_index = match.end()
            else:  # inserted
                end_index = match.start()
        result = ''
        if start_index is not None and end_index is not None:
            extracted_content = text[start_index:end_index].strip()
            extracted_content = extracted_content.replace(text_to_remove, '')
            extracted_content = replace_multiple_blank_lines(extracted_content)
            extracted_lines = [line for line in extracted_content.split('\n') if line.strip()]
            filtered_lines = [line for line in extracted_lines if all((keyword not in line for keyword in filter_list))]
            filtered_lines = [line for line in filtered_lines if all((keyword not in line for keyword in filtered_lines))]
            filter_list_lower = [keyword.lower() for keyword in filter_list]
            filter_list_upper = [keyword.upper() for keyword in filter_list]
            filtered_lines = [line for line in filtered_lines if all((keyword not in line for keyword in filtered_lines))]
            filtered_lines = [line for line in filtered_lines if all((keyword not in line for keyword in filtered_lines))]
            filtered_lines = [line for line in filtered_lines if all((keyword not in line for keyword in filtered_lines))]
            filtered_lines = [line for line in filtered_lines if all((keyword not in line for keyword in filtered_lines))]
            for remove_string in remove_string_lst:
                filtered_lines = [line for line in filtered_lines if line!= remove_string]
            filtered_lines = [line for line in filtered_lines if line[:2]!= 'p/']
            filtered_lines = [line for line in filtered_lines if len(line) >= 3]
            filtered_lines.insert(0, '=================================================================================================================')
            filtered_lines.append('=================================================================================================================')
            filtered_lines.insert(0, '+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
            filtered_lines.append('+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++')
            result = '\n'.join(filtered_lines)
            result = re.sub('(\\++\\n)+', '+++++++++++++++++++++++++++++++++\\n', result)
        return result

    def Create_section(text):
        split_text = re.split('\\+{30,}', text)
        split_text = [s.strip() for s in split_text if s.strip()]
        SECTION_LST = []
        for idx, section in enumerate(split_text):
            section = re.sub('\\={2,}', '', section)
            section = re.sub('\\-{2,}', '', section)
            section = re.sub('\\n\\s*\\n', '\n', section)
            SECTION_LST.append(f'{section}')
        return SECTION_LST
    filter_list = ['warning:', 'tmp', 'Collecting', 'group:', '.....', 'quit', 'amos', 'get', 'INFO:', 'Checking', 'Connected', 'Getting', 'BASIC', 'HELP', 'Bye', 'Framework', 'Ericsson', 'moshell', 'COMMANDS', 'lt all', 'rbs', '$ssh', '|', 'Node:']
    remove_string_lst = ['.', '..', '...', '....', 'OK', 'OK.']
    command_llst = ['cabex', 'altk']
    text_to_remove = '__  __  ____   _____\n                  /\\   |  \\/  |/ __ \\ / ____|\n                 /  \\  | \\  / | |  | | (___  \n                / /\\ \\ | |\\/| | |  | |\\___ \\ \n               / ____ \\| |  | | |__| |____) |\n              /_/    \\_\\_|  |_|\\____/|_____/'

    def clentext_b2(text):
        lines = text.split('\n')
        cleane_text = []
        for ll in lines:
            try:
                cleane_text.append(ll.strip())
                if '------' in str(cleane_text[(-1)].strip()) and str(cleane_text[(-2)].strip())!= '' and (str(cleane_text[(-3)].strip()) == ''):
                    del cleane_text[(-2)]
            except:
                cleane_text.append(ll.strip())
        cleaned_input_string = '\n'.join(cleane_text)
        return cleaned_input_string
    input_string = clentext_b2(input_string)
    Process_logs = process_log(input_string, filter_list, remove_string_lst, command_llst, text_to_remove)
    Final_log_lst2 = Create_section(Process_logs)
    return Final_log_lst2

def process_clean_df(result_df, special_cmd):
    mod_df = result_df.copy()
    mod_df.columns = mod_df.columns.str.strip()
    mod_df = mod_df.dropna(axis=1, how='all')
    mod_df = mod_df.rename(columns=lambda x: x.strip())
    h_lst = mod_df.columns.tolist()
    for header in h_lst:
        for dfdf in special_cmd:
            if dfdf in header:
                try:
                    mod_df = mod_df.replace('', pd.NA).dropna(subset=[header])
                except:
                    pass
                try:
                    mod_df[header] = mod_df[header].str.split('_').str[0].astype(float)
                except:
                    pass
    mod_df = mod_df.dropna(axis=1, how='all')
    mod_df = mod_df.rename(columns=lambda x: x.strip())
    try:
        del mod_df['Command']
        return mod_df
    except:
        return mod_df

def convert_to_excel(excel_df_LST, OUTPATH, special_cmd):
    result_df = pd.concat(excel_df_LST, axis=0, ignore_index=True)
    cmd_filter_lst = result_df['Command'].unique()
    with pd.ExcelWriter(OUTPATH + 'Processed Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', engine='xlsxwriter') as writer2:
        for cmmd in cmd_filter_lst:
            mod_df = result_df.copy()
            mod_df = mod_df.loc[mod_df['Command'] == cmmd]
            mod_df.columns = mod_df.columns.str.strip()
            mod_df = mod_df.dropna(axis=1, how='all')
            mod_df = mod_df.rename(columns=lambda x: x.strip())
            h_lst = mod_df.columns.tolist()
            for header in h_lst:
                for dfdf in special_cmd:
                    if dfdf in header:
                        try:
                            mod_df = mod_df.replace('', pd.NA).dropna(subset=[header])
                        except:
                            pass
            mod_df = mod_df.dropna(axis=1, how='all')
            mod_df = mod_df.rename(columns=lambda x: x.strip())
            try:
                del mod_df['Command']
            except:
                pass
            mod_df_lst = mod_df.columns.tolist()
            mod_df = mod_df.drop_duplicates(subset=mod_df_lst, keep='first')
            mod_df.to_excel(writer2, sheet_name=cmmd, index=False, encoding='utf-8-sig')
            headr_colour_multi(writer2, mod_df, cmmd, '#facf' + str('91'), ['KPI'], '#facf' + str('91'))

def headr_colour_multi(writer2, dfff, sh_name, colour, hed_lst, tab):
    workbook = writer2.book
    worksheet = writer2.sheets[sh_name]
    worksheet.set_tab_color(tab)
    workbook.formats[0].set_font_size(9)
    workbook.formats[0].set_align('center')
    workbook.formats[0].set_font('Calibri')
    header_format = workbook.add_format({'bold': True, 'font_name': 'Calibri', 'font_size': 10, 'font_color': '#FFFFFF', 'bg_color': '#262626', 'align': 'left', 'valign': 'top', 'text_wrap': False})
    header_format_row = workbook.add_format({'font_name': 'Calibri', 'font_size': 9, 'align': 'left', 'valign': 'top', 'text_wrap': False})
    for col_num, value in enumerate(dfff.columns.values):
        worksheet.write(0, col_num, value, header_format)
    num_rows = worksheet.dim_rowmax
    for row_num in range(num_rows + 1):
        worksheet.set_row(row_num, None, header_format_row)
    for column in dfff:
        column_length = max(dfff[column].astype(str).apply(len).max(), len(column))
        col_idx = dfff.columns.get_loc(column)
        writer2.sheets[sh_name].set_column(col_idx, col_idx, column_length)
    return dfff

def crate_txt_df(input_string, cmd_node, cmd_ex):
    try:
        linesi = input_string.split('\n')
        finalstr = []
        for kkkk in linesi:
            if '-admitted' in kkkk:
                pass  # postinserted
            if 'coli>' in kkkk:
                continue
            finalstr.append(kkkk)
        input_string = '\n'.join(finalstr)
    except:
        pass
    try:
        def replace_spaces_with_na(text):
            lines = text.split('\n')
            modified_lines = []
            for index, line in enumerate(lines):
                if len(line.strip()) > 0:
                    modified_line = line.replace('            ', ' deletestring ')
                    modified_lines.append(modified_line)
                else:  # inserted
                    modified_lines.append(line)
            return '\n'.join(modified_lines)
        combined_cmd = '\n'.join([line.replace('Date & Time', 'Date&Time') for line in input_string.split('\n')])
        combined_cmd = '\n'.join([line.replace('_(UTC)', ' UTC') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('Specific Problem', 'Specific_Problem') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('MO_(Cause/AdditionalInfo)', 'MO(Cause/AdditionalInfo)') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('Adm State', 'AdmState') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('Op. State', 'Op.State') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace(' (', '_(') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('T S OpMode', 'T.S.OpMode') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('AutoNeg MacAddress', 'AutoNegMacAddress') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('TEMP MO', 'TEMPMO') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace(' _(', '_(') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('coli>', '') for line in combined_cmd.split('\n')])
        combined_cmd = '\n'.join([line.replace('NRAT is dormant.', '') for line in combined_cmd.split('\n')])
        combined_heder = re.sub(' {2,}', ' ', combined_cmd.split('\n')[0])
        combined_heder_count = combined_heder.split(' ')
        combined_heder_count = len(list(filter(lambda x: x!= '', combined_heder_count)))
        new_text = []
        for lll in combined_cmd.split('\n'):
            lll22 = re.sub(' {2,}', '  ', lll)
            com_he = lll22.split(' ')
            com_he = list(filter(lambda x: x!= '', com_he))
            if len(com_he) < combined_heder_count:
                lll = replace_spaces_with_na(lll)
                pattern = re.compile('\\bdeletestring\\s+deletestring\\b')
                lll = pattern.sub('', lll)
                lll = re.sub(' {2,}', '  ', lll)
                new_text.append(lll)
            else:  # inserted
                lll = re.sub(' {2,}', '  ', lll)
                new_text.append(lll)
        input_string = '\n'.join(new_text)
        input_string = '\n'.join([line.replace('Specific_Problem', 'Specific Problem') for line in input_string.split('\n')])
        if 'S Specific Problem' in input_string:
            lines = input_string.split('\n')
            lines = [i for i in lines if i]
            lines = [item for item in lines if '====' not in item]
            lines = [item for item in lines if 'Total:' not in item]
            lines = [item for item in lines if 'Added' not in item]
            header = re.split('\\s{2,}', lines[0].strip())
            if 'S Specific Problem' in header:
                header = ['Date & Time (UTC)', 'Severty ', 'Specific Problem           MO (Cause/AdditionalInfo)']
            columns = {col: [] for col in header}
            for line in lines[1:]:
                match = re.match('(\\d{4}-\\d{2}-\\d{2} \\d{2}:\\d{2}:\\d{2}) (\\w) (.+)$', line)
                if match:
                    for i, col in enumerate(header):
                        columns[col].append(match.group(i + 1).strip())
            df = pd.DataFrame(columns)
            try:
                df[['Specific Problem', 'MO (Cause/AdditionalInfo)']] = df['Specific Problem           MO (Cause/AdditionalInfo)'].str.split('\\s{2,}', n=1, expand=True)
                del df['Specific Problem           MO (Cause/AdditionalInfo)']
            except:
                try:
                    df = df.rename(columns={'Specific Problem           MO (Cause/AdditionalInfo)': 'Specific Problem'})
                    if 'MO (Cause/AdditionalInfo)' not in df.columns:
                        df['MO (Cause/AdditionalInfo)'] = np.nan
                except:
                    pass
            crate_txt_df.df_fram = 'ok'
            return df
        else:  # inserted
            if cmd_ex.lower() == 'cell':
                text_list = input_string.strip().splitlines()
                text_listm = []
                i = 1
                for line in text_list:
                    if i == 1:
                        line = re.sub('\\b(\\S+)\\s(\\S+)\\b', '\\1_\\2', line)
                    line1 = line.replace('            ', '  nonenone  ')
                    line1 = line1.replace(' (', '_(')
                    text_listm.append(line1)
                    i = i + 1
                text = '\n'.join(text_listm)
                rows = text.strip().split('\n')
                column_count = len(rows[0].split())
                column_widths = [max((len(row.split()[i].strip()) if len(row.split()) > i else 0 for row in range(column_count))) for i in range(column_count)]
                columns = []
                for row in rows:
                    column_data = []
                    split_row = row.split()
                    for i, width in enumerate(column_widths):
                        if len(split_row) > i:
                            column_data.append(split_row[i].strip().ljust(width))
                        else:  # inserted
                            column_data.append(' ' * width)
                    columns.append(column_data)
                header = columns[0]
                data = columns[1:]
                df = pd.DataFrame(data, columns=header)
                df = df.dropna(how='all')
                df = df[~df.apply(lambda row: row.str.isspace().all(), axis=1)]
                try:
                    df = df.applymap(lambda x: None if x == 'nonenone    ' else x)
                except:
                    pass
                try:
                    df.columns = df.columns.str.replace('_', ' ')
                except:
                    pass
                try:
                    df = df.applymap(lambda x: str(x).replace('_', ' ') if x is not None else x)
                except:
                    pass
                crate_txt_df.df_fram = 'ok'
                return df
            else:  # inserted
                if 'st ret' in cmd_ex.lower():
                    try:
                        def remove_extra_spaces_line_by_line(text):
                            lines = text.splitlines()
                            cleaned_lines = []
                            for line in lines:
                                cleaned_line = re.sub('\\s+', ' ', line).strip()
                                cleaned_lines.append(cleaned_line)
                            return '\n'.join(cleaned_lines)
                        text_data = input_string.replace(' (', '_(')
                        flines = text_data.splitlines()
                        header_line = flines[0]
                        header_line = re.sub('\\s{4,}', '  ', header_line)
                        header_lst = re.split('\\s{2,}', header_line)
                        text_data = re.sub('\\s{4,}', '  NA  ', text_data)
                        text_data = text_data.splitlines()
                        text_data[0] = header_line
                        text_data = '\n'.join(text_data)
                        text_data = remove_extra_spaces_line_by_line(text_data)
                        lines = text_data.strip().split('\n')
                        data = []
                        for line in lines:
                            parts = line.split(None, 3)
                            if len(parts) == 4:
                                data.append(parts)
                            else:  # inserted
                                data[(-1)][(-1)] += ' ' + line
                        df = pd.DataFrame(data, columns=header_lst)
                        df = df.drop(df.index[0])
                        crate_txt_df.df_fram = 'ok'
                        return df
                    except:
                        df = pd.DataFrame()
                        return df
                else:  # inserted
                    if cmd_ex.lower() == 'ret':
                        try:
                            def remove_extra_spaces_line_by_line(text):
                                lines = text.splitlines()
                                cleaned_lines = []
                                for line in lines:
                                    cleaned_line = re.sub('\\s+', ' ', line).strip()
                                    cleaned_lines.append(cleaned_line)
                                return '\n'.join(cleaned_lines)
                            text_data = input_string.replace(' (', '_(')
                            flines = text_data.splitlines()
                            header_line = flines[0]
                            header_line = re.sub('\\s{4,}', '  ', header_line)
                            header_lst = re.split('\\s{2,}', header_line)
                            text_data = re.sub('\\s{4,}', '  NA  ', text_data)
                            text_data = text_data.splitlines()
                            text_data[0] = header_line
                            text_data = '\n'.join(text_data)
                            text_data = remove_extra_spaces_line_by_line(text_data)
                            lines = text_data.strip().split('\n')
                            data = []
                            for line in lines:
                                parts = line.split(None, 3)
                                if len(parts) == 4:
                                    data.append(parts)
                                else:  # inserted
                                    data[(-1)][(-1)] += ' ' + line
                            df = pd.DataFrame(data, columns=header_lst)
                            df = df.drop(df.index[0])
                            crate_txt_df.df_fram = 'ok'
                            return df
                        except:
                            df = pd.DataFrame()
                            return df
                    else:  # inserted
                        if 'pmTaInit2Distr'.lower() in cmd_ex.lower():
                            try:
                                lines = input_string.split('\n')[1:]
                                text_without_first_row = '\n'.join(lines)
                                df = pd.read_csv(StringIO(text_without_first_row), delimiter=' ', names=['Object', 'pmTaInit2Distr'])
                                split_columns = df['pmTaInit2Distr'].str.split(',', expand=True)
                                split_columns = split_columns[split_columns.columns[::2]]
                                split_columns.columns = [f'TA {i + 1 - 1}' for i in range(split_columns.shape[1])]
                                df = pd.concat([df, split_columns], axis=1)
                                df = df.drop('pmTaInit2Distr', axis=1)
                                crate_txt_df.df_fram = 'ok'
                                return df
                            except:
                                df = pd.DataFrame()
                                return df
                        if 'propagationdelay' in cmd_ex.lower():
                            try:
                                lines = input_string.split('\n')[1:]
                                text_without_first_row = '\n'.join(lines)
                                df = pd.read_csv(StringIO(text_without_first_row), delimiter=' ', names=['Object', 'PropagationDelay', 'PrachPropagationDelay'])
                                split_columns = df['PrachPropagationDelay'].str.split(',', expand=True)
                                split_columns.columns = [f'TA {i + 1 - 1}' for i in range(split_columns.shape[1])]
                                df = pd.concat([df, split_columns], axis=1)
                                df = df.drop('PrachPropagationDelay', axis=1)
                                crate_txt_df.df_fram = 'ok'
                                return df
                            except:
                                df = pd.DataFrame()
                                return df
                        rows = [row for row in input_string.split('\n') if row.strip()]
                        rows = [item for item in rows if not item.startswith('Last MO:')]
                        try:
                            first_row_has_semicolon = ';' in rows[0]
                        except:
                            first_row_has_semicolon = False
                        if first_row_has_semicolon == False:
                            try:
                                first_row_has_semicolon = ';' in rows[1]
                            except:
                                first_row_has_semicolon = False
                        if first_row_has_semicolon == True:
                            try:
                                lines = input_string.strip().split('\n')
                                rows = [line.split(';') for line in lines]
                                df = pd.DataFrame(rows[1:], columns=rows[0])
                                try:
                                    df = df.applymap(lambda x: None if x == 'nonenone    ' else x)
                                except:
                                    pass
                                try:
                                    df.columns = df.columns.str.replace('_', ' ')
                                except:
                                    pass
                                try:
                                    df = df.applymap(lambda x: str(x).replace('_', ' ') if x is not None else x)
                                except:
                                    pass
                                crate_txt_df.df_fram = 'ok'
                                return df
                            except:
                                return
                        if cmd_ex.lower()!= 'st cell':
                            total_pattern = re.compile('Total: \\d+ MOs')
                            group_pattern = re.compile('Added \\d+ MOs to group: \\w+')
                            alarm_pattern = re.compile('>>> Total: \\d+ Alarms \\(0 Critical, 0 Major\\)')
                            filtered_lines = [line for line in input_string.split('\n') if not total_pattern.match(line) and (not group_pattern.match(line)) and (not alarm_pattern.match(line))]
                            input_string = '\n'.join(filtered_lines)
                            text = input_string.replace(' (', '_(')
                            rows = text.strip().split('\n')
                            column_count = len(rows[0].split())
                            column_widths = [max((len(row.split()[i].strip()) if len(row.split()) > i else 0 for row in range(column_count))) for i in range(column_count)]
                            columns = []
                            for row in rows:
                                column_data = []
                                split_row = row.split()
                                for i, width in enumerate(column_widths):
                                    if len(split_row) > i:
                                        column_data.append(split_row[i].strip().ljust(width))
                                    else:  # inserted
                                        column_data.append(' ' * width)
                                columns.append(column_data)
                            header = columns[0]
                            data = columns[1:]
                            df = pd.DataFrame(data, columns=header)
                            df = df.dropna(how='all')
                            df = df[~df.apply(lambda row: row.str.isspace().all(), axis=1)]
                            crate_txt_df.df_fram = 'ok'
                            return df
                        else:  # inserted
                            text_list = input_string.strip().splitlines()
                            text_listm = []
                            i = 1
                            for line in text_list:
                                if i == 1:
                                    line = re.sub('\\b(\\S+)\\s(\\S+)\\b', '\\1_\\2', line)
                                line1 = line.replace('            ', '  nonenone  ')
                                line1 = line1.replace(' (', '_(')
                                text_listm.append(line1)
                                i = i + 1
                            text = '\n'.join(text_listm)
                            rows = text.strip().split('\n')
                            column_count = len(rows[0].split())
                            column_widths = [max((len(row.split()[i].strip()) if len(row.split()) > i else 0 for row in range(column_count))) for i in range(column_count)]
                            columns = []
                            for row in rows:
                                column_data = []
                                split_row = row.split()
                                for i, width in enumerate(column_widths):
                                    if len(split_row) > i:
                                        column_data.append(split_row[i].strip().ljust(width))
                                    else:  # inserted
                                        column_data.append(' ' * width)
                                columns.append(column_data)
                            header = columns[0]
                            data = columns[1:]
                            df = pd.DataFrame(data, columns=header)
                            df = df.dropna(how='all')
                            df = df[~df.apply(lambda row: row.str.isspace().all(), axis=1)]
                            try:
                                df = df.applymap(lambda x: None if x == 'nonenone    ' else x)
                            except:
                                pass
                            try:
                                df.columns = df.columns.str.replace('_', ' ')
                            except:
                                pass
                            try:
                                df = df.applymap(lambda x: str(x).replace('_', ' ') if x is not None else x)
                            except:
                                pass
                            crate_txt_df.df_fram = 'ok'
                            return df
    except:
        return None

def RRSG(self, Login_comm_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, PATHHHH_TIME_mokgt, tecno_mode, RBS_RNC):
    if VPN_select_METHOD_F[0] == 'VPN-CAS':
        try:
            ssh.close()
            RRSG_auto_login.remote_conn.close()
            del RRSG_auto_login.remote_conn
        except:
            pass
    if MOTASK == 'MOBATCH':
        try:
            os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt))
        except:
            pass
    if MOTASK == 'KGET':
        try:
            os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KGET_') + str(PATHHHH_TIME_mokgt))
        except:
            pass
    if 'login_error' in LOGINERROR_LST:
        thread = threading.Thread(target=enabled_button, args=(self,))
        thread.start()
        thread.join()
        return
    check_lst = []

    def read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK):
        commnd_sh = []
        CIPRILST = []
        commnd_sh.append(cmd_ex)

        def remove_color_codes(input_string):
            ansi_escape = re.compile('\\x1b[^m]*m')
            clean_string = ansi_escape.sub('', input_string)
            html_color = re.compile('<[^>]*>')
            clean_string = html_color.sub('', clean_string)
            return clean_string
        if MOTASK == 'KGET':
            try:
                os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KGET_') + str(PATHHHH_TIME_mokgt))
            except:
                pass
            read_received_data_auto.out = ''
            output_B = b''
            while not output_B.endswith(b'$'):
                output_B = remote_conn.recv(999999)
                write_t = output_B.decode(encoding='utf-8')
                write_te_1 = remove_color_codes(write_t)
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('\a', '')
                self.update_textbox(f'{write_te_1}\n')
                try:
                    self.reset_timer(None)
                except:
                    pass
                if Save_logfile == 'YES':
                    with open(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KGET_') + str(PATHHHH_TIME_mokgt) + '\\' + str(cmd_node) + '_KGET.Log', 'a+', encoding='utf-8') as f:
                        f.writelines(f'{write_te_1}\n\n')
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(cmd_node) + '__' + str(command) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                last_line = str(write_te_1.strip().split('\n')[(-1)]).split(' ')[0].strip()
                try:
                    last_line2 = str(write_te_1.strip().split('\n')[(-1)])[(-1)]
                except:
                    last_line2 = '=============='
                if last_line == 'Total:':
                    break
                if last_line == (str(cmd_node).strip() + '>').strip():
                    break
                if last_line2 == '$':
                    break
                if last_line == 'e:':
                    break
                if last_line2 == ':':
                    break
            read_received_data_auto.fil = OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KGET_') + str(PATHHHH_TIME_mokgt) + '\\' + str(cmd_node) + '_KGET.Log'
        if MOTASK == 'MOBATCH':
            read_received_data_auto.out = ''
            output_B = b''
            combined_cmd = ''
            command_not_found = []
            while not (output_B.endswith(b'> ') or output_B.endswith(b'<') or output_B.endswith(b': ') or output_B.endswith(b'<') or output_B.endswith(b'< ') or output_B.endswith(b':') or output_B.endswith(b'$ ')):
                output_B += remote_conn.recv(99999)
                write_t = output_B.decode(encoding='utf-8')
                write_te_1 = remove_color_codes(write_t)
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('', '')
                write_te_1 = write_te_1.replace('\a', '')
                self.update_textbox(f'{write_te_1}\n')
                if cmd_ex == 'sdir' and 'CPRI links' in write_te_1 and (not CIPRILST):
                    CIPRILST.append(write_te_1)
                try:
                    self.reset_timer(None)
                except:
                    pass
                if 'no such command' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'command not found' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                write_te_1 = '\n'.join([line.replace('coli>/rc/nrat/ue print -admitted', '') for line in write_te_1.split('\n')])
                write_te_1 = '\n'.join([line.replace('NRAT is dormant.', '') for line in write_te_1.split('\n')])
                if 'ue print -admitted'.lower() in command.lower():
                    write_te_1 = '\n'.join([line.replace('coli>', '') for line in write_te_1.split('\n')])
                if 'coli>' in write_te_1 and 'no such command' not in command_not_found:
                    command_not_found.append('no such command')
                if 'amos' in command and 'AMOS error' in write_te_1 and ('no such command' not in command_not_found):
                    command_not_found.append('no such command')
                last_line = output_B.decode(encoding='utf-8').split('\n')[(-1)]
                last_line = ' '.join(last_line.split()).strip()
                if check_lst and last_line == check_lst[0]:
                    read_received_data_auto.brek = 'break'
                try:
                    if last_line[(-1)].strip() == '>':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '<':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == ':':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
                try:
                    if last_line[(-1)].strip() == '$':
                        check_lst.clear()
                        check_lst.append(last_line)
                except IndexError:
                    pass
            if not command_not_found:
                write_text1 = output_B.decode(encoding='utf-8')
                write_text = remove_color_codes(write_text1)
                write_text = write_text.replace('', '')
                write_text = write_text.replace('', '')
                write_text = write_text.replace('\a', '')
                if Save_logfile == 'YES':
                    if chk_oss!= 'LOGIN_CNM' and chk_oss!= 'no':
                        with open(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt) + '\\' + str(cmd_node) + '.Log', 'a+', encoding='utf-8') as f:
                            f.writelines(f'{write_text}\n\n')
                        with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(cmd_node) + '__' + str(command) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    if chk_oss == 'OSS_CNM':
                        combined_cmd += f'{write_text}\n'
                crate_txt_df.df_fram = ''
                if chk_oss == 'OSS_CNM' and command!= 'q' and (command!= 'quit'):
                    def remove_line_containing(text, substring):
                        lines = text.split('\n')
                        filtered_lines = [line for line in lines if substring not in line]
                        return '\n'.join(filtered_lines)
                    combined_cmd = combined_cmd.replace('It is recommended to remove duplicate counter definitions from the PM scanners.', '')
                    combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                    combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo)', 'Date & Time (UTC)   S Specific Problem                    MO (Cause/AdditionalInfo)')
                    combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo) Operator', '')
                    combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                    combined_cmd = remove_line_containing(combined_cmd, 'ACKNOWLEDGED ALARMS')
                    if 'pmr' in combined_cmd.lower():
                        combined_cmd = remove_line_containing(combined_cmd, 'pmr')
                        combined_cmd = remove_line_containing(combined_cmd, 'PMR')
                    combined_cmd = combined_cmd.strip()
                    if '16Qam'.lower() not in command.lower():
                        combined_cmd = log_clening(combined_cmd, cmd_ex, cmd_node)
                    Final_log_lst = final_log_process(combined_cmd)
                    try:
                        combined_cmd = ''
                    except:
                        pass
                    Final_log_lst = [item.strip() for item in Final_log_lst if item.strip()]
                    try:
                        if Final_log_lst[0] == '\n':
                            Final_log_lst = []
                    except:
                        pass
                    if Final_log_lst:
                        i = 1
                        Final_log_lst = [item for item in Final_log_lst if not item.startswith('+')]
                        Final_log_lst = [item for item in Final_log_lst if not item.startswith('\n+')]
                        for section in Final_log_lst:
                            lines = section.strip().split('\n')
                            data = [line.split('\t') for line in lines]
                            if len(data) > 1:
                                crate_df = crate_txt_df(section, cmd_node, cmd_ex)
                                try:
                                    crate_df.replace('deletestring', '', inplace=True)
                                except:
                                    pass
                                if crate_df is not None and (not crate_df.empty):
                                    crate_df = crate_df[~crate_df.apply(lambda row: any(row.astype(str).str.contains('Total:')), axis=1)]
                                    crate_df.columns = crate_df.columns.str.strip()
                                    if not crate_df.empty:
                                        try:
                                            try:
                                                crate_df['VSWR (RL)'] = crate_df['VSWR (RL)'].str.strip()
                                            except:
                                                pass
                                            VSWR = [col for col in crate_df.columns if 'VSWR' in col]
                                            if VSWR:
                                                try:
                                                    crate_df[['VSWR', 'RL']] = crate_df['VSWR_(RL)'].str.split('_', expand=True)
                                                except:
                                                    crate_df[['VSWR', 'RL']] = crate_df['VSWR (RL)'].str.split(' ', expand=True)
                                                crate_df['RL'] = crate_df['RL'].str.strip('()')
                                                crate_df['VSWR'] = crate_df['VSWR'].astype(float)
                                                crate_df['RL'] = crate_df['RL'].astype(float)
                                                try:
                                                    del crate_df['VSWR_(RL)']
                                                except:
                                                    try:
                                                        del crate_df['VSWR (RL)']
                                                    except:
                                                        pass
                                        except:
                                            pass
                                        try:
                                            crate_df = crate_df.applymap(lambda x: x.replace('_', ' '))
                                        except:
                                            pass
                                        try:
                                            crate_df['Node Id'] = cmd_node
                                        except:
                                            pass
                                        try:
                                            crate_df.insert(0, 'Node Id', crate_df.pop('Node Id'))
                                        except:
                                            pass
                                        excel_df = crate_df.copy()
                                        counter = [col for col in excel_df.columns if 'counter' in col.lower()]
                                        if counter:
                                            excel_df = melted_conter(excel_df)
                                        try:
                                            excel_df = excel_df.applymap(lambda x: str(x).strip())
                                        except:
                                            pass
                                        try:
                                            del crate_df['Command']
                                        except:
                                            pass
                                        try:
                                            del crate_df['Baseline Check']
                                        except:
                                            pass
                                        if crate_txt_df.df_fram!= 'error':
                                            tabulate.PRESERVE_WHITESPACE = True
                                            data_list = crate_df.to_dict(orient='records')
                                            table = tabulate(data_list, headers='keys', tablefmt='presto')
                                            table = cleeend_text(table)
                                            self.update_textbox_big_window(table, iii)
                                            try:
                                                self.reset_timer(None)
                                            except:
                                                pass
                                            if 'done' not in processdone:
                                                processdone.append('done')
                                            if not excel_df.empty:
                                                Save_log_switch = self.swich_Sav_loss.get()
                                                if Save_log_switch == 1:
                                                    Save_logfile = 'YES'
                                                else:  # inserted
                                                    Save_logfile = 'NO'
                                                if Save_logfile == 'YES' and chk_oss!= 'LOGIN_CNM' and (chk_oss!= 'no'):
                                                    try:
                                                        del excel_df['Command']
                                                    except:
                                                        pass
                                                    try:
                                                        try:
                                                            if len(Final_log_lst) > 1:
                                                                cmd_ex = commnd_sh[0] + str(i)
                                                            else:  # inserted
                                                                pass
                                                        except:
                                                            pass
                                                        try:
                                                            if commnd_sh[0] == 'sdir':
                                                                cmd_ex = 'sdir'
                                                        except:
                                                            pass
                                                        try:
                                                            if commnd_sh[0] == 'cabex':
                                                                cmd_ex = 'cabex'
                                                        except:
                                                            pass
                                                        try:
                                                            if commnd_sh[0] == 'sdi':
                                                                cmd_ex = 'sdi'
                                                        except:
                                                            pass
                                                        if tecno_mode == 'WCDMA':
                                                            if RBS_RNC!= 'KGET':
                                                                write_excel_new(excel_df, 'Mobatch_Output ' + str(RBS_RNC) + ' ' + datetime.now().strftime('%d%m%Y') + '.xlsx', cmd_ex, OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt) + '\\')
                                                            if RBS_RNC == 'KGET':
                                                                write_excel_new(excel_df, 'Mobatch_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', cmd_ex, OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt) + '\\')
                                                        else:  # inserted
                                                            write_excel_new(excel_df, 'Mobatch_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', cmd_ex, OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt) + '\\')
                                                        try:
                                                            del excel_df
                                                        except:
                                                            pass
                                                    except:
                                                        pass
                                        i = i + 1
                        try:
                            if command.strip().lower() == 'sdir':
                                CIPRI_DFF = CIPRIDFFF(CIPRILST[0])
                                CIPRILST.clear()
                                if CIPRI_DFF is not None and (not CIPRI_DFF.empty):
                                    try:
                                        CIPRI_DFF['Node Id'] = cmd_node
                                    except:
                                        pass
                                    try:
                                        CIPRI_DFF.insert(0, 'Node Id', CIPRI_DFF.pop('Node Id'))
                                    except:
                                        pass
                                    if not CIPRI_DFF.empty:
                                        try:
                                            CIPRI_DFF = CIPRI_DFF.applymap(lambda x: str(x).strip())
                                        except:
                                            pass
                                        try:
                                            def conditione(row):
                                                if int(row['CPRI Link NOK']) > 0:
                                                    return 'Fail'
                                                if int(row['CPRI Link OKW']) > 0:
                                                    return 'Fail'
                                                return 'Pass'
                                            CIPRI_DFF['Baseline Check'] = CIPRI_DFF.apply(conditione, axis=1)
                                        except:
                                            pass
                                        if tecno_mode == 'WCDMA':
                                            if RBS_RNC!= 'KGET':
                                                write_excel_new(CIPRI_DFF, 'Mobatch_Output ' + str(RBS_RNC) + ' ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'CPRI_link', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt) + '\\')
                                            if RBS_RNC == 'KGET':
                                                write_excel_new(CIPRI_DFF, 'Mobatch_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'CPRI_link', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt) + '\\')
                                        else:  # inserted
                                            write_excel_new(CIPRI_DFF, 'Mobatch_Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'CPRI_link', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt) + '\\')
                                        tabulate.PRESERVE_WHITESPACE = True
                                        data_list = CIPRI_DFF.to_dict(orient='records')
                                        table = tabulate(data_list, headers='keys', tablefmt='presto')
                                        table = cleeend_text(table)
                                        CIPRI_PRINT(self, table)
                                    try:
                                        del CIPRI_DFF
                                    except:
                                        pass
                        except:
                            pass
                        try:
                            del Final_log_lst
                        except:
                            pass
                        try:
                            del command_not_found
                        except:
                            pass
                        combined_cmd = ''

    def send_command(remote_conn, command, chk_oss, iii, cmd_node, Save_logfile, cmd_ex, MOTASK):
        read_received_data_auto.fil = 'None'
        remote_conn.send(command + '\n')
        read_received_data_auto(remote_conn, chk_oss, iii, command, cmd_node, Save_logfile, cmd_ex, MOTASK)
        send_command.fil = read_received_data_auto.fil

    def main(logcommand_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK):
        iii = 1
        paaaawe = None
        try:
            llll = RRSG_auto_login.remote_conn
            paaaawe = 'ok'
        except:
            try:
                con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                del Rdf['indexxx']
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                PID_LST = Rdf['value'].tolist()
                remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
            except:
                RSG_LST_N = []
            if not RSG_LST_N:
                if VPN_select_METHOD_F[0]!= 'VPN-CAS':
                    showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 10000)
            else:  # inserted
                if not VPN_select_METHOD_F:
                    if VPN_select_METHOD_F[0] == 'VPN-CAS':
                        showMessage_qt('Please update VPN/CAS details before continue.', 10000)
                        thread = threading.Thread(target=enabled_button, args=(self,))
                        thread.start()
                        thread.join()
                else:  # inserted
                    paaaawe = OTPPP_tets()
            if paaaawe == None and 'cancle' not in processdone:
                processdone.append('cancle')
            try:
                if VPN_select_METHOD_F[0]!= 'VPN-CAS':
                    if is_non_numeric_string(paaaawe):
                        showMessage_qt('Wrong OTP password please try again.', 10000)
                    else:  # inserted
                        if paaaawe:
                            self.Auto_login_sc(paaaawe, RSG_LST_N)
                else:  # inserted
                    if paaaawe:
                        self.Auto_login_sc(paaaawe, RSG_LST_N)
            except:
                pass
        try:
            chk_rssg = RRSG_auto_login.remote_conn
            chk_rssg_chk = 'OK'
        except:
            chk_rssg_chk = 'NOK'
        if chk_rssg_chk == 'NOK':
            clear_all_lists()
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Something wrong in RSG/VPN connection please try again after some time.') + '\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
        else:  # inserted
            if not CHECK_ISF_SEQUENCE:
                sig = os.getlogin().lower()
                priority = 'High'
                ISF_API_Based = 'SharePoint'
                KEY = '5b97555705ce40128131d58c8d3596f9'
                try:
                    con = sqlite3.connect('./res/ISFdetails.db')
                    Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                    con.close()
                    del Rdf['indexxx']
                    ppPID_LST = Rdf['value'].tolist()
                    ppPID_LST2 = []
                    for ppp in ppPID_LST:
                        ppPID_LST2.append(decrypt(ppp, 3))
                    del Rdf['value']
                    Rdf['value'] = ppPID_LST2
                    PID_LST = Rdf['value'].tolist()
                    remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                    ISF_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                    try:
                        del Rdf
                    except:
                        pass
                except:
                    ISF_LST = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                chk_line = []
                for llll in ISF_LST:
                    if 'line' in llll.lower():
                        chk_line.append('stop_process')
                        break
                    if 'start_process' not in chk_line:
                        chk_line.append('start_process')
                if 'start_process' in chk_line:
                    projectID = ISF_LST[0]
                    type1 = ISF_LST[1]
                    nodname = ISF_LST[2]
                    wOName = ISF_LST[3]
                    executionPlanName = ISF_LST[4]
                    lastModifiedBy = ISF_LST[5]
                    ISF_STEP_START_stepID = ISF_LST[6]
                    ISF_STEP_START_Task_Id = ISF_LST[7]
                    if not ISF_STEP_START_Task_Id_lst:
                        ISF_STEP_START_Task_Id_lst.append(ISF_STEP_START_Task_Id)
                    ISF_WO_CREATE_woId, STEP_START = ISF_API(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY, ISF_STEP_START_stepID, ISF_STEP_START_Task_Id)
                    WO_LSTT = [ISF_WO_CREATE_woId, STEP_START]
                    ISF_STTUS_MSG = 'sucess'
                    if WO_LSTT[0] == 'Some issue in isf workflow':
                        ISF_STTUS_MSG = 'ISF details not correct. please check and try again.'
                    if ISF_STTUS_MSG == 'sucess' and WO_LSTT[1]!= 'sucess':
                        ISF_STTUS_MSG = 'TaskID and StepID not correct. please provide the valid taskID for the StepID.'
            else:  # inserted
                ISF_STTUS_MSG = 'sucess'
            if ISF_STTUS_MSG == 'sucess' and (not CHECK_ISF_SEQUENCE):
                if 'work_order_raised' not in CHECK_ISF_SEQUENCE:
                    CHECK_ISF_SEQUENCE.append('work_order_raised')
                if MOTASK == 'MOBATCH':
                    def sucss_msg(self):
                        self.textbox2.delete(1.0, 'end')
                        self.after(0, self.textbox2.insert, 'end', 'ISF Work Order created.\n\n')
                        self.after(0, self.textbox2.insert, 'end', 'Signum    : ' + str(sig) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'Nodtype   : ' + str(type1) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'Nodname   : ' + str(nodname) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'ProjectID : ' + str(projectID) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'WOID      : ' + str(ISF_WO_CREATE_woId) + '\n')
                        try:
                            self.after(0, self.textbox2.insert, 'end', 'Time      : ' + str(ISF_API.timestr) + '\n\n')
                        except:
                            pass
                        try:
                            self.after(0, self.textbox2.insert, 'end', str(ISF_API.urlcreateWO) + '\n\n')
                        except:
                            return None
                    with open('./res/CHATLOGS/' + str(sig) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Signum__' + str(sig) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Nodtype__' + str(type1) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Nodname__' + str(nodname) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'ProjectID__' + str(projectID) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'WOID__' + str(ISF_WO_CREATE_woId) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Time__' + str(ISF_API.timestr) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    threading.Thread(target=sucss_msg, args=(self,)).start()
                if MOTASK == 'KGET':
                    def sucss_msg2(self):
                        self.textbox2.delete(1.0, 'end')
                        self.after(0, self.textbox2.insert, 'end', 'ISF Work Order created.\n\n')
                        self.after(0, self.textbox2.insert, 'end', 'Signum    : ' + str(sig) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'Nodtype   : ' + str(type1) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'Nodname   : ' + str(nodname) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'ProjectID : ' + str(projectID) + '\n')
                        self.after(0, self.textbox2.insert, 'end', 'WOID      : ' + str(ISF_WO_CREATE_woId) + '\n')
                        try:
                            self.after(0, self.textbox2.insert, 'end', 'Time      : ' + str(ISF_API.timestr) + '\n\n')
                        except:
                            pass
                        try:
                            self.after(0, self.textbox2.insert, 'end', str(ISF_API.urlcreateWO) + '\n\n')
                        except:
                            return None
                    with open('./res/CHATLOGS/' + str(sig) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Signum__' + str(sig) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Nodtype__' + str(type1) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Nodname__' + str(nodname) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'ProjectID__' + str(projectID) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'WOID__' + str(ISF_WO_CREATE_woId) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(MOTASK) + '__' + 'Time__' + str(ISF_API.timestr) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    threading.Thread(target=sucss_msg2, args=(self,)).start()
            if 'work_order_raised' in CHECK_ISF_SEQUENCE:
                KGETFILE = []
                if paaaawe!= None:
                    send_command.fil = 'None'
                    if 'FAIL' not in establish_ssh_conn_lst:
                        for cmd_node in Node_cmd_lst:
                            self.textbox_LIVECMD.delete(1.0, 'end')
                            chk_oss = 'NODE_CNM'
                            cmd_ex_sh = ''
                            send_command(RRSG_auto_login.remote_conn, 'amos ' + str(cmd_node), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            send_command(RRSG_auto_login.remote_conn, 'lt all', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            RBS_RBS_switch = self.swich_Sec_rbs.get()
                            if RBS_RBS_switch == 1:
                                RBS_RBS_switch_loe = 'YES'
                            else:  # inserted
                                RBS_RBS_switch_loe = 'NO'
                            if RBS_RBS_switch_loe == 'YES':
                                if RBSPASS_CRED_CRREDD[0] == 'rbs':
                                    user_rbs = 'rbs'
                                else:  # inserted
                                    try:
                                        user_rbs = RBSPASS_CRED_CRREDD[0]
                                    except:
                                        user_rbs = 'rbs'
                                if RBSPASS_CRED_CRREDD[1] == 'rbs':
                                    pass_rbs = 'rbs'
                                else:  # inserted
                                    try:
                                        pass_rbs = RBSPASS_CRED_CRREDD[1]
                                    except:
                                        pass_rbs = 'rbs'
                                print('ok4')
                                send_command(RRSG_auto_login.remote_conn, str(user_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                                send_command(RRSG_auto_login.remote_conn, str(pass_rbs), chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            for cmd_ex in oss_cmd_lst:
                                cmd_ex_sh = clean_sheet_name(cmd_ex, max_length=28)
                                chk_oss = 'OSS_CNM'
                                send_command(RRSG_auto_login.remote_conn, cmd_ex, chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                                iii = iii + 1
                            send_command(RRSG_auto_login.remote_conn, 'quit', chk_oss, iii, cmd_node, Save_logfile, cmd_ex_sh, MOTASK)
                            if MOTASK == 'KGET':
                                KGETFILE.append(str(cmd_node) + '$' + send_command.fil)
                        if MOTASK == 'KGET':
                            if KGETFILE:
                                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait KGET processing ongoing !\n')
                                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                start_index = '1.0'
                                while True:
                                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                    if not start_index:
                                        break
                                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                    start_index = end_index
                                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                threads = []
                                for kk in KGETFILE:
                                    cmd_node = str(kk).split('$')[0].strip()
                                    node_path = str(kk).split('$')[(-1)].strip()
                                    self.update_textbox(f'Please wait {cmd_node} Kget underprocessing. \n\n')
                                    try:
                                        self.reset_timer(None)
                                    except:
                                        pass
                                    t = threading.Thread(target=KGET_PARSER, args=(self, node_path, cmd_node))
                                    t.start()
                                    threads.append(t)
                                for t in threads:
                                    t.join()
                                dump_path = os.path.dirname(node_path)
                                combined_xl(dump_path)
                                thread_combined_xl = threading.Thread(target=combined_xl_audit, args=dump_path)
                                thread_combined_xl.start()
                                thread_combined_xl.join()
                                if KGETFILE:
                                    pass  # postinserted
                            self.update_textbox('Kget logs converted into excel check in output folder. \n\n')
                            try:
                                self.reset_timer(None)
                            except:
                                return
            else:  # inserted
                self.textbox2.delete(1.0, 'end')
                self.after(0, self.textbox2.insert, 'end', ISF_STTUS_MSG + '\n\n')
    if SQL_METHOD == 'LOCAL':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check(user_signum, command_to_barred_required, command_to_barred)
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)
    if SQL_METHOD == 'AZURE_CLOUDE':
        user_signum = os.getlogin().lower()
        command_to_barred_required = 'YES'
        oss_cmd_lst_v2 = []
        for command_to_barred in oss_cmd_lst:
            if change_User:
                ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check_AZURE(change_User, command_barred, user_signum, command_to_barred_required, command_to_barred)
            else:  # inserted
                if 'amos' in command_to_barred.lower():
                    ucercmd_AUTHENTICATION = 'User_Authenticated'
                else:  # inserted
                    if 'rbs' in command_to_barred.lower():
                        ucercmd_AUTHENTICATION = 'User_Authenticated'
                    else:  # inserted
                        if 'lt all' in command_to_barred.lower():
                            ucercmd_AUTHENTICATION = 'User_Authenticated'
                        else:  # inserted
                            ucercmd_AUTHENTICATION = 'NO'
            if ucercmd_AUTHENTICATION == 'User_Authenticated':
                oss_cmd_lst_v2.append(command_to_barred)
        main(Login_comm_lst, Node_cmd_lst, oss_cmd_lst_v2, router, username, password, Save_logfile, MOTASK)

def cmd_summary(self, clean_df1, special_cmd, vswr_threshold):
    h_lst = clean_df1.columns.tolist()
    for header in h_lst:
        for dfdf in special_cmd:
            if dfdf in header:
                clean_df1 = clean_df1[clean_df1['VSWR_(RL)'] >= vswr_threshold]
                if not clean_df1.empty:
                    unique_cells = clean_df1['Sector/AntennaGroup/Cells_(State:CellIds:PCIs)'].unique()
                    self.update_textbox_big_window_enter_cmd_summ('========================', '1')
                    try:
                        self.reset_timer(None)
                    except:
                        pass
                    self.update_textbox_big_window_enter_cmd_summ('Summary:', '1')
                    try:
                        self.reset_timer(None)
                    except:
                        pass
                    for cell_name in unique_cells:
                        filterdf = clean_df1[clean_df1['Sector/AntennaGroup/Cells_(State:CellIds:PCIs)'] == cell_name]
                        self.update_textbox_big_window_enter_cmd_summ('-------------------', '1')
                        try:
                            self.reset_timer(None)
                        except:
                            pass
                        if not filterdf.empty:
                            max_vswr_value = filterdf['VSWR_(RL)'].max()
                            self.update_textbox_big_window_enter_cmd_summ(f'Cell: {cell_name}', '1')
                            try:
                                self.reset_timer(None)
                            except:
                                pass
                            self.update_textbox_big_window_enter_cmd_summ('Impact: High VSWR', '1')
                            try:
                                self.reset_timer(None)
                            except:
                                pass
                            self.update_textbox_big_window_enter_cmd_summ(f'VSWR value: {max_vswr_value}', '1')
                            try:
                                self.reset_timer(None)
                            except:
                                continue
                    self.update_textbox_big_window_enter_cmd_summ('========================', '1')
                    try:
                        self.reset_timer(None)
                    except:
                        pass
                    self.update_textbox_big_window_enter_cmd_summ('\n', '1')
                    try:
                        self.reset_timer(None)
                    except:
                        continue
                else:  # inserted
                    self.update_textbox_big_window_enter_cmd_summ('VSWR ok for all cells', '1')
                    try:
                        self.reset_timer(None)
                    except:
                        continue

def add_commad_to_db(self):
    KEYYY = KEYWORD_ADD[0]
    keyword = KEYYY.split('_')[0].strip()
    command = KEYYY.split('_')[1].strip()
    KEYWORD_ADD.clear()
    SKIP_COMMANDDDD_NEW = 'YES'
    if SQL_METHOD == 'LOCAL':
        SKPI_COMMAND = 'NO'
        try:
            SPLIT_txt = list(filter(None, command.split(':')))
            SPLIT_txt = SPLIT_txt[0].upper().strip()
            if SPLIT_txt[:2] in GSMCMD_lst:
                SKPI_COMMAND = 'YES'
        except:
            pass
        if SKPI_COMMAND == 'NO':
            try:
                conn = sqlite3.connect('./res/conversations.db')
                cursor = conn.cursor()
                cursor.execute('INSERT INTO conversations (text) VALUES (?)', (str(keyword).strip() + '_' + str(command).strip(),))
                conn.commit()
                cursor.execute('SELECT text FROM conversations')
                SKIP_COMMANDDDD_NEW = 'NO'
            except:
                pass
    if SQL_METHOD == 'AZURE_CLOUDE':
        try:
            con = sqlite3.connect('./res/AZURE_CRED.db')
            Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
            con.close()
            del Rdf['indexxx']
            Rdf = Rdf.drop_duplicates(subset=['value'], keep='first')
            ppPID_LST = Rdf['value'].tolist()
            ppPID_LST2 = []
            for ppp in ppPID_LST:
                ppPID_LST2.append(decrypt(ppp, 3))
            del Rdf['value']
            Rdf['value'] = ppPID_LST2
            PID_LST = Rdf['value'].tolist()
            remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5']
            AZ_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
            AZ_LST.insert(1, os.getlogin().lower())
            try:
                del Rdf
            except:
                pass
        except:
            AZ_LST = []
        try:
            azure_cred = 'YES'
            if AZ_LST:
                for kl in AZ_LST:
                    if 'LINE' in kl:
                        azure_cred = 'NO'
                        break
            if azure_cred == 'YES':
                gateway_host = AZ_LST[0]
                remote_host = AZ_LST[3]
                remote_username = AZ_LST[4]
                remote_password = AZ_LST[5]
                gateway_user = os.getlogin().lower()
                gateway_pass = AZ_LST[2]
                if my_transport is None or not my_transport.is_active():
                    if not remote_ssh_client:
                        gateway_Token = Azure_OTP()
                    else:  # inserted
                        gateway_Token = 'Not required'
                else:  # inserted
                    gateway_Token = 'Not required'
                if gateway_Token == None:
                    thread = threading.Thread(target=enabled_button, args=(self,))
                    thread.start()
                    thread.join()
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Unable to update Azure database, connection to Azure cloud refused.') + '\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                else:  # inserted
                    remote_file_path = '/home/inputdata_TIGO-TZ/Input/Ericsson/4G/RCA/Test_data/conversations.db'
                    SQL_TABLE = 'SELECT text FROM conversations'
                    SKPI_COMMAND = 'NO'
                    try:
                        SPLIT_txt = list(filter(None, command.split(':')))
                        SPLIT_txt = SPLIT_txt[0].upper().strip()
                        if SPLIT_txt[:2] in GSMCMD_lst:
                            SKPI_COMMAND = 'YES'
                    except:
                        pass
                    if SKPI_COMMAND == 'NO':
                        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait command updating in the cloud database !\n')
                        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                        start_index = '1.0'
                        while True:
                            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                            if not start_index:
                                break
                            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                            start_index = end_index
                        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                        threading.Thread(target=disable_button, args=(self,)).start()
                        COMMAND_LS_WRT = conn_ssh_sqlwrite(gateway_pass, gateway_Token, gateway_user, gateway_host, remote_host, remote_username, remote_password, remote_file_path, SQL_TABLE, keyword, command)
                        thread = threading.Thread(target=enabled_button, args=(self,))
                        thread.start()
                        thread.join()
                        SKIP_COMMANDDDD_NEW = 'YES'
                    else:  # inserted
                        threading.Thread(target=check_GSM_CMD, args=(self,)).start()
        except:
            pass
    if SKIP_COMMANDDDD_NEW == 'NO':
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : I\'ll remember that for next time; learning more commands will be valuable.\n')
        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Command updated..\n')
        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
        start_index = '1.0'
        while True:
            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
            if not start_index:
                break
            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
            start_index = end_index
        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
    try:
        KEYWORD_ADD.clear()
    except:
        return None

def RRSG_enter_command(self, Login_comm_lst, test_buddy):
    if 'login_error' in LOGINERROR_LST:
        thread = threading.Thread(target=enabled_button, args=(self,))
        thread.start()
        thread.join()
        return
    check_lst = []
    CIPRILST = []
    if 'stoppp' not in combined_cmd_window_chk:
        combined_cmd_window_chk.append('stoppp')

    def read_received_data_enter(remote_conn, chk_oss, iii, command, cmd_node, cmd_ex_sh):
        commnd_sh = []
        commnd_sh.append(cmd_ex_sh)

        def remove_color_codes(input_string):
            ansi_escape = re.compile('\\x1b[^m]*m')
            clean_string = ansi_escape.sub('', input_string)
            html_color = re.compile('<[^>]*>')
            clean_string = html_color.sub('', clean_string)
            return clean_string
        command_not_found = []
        read_received_data_enter.out = ''
        output_B = b''
        combined_cmd = ''
        while not (output_B.endswith(b'> ') or output_B.endswith(b'<') or output_B.endswith(b': ') or output_B.endswith(b'<') or output_B.endswith(b'< ') or output_B.endswith(b':') or output_B.endswith(b'$ ')):
            output_B += remote_conn.recv(9999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999)
            write_t = output_B.decode(encoding='utf-8')
            write_te_1 = remove_color_codes(write_t)
            write_te_1 = write_te_1.replace('', '')
            write_te_1 = write_te_1.replace('', '')
            write_te_1 = write_te_1.replace('\a', '')
            if cmd_ex_sh == 'sdir' and 'CPRI links' in write_te_1:
                if not CIPRILST:
                    CIPRILST.append(write_te_1)
            if 'no such command' in write_te_1 and 'no such command' not in command_not_found:
                command_not_found.append('no such command')
            if 'command not found' in write_te_1 and 'no such command' not in command_not_found:
                command_not_found.append('no such command')
            write_te_1 = '\n'.join([line.replace('coli>/rc/nrat/ue print -admitted', '') for line in write_te_1.split('\n')])
            write_te_1 = '\n'.join([line.replace('NRAT is dormant.', '') for line in write_te_1.split('\n')])
            if 'ue print -admitted'.lower() in command.lower():
                write_te_1 = '\n'.join([line.replace('coli>', '') for line in write_te_1.split('\n')])
            if 'coli>' in write_te_1 and 'no such command' not in command_not_found:
                command_not_found.append('no such command')
            self.update_textbox(f'{write_te_1}\n')
            try:
                self.reset_timer(None)
            except:
                pass
            check_textbox.clear()
            check_textbox.append('updating')
            if 'amos' in command:
                if 'AMOS error' in write_te_1:
                    checkamos.append('quit_amos')
                if 'java version' in write_te_1:
                    checkamos_java.append('Java_version')
            combined_cmd_window.append(f'{write_te_1}')
            last_line = output_B.decode(encoding='utf-8').split('\n')[(-1)]
            last_line = ' '.join(last_line.split()).strip()
            if check_lst and last_line == check_lst[0]:
                read_received_data_enter.brek = 'break'
            try:
                if last_line[(-1)].strip() == '>':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
            try:
                if last_line[(-1)].strip() == '<':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
            try:
                if last_line[(-1)].strip() == ':':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
            try:
                if last_line[(-1)].strip() == '$':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
        check_textbox.clear()
        check_textbox.append('done_ok')
        if not command_not_found:
            write_text1 = output_B.decode(encoding='utf-8')
            write_text = remove_color_codes(write_text1)
            write_text = write_text.replace('', '')
            write_text = write_text.replace('', '')
            write_text = write_text.replace('\a', '')
            Save_log_switch = self.swich_Sav_loss.get()
            if Save_log_switch == 1:
                Save_logfile = 'YES'
            else:  # inserted
                Save_logfile = 'NO'
            if Save_logfile == 'YES' and chk_oss!= 'LOGIN_CNM' and (chk_oss!= 'no') and ('amos' not in command) and ('lt all' not in command) and ('rbs' not in command) and (command!= 'q') and (command!= 'quit'):
                with open(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str(cmd_node) + '.Log', 'a+', encoding='utf-8') as f:
                    f.writelines(f'{write_text}\n\n')
                with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                    f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str(cmd_node) + '__' + str(command) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            if 'amos' not in command and 'lt all' not in command and ('rbs' not in command):
                combined_cmd += f'{write_text}\n'
            if 'amos' not in command and 'lt all' not in command and ('rbs' not in command) and (command!= 'q') and (command!= 'quit'):
                if KEYWORD_ADD:
                    threading.Thread(target=add_commad_to_db, args=(self,)).start()

                def remove_line_containing(text, substring):
                    lines = text.split('\n')
                    filtered_lines = [line for line in lines if substring not in line]
                    return '\n'.join(filtered_lines)
                combined_cmd = combined_cmd.replace('It is recommended to remove duplicate counter definitions from the PM scanners.', '')
                combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo)', 'Date & Time (UTC)   S Specific Problem                    MO (Cause/AdditionalInfo)')
                combined_cmd = combined_cmd.replace('Date & Time (Local) S Specific Problem                    MO (Cause/AdditionalInfo) Operator', '')
                combined_cmd = combined_cmd.replace('To print the duplicated counters, run command \"pmxs\".', '')
                combined_cmd = remove_line_containing(combined_cmd, 'ACKNOWLEDGED ALARMS')
                if 'pmr' in combined_cmd.lower():
                    combined_cmd = remove_line_containing(combined_cmd, 'pmr')
                    combined_cmd = remove_line_containing(combined_cmd, 'PMR')
                combined_cmd = combined_cmd.strip()
                combined_cmd = log_clening(combined_cmd, cmd_ex_sh, cmd_node)
                Final_log_lst = final_log_process(combined_cmd)
                try:
                    combined_cmd = ''
                except:
                    pass
                Final_log_lst = [item.strip() for item in Final_log_lst if item.strip()]
                try:
                    if Final_log_lst[0] == '\n':
                        Final_log_lst = []
                except:
                    pass
                if Final_log_lst:
                    i = 1
                    Final_log_lst = [item for item in Final_log_lst if not item.startswith('+')]
                    Final_log_lst = [item for item in Final_log_lst if not item.startswith('\n+')]
                    for section in Final_log_lst:
                        lines = section.strip().split('\n')
                        data = [line.split('\t') for line in lines]
                        if len(data) > 1:
                            crate_df = crate_txt_df(section, cmd_node, cmd_ex_sh)
                            try:
                                crate_df.replace('deletestring', '', inplace=True)
                            except:
                                pass
                            if crate_df is not None and (not crate_df.empty):
                                crate_df = crate_df[~crate_df.apply(lambda row: any(row.astype(str).str.contains('Total:')), axis=1)]
                                crate_df.columns = crate_df.columns.str.strip()
                                if not crate_df.empty:
                                    try:
                                        try:
                                            crate_df['VSWR (RL)'] = crate_df['VSWR (RL)'].str.strip()
                                        except:
                                            pass
                                        VSWR = [col for col in crate_df.columns if 'VSWR' in col]
                                        if VSWR:
                                            try:
                                                crate_df[['VSWR', 'RL']] = crate_df['VSWR_(RL)'].str.split('_', expand=True)
                                            except:
                                                crate_df[['VSWR', 'RL']] = crate_df['VSWR (RL)'].str.split(' ', expand=True)
                                            crate_df['RL'] = crate_df['RL'].str.strip('()')
                                            crate_df['VSWR'] = crate_df['VSWR'].astype(float)
                                            crate_df['RL'] = crate_df['RL'].astype(float)
                                            try:
                                                del crate_df['VSWR_(RL)']
                                            except:
                                                try:
                                                    del crate_df['VSWR (RL)']
                                                except:
                                                    pass
                                    except:
                                        pass
                                    try:
                                        crate_df = crate_df.applymap(lambda x: x.replace('_', ' '))
                                    except:
                                        pass
                                    try:
                                        crate_df['Node Id'] = cmd_node
                                    except:
                                        pass
                                    try:
                                        crate_df.insert(0, 'Node Id', crate_df.pop('Node Id'))
                                    except:
                                        pass
                                    excel_df = crate_df.copy()
                                    counter = [col for col in excel_df.columns if 'counter' in col.lower()]
                                    if counter:
                                        excel_df = melted_conter(excel_df)
                                    try:
                                        del crate_df['Command']
                                    except:
                                        pass
                                    try:
                                        del crate_df['Baseline Check']
                                    except:
                                        pass
                                    if crate_txt_df.df_fram!= 'error':
                                        tabulate.PRESERVE_WHITESPACE = True
                                        data_list = crate_df.to_dict(orient='records')
                                        table = tabulate(data_list, headers='keys', tablefmt='presto')
                                        table = cleeend_text(table)
                                        if 'done' not in processdone:
                                            processdone.append('done')
                                        self.update_textbox_big_window_enter_cmd(table, iii)
                                        if not excel_df.empty:
                                            Save_log_switch = self.swich_Sav_loss.get()
                                            if Save_log_switch == 1:
                                                Save_logfile = 'YES'
                                            else:  # inserted
                                                Save_logfile = 'NO'
                                            if Save_logfile == 'YES' and chk_oss!= 'LOGIN_CNM' and (chk_oss!= 'no'):
                                                try:
                                                    del excel_df['Command']
                                                except:
                                                    pass
                                                try:
                                                    os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\Processed_logs')
                                                except:
                                                    pass
                                                with open(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\Processed_logs\\' + str(cmd_node) + '_trim.Log', 'a+', encoding='utf-8') as f:
                                                    f.writelines(f'{table}\n\n')
                                                try:
                                                    try:
                                                        if len(Final_log_lst) > 1:
                                                            cmd_ex_sh = commnd_sh[0] + str(i)
                                                        else:  # inserted
                                                            pass
                                                    except:
                                                        pass
                                                    try:
                                                        if commnd_sh[0] == 'sdir':
                                                            cmd_ex_sh = 'sdir'
                                                    except:
                                                        pass
                                                    try:
                                                        if commnd_sh[0] == 'cabex':
                                                            cmd_ex_sh = 'cabex'
                                                    except:
                                                        pass
                                                    try:
                                                        if commnd_sh[0] == 'sdi':
                                                            cmd_ex_sh = 'sdi'
                                                    except:
                                                        pass
                                                    try:
                                                        excel_df = excel_df.applymap(lambda x: str(x).strip())
                                                    except:
                                                        pass
                                                    write_excel_new(excel_df, 'Processed Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', cmd_ex_sh, OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\')
                                                    try:
                                                        del excel_df
                                                    except:
                                                        pass
                                                except:
                                                    pass
                                    i = i + 1
                    try:
                        if command.strip().lower() == 'sdir':
                            CIPRI_DFF = CIPRIDFFF(CIPRILST[0])
                            CIPRILST.clear()
                            if CIPRI_DFF is not None and (not CIPRI_DFF.empty):
                                try:
                                    CIPRI_DFF['Node Id'] = cmd_node
                                except:
                                    pass
                                try:
                                    CIPRI_DFF.insert(0, 'Node Id', CIPRI_DFF.pop('Node Id'))
                                except:
                                    pass
                                if not CIPRI_DFF.empty:
                                    try:
                                        CIPRI_DFF = CIPRI_DFF.applymap(lambda x: str(x).strip())
                                    except:
                                        pass
                                    try:
                                        def conditione(row):
                                            if int(row['CPRI Link NOK']) > 0:
                                                return 'Fail'
                                            if int(row['CPRI Link OKW']) > 0:
                                                return 'Fail'
                                            return 'Pass'
                                        CIPRI_DFF['Baseline Check'] = CIPRI_DFF.apply(conditione, axis=1)
                                    except:
                                        pass
                                    write_excel_new(CIPRI_DFF, 'Processed Output ' + datetime.now().strftime('%d%m%Y') + '.xlsx', 'CPRI_link', OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\')
                                    tabulate.PRESERVE_WHITESPACE = True
                                    data_list = CIPRI_DFF.to_dict(orient='records')
                                    table = tabulate(data_list, headers='keys', tablefmt='presto')
                                    table = cleeend_text(table)
                                    CIPRI_PRINT(self, table)
                                try:
                                    del CIPRI_DFF
                                except:
                                    pass
                    except:
                        pass
                    try:
                        del Final_log_lst
                    except:
                        pass
                    try:
                        del command_not_found
                    except:
                        pass
                    combined_cmd = ''
                    thread = threading.Thread(target=enabled_button, args=(self,))
                    thread.start()
                    thread.join()

    def send_command_enter(remote_conn, command, chk_oss, iii, cmd_node, cmd_ex_sh):
        LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
        if LIVECMD_TEXT.strip()!= str(cmd_node) + '>' and 'your choice' not in LIVECMD_TEXT.strip().lower():
            if command.lower() == 'q':
                ltall_nod2.clear()
            if command.lower() == 'q':
                amos_node.clear()
            if command.lower() == 'quit':
                amos_nod2.clear()
            if command.lower() == 'quit':
                ltall_nod2.clear()
            if command.lower() == 'quit':
                amos_node.clear()
            if command.lower() == 'quit':
                amos_nod2.clear()
        try:
            cmd_node = amos_node[(-1)]
        except:
            pass
        if 'amos' in command.lower():
            if command.lower() not in amos_nod2:
                remote_conn.send(command + '\n')
                if 'amos' in command:
                    amos_node.append(command.split(' ')[(-1)].strip().upper())
                    amos_nod2.append(command.lower())
                return read_received_data_enter(remote_conn, chk_oss, iii, command, str(amos_node[0]), cmd_ex_sh)
        else:  # inserted
            if test_buddy == 'chatbuddy':
                if 'lt all' in command.lower():
                    if command.lower() not in ltall_nod2:
                        remote_conn.send(command + '\n')
                        if 'lt all' in command.lower():
                            ltall_nod2.append(command.lower())
                        return read_received_data_enter(remote_conn, chk_oss, iii, command, cmd_node, cmd_ex_sh)
                else:  # inserted
                    remote_conn.send(command + '\n')
                    return read_received_data_enter(remote_conn, chk_oss, iii, command, cmd_node, cmd_ex_sh)
            else:  # inserted
                remote_conn.send(command + '\n')
                return read_received_data_enter(remote_conn, chk_oss, iii, command, cmd_node, cmd_ex_sh)

    def main_enter(logcommand_lst):
        iii = 1
        paaaawe = None
        cmd_ex_sh = logcommand_lst[0].strip()
        if 'FAIL' not in establish_ssh_conn_lst:
            for cmd_node in logcommand_lst:
                if 'none' in cmd_node.lower():
                    continue
                try:
                    llll = RRSG_auto_login.remote_conn
                except:
                    try:
                        con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                        Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                        con.close()
                        del Rdf['indexxx']
                        ppPID_LST = Rdf['value'].tolist()
                        ppPID_LST2 = []
                        for ppp in ppPID_LST:
                            ppPID_LST2.append(decrypt(ppp, 3))
                        del Rdf['value']
                        Rdf['value'] = ppPID_LST2
                        PID_LST = Rdf['value'].tolist()
                        remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                        RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                    except:
                        RSG_LST_N = []
                    if not RSG_LST_N:
                        showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 5000)
                    else:  # inserted
                        paaaawe = OTPPP_tets()
                if paaaawe == None and 'cancle' not in processdone:
                    processdone.append('cancle')
                try:
                    if 'eoclosed' in WORKORDER_CLOSE:
                        self.textbox_LIVECMD.delete(1.0, 'end')
                        paaaawe = '12345'
                    if VPN_select_METHOD_F[0]!= 'VPN-CAS':
                        if is_non_numeric_string(paaaawe):
                            showMessage_qt('Wrong OTP password please try again.', 10000)
                        else:  # inserted
                            if paaaawe:
                                self.Auto_login_sc(paaaawe, RSG_LST_N)
                    else:  # inserted
                        if paaaawe:
                            self.Auto_login_sc(paaaawe, RSG_LST_N)
                except:
                    pass
                try:
                    chk_rssg = RRSG_auto_login.remote_conn
                    chk_rssg_chk = 'OK'
                except:
                    chk_rssg_chk = 'NOK'
                if chk_rssg_chk == 'NOK':
                    clear_all_lists()
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Something wrong in RSG/VPN connection please try again after some time.') + '\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                else:  # inserted
                    WORKORDER_CLOSE.clear()
                    if not CHECK_ISF_SEQUENCE:
                        sig = os.getlogin().lower()
                        priority = 'High'
                        ISF_API_Based = 'SharePoint'
                        KEY = '5b97555705ce40128131d58c8d3596f9'
                        try:
                            con = sqlite3.connect('./res/ISFdetails.db')
                            Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                            con.close()
                            del Rdf['indexxx']
                            ppPID_LST = Rdf['value'].tolist()
                            ppPID_LST2 = []
                            for ppp in ppPID_LST:
                                ppPID_LST2.append(decrypt(ppp, 3))
                            del Rdf['value']
                            Rdf['value'] = ppPID_LST2
                            PID_LST = Rdf['value'].tolist()
                            remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                            ISF_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                            try:
                                del Rdf
                            except:
                                pass
                        except:
                            ISF_LST = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                        chk_line = []
                        for llll in ISF_LST:
                            if 'line' in llll.lower():
                                chk_line.append('stop_process')
                                break
                            if 'start_process' not in chk_line:
                                chk_line.append('start_process')
                        if 'start_process' in chk_line:
                            projectID = ISF_LST[0]
                            type1 = ISF_LST[1]
                            nodname = ISF_LST[2]
                            wOName = ISF_LST[3]
                            executionPlanName = ISF_LST[4]
                            lastModifiedBy = ISF_LST[5]
                            ISF_STEP_START_stepID = ISF_LST[6]
                            ISF_STEP_START_Task_Id = ISF_LST[7]
                            if not ISF_STEP_START_Task_Id_lst:
                                ISF_STEP_START_Task_Id_lst.append(ISF_STEP_START_Task_Id)
                            ISF_WO_CREATE_woId, STEP_START = ISF_API(sig, type1, nodname, projectID, priority, wOName, lastModifiedBy, ISF_API_Based, executionPlanName, KEY, ISF_STEP_START_stepID, ISF_STEP_START_Task_Id)
                            WO_LSTT = [ISF_WO_CREATE_woId, STEP_START]
                            ISF_STTUS_MSG = 'sucess'
                            if WO_LSTT[0] == 'Some issue in isf workflow':
                                ISF_STTUS_MSG = 'ISF details not correct. please check and try again.'
                            if ISF_STTUS_MSG == 'sucess' and WO_LSTT[1]!= 'sucess':
                                ISF_STTUS_MSG = 'TaskID and StepID not correct. please provide the valid taskID for the StepID.'
                    else:  # inserted
                        ISF_STTUS_MSG = 'sucess'
                    if ISF_STTUS_MSG == 'sucess' and (not CHECK_ISF_SEQUENCE):
                        if 'work_order_raised' not in CHECK_ISF_SEQUENCE:
                            CHECK_ISF_SEQUENCE.append('work_order_raised')

                        def sucss_msg(self):
                            self.textbox2.delete(1.0, 'end')
                            self.after(0, self.textbox2.insert, 'end', 'ISF Work Order created.\n\n')
                            self.after(0, self.textbox2.insert, 'end', 'Signum    : ' + str(sig) + '\n')
                            self.after(0, self.textbox2.insert, 'end', 'Nodtype   : ' + str(type1) + '\n')
                            self.after(0, self.textbox2.insert, 'end', 'Nodname   : ' + str(nodname) + '\n')
                            self.after(0, self.textbox2.insert, 'end', 'ProjectID : ' + str(projectID) + '\n')
                            self.after(0, self.textbox2.insert, 'end', 'WOID      : ' + str(ISF_WO_CREATE_woId) + '\n')
                            try:
                                self.after(0, self.textbox2.insert, 'end', 'Time      : ' + str(ISF_API.timestr) + '\n\n')
                            except:
                                pass
                            try:
                                self.after(0, self.textbox2.insert, 'end', str(ISF_API.urlcreateWO) + '\n\n')
                            except:
                                return None
                        with open('./res/CHATLOGS/' + str(sig) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('chatbuddy') + '__' + 'Signum__' + str(sig) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('chatbuddy') + '__' + 'Nodtype__' + str(type1) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('chatbuddy') + '__' + 'Nodname__' + str(nodname) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('chatbuddy') + '__' + 'ProjectID__' + str(projectID) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('chatbuddy') + '__' + 'WOID__' + str(ISF_WO_CREATE_woId) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('chatbuddy') + '__' + 'Time__' + str(ISF_API.timestr) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                        threading.Thread(target=sucss_msg, args=(self,)).start()
                    if 'work_order_raised' in CHECK_ISF_SEQUENCE:
                        chk_oss = 'NODE_CNM'
                        try:
                            send_command_enter(RRSG_auto_login.remote_conn, cmd_node, chk_oss, iii, cmd_node, cmd_ex_sh)
                        except:
                            continue
                    else:  # inserted
                        self.textbox2.delete(1.0, 'end')
                        self.after(0, self.textbox2.insert, 'end', ISF_STTUS_MSG + '\n\n')
    main_enter(Login_comm_lst)
    combined_cmd_window_chk.clear()
    threading.Thread(target=check_empty_textbox_live, args=(self,)).start()

def RRSG_auto_login(self, login_command_lst, router, username, password):
    if VPN_select_METHOD_F[0] == 'VPN-CAS':
        username = str(VPN_RSSG_CRREDD[0])
        password = str(VPN_RSSG_CRREDD[1])
        login_command_lst.clear()
    if 'ok' in chek_first_time_rsg:
        clear_all_lists()
    check_lst = []

    def read_received_data(remote_conn, chk_oss, iii, command, cmd_node):
        read_received_data.out = ''

        def remove_color_codes(input_string):
            ansi_escape = re.compile('\\x1b[^m]*m')
            clean_string = ansi_escape.sub('', input_string)
            html_color = re.compile('<[^>]*>')
            clean_string = html_color.sub('', clean_string)
            return clean_string
        read_received_data.out = ''
        output_B = b''
        while not (output_B.endswith(b'> ') or output_B.endswith(b'<') or output_B.endswith(b': ') or output_B.endswith(b'<') or output_B.endswith(b'< ') or output_B.endswith(b':') or output_B.endswith(b'$ ')):
            output_B += remote_conn.recv(9999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999999)
            write_t = output_B.decode(encoding='utf-8')
            write_te_1 = remove_color_codes(write_t)
            write_te_1 = write_te_1.replace('', '')
            write_te_1 = write_te_1.replace('', '')
            write_te_1 = write_te_1.replace('\a', '')
            self.update_textbox(f'{write_te_1}\n')
            time.sleep(0.5)
            error_lst = ['Permission denied', 'Operation timed out']
            for eerr in error_lst:
                if eerr.lower() in write_te_1.lower():
                    LOGINERROR_LST.clear()
                    LOGINERROR_LST.append('login_error')
            try:
                self.reset_timer(None)
            except:
                pass
            last_line = output_B.decode(encoding='utf-8').split('\n')[(-1)]
            last_line = ' '.join(last_line.split()).strip()
            if check_lst and last_line == check_lst[0]:
                read_received_data.brek = 'break'
            try:
                if last_line[(-1)].strip() == '>':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
            try:
                if last_line[(-1)].strip() == '<':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
            try:
                if last_line[(-1)].strip() == ':':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
            try:
                if last_line[(-1)].strip() == '$':
                    check_lst.clear()
                    check_lst.append(last_line)
            except IndexError:
                pass
        last_character = output_B.decode(encoding='utf-8').split('\n')[(-1)].strip()[(-1)]
        if last_character!= '$':
            LOGINERROR_LST.clear()

    def send_command(remote_conn, command, chk_oss, iii, cmd_node):
        remote_conn.send(command + '\n')
        return read_received_data(remote_conn, chk_oss, iii, command, cmd_node)

    def establish_ssh_connection(hostname, username, password):
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(hostname, username=username, password=password)
            return ssh
        except paramiko.AuthenticationException:
            self.after(0, self.textbox2.insert, 'end', 'RSG Authentication failed. Please check your credentials.\n\n')
            establish_ssh_conn_lst.append('FAIL')
            threading.Thread(target=autenti_fail, args=(self,)).start()
            clear_all_lists()
        except paramiko.SSHException as e:
            self.after(0, self.textbox2.insert, 'end', f'Unable to establish SSH connection: {e}' + '\n\n')

    def main_auto(logcommand_lst, router, username, password):
        global ssh  # inserted
        ssh = establish_ssh_connection(router, username, password)
        if ssh is None:
            return
        self.textbox_LIVECMD.delete(1.0, 'end')
        self.textbox_LIVECMD.insert('0.0', f"{'Please wait connecting RSG connection..'}\n\n")
        chk_oss = 'no'
        iii = 1
        cmd_node = ''
        RRSG_auto_login.remote_conn = ssh.invoke_shell()
        read_received_data(RRSG_auto_login.remote_conn, chk_oss, iii, '', cmd_node)
        read_received_data.brek = ''
        read_received_data.out = ''
        for cmd in logcommand_lst:
            chk_oss = 'LOGIN_CNM'
            if read_received_data.brek == 'break':
                self.textbox2.insert('0.0', f"{'wrong RSG credential please check and try again..'}\n")
                return
            send_command(RRSG_auto_login.remote_conn, cmd, chk_oss, iii, cmd_node)
    main_auto(login_command_lst, router, username, password)

def GSMGMRRTool(INPATH, OUTPATH, MODEEE):
    mylist = ['green']
    new_theme = random.choice(mylist)
    customtkinter.set_appearance_mode(MODEEE)
    customtkinter.set_default_color_theme(new_theme)

    class App(customtkinter.CTk):
        def __init__(self):
            super().__init__()
            threading.Thread(target=check_download_completion, args=(self,)).start()
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
            self.title('DEVELOPED BY : MOHNISH SAINI  |  BOT ID-216477')
            if MODEEE == 'Dark':
                butoncolour = '#7b6d48'
            if MODEEE == 'light':
                butoncolour = '#A8A970'
            if MODEEE == 'Dark':
                border_colorii = '#919492'
            if MODEEE == 'light':
                border_colorii = 'white'
            if MODEEE == 'Dark':
                bigbutoncolour = '#76a8af'
            if MODEEE == 'light':
                bigbutoncolour = '#a2c4c9'
            if MODEEE == 'Dark':
                bigtextcol = '#FAF9F6'
            if MODEEE == 'light':
                bigtextcol = '#404040'
            if MODEEE == 'Dark':
                bigbordercol = '#919492'
            if MODEEE == 'light':
                bigbordercol = 'white'
            scaleFactor = ctypes.windll.shcore.GetScaleFactorForDevice(0) / 100
            if scaleFactor == 1.5:
                self.geometry(f'{1232}x{560}')
                self.tk.call('tk', 'scaling', 2.0)
                self.resizable(False, False)
                windowWidth = self.winfo_reqwidth()
                windowHeight = self.winfo_reqheight()
                positionRight = int(self.winfo_screenwidth() / 20 - windowWidth / 5.5)
                positionDown = int(self.winfo_screenheight() / 3.5 - windowHeight / 1.63)
                self.geometry('+{}+{}'.format(positionRight, positionDown))
            else:  # inserted
                if scaleFactor == 1.25:
                    self.geometry(f'{1476}x{735}')
                    self.tk.call('tk', 'scaling', 2.0)
                    self.resizable(False, False)
                    windowWidth = self.winfo_reqwidth()
                    windowHeight = self.winfo_reqheight()
                    positionRight = int(self.winfo_screenwidth() / 20 - windowWidth / 5.2)
                    positionDown = int(self.winfo_screenheight() / 4.4 - windowHeight / 1.5)
                    self.geometry('+{}+{}'.format(positionRight, positionDown))
                    new_scaling_float = int(str(110).replace('%', '')) / 100
                    customtkinter.set_widget_scaling(new_scaling_float)
                else:  # inserted
                    self.geometry(f'{1476}x{735}')
                    self.tk.call('tk', 'scaling', 2.0)
                    self.resizable(False, False)
                    windowWidth = self.winfo_reqwidth()
                    windowHeight = self.winfo_reqheight()
                    positionRight = int(self.winfo_screenwidth() / 27 - windowWidth / 5.1)
                    positionDown = int(self.winfo_screenheight() / 5 - windowHeight / 1.5)
                    self.geometry('+{}+{}'.format(positionRight, positionDown))
                    new_scaling_float = int(str(110).replace('%', '')) / 100
                    customtkinter.set_widget_scaling(new_scaling_float)
            try:
                self.iconbitmap('./res/logo.ico')
            except:
                pass
            self.sidebar_frame = customtkinter.CTkFrame(self, width=145, corner_radius=0)
            self.sidebar_frame.grid(row=0, column=0, rowspan=16, sticky='nsew')
            self.sidebar_frame.grid_rowconfigure(16, weight=0)
            self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text='RAN OSS BUDDY', font=customtkinter.CTkFont('Roboto', size=15, weight='bold'))
            if scaleFactor == 1.5:
                self.logo_label.grid(row=0, column=0, padx=2, pady=(2, 0))
            else:  # inserted
                self.logo_label.grid(row=0, column=0, padx=2, pady=(8, 0))
            self.switch_value1 = customtkinter.IntVar()
            self.switch_value2 = customtkinter.IntVar()
            self.switch_value3 = customtkinter.IntVar()
            self.switch_aut_scri = customtkinter.IntVar()
            self.swich_Sav_loss = customtkinter.IntVar()
            self.swich_Sec_rbs = customtkinter.IntVar()
            self.activity_BU = customtkinter.IntVar()
            self.radio_var = customtkinter.IntVar(value=3)
            self.switch_value3 = customtkinter.IntVar()
            self.sidebar_node = customtkinter.IntVar()
            self.sidebar_command = customtkinter.IntVar()
            self.sidebar_Project = customtkinter.CTkOptionMenu(self.sidebar_frame, state='disabled', values=PROJECT_LST, font=customtkinter.CTkFont(size=10, weight='bold'), width=116, command=self.project_select_button)
            if scaleFactor == 1.5:
                self.sidebar_Project.grid(row=2, column=0, padx=12, pady=(3, 0), sticky='w')
            else:  # inserted
                self.sidebar_Project.grid(row=2, column=0, padx=14, pady=(12, 0), sticky='w')
            self.sidebar_Project.set('Select Project')
            self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame, state='normal', text='Auto Login', fg_color='#CD5C5C', font=customtkinter.CTkFont(size=10, weight='bold'), width=116, command=self.Auto_login_button)
            if scaleFactor == 1.5:
                self.sidebar_button_3.grid(row=3, column=0, padx=12, pady=(6, 0), sticky='w')
            else:  # inserted
                self.sidebar_button_3.grid(row=3, column=0, padx=14, pady=(14, 0), sticky='w')
            self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, state='normal', text='Tool Setting', font=customtkinter.CTkFont(size=10, weight='bold'), width=116, command=self.RSG_Login_button_event)
            if scaleFactor == 1.5:
                self.sidebar_button_1.grid(row=4, column=0, padx=12, pady=(6, 0), sticky='w')
            else:  # inserted
                self.sidebar_button_1.grid(row=4, column=0, padx=14, pady=(9, 0), sticky='w')
            self.switch_VPPN_CASS = customtkinter.CTkOptionMenu(self.sidebar_frame, state='normal', values=['RSG', 'VPN-CAS'], font=customtkinter.CTkFont(size=10, weight='bold'), width=116, command=self.VPN_select_button)
            if scaleFactor == 1.5:
                self.switch_VPPN_CASS.grid(row=5, column=0, padx=12, pady=(6, 0), sticky='w')
            else:  # inserted
                self.switch_VPPN_CASS.grid(row=5, column=0, padx=12, pady=(9, 0), sticky='w')
            if VPN_select_METHOD_F:
                VPN_select_METHOD_val = VPN_select_METHOD_F[0]
            else:  # inserted
                VPN_select_METHOD_val = 'Conn Method'
            self.switch_VPPN_CASS.set(VPN_select_METHOD_val)
            self.switch3 = customtkinter.CTkSwitch(self.sidebar_frame, state='normal', text='Save log file', font=customtkinter.CTkFont(size=12), command=self.Save_logsss, variable=self.swich_Sav_loss)
            if scaleFactor == 1.5:
                self.switch3.grid(row=6, column=0, padx=12, pady=(5, 0), sticky='w')
            else:  # inserted
                self.switch3.grid(row=6, column=0, padx=14, pady=(20, 0), sticky='w')
            self.switch3.select()
            if scaleFactor == 1.5:
                try:
                    gif_path = './res/RPA_4.png'
                    IMAGE_WIDTH = 80
                    IMAGE_HEIGHT = 80
                    your_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(gif_path)), size=(IMAGE_WIDTH, IMAGE_HEIGHT))
                    self.gif_label = customtkinter.CTkLabel(master=self.sidebar_frame, image=your_image, text='')
                    self.gif_label.grid(row=7, column=0, padx=25, pady=(3, 0), sticky='w')
                except:
                    pass
            else:  # inserted
                try:
                    gif_path = './res/RPA_4.png'
                    IMAGE_WIDTH = 90
                    IMAGE_HEIGHT = 90
                    your_image = customtkinter.CTkImage(light_image=Image.open(os.path.join(gif_path)), size=(IMAGE_WIDTH, IMAGE_HEIGHT))
                    self.gif_label = customtkinter.CTkLabel(master=self.sidebar_frame, image=your_image, text='')
                    self.gif_label.grid(row=7, column=0, padx=25, pady=(10, 0), sticky='w')
                except:
                    pass
            self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text='-----------------------', font=customtkinter.CTkFont(size=13, weight='bold'))
            if scaleFactor == 1.5:
                self.logo_label.grid(row=8, column=0, padx=10, pady=(0, 0), sticky='w')
            else:  # inserted
                self.logo_label.grid(row=8, column=0, padx=15, pady=(5, 0), sticky='w')
            self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text='MOBATCH & KGET', state='disabled', font=customtkinter.CTkFont('Consolas', size=15, weight='bold'))
            if scaleFactor == 1.5:
                self.logo_label.grid(row=9, column=0, padx=15, pady=(2, 0), sticky='nw')
            else:  # inserted
                self.logo_label.grid(row=9, column=0, padx=15, pady=(3, 0), sticky='nw')
            self.switch_mo = customtkinter.CTkSwitch(self.sidebar_frame, state='normal', text='Required', font=customtkinter.CTkFont(size=12), command=self.Save_logsss, variable=self.activity_BU)
            if scaleFactor == 1.5:
                self.switch_mo.grid(row=10, column=0, padx=16, pady=(3, 0), sticky='w')
            else:  # inserted
                self.switch_mo.grid(row=10, column=0, padx=16, pady=(8, 0), sticky='w')
            self.radio_button_1 = customtkinter.CTkRadioButton(self.sidebar_frame, font=customtkinter.CTkFont(size=10), text=' MO BATCH', variable=self.radio_var, value=0)
            if scaleFactor == 1.5:
                self.radio_button_1.grid(row=11, column=0, padx=16, pady=(3, 0), sticky='w')
            else:  # inserted
                self.radio_button_1.grid(row=11, column=0, padx=16, pady=(10, 0), sticky='w')
            self.radio_button_2 = customtkinter.CTkRadioButton(self.sidebar_frame, text=' KGET', font=customtkinter.CTkFont(size=10), variable=self.radio_var, value=1)
            if scaleFactor == 1.5:
                self.radio_button_2.grid(row=12, column=0, padx=16, pady=(3, 0), sticky='w')
            else:  # inserted
                self.radio_button_2.grid(row=12, column=0, padx=16, pady=(10, 0), sticky='w')
            self.tecno_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=['GSM', 'WCDMA', 'LTE', 'NR'], width=116, font=customtkinter.CTkFont(size=10), command=self.change_appearance_mode_event)
            if scaleFactor == 1.5:
                self.tecno_mode_optionemenu.grid(row=13, column=0, padx=12, pady=(6, 0), sticky='w')
            else:  # inserted
                self.tecno_mode_optionemenu.grid(row=13, column=0, padx=12, pady=(15, 0), sticky='w')
            self.tecno_mode_optionemenu.set('LTE')
            self.sidebar_node = customtkinter.CTkButton(self.sidebar_frame, state='disabled', text='NODE List', width=116, font=customtkinter.CTkFont(size=10, weight='bold'), command=self.NODE_List)
            self.sidebar_node.grid(row=14, column=0, padx=12, pady=(8, 0), sticky='w')
            self.sidebar_command = customtkinter.CTkButton(self.sidebar_frame, state='disabled', text='Command List', width=116, font=customtkinter.CTkFont(size=10, weight='bold'), command=self.Command_List)
            self.sidebar_command.grid(row=15, column=0, padx=12, pady=(8, 0), sticky='w')
            self.main_button_1 = customtkinter.CTkButton(self.sidebar_frame, state='disabled', fg_color='#CD5C5C', border_width=1, text_color='white', border_color='white', text='RUN-MOBATCH-KGET', width=116, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_PROGRAMM)
            if scaleFactor == 1.5:
                self.main_button_1.grid(row=16, column=0, padx=12, pady=(5, 0), sticky='nw')
            else:  # inserted
                self.main_button_1.grid(row=16, column=0, padx=12, pady=(18, 0), sticky='nw')
            if scaleFactor == 1.5:
                self.sidebar_frame2 = customtkinter.CTkFrame(self, width=120, corner_radius=0)
                self.sidebar_frame2.grid(row=0, column=3, rowspan=3, padx=(5, 0), pady=(0, 0), sticky='nw')
                self.logo_label = customtkinter.CTkLabel(self.sidebar_frame2, text='OSS TERMINAL', font=customtkinter.CTkFont('Roboto', size=15, weight='bold'))
                self.logo_label.grid(row=0, column=0, padx=(178, 0), pady=(1, 0), sticky='nw')
                if MODEEE == 'Dark':
                    self.textbox_LIVECMD = customtkinter.CTkTextbox(self.sidebar_frame2, wrap='none', width=300, height=420, corner_radius=5, border_width=1, border_color='black', font=customtkinter.CTkFont('Calibri', size=10))
                    self.textbox_LIVECMD.grid(row=1, column=0, padx=(3, 0), pady=(5, 0), sticky='nw')
                    self.textbox_LIVECMD.insert('0.10', 'OSS CMD TERMINAL\n\n')
                    self.textbox8 = customtkinter.CTkTextbox(self.sidebar_frame2, width=300, height=70, text_color='#000000', fg_color='#FAF9F6', border_width=1, corner_radius=5, font=customtkinter.CTkFont(family='Segoe UI', size=13))
                    self.textbox8.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                else:  # inserted
                    self.textbox_LIVECMD = customtkinter.CTkTextbox(self.sidebar_frame2, wrap='none', width=300, height=420, fg_color='#e6f7f6', corner_radius=5, border_width=1, border_color='white', font=customtkinter.CTkFont('Calibri', size=10))
                    self.textbox_LIVECMD.grid(row=1, column=0, padx=(3, 0), pady=(5, 0), sticky='nw')
                    self.textbox_LIVECMD.insert('0.10', 'OSS CMD TERMINAL\n\n')
                    self.textbox8 = customtkinter.CTkTextbox(self.sidebar_frame2, width=300, height=70, fg_color='#f2f5f3', corner_radius=5, font=customtkinter.CTkFont(family='Segoe UI', size=13))
                    self.textbox8.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                self.textbox8.insert('0.10', 'Please wait loading data...')
                self.textbox8.bind('<Button-1>', self.clear_text_crt)
                self.textbox8.bind('<Return>', self.START_enter)
            else:  # inserted
                self.sidebar_frame2 = customtkinter.CTkFrame(self, width=120, corner_radius=0)
                self.sidebar_frame2.grid(row=0, column=3, rowspan=3, padx=(5, 0), pady=(0, 0), sticky='nw')
                self.logo_label = customtkinter.CTkLabel(self.sidebar_frame2, text='  OSS TERMINAL', font=customtkinter.CTkFont('Roboto', size=14, weight='bold'))
                self.logo_label.grid(row=0, column=0, padx=(178, 0), pady=(1, 0), sticky='nw')
                if MODEEE == 'Dark':
                    self.textbox_LIVECMD = customtkinter.CTkTextbox(self.sidebar_frame2, wrap='none', width=300, height=510, corner_radius=5, border_width=1, border_color='black', font=customtkinter.CTkFont('Calibri', size=10))
                    self.textbox_LIVECMD.grid(row=1, column=0, padx=(3, 0), pady=(5, 0), sticky='nw')
                    self.textbox_LIVECMD.insert('0.10', 'OSS CMD TERMINAL\n\n')
                    self.textbox8 = customtkinter.CTkTextbox(self.sidebar_frame2, width=300, height=90, text_color='#000000', fg_color='#FAF9F6', border_width=1, corner_radius=5, font=customtkinter.CTkFont(family='Segoe UI', size=13))
                    self.textbox8.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                else:  # inserted
                    self.textbox_LIVECMD = customtkinter.CTkTextbox(self.sidebar_frame2, wrap='none', width=300, height=510, fg_color='#e6f7f6', corner_radius=5, border_width=1, border_color='white', font=customtkinter.CTkFont('Calibri', size=10))
                    self.textbox_LIVECMD.grid(row=1, column=0, padx=(3, 0), pady=(5, 0), sticky='nw')
                    self.textbox_LIVECMD.insert('0.10', 'OSS CMD TERMINAL\n\n')
                    self.textbox8 = customtkinter.CTkTextbox(self.sidebar_frame2, width=300, height=90, fg_color='#f2f5f3', corner_radius=5, font=customtkinter.CTkFont(family='Segoe UI', size=13))
                    self.textbox8.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                self.textbox8.insert('0.10', 'Please wait loading data...')
                self.textbox8.bind('<Button-1>', self.clear_text_crt)
                self.textbox8.bind('<Return>', self.START_enter)
            if MODEEE == 'Dark':
                self.main_button_shot = customtkinter.CTkButton(master=self, state='normal', fg_color=bigbutoncolour, border_width=1, text_color='#3b3b3b', border_color=bigbordercol, text='CMD-SHORTCUT', width=88, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_command)
                self.main_button_shot.grid(row=3, column=3, padx=(9, 0), pady=(0, 0), sticky='nw')
            else:  # inserted
                self.main_button_shot = customtkinter.CTkButton(master=self, state='normal', fg_color=bigbutoncolour, border_width=1, text_color=bigtextcol, border_color=bigbordercol, text='CMD-SHORTCUT', width=88, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_command)
                self.main_button_shot.grid(row=3, column=3, padx=(9, 0), pady=(0, 0), sticky='nw')
            self.main_button_amos = customtkinter.CTkButton(master=self, state='normal', fg_color=butoncolour, border_width=1, text_color='white', border_color=border_colorii, text='amos', width=50, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_amos)
            if scaleFactor == 1.5:
                self.main_button_amos.grid(row=3, column=3, padx=(206, 0), pady=(0, 0), sticky='nw')
            else:  # inserted
                self.main_button_amos.grid(row=3, column=3, padx=(204, 0), pady=(0, 0), sticky='nw')
            self.main_button_quit = customtkinter.CTkButton(master=self, state='normal', fg_color=butoncolour, border_width=1, text_color='white', border_color=border_colorii, text='Quit', width=50, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_QUIT)
            if scaleFactor == 1.5:
                self.main_button_quit.grid(row=3, column=3, padx=(257, 0), pady=(0, 0), sticky='nw')
            else:  # inserted
                self.main_button_quit.grid(row=3, column=3, padx=(257, 0), pady=(0, 0), sticky='nw')
            if scaleFactor == 1.5:
                self.frame3 = customtkinter.CTkFrame(self, width=770, height=300, corner_radius=0)
                self.frame3.grid(row=0, column=1, sticky='nw', padx=(5, 0), pady=(0, 0))
                if MODEEE == 'Dark':
                    self.textbox2 = customtkinter.CTkTextbox(self.frame3, wrap='none', width=770, height=300, corner_radius=0, font=customtkinter.CTkFont('Consolas', size=11), border_width=0, border_color='#808080')
                    self.textbox2.grid(row=0, column=0, padx=(5, 0), pady=(0, 0), sticky='nw')
                    self.textbox2.pack(expand=True, side='right')
                else:  # inserted
                    self.textbox2 = customtkinter.CTkTextbox(self.frame3, wrap='none', width=770, height=300, corner_radius=0, font=customtkinter.CTkFont('Consolas', size=11), border_width=1, border_color='#D9D9D9')
                    self.textbox2.grid(row=0, column=0, padx=(5, 0), pady=(0, 0), sticky='nw')
                    self.textbox2.pack(expand=True, side='right')
            else:  # inserted
                self.frame3 = customtkinter.CTkFrame(self, width=872, height=375, corner_radius=0)
                self.frame3.grid(row=0, column=1, sticky='nw', padx=(5, 0), pady=(0, 0))
                if MODEEE == 'Dark':
                    self.textbox2 = customtkinter.CTkTextbox(self.frame3, wrap='none', width=872, height=375, corner_radius=0, font=customtkinter.CTkFont('Consolas', size=11), border_width=1, border_color='black')
                    self.textbox2.grid(row=0, column=0, padx=(5, 0), pady=(0, 0), sticky='nw')
                    self.textbox2.pack(expand=True, side='right')
                else:  # inserted
                    self.textbox2 = customtkinter.CTkTextbox(self.frame3, wrap='none', width=872, height=375, corner_radius=0, font=customtkinter.CTkFont('Consolas', size=11), border_width=1, border_color='#D9D9D9')
                    self.textbox2.grid(row=0, column=0, padx=(5, 0), pady=(0, 0), sticky='nw')
                    self.textbox2.pack(expand=True, side='right')
            original_text = '  ___     _____    _____\n / _ \\   / ____|  / ____|\n| | | | | (___   | (___  \n| | | |  \\___ \\   \\___ \\ \n| |_| |  ____) |  ____) |\n \\___/  |_____/  |_____/'
            ascii_art_variable = '                ______   ___   _   _     _____  _____  _____     _____  _   _   ___   _____   ______  _____  _____ \n               | ___ \\ / _ \\ | \\ | |   |  _  |/  ___|/  ___|   /  __ \\| | | | / _ \\ |_   _|  | ___ \\|  _  ||_   _|\n               | |_/ // /_\\ \\|  \\| |   | | | |\\ `--. \\ `--.    | /  \\/| |_| |/ /_\\ \\  | |    | |_/ /| | | |  | |  \n               |    / |  _  || . ` |   | | | | `--. \\ `--. \\   | |    |  _  ||  _  |  | |    | ___ \\| | | |  | |  \n               | |\\ \\ | | | || |\\  |   \\ \\_/ //\\__/ //\\__/ /   | \\__/\\| | | || | | |  | |    | |_/ /\\ \\_/ /  | |  \n               \\_| \\_|\\_| |_/\\_| \\_/    \\___/ \\____/ \\____/     \\____/\\_| |_/\\_| |_/  \\_/    \\____/  \\___/   \\_/  \n                                                                                                                  \n                                                                                                                  \n                      ______  _____  _____    _____ ______            _____  __    ____    ___  ______ ______\n                      | ___ \\|  _  ||_   _|  |_   _||  _  \\          / __  \\/  |  / ___|  /   ||___  /|___  /\n                      | |_/ /| | | |  | |      | |  | | | |  ______  `\' / /\'`| | / /___  / /| |   / /    / / \n                      | ___ \\| | | |  | |      | |  | | | | |______|   / /   | | | ___ \\/ /_| |  / /    / /  \n                      | |_/ /\\ \\_/ /  | |     _| |_ | |/ /           ./ /____| |_| \\_/ |\\___  |./ /   ./ /   \n                      \\____/  \\___/   \\_/     \\___/ |___/            \\_____/\\___/\\_____/    |_/\\_/    \\_/    \n                                                                                                             \n                                                                                                             \n'[1:(-1)]
            original_lines = original_text.split('\n')
            ascii_lines = ascii_art_variable.split('\n\n\n\n')
            max_length = max((len(line) for line in original_lines))
            ascii_max_length = max((len(line) for line in ascii_lines))
            spacing = ' ' * ((max_length - ascii_max_length) // 2)
            oss_chat_buddy_as = ''.join((spacing + line for line in ascii_lines))
            oss_chat_buddy_ascii = oss_chat_buddy_as
            var2 = 'RAN OSS application by integrating a built-in chatbot, empowered by an advanced NLP model, guaranteeing intelligent and \nnatural language interactions. Furthermore, leverage the Smart Cloud Authenticator to achieve secure and seamless \nauthentication. Users can effortlessly navigate the solution without the need for command knowledge, thanks to the intuitive \nguidance provided by the chatbot.'
            if scaleFactor == 1.5:
                final_var = '\n' + oss_chat_buddy_ascii + '\n' + var2 + '\n' + '                                                                                                                 Mohnish Saini'
                self.textbox2.insert('0.1', final_var)
            else:  # inserted
                final_var = '\n' + oss_chat_buddy_ascii + '\n\n' + var2 + '\n\n\n\n' + '                                                                                                                                      Mohnish Saini'
                self.textbox2.insert('0.1', final_var)
            if scaleFactor == 1.5:
                self.frame4 = customtkinter.CTkTabview(self, width=510, height=230, corner_radius=0)
                self.frame4.grid(row=2, column=1, sticky='nw', padx=(5, 0), pady=(0, 0))
                self.main_button_CLEAR = customtkinter.CTkButton(master=self, state='normal', fg_color=butoncolour, border_width=1, text_color='white', border_color=border_colorii, text='ABOUT', width=60, font=customtkinter.CTkFont(size=8, weight='bold'), command=self.Clear_textbox2)
                self.main_button_CLEAR.grid(row=0, column=3, padx=(0, 0), pady=(0, 0), sticky='nw')
                self.reset_button1 = customtkinter.CTkButton(master=self, text='Reset', fg_color='#CD5C5C', font=customtkinter.CTkFont(size=8, weight='bold'), width=60, command=self.button_function4_reset)
                self.reset_button1.grid(row=0, column=3, padx=(62, 0), pady=(0, 0), sticky='nw')
                self.frame4.add('Chat Buddy')
                if MODEEE == 'Dark':
                    self.CHATBUDDY_textbox3 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=765, height=157, corner_radius=0, font=customtkinter.CTkFont(family='Segoe UI', size=13), border_width=0, border_color='#808080')
                    self.CHATBUDDY_textbox3.grid(row=0, column=0, padx=(3, 0), pady=(0, 0), sticky='nw')
                else:  # inserted
                    self.CHATBUDDY_textbox3 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=765, height=157, corner_radius=0, font=customtkinter.CTkFont(family='Segoe UI', size=13), border_width=1, border_color='#D9D9D9')
                    self.CHATBUDDY_textbox3.grid(row=0, column=0, padx=(3, 0), pady=(0, 0), sticky='nw')
                if MODEEE == 'Dark':
                    self.CHATBUDDY_textbox4 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=725, height=7, corner_radius=5, text_color='#000000', fg_color='#FAF9F6', font=customtkinter.CTkFont(family='Arial', size=13), border_width=1, border_color='#808080')
                    self.CHATBUDDY_textbox4.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                else:  # inserted
                    self.CHATBUDDY_textbox4 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=725, height=7, corner_radius=5, text_color='#000000', font=customtkinter.CTkFont(family='Arial', size=13), border_width=1, border_color='#D9D9D9')
                    self.CHATBUDDY_textbox4.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                self.CHATBUDDY_textbox4.insert('0.10', 'Please wait loading data...')
                self.CHATBUDDY_textbox4.bind('<Button-1>', self.clear_text_buddy)
                self.CHATBUDDY_textbox4.bind('<Return>', self.CHAT_BUTTON_enter)
                self.CHATBUDDY_BUTTON = customtkinter.CTkButton(self.frame4.tab('Chat Buddy'), state='normal', fg_color='#61C671', border_width=0, text_color='white', border_color='black', text='  >  ', width=0, font=customtkinter.CTkFont(size=10, weight='bold'), command=self.CHAT_BUTTON_enter)
                self.CHATBUDDY_BUTTON.grid(row=2, column=0, padx=(733, 0), pady=(3, 0), sticky='nw')
            else:  # inserted
                self.frame4 = customtkinter.CTkTabview(self, width=400, height=260, corner_radius=0)
                self.frame4.grid(row=2, column=1, sticky='nw', padx=(5, 0), pady=(0, 0))
                self.main_button_CLEAR = customtkinter.CTkButton(master=self, state='normal', fg_color=butoncolour, border_width=1, text_color='white', border_color=border_colorii, text='ABOUT', width=60, font=customtkinter.CTkFont(size=8, weight='bold'), command=self.Clear_textbox2)
                self.main_button_CLEAR.grid(row=0, column=3, padx=(0, 0), pady=(0, 0), sticky='nw')
                self.reset_button1 = customtkinter.CTkButton(master=self, text='Reset', fg_color='#CD5C5C', font=customtkinter.CTkFont(size=8, weight='bold'), width=60, command=self.button_function4_reset)
                self.reset_button1.grid(row=0, column=3, padx=(62, 0), pady=(0, 0), sticky='nw')
                self.frame4.add('Chat Buddy')
                if MODEEE == 'Dark':
                    self.CHATBUDDY_textbox3 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=872, height=195, corner_radius=0, font=customtkinter.CTkFont(family='Segoe UI', size=13), border_width=1, border_color='#000000')
                    self.CHATBUDDY_textbox3.grid(row=0, column=0, padx=(3, 0), pady=(0, 0), sticky='nw')
                    self.CHATBUDDY_textbox4 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=835, fg_color='#FAF9F6', height=7, corner_radius=5, text_color='#000000', font=customtkinter.CTkFont(family='Arial', size=13), border_width=1)
                    self.CHATBUDDY_textbox4.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                else:  # inserted
                    self.CHATBUDDY_textbox3 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=872, height=195, corner_radius=0, font=customtkinter.CTkFont(family='Segoe UI', size=13), border_width=1, border_color='#D9D9D9')
                    self.CHATBUDDY_textbox3.grid(row=0, column=0, padx=(3, 0), pady=(0, 0), sticky='nw')
                    self.CHATBUDDY_textbox4 = customtkinter.CTkTextbox(self.frame4.tab('Chat Buddy'), width=835, height=7, corner_radius=5, font=customtkinter.CTkFont(family='Arial', size=13), border_width=1, border_color='#D9D9D9')
                    self.CHATBUDDY_textbox4.grid(row=2, column=0, padx=(3, 0), pady=(3, 0), sticky='nw')
                self.CHATBUDDY_textbox4.insert('0.10', 'Please wait loading data...')
                self.CHATBUDDY_textbox4.bind('<Button-1>', self.clear_text_buddy)
                self.CHATBUDDY_textbox4.bind('<Return>', self.CHAT_BUTTON_enter)
                self.CHATBUDDY_BUTTON = customtkinter.CTkButton(self.frame4.tab('Chat Buddy'), state='normal', fg_color='#61C671', border_width=0, text_color='white', border_color='black', text='  >  ', width=0, font=customtkinter.CTkFont(size=10, weight='bold'), command=self.CHAT_BUTTON_enter)
                self.CHATBUDDY_BUTTON.grid(row=2, column=0, padx=(841, 0), pady=(3, 0), sticky='nw')
            self.main_button_module = customtkinter.CTkButton(master=self, state='normal', fg_color=butoncolour, border_width=1, border_color=border_colorii, text='Modules', width=100, text_color='white', font=customtkinter.CTkFont(size=9, weight='bold'), command=self.module_button_event)
            self.main_button_module.grid(row=3, column=1, padx=(8, 0), pady=(0, 0), sticky='nw')
            if MODEEE == 'Dark':
                self.main_button_big = customtkinter.CTkButton(master=self, state='normal', fg_color=bigbutoncolour, border_width=1, text_color='#3b3b3b', border_color=bigbordercol, text='CMD Window', width=100, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_bigwindow)
                self.main_button_big.grid(row=3, column=1, padx=(219, 0), pady=(0, 0), sticky='nw')
                self.main_button_big2 = customtkinter.CTkButton(master=self, state='normal', fg_color=bigbutoncolour, border_width=1, text_color='#3b3b3b', border_color=bigbordercol, text='Main Window', width=100, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_bigwindow2)
                self.main_button_big2.grid(row=3, column=1, padx=(114, 0), pady=(0, 0), sticky='nw')
            else:  # inserted
                self.main_button_big = customtkinter.CTkButton(master=self, state='normal', fg_color=bigbutoncolour, border_width=1, text_color=bigtextcol, border_color=bigbordercol, text='CMD Window', width=100, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_bigwindow)
                self.main_button_big.grid(row=3, column=1, padx=(219, 0), pady=(0, 0), sticky='nw')
                self.main_button_big2 = customtkinter.CTkButton(master=self, state='normal', fg_color=bigbutoncolour, border_width=1, text_color=bigtextcol, border_color=bigbordercol, text='Main Window', width=100, font=customtkinter.CTkFont(size=9, weight='bold'), command=self.START_bigwindow2)
                self.main_button_big2.grid(row=3, column=1, padx=(114, 0), pady=(0, 0), sticky='nw')
                self.switch_Secondary_pass_rbs = customtkinter.CTkSwitch(master=self, state='normal', text='Secondary pass rbs', font=customtkinter.CTkFont(size=12), command=self.SSecondary_pass_rbs, variable=self.swich_Sec_rbs)
                if scaleFactor == 1.5:
                    self.switch_Secondary_pass_rbs.grid(row=3, column=1, padx=(625, 0), pady=(1, 0), sticky='nw')
                else:  # inserted
                    self.switch_Secondary_pass_rbs.grid(row=3, column=1, padx=(732, 0), pady=(3, 0), sticky='nw')
                self.switch_Secondary_pass_rbs.select()
                self.switch_Secondary_pass_rbs.configure(state='disabled')
            self.activity_BU.trace_add('write', self.update_radio_buttons_state)
            self.update_radio_buttons_state()
            self.radio_var.trace_add('write', self.update_radio_2)
            self.update_radio_2()
            threading.Thread(target=disable_button2, args=(self,)).start()

            def start_buddy():
                self.after(0, self.CHATBUDDY_textbox3.delete, '1.0', 'end')
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait loading data it may takes few seconds...\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            threading.Thread(target=start_buddy).start()
            self.protocol('WM_DELETE_WINDOW', self.on_closinggg)
            self.attributes('-topmost', True)
            self.attributes('-topmost', False)
            self.bind('<Any-KeyPress>', self.reset_timer)
            self.bind('<Any-ButtonPress>', self.reset_timer)
            self.textbox_LIVECMD.bind('<Key>', self.reset_timer)
            self.inactivity_timeout = AUTOCLOSE_TIME
            self.last_activity_time = time.time()
            self.check_inactivity()

        def button_function4_reset(self):
            try:
                eyye = RRSG_auto_login.remote_conn
                RRSG_enter_command(self, ['quit'], 'chatbuddy')
                clear_all_lists()
            except:
                pass  # postinserted
            try:
                ssh.close()
                RRSG_auto_login.remote_conn.close()
                del RRSG_auto_login.remote_conn
            except:
                pass
            try:
                combined_cmd = ''
            except:
                pass
            clear_all_lists()
            try:
                RRSG_auto_login.remote_conn
            except:
                pass
            self.radio_var.set((-1))
            self.activity_BU.set(0)
            threading.Thread(target=enabled_button, args=(self,)).start()
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Reset completed.\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            try:
                self.textbox_LIVECMD.delete(1.0, 'end')
            except:
                return None

        def on_closinggg22(self):
            def delayed_destroy3():
                if self.winfo_exists():
                    try:
                        loggs_up()
                    except:
                        pass
                    self.destroy()

            def delayed_destroy9():
                try:
                    tool_setting.app.destroy()
                except:
                    return None
            try:
                ssh.close()
                RRSG_auto_login.remote_conn.close()
                del RRSG_auto_login.remote_conn
            except:
                pass
            try:
                os.remove(REPORT_INPATH22 + 'User_access.db')
            except:
                pass
            try:
                delete_empty_folder(OUTPATH + 'OSS_logs_' + str(mfolder_create))
            except:
                pass
            try:
                ISF_STEP_CLOSE(ISF_API.API_Based, ISF_STEP_START_Task_Id_lst[0], ISF_API.woid, ISF_API.headers)
            except:
                pass
            try:
                ISF_WO_CLOSE(ISF_API.API_Based, ISF_API.headers, ISF_API.woid)
            except:
                pass
            try:
                self.after(100, delayed_destroy9)
            except:
                pass
            try:
                def is_folder_empty(folder_path):
                    return len(os.listdir(folder_path)) == 0
                if is_folder_empty('./res/CHATLOGS'):
                    break
                try:
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('AUTOWOClose WOID') + '__' + str('Close') + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                except:
                    pass
                self.after(100, delayed_destroy3)
            except:
                return None

        def on_closinggg(self):
            def delayed_destroy3():
                try:
                    clear_all_lists()
                except:
                    pass
                if self.winfo_exists():
                    try:
                        loggs_up()
                    except:
                        pass
                    try:
                        self.destroy()
                    except:
                        pass
                    try:
                        kill_process_by_name('OSS Chatbot 1.16.exe')
                    except:
                        pass
                    try:
                        sys.exit()
                    except:
                        return None

            def delayed_destroy9():
                try:
                    tool_setting.app.destroy()
                except:
                    return None
            response = askYesNoStylish('Are you sure you want to close the application?')
            if response:
                try:
                    try:
                        break
                    except:
                        pass
                    try:
                        ssh.close()
                    except:
                        pass
                    try:
                        RRSG_auto_login.remote_conn.close()
                    except:
                        pass
                    try:
                        del RRSG_auto_login.remote_conn
                    except:
                        pass
                except:
                    pass
                try:
                    try:
                        break
                    except:
                        pass
                    try:
                        break
                    except:
                        pass
                    try:
                        remote_ssh_client.close()
                    except:
                        pass
                    try:
                        my_transport.close()
                    except:
                        pass
                except:
                    pass
                try:
                    os.remove(REPORT_INPATH22 + 'User_access.db')
                except:
                    pass
                try:
                    delete_empty_folder(OUTPATH + 'OSS_logs_' + str(mfolder_create))
                except:
                    pass
                try:
                    ISF_STEP_CLOSE(ISF_API.API_Based, ISF_STEP_START_Task_Id_lst[0], ISF_API.woid, ISF_API.headers)
                except:
                    pass
                try:
                    ISF_WO_CLOSE(ISF_API.API_Based, ISF_API.headers, ISF_API.woid)
                except:
                    pass
                try:
                    self.after(100, delayed_destroy9)
                except:
                    pass
                try:
                    def is_folder_empty(folder_path):
                        return len(os.listdir(folder_path)) == 0
                    if is_folder_empty('./res/CHATLOGS'):
                        break
                    try:
                        with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                            f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str('Multi') + '__' + str('ChatbuddyWOClose WOID') + '__' + str('Close') + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
                    except:
                        pass
                except:
                    pass
                try:
                    self.after(100, delayed_destroy3)
                except:
                    return None

        def reset_timer(self, event):
            self.last_activity_time = time.time()

        def check_inactivity(self):
            current_time = time.time()
            if current_time - self.last_activity_time >= self.inactivity_timeout:
                Exit_msg = askYesNoStylishAutoClose_reset('No activity since long Are you sure you want to close the application otherwise it will auto close after 10 secound.?')
                if 'True' in chek_exit:
                    Exit_msg = True
                chek_exit.clear()
                if Exit_msg == True:
                    self.on_closinggg22()
            else:  # inserted
                self.after(1000, self.check_inactivity)

        def update_radio_2(self, *args):
            switch_stat = self.radio_var.get()
            state = 'normal' if switch_stat >= 0 else 'disabled'
            self.sidebar_node.configure(state=state)
            self.sidebar_command.configure(state=state)
            self.main_button_1.configure(state=state)
            self.tecno_mode_optionemenu.configure(state=state)
            self.tecno_mode_optionemenu.set('LTE')

        def update_radio_buttons_state(self, *args):
            switch_state = self.activity_BU.get()
            state = 'normal' if switch_state else 'disabled'
            state2 = 'disabled' if switch_state else 'normal'
            self.radio_button_1.configure(state=state)
            self.radio_button_2.configure(state=state)
            self.CHATBUDDY_textbox4.configure(state=state2)
            self.textbox8.configure(state=state2)
            self.main_button_shot.configure(state=state2)
            self.main_button_amos.configure(state=state2)
            self.main_button_quit.configure(state=state2)
            self.main_button_big.configure(state=state2)
            self.main_button_big2.configure(state=state2)
            self.main_button_module.configure(state=state2)
            if switch_state == 0:
                self.radio_var.set((-1))

        def clear_text_buddy(self, event):
            self.CHATBUDDY_textbox4.delete(1.0, 'end')

        def clear_text_crt(self, event):
            B_TEXT = self.textbox8.get('1.0', 'end').replace('\n', '').strip()
            self.textbox8.delete(1.0, 'end')

        def NODE_List(self):
            threading.Thread(target=self.execute_NODE_SELECT).start()

        def execute_NODE_SELECT(self):
            NODE_LSSTT = NODE_SELECT()
            self.NODE_Llist = NODE_LSSTT

        def Command_List(self):
            threading.Thread(target=self.execute_command_SELECT).start()

        def execute_command_SELECT(self):
            GSM_LST, WCDMA_LST, LTE_LST, NR_LST = COMMAND_SELECT()
            self.sidebar_command.configure(state='normal')

        def change_appearance_mode_event(self, new_appearance_mode: str):
            customtkinter.set_appearance_mode(new_appearance_mode)

        def change_scaling_event(self, new_scaling: str):
            new_scaling_float = int(new_scaling.replace('%', '')) / 100
            customtkinter.set_widget_scaling(new_scaling_float)

        def Save_logsss(self):
            Save_log_switch = self.swich_Sav_loss.get()
            return Save_log_switch

        def SSecondary_pass_rbs(self):
            RBS_RBS_switch = self.swich_Sec_rbs.get()
            return RBS_RBS_switch

        def RSG_Login_button_event(self):
            threading.Thread(target=disable_button, args=(self,)).start()
            tool_setting(self)
            thread = threading.Thread(target=enabled_button, args=(self,))
            thread.start()
            thread.join()

        def module_button_event(self):
            threading.Thread(target=disable_button, args=(self,)).start()
            module_setting(self)
            thread = threading.Thread(target=enabled_button, args=(self,))
            thread.start()
            thread.join()

        def Auto_Script(self):
            Auto_Script_switch = self.switch_aut_scri.get()
            return Auto_Script_switch

        def VPN_select_button(self, selected_value):
            VPN_select_METHOD_F.clear()
            VPN_select_METHOD_F.append(str(selected_value))
            DF = pd.DataFrame()
            DF['value'] = VPN_select_METHOD_F
            try:
                os.remove(os.path.join('./res/VPN_CRED_METHOD.db'))
            except:
                pass
            con = sqlite3.connect(os.path.join('./res/VPN_CRED_METHOD.db'))
            try:
                DF.to_sql('Credd', con, index=False, if_exists='replace')
            except:
                return
            finally:  # inserted
                con.close()

        def project_select_button(self, selected_value):
            SELECTED_PROJECT.clear()
            SELECTED_PROJECT.append(str(selected_value))
            try:
                os.remove(REPORT_INPATH22 + SELECTED_PROJECT[0] + '_PROJECT_MAPPING.db')
            except:
                pass
            PROJECT_WISE_MAPPING(SELECTED_PROJECT[0], REPORT_INPATH22)
            PROJECT_M = PROJECT_MAPPING_LST2(REPORT_INPATH22 + SELECTED_PROJECT[0] + '_PROJECT_MAPPING.db', SELECTED_PROJECT[0])
            try:
                PROJECT_LST.remove('0')
            except:
                pass
            USER_MAPPED_IN_PROJ.clear()
            USER_MAPPED_ROLE.clear()
            for uuser in PROJECT_USER_LST:
                if os.getlogin().lower() == uuser.lower():
                    USER_MAPPED_IN_PROJ.append('USER_MAPPED')
            if USER_MAPPED_IN_PROJ:
                try:
                    os.remove(PID_ISF_FILE + 'ISFdetails.db')
                except:
                    pass
                threading.Thread(target=PROJECT_ISF_DL, args=(PID_ISF_FILE,)).start()
                threading.Thread(target=greet_based_on_time, args=(self,)).start()
                threading.Thread(target=check_prroject, args=(self,)).start()
                chekkklsstt = []
                for uuser in PROJECT_SME_LST:
                    if SELECTED_PROJECT[0] == uuser.split('_')[0].lower():
                        if os.getlogin().lower() == uuser.split('_')[1].lower():
                            if chekkklsstt:
                                continue
                            USER_MAPPED_ROLE.append('SME')
                            chekkklsstt.append('done')
                if chekkklsstt:
                    break
                USER_MAPPED_ROLE.append('ENGG')
                chekkklsstt.append('done')
                if os.getlogin().lower() == 'emshsni':
                    USER_MAPPED_ROLE.clear()
                    chekkklsstt.clear()
                    USER_MAPPED_ROLE.append('SME')
                    chekkklsstt.append('done')
            if USER_MAPPED_IN_PROJ:
                return
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : You are not mapped to the selected Project ID. Please contact the administrator for access.\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            threading.Thread(target=decheck_prroject, args=(self,)).start()

        def Auto_login_button(self):
            try:
                ssh.close()
                RRSG_auto_login.remote_conn.close()
                del RRSG_auto_login.remote_conn
            except:
                pass
            try:
                con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                del Rdf['indexxx']
                ppPID_LST = Rdf['value'].tolist()
                ppPID_LST2 = []
                for ppp in ppPID_LST:
                    ppPID_LST2.append(decrypt(ppp, 3))
                del Rdf['value']
                Rdf['value'] = ppPID_LST2
                PID_LST = Rdf['value'].tolist()
                remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
            except:
                RSG_LST_N = []
            if not RSG_LST_N:
                showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 5000)
                return
            paaaawe = OTPPP_tets()
            try:
                if VPN_select_METHOD_F[0]!= 'VPN-CAS':
                    if is_non_numeric_string(paaaawe):
                        showMessage_qt('Wrong OTP password please try again.', 5000)
                        return
                    if paaaawe:
                        threading.Thread(target=self.Auto_login_sc, args=(paaaawe, RSG_LST_N)).start()
                else:  # inserted
                    if paaaawe:
                        threading.Thread(target=self.Auto_login_sc, args=(paaaawe, RSG_LST_N)).start()
            except:
                return None

        def Clear_textbox2(self):
            class MainWindow(QMainWindow):
                def __init__(self):
                    super().__init__()
                    self.setWindowTitle('RAN OSS CHAT BOT')
                    self.setGeometry(100, 100, 300, 300)
                    central_widget = QWidget(self)
                    self.setCentralWidget(central_widget)
                    layout = QVBoxLayout(central_widget)
                    self.image_paths = ['./res/Picture1.PNG', './res/Picture2.PNG', './res/Picture3.PNG', './res/Picture4.PNG', './res/Picture5.PNG']
                    self.current_image_index = 0
                    self.image_label = QLabel(self)
                    self.show_image()
                    layout.addWidget(self.image_label)
                    next_picture_button = QPushButton('Next', self)
                    next_picture_button.clicked.connect(self.next_picture)
                    next_picture_button.setStyleSheet('QPushButton {   background-color: #4CAF50;   border: none;   color: white;   padding: 10px 20px;   text-align: center;   text-decoration: none;   display: inline-block;   font-size: 16px;   margin: 4px 2px;   transition-duration: 0.4s;   cursor: pointer;   border-radius: 8px;}QPushButton:hover {   background-color: #45a049;}')
                    layout.addWidget(next_picture_button)
                    self.setWindowFlag(Qt.WindowStaysOnTopHint)
                    self.center_on_screen()

                def center_on_screen(self):
                    screen_geometry = QApplication.desktop().screenGeometry()
                    x = (screen_geometry.width() - self.width()) // 6
                    y = (screen_geometry.height() - self.height()) // 7
                    self.move(x, y)

                def show_image(self):
                    pixmap = QPixmap(self.image_paths[self.current_image_index])
                    self.image_label.setPixmap(pixmap)
                    self.image_label.setAlignment(Qt.AlignCenter)

                def next_picture(self):
                    self.current_image_index = (self.current_image_index + 1) % len(self.image_paths)
                    self.show_image()
            if __name__ == '__main__':
                app = QApplication([])
                rsg_app = MainWindow()
                rsg_app.show()
                app.exec_()

        def audit_response(self, response_1):
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str(response_1) + '\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')

        def CHAT_BUTTON_enter(self, event=None):
            BUDDY_TEXT = self.CHATBUDDY_textbox4.get('1.0', 'end').replace('\n', '').strip()
            BUDDY_TEXT = BUDDY_TEXT.replace('Type your message..', '').strip()
            BUDDY_TEXT = BUDDY_TEXT.replace('Please wait for response..', '').strip()
            if BUDDY_TEXT:
                threading.Thread(target=self.CHAT_BUDDY, args=(BUDDY_TEXT,)).start()
            else:  # inserted
                self.after(0, self.CHATBUDDY_textbox4.delete, '1.0', 'end')
                self.after(0, self.CHATBUDDY_textbox4.insert, 'end', 'Type your message..')
                self.after(0, self.CHATBUDDY_textbox4.see, 'end')

        def CHAT_BUDDY(self, BUDDY_TEXT):
            self.CHATBUDDY_textbox4.delete(1.0, 'end')
            responseee_lst = []

            def message_probability(user_message, recognised_words, single_response=False, required_words=[]):
                message_certainty = 0
                has_required_words = True
                for word in user_message:
                    if word in recognised_words:
                        message_certainty += 1
                percentage = float(message_certainty) / float(len(recognised_words))
                for word in required_words:
                    if word not in user_message:
                        has_required_words = False
                        break
                if has_required_words or single_response:
                    return int(percentage * 100)
                return 0

            def check_all_messages(message):
                highest_prob_list = {}

                def response(bot_response, list_of_words, single_response=False, required_words=[]):
                    highest_prob_list[bot_response] = message_probability(message, list_of_words, single_response, required_words)
                response('Hey! Please wait data fetching in progress !', ['node_entered'], single_response=True)
                response('satrt_fetch', ['waiting_for_responce'], single_response=True)
                response(long.R_ADVICE, ['give', 'advice'], required_words=['advice'])
                response(long.R_EATING, ['what', 'you', 'eat'], required_words=['you', 'eat'])
                best_match = max(highest_prob_list, key=highest_prob_list.get)
                if highest_prob_list[best_match] < 1:
                    return long.unknown()
                return best_match

            def get_response(user_input):
                split_message = [user_input.lower()]
                Buddy_request_lst.append(user_input.lower().strip())
                response = check_all_messages(split_message)
                responseee_lst.append(response)
                if 'Hey! Please wait data fetching in progress !' in response:
                    buddy_node.append(user_input)
                    buddy_node_curr.append(user_input)
                if response in ['satrt_fetch']:
                    def rum_command(BUDDY_TEXT):
                        ops_command = cmmmddd[0].split('_')[(-1)].strip()
                        nodiid = siteddd[0]
                        if nodiid not in checknxtsit:
                            checknxtsit.clear()
                            checknxtsit.append(nodiid)
                            LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
                            if LIVECMD_TEXT.strip().endswith('>') and LIVECMD_TEXT.strip()!= str(nodiid.upper()) + '>':
                                RRSG_enter_command(self, ['quit'], 'chatbuddy')
                                clear_all_lists()
                            RRSG_enter_command(self, ['amos ' + str(nodiid.upper())], 'chatbuddy')
                            if 'quit_amos' in checkamos:
                                cmmmddd.clear()
                                siteddd.clear()
                                checknxtsit.clear()
                                processdone.clear()
                                if 'java version' in checkamos_java:
                                    checkamos_java.clear()
                                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Sorry, couldn\'t login amos error message-- amos misconfigured - java version must be 1.4.1_09 or greater\n')
                                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                    start_index = '1.0'
                                    while True:
                                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                        if not start_index:
                                            break
                                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                        start_index = end_index
                                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                else:  # inserted
                                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please check NodeId and try again !\n')
                                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                    start_index = '1.0'
                                    while True:
                                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                        if not start_index:
                                            break
                                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                        start_index = end_index
                                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                            if 'quit_amos' not in checkamos:
                                try:
                                    checkamos.clear()
                                    if RRSG_auto_login.remote_conn:
                                        RRSG_enter_command(self, ['lt all'], 'chatbuddy')
                                        RBS_RBS_switch = self.swich_Sec_rbs.get()
                                        if RBS_RBS_switch == 1:
                                            RBS_RBS_switch_loe = 'YES'
                                        else:  # inserted
                                            RBS_RBS_switch_loe = 'NO'
                                        if RBS_RBS_switch_loe == 'YES':
                                            if RBSPASS_CRED_CRREDD[0] == 'rbs':
                                                user_rbs = 'rbs'
                                            else:  # inserted
                                                try:
                                                    user_rbs = RBSPASS_CRED_CRREDD[0]
                                                except:
                                                    user_rbs = 'rbs'
                                            if RBSPASS_CRED_CRREDD[1] == 'rbs':
                                                pass_rbs = 'rbs'
                                            else:  # inserted
                                                try:
                                                    pass_rbs = RBSPASS_CRED_CRREDD[1]
                                                except:
                                                    pass_rbs = 'rbs'
                                            RRSG_enter_command(self, [str(user_rbs)], 'nochatbuddy')
                                            RRSG_enter_command(self, [str(pass_rbs)], 'nochatbuddy')
                                        RRSG_enter_command(self, [str(ops_command.strip())], 'chatbuddy')
                                except:
                                    pass
                        else:  # inserted
                            try:
                                if RRSG_auto_login.remote_conn:
                                    RRSG_enter_command(self, [str(ops_command.strip())], 'chatbuddy')
                            except:
                                pass
                        if 'done' in processdone:
                            rum_command.res = 'Done! data has been successfully fetched. Feel free to ask for more assistance or information.'
                            processdone.clear()
                            thread = threading.Thread(target=enabled_button, args=(self,))
                            thread.start()
                            thread.join()
                        else:  # inserted
                            rum_command.res = 'Feel free to ask for more assistance or information.'
                            processdone.clear()
                            thread = threading.Thread(target=enabled_button, args=(self,))
                            thread.start()
                            thread.join()
                    rum_command(BUDDY_TEXT)
                    response = rum_command.res
                return response
            quit_lst = ['next site', 'quit', 'q', 'refresh', 'exit', 'stop']
            BUDDY_TEXT = BUDDY_TEXT.replace('Type your message..', '').strip()
            BUDDY_TEXT = BUDDY_TEXT.replace('Please wait for response it may take few seconds.', '').strip()

            def audit_response(self, response_1):
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str(response_1) + '\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            if BUDDY_TEXT:
                if BUDDY_TEXT.lower() == 'stop':
                    check_yesno.clear()
                if 'updating' in check_textbox:
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Hey! Please wait data fetching in progress !\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        else:  # inserted
                            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                            start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                    return
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'You : ' + BUDDY_TEXT + '\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('You :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 3 chars')
                    self.CHATBUDDY_textbox3.tag_add('You', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('You', foreground='red')
                time.sleep(0.5)
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait for response it may take few seconds.\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                threading.Thread(target=disable_button, args=(self,)).start()
                if SQL_METHOD == 'LOCAL':
                    user_signum = os.getlogin().lower()
                    command_to_barred_required = 'YES'
                    command_to_barred = BUDDY_TEXT.lower()
                    ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check(user_signum, command_to_barred_required, command_to_barred)
                if SQL_METHOD == 'AZURE_CLOUDE':
                    if change_User:
                        user_signum = os.getlogin().lower()
                        command_to_barred_required = 'YES'
                        command_to_barred = BUDDY_TEXT.lower()
                        ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check_AZURE(change_User, command_barred, user_signum, command_to_barred_required, command_to_barred)
                    else:  # inserted
                        if 'amos' in BUDDY_TEXT.lower():
                            ucercmd_AUTHENTICATION = 'User_Authenticated'
                        else:  # inserted
                            if 'rbs' in BUDDY_TEXT.lower():
                                ucercmd_AUTHENTICATION = 'User_Authenticated'
                            else:  # inserted
                                if 'lt all' in BUDDY_TEXT.lower():
                                    ucercmd_AUTHENTICATION = 'User_Authenticated'
                                else:  # inserted
                                    ucercmd_AUTHENTICATION = 'NO'
                if ucercmd_AUTHENTICATION == 'User_Authenticated':
                    checkamos.clear()
                    threading.Thread(target=check_empty_textbox_thread, args=(self,)).start()
                    if BUDDY_TEXT == 'no':
                        BUDDY_TEXT = 'stop'
                    if BUDDY_TEXT == 'stop':
                        response = 'close_chat'
                    else:  # inserted
                        if BUDDY_TEXT.lower().strip() in quit_lst:
                            response = None
                        else:  # inserted
                            if BUDDY_TEXT == 'yes':
                                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy :Thanks! how Chat Buddy will assist you.\n')
                                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                start_index = '1.0'
                                while True:
                                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                    if not start_index:
                                        break
                                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                    start_index = end_index
                                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                response = 'close_chat'
                            else:  # inserted
                                if chk_commad_add:
                                    response = None
                                else:  # inserted
                                    response = find_keyword(self, BUDDY_TEXT, DEFAULT_RESPONSES_lst)
                                    if response == 'quit':
                                        response = 'none'
                    if response:
                        response = str(response)
                    else:  # inserted
                        response = 'none'
                    try:
                        PROCEED_TO_AUDIT = response.lower()
                    except:
                        PROCEED_TO_AUDIT = 'none'
                    if PROCEED_TO_AUDIT == 'audit_audit':
                        DATABASE = 'AUDDIT_START'
                        INSERT_COM = ['Audit_Audit']
                        DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)
                        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Welcome! Ready for the health check and audit.') + '\n')
                        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                        start_index = '1.0'
                        while True:
                            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                            if not start_index:
                                break
                            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                            start_index = end_index
                        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                        response342 = ask_yes_or_no_qt('Welcome! Ready for the health check and audit?')
                        if response342 == True:
                            response_1 = 'yes'
                        else:  # inserted
                            response_1 = 'no'
                        if response_1 == 'yes':
                            try:
                                if amos_node:
                                    try:
                                        eyye = RRSG_auto_login.remote_conn
                                        threading.Thread(target=RRSG_enter_command, args=(self, ['quit'], 'chatbuddy')).start()
                                    except:
                                        pass
                            except:
                                pass
                            COMMAND_DEFALT_AUTO = []
                            threading.Thread(target=disable_button, args=(self,)).start()
                            selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST = AUDIT_main_SELECT(COMMAND_DEFALT_AUTO)
                            try:
                                GSM_NODE_LST = [item.upper() for item in GSM_NODE_LST]
                            except:
                                pass
                            try:
                                WCDMA_NODE_LST = [item.upper() for item in WCDMA_NODE_LST]
                            except:
                                pass
                            try:
                                LTE_NODE_LST = [item.upper() for item in LTE_NODE_LST]
                            except:
                                pass
                            try:
                                NR_NODE_LST = [item.upper() for item in NR_NODE_LST]
                            except:
                                pass
                            DATABASE = 'CAUDIT_CLOSE_WINDOW'
                            DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                            if 'close' not in DATBASE_VALUE_READ:
                                try:
                                    DATBASE_VALUE_READ.clear()
                                    del DATBASE_VALUE_READ
                                except:
                                    pass
                                DATABASE = 'ACTIVITYY'
                                DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                                RRSG_Audit_Module(self, selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST, DATBASE_VALUE_READ[0], 'CHEATBOOT')
                                try:
                                    DATBASE_VALUE_READ.claer()
                                    del DATBASE_VALUE_READ
                                except:
                                    return None
                            else:  # inserted
                                thread = threading.Thread(target=enabled_button, args=(self,))
                                thread.start()
                                thread.join()
                                response_1 = site_audit('quit')
                                audit_response(self, response_1)
                                DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                SQLITE_DB_CLEAR(DATABASE)
                                DATABASE = 'ACTIVITYY'
                                SQLITE_DB_CLEAR(DATABASE)
                                clear_all_lists()
                        else:  # inserted
                            thread = threading.Thread(target=enabled_button, args=(self,))
                            thread.start()
                            thread.join()
                            response_1 = site_audit('quit')
                            audit_response(self, response_1)
                            DATABASE = 'CAUDIT_CLOSE_WINDOW'
                            SQLITE_DB_CLEAR(DATABASE)
                            DATABASE = 'ACTIVITYY'
                            SQLITE_DB_CLEAR(DATABASE)
                            cmmmddd.clear()
                            siteddd.clear()
                            checknxtsit.clear()
                            processdone.clear()
                    else:  # inserted
                        if PROCEED_TO_AUDIT == 'parameter audit_parameter audit':
                            DATABASE = 'AUDDIT_START'
                            INSERT_COM = ['Audit_Audit']
                            DATBASE_VALUE = SQLITE_DB(DATABASE, INSERT_COM)
                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Welcome! Ready for the Parameter Audit?') + '\n')
                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                            start_index = '1.0'
                            while True:
                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                if not start_index:
                                    break
                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                start_index = end_index
                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                            response342 = ask_yes_or_no_qt('Welcome! Ready for the Parameter audit?')
                            if response342 == True:
                                response_1 = 'yes'
                            else:  # inserted
                                response_1 = 'no'
                            if response_1 == 'yes':
                                try:
                                    if amos_node:
                                        try:
                                            eyye = RRSG_auto_login.remote_conn
                                            threading.Thread(target=RRSG_enter_command, args=(self, ['quit'], 'chatbuddy')).start()
                                        except:
                                            pass
                                except:
                                    pass
                                COMMAND_DEFALT_AUTO = []
                                selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST = AUDIT_main_SELECT(COMMAND_DEFALT_AUTO)
                                try:
                                    GSM_NODE_LST = [item.upper() for item in GSM_NODE_LST]
                                except:
                                    pass
                                try:
                                    WCDMA_NODE_LST = [item.upper() for item in WCDMA_NODE_LST]
                                except:
                                    pass
                                try:
                                    LTE_NODE_LST = [item.upper() for item in LTE_NODE_LST]
                                except:
                                    pass
                                try:
                                    NR_NODE_LST = [item.upper() for item in NR_NODE_LST]
                                except:
                                    pass
                                DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                                if 'close' not in DATBASE_VALUE_READ:
                                    try:
                                        DATBASE_VALUE_READ.clear()
                                        del DATBASE_VALUE_READ
                                    except:
                                        pass
                                    DATABASE = 'ACTIVITYY'
                                    DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                                    RRSG_Audit_Module(self, selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST, DATBASE_VALUE_READ[0], 'CHEATBOOT')
                                    try:
                                        DATBASE_VALUE_READ.claer()
                                        del DATBASE_VALUE_READ
                                    except:
                                        return None
                                else:  # inserted
                                    thread = threading.Thread(target=enabled_button, args=(self,))
                                    thread.start()
                                    thread.join()
                                    response_1 = site_audit('quit')
                                    audit_response(self, response_1)
                                    DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                    SQLITE_DB_CLEAR(DATABASE)
                                    DATABASE = 'ACTIVITYY'
                                    SQLITE_DB_CLEAR(DATABASE)
                                    clear_all_lists()
                            else:  # inserted
                                thread = threading.Thread(target=enabled_button, args=(self,))
                                thread.start()
                                thread.join()
                                response_1 = site_audit('quit')
                                audit_response(self, response_1)
                                DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                SQLITE_DB_CLEAR(DATABASE)
                                DATABASE = 'ACTIVITYY'
                                SQLITE_DB_CLEAR(DATABASE)
                                cmmmddd.clear()
                                siteddd.clear()
                                checknxtsit.clear()
                                processdone.clear()
                        else:  # inserted
                            if PROCEED_TO_AUDIT == 'kpi monitoring_kpi monitoring':
                                response342 = ask_yes_or_no_qt('Welcome! Ready for the KPI Monitoring?')
                                if response342 == True:
                                    response_1 = 'yes'
                                else:  # inserted
                                    response_1 = 'no'
                                if response_1 == 'yes':
                                    execution_count = [0]
                                    try:
                                        if amos_node:
                                            try:
                                                eyye = RRSG_auto_login.remote_conn
                                                threading.Thread(target=RRSG_enter_command, args=(self, ['quit'], 'chatbuddy')).start()
                                            except:
                                                pass
                                    except:
                                        pass
                                    DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                    SQLITE_DB_CLEAR(DATABASE)
                                    DATABASE = 'AUDDIT_START'
                                    SQLITE_DB_CLEAR(DATABASE)
                                    DATABASE = 'ACTIVITYY'
                                    SQLITE_DB_CLEAR(DATABASE)

                                    def START_FUNNNN():
                                        COMMAND_DEFALT_AUTO = []
                                        FREQUENCY_count_LST.clear()
                                        FREQUENCY_LST.clear()
                                        EMAIL_METHOOD_LST.clear()
                                        selected_checkboxes, GSM_LST, WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST = Monitoring_main_SELECT(COMMAND_DEFALT_AUTO)
                                        try:
                                            GSM_NODE_LST = [item.upper() for item in GSM_NODE_LST]
                                        except:
                                            pass
                                        try:
                                            WCDMA_NODE_LST = [item.upper() for item in WCDMA_NODE_LST]
                                        except:
                                            pass
                                        try:
                                            LTE_NODE_LST = [item.upper() for item in LTE_NODE_LST]
                                        except:
                                            pass
                                        try:
                                            NR_NODE_LST = [item.upper() for item in NR_NODE_LST]
                                        except:
                                            pass
                                        DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                        DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                                        if 'close' not in DATBASE_VALUE_READ:
                                            try:
                                                DATBASE_VALUE_READ.clear()
                                                del DATBASE_VALUE_READ
                                            except:
                                                pass
                                            execution_count[0] = 0
                                            NEW_WCDMA_LST = []
                                            if WCDMA_LST:
                                                for nod in WCDMA_NODE_LST:
                                                    if nod not in NEW_WCDMA_LST:
                                                        NEW_WCDMA_LST.append(WCDMA_LST[0] + nod.split('@')[(-1)])
                                            NEW_WCDMA_NODE_LST = []
                                            if WCDMA_NODE_LST:
                                                for nod in WCDMA_NODE_LST:
                                                    if nod not in NEW_WCDMA_NODE_LST:
                                                        NEW_WCDMA_NODE_LST.append(nod.split('@')[0])

                                            def START_FUNNNN_22():
                                                RRSG_KPI_MONITORING_Module(self, selected_checkboxes, GSM_LST, NEW_WCDMA_LST, LTE_LST, NR_LST, GSM_NODE_LST, NEW_WCDMA_NODE_LST, LTE_NODE_LST, NR_NODE_LST, 'KPI MONITORING', 'UIBOOT')
                                                if len(FREQUENCY_LST) > 1 and FREQUENCY_LST[0]!= 'Not required':
                                                    execution_count[0] += 1
                                                    if execution_count[0] < int(FREQUENCY_count_LST[0]):
                                                        current_time = datetime.now()
                                                        new_time = current_time + timedelta(minutes=int(FREQUENCY_LST[0]))
                                                        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Please wait next execution scheduled on ' + new_time.strftime('%H:%M:%S') + ' Hour') + '\n')
                                                        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                                        start_index = '1.0'
                                                        while True:
                                                            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                            if not start_index:
                                                                break
                                                            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                            start_index = end_index
                                                        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                            if len(FREQUENCY_LST) > 1 and FREQUENCY_LST[0] == 'Not required':
                                                START_FUNNNN_22()
                                                FREQUENCY_count_LST.clear()
                                                FREQUENCY_LST.clear()
                                                EMAIL_METHOOD_LST.clear()
                                                DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                                SQLITE_DB_CLEAR(DATABASE)
                                                DATABASE = 'AUDDIT_START'
                                                SQLITE_DB_CLEAR(DATABASE)
                                                DATABASE = 'ACTIVITYY'
                                                SQLITE_DB_CLEAR(DATABASE)
                                                GSM_LST.clear()
                                                LTE_LST.clear()
                                                WCDMA_LST.clear()
                                                NR_LST.clear()
                                                selected_checkboxes.clear()
                                                GSM_NODE_LST.clear()
                                                LTE_NODE_LST.clear()
                                                WCDMA_NODE_LST.clear()
                                                NR_NODE_LST.clear()
                                                COMMAND_DEFALT_AUTO.clear()
                                                return
                                            START_FUNNNN_22()
                                            if len(FREQUENCY_LST) > 1:
                                                schedule.every(int(FREQUENCY_LST[0])).minutes.do(START_FUNNNN_22)
                                                while execution_count[0] < int(FREQUENCY_count_LST[0]):
                                                    schedule.run_pending()
                                                    time.sleep(1)
                                            FREQUENCY_count_LST.clear()
                                            FREQUENCY_LST.clear()
                                            EMAIL_METHOOD_LST.clear()
                                            DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            DATABASE = 'AUDDIT_START'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            DATABASE = 'ACTIVITYY'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            GSM_LST.clear()
                                            LTE_LST.clear()
                                            WCDMA_LST.clear()
                                            NR_LST.clear()
                                            selected_checkboxes.clear()
                                            GSM_NODE_LST.clear()
                                            LTE_NODE_LST.clear()
                                            WCDMA_NODE_LST.clear()
                                            NR_NODE_LST.clear()
                                            COMMAND_DEFALT_AUTO.clear()
                                            self.radio_var.set((-1))
                                            self.activity_BU.set(0)
                                        else:  # inserted
                                            thread = threading.Thread(target=enabled_button, args=(self,))
                                            thread.start()
                                            thread.join()
                                            DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            DATABASE = 'AUDDIT_START'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            DATABASE = 'ACTIVITYY'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            GSM_LST.clear()
                                            LTE_LST.clear()
                                            WCDMA_LST.clear()
                                            NR_LST.clear()
                                            selected_checkboxes.clear()
                                            GSM_NODE_LST.clear()
                                            LTE_NODE_LST.clear()
                                            WCDMA_NODE_LST.clear()
                                            NR_NODE_LST.clear()
                                            COMMAND_DEFALT_AUTO.clear()
                                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('Thanks! how Chat Buddy will assist you !') + '\n')
                                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                            start_index = '1.0'
                                            while True:
                                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                if not start_index:
                                                    break
                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                start_index = end_index
                                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                    threading.Thread(target=START_FUNNNN).start()
                                self.radio_var.set((-1))
                                self.activity_BU.set(0)
                            else:  # inserted
                                if len(PROCEED_TO_AUDIT.split('_')) > 1 and PROCEED_TO_AUDIT.split('_')[1].lower().strip() == 'pmx erca':
                                    response342 = ask_yes_or_no_qt('Welcome! Ready for the ERCA PM Report?')
                                    if response342 == True:
                                        response_1 = 'yes'
                                    else:  # inserted
                                        response_1 = 'no'
                                    if response_1 == 'yes':
                                        execution_count = [0]
                                        try:
                                            if amos_node:
                                                try:
                                                    eyye = RRSG_auto_login.remote_conn
                                                    threading.Thread(target=RRSG_enter_command, args=(self, ['quit'], 'chatbuddy')).start()
                                                except:
                                                    pass
                                        except:
                                            pass
                                        COMMAND_DEFALT_AUTO = []
                                        DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                        DATBASE_VALUE_READ = SQLITE_DB_READ(DATABASE)
                                        if 'close' not in DATBASE_VALUE_READ:
                                            try:
                                                DATBASE_VALUE_READ.clear()
                                                del DATBASE_VALUE_READ
                                            except:
                                                pass
                                            chekkkkkkk = ['dddl']

                                            def ERCA_PM_Tool_new():
                                                try:
                                                    lte_pmx_value, node_list_value, hourly_combo_value, start_hour_value, end_hour_value, start_date_value, end_date_value, selected_radio_button = ERCA_PM_Tool()
                                                    chekkkkkkk.clear()
                                                    RRSG_PMX_Module(self, lte_pmx_value, node_list_value, hourly_combo_value, start_hour_value, end_hour_value, start_date_value, end_date_value, selected_radio_button, mfolder_create, 'UIBOOT')
                                                except:
                                                    pass
                                                thread = threading.Thread(target=enabled_button, args=(self,))
                                                thread.start()
                                                thread.join()
                                            thread_combined_xl = threading.Thread(target=ERCA_PM_Tool_new)
                                            thread_combined_xl.start()
                                            thread_combined_xl.join()
                                            if chekkkkkkk:
                                                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Thanks! how Chat Buddy will assist you !\n')
                                                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                                start_index = '1.0'
                                                while True:
                                                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                    if not start_index:
                                                        break
                                                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                    start_index = end_index
                                                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                        else:  # inserted
                                            thread = threading.Thread(target=enabled_button, args=(self,))
                                            thread.start()
                                            thread.join()
                                            response_1 = site_audit('quit')
                                            audit_response(self, response_1)
                                            DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            DATABASE = 'ACTIVITYY'
                                            SQLITE_DB_CLEAR(DATABASE)
                                            COMMAND_DEFALT_AUTO.clear()
                                            cmmmddd.clear()
                                            siteddd.clear()
                                            checknxtsit.clear()
                                            processdone.clear()
                                            return
                                    else:  # inserted
                                        thread = threading.Thread(target=enabled_button, args=(self,))
                                        thread.start()
                                        thread.join()
                                        response_1 = site_audit('quit')
                                        audit_response(self, response_1)
                                        DATABASE = 'CAUDIT_CLOSE_WINDOW'
                                        SQLITE_DB_CLEAR(DATABASE)
                                        DATABASE = 'ACTIVITYY'
                                        SQLITE_DB_CLEAR(DATABASE)
                                        cmmmddd.clear()
                                        siteddd.clear()
                                        checknxtsit.clear()
                                        processdone.clear()
                                        return
                                else:  # inserted
                                    if str(response) in DEFAULT_RESPONSES_lst.values():
                                        response_pri = str(response)
                                        response = 'DEFAULT_RESPONSES'
                                    exclude_llsstt = ['None_None', 'DEFAULT_RESPONSES']
                                    if response not in exclude_llsstt:
                                        if BUDDY_TEXT.lower() in quit_lst:
                                            response = 'next site'
                                        if str(response) == 'DEFAULT_RESPONSES':
                                            break
                                        if str(response).lower() in quit_lst:
                                            LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
                                            if LIVECMD_TEXT.strip().endswith('>'):
                                                RRSG_enter_command(self, ['quit'], 'chatbuddy')
                                                clear_all_lists()
                                            cmmmddd.clear()
                                            siteddd.clear()
                                            checknxtsit.clear()
                                            processdone.clear()
                                            response = 'STOP_LOOP'
                                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please feel free to ask for more assistance or information !\n')
                                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                            start_index = '1.0'
                                            while True:
                                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                if not start_index:
                                                    break
                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                start_index = end_index
                                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                        if response!= 'STOP_LOOP' and response!= None and (response!= 'close_chat'):
                                            if len(cmmmddd) >= 1 and str(response) not in cmmmddd:
                                                cmmmddd.clear()
                                            if len(cmmmddd) < 1:
                                                cmmmddd.append(str(response))
                                                BUDDY_TEXT = str(response).lower()
                                                if len(siteddd) < 1:
                                                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please enter Node id !\n')
                                                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                                    start_index = '1.0'
                                                    while True:
                                                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                        if not start_index:
                                                            break
                                                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                        start_index = end_index
                                                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                                    NODEID = get_nodeid()
                                                    if str(NODEID)!= 'None':
                                                        siteddd.append(str(NODEID).upper())
                                                        BUDDY_TEXT = str(NODEID).upper()
                                                        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'You : ' + str(NODEID) + '\n')
                                                        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                                        start_index = '1.0'
                                                        while True:
                                                            start_index = self.CHATBUDDY_textbox3.search('You :', start_index, 'end')
                                                            if not start_index:
                                                                break
                                                            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 3 chars')
                                                            self.CHATBUDDY_textbox3.tag_add('You', start_index, end_index)
                                                            start_index = end_index
                                                        self.CHATBUDDY_textbox3.tag_config('You', foreground='red')
                                                    else:  # inserted
                                                        response = 'STOP_LOOP'
                                                        cmmmddd.clear()
                                                        siteddd.clear()
                                                        checknxtsit.clear()
                                                        processdone.clear()
                                                        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Hello! How can I assist you?\n')
                                                        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                                        start_index = '1.0'
                                                        while True:
                                                            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                            if not start_index:
                                                                break
                                                            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                            start_index = end_index
                                                        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                                        cmmmddd.clear()
                                                        siteddd.clear()
                                        if response == None:
                                            thread = threading.Thread(target=enabled_button, args=(self,))
                                            thread.start()
                                            thread.join()
                                            chk_commad_add.append('chk_commad_add')
                                            if 'startloop' not in check_yesno:
                                                check_yesno.append('startloop')
                                            start_index = '1.0'
                                            while True:
                                                start_index = self.CHATBUDDY_textbox3.search('You :', start_index, 'end')
                                                if not start_index:
                                                    break
                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 3 chars')
                                                self.CHATBUDDY_textbox3.tag_add('You', start_index, end_index)
                                                start_index = end_index
                                            self.CHATBUDDY_textbox3.tag_config('You', foreground='red')
                                            start_index = '1.0'
                                            while True:
                                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                if not start_index:
                                                    break
                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                start_index = end_index
                                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                        else:  # inserted
                                            chk_commad_add.clear()
                                            if response!= 'STOP_LOOP':
                                                if response!= None and response!= 'close_chat':
                                                    try:
                                                        self.textbox8.delete(1.0, 'end')
                                                    except:
                                                        pass
                                                    bot_response = 'Buddy : ' + get_response('node_entered')
                                                    if bot_response!= 'Buddy : clear_llloop':
                                                        self.after(0, self.CHATBUDDY_textbox3.insert, 'end', bot_response + '\n')
                                                        self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                                        start_index = '1.0'
                                                        while True:
                                                            start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                            if not start_index:
                                                                break
                                                            end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                            self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                            start_index = end_index
                                                        self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                                    if 'wait' in bot_response:
                                                        bot_response = 'Buddy : ' + get_response('waiting_for_responce')
                                                        if bot_response!= 'Buddy : clear_llloop':
                                                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', bot_response + '\n')
                                                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                                            start_index = '1.0'
                                                            while True:
                                                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                                if not start_index:
                                                                    break
                                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                                start_index = end_index
                                                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                            else:  # inserted
                                                thread = threading.Thread(target=enabled_button, args=(self,))
                                                thread.start()
                                                thread.join()
                                    else:  # inserted
                                        if str(response) == 'DEFAULT_RESPONSES':
                                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'You : ' + str(BUDDY_TEXT) + '\n')
                                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                            start_index = '1.0'
                                            while True:
                                                start_index = self.CHATBUDDY_textbox3.search('You :', start_index, 'end')
                                                if not start_index:
                                                    break
                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 3 chars')
                                                self.CHATBUDDY_textbox3.tag_add('You', start_index, end_index)
                                                start_index = end_index
                                            self.CHATBUDDY_textbox3.tag_config('You', foreground='red')
                                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str(response_pri) + '\n')
                                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                            start_index = '1.0'
                                            while True:
                                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                if not start_index:
                                                    break
                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                start_index = end_index
                                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                            thread = threading.Thread(target=enabled_button, args=(self,))
                                            thread.start()
                                            thread.join()
                                        if response!= 'DEFAULT_RESPONSES':
                                            check_yesno.clear()
                                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Hello! How can I assist you?\n\n')
                                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                                            start_index = '1.0'
                                            while True:
                                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                                if not start_index:
                                                    break
                                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                                start_index = end_index
                                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                                            thread = threading.Thread(target=enabled_button, args=(self,))
                                            thread.start()
                                            thread.join()
                else:  # inserted
                    thread = threading.Thread(target=enabled_button, args=(self,))
                    thread.start()
                    thread.join()
                    self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : ' + str('You are not authorized to run this command.') + '\n')
                    self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                    start_index = '1.0'
                    while True:
                        start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                        if not start_index:
                            break
                        end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                        self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                        start_index = end_index
                    self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')

        def Auto_login_sc(self, paaaawe, RSG_LST_N):
            try:
                db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                con.close()
                Rdf = Rdf.drop_duplicates(keep='first')
                PID_LST = Rdf['value'].tolist()
                HOSTIP = PID_LST[0]
            except:
                HOSTIP = '148.135.15.71'
            login_command_lst = RSG_LST_N
            router = HOSTIP
            username = os.getlogin().lower()
            password = paaaawe
            RRSG_auto_login(self, login_command_lst, router, username, password)
            if 'FAIL' not in establish_ssh_conn_lst:
                chek_first_time_rsg.append('ok')

        def START_PROGRAMM(self):
            threading.Thread(target=disable_button, args=(self,)).start()
            Radio_option = ''
            Radio_opt = self.radio_var.get()
            tecno_mode = self.tecno_mode_optionemenu.get()
            if Radio_opt == 0:
                Radio_option = 'MO_DATCH_CLICK'
            if Radio_opt == 1:
                Radio_option = 'KGET_CLICK'
            if Radio_option == 'MO_DATCH_CLICK':
                try:
                    con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                    Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                    con.close()
                    del Rdf['indexxx']
                    ppPID_LST = Rdf['value'].tolist()
                    ppPID_LST2 = []
                    for ppp in ppPID_LST:
                        ppPID_LST2.append(decrypt(ppp, 3))
                    del Rdf['value']
                    Rdf['value'] = ppPID_LST2
                    PID_LST = Rdf['value'].tolist()
                    remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                    RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                except:
                    RSG_LST_N = []
                if not RSG_LST_N:
                    showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 10000)
                else:  # inserted
                    retunmsg = 'noooo'
                    try:
                        con = sqlite3.connect('./res/Node_list.db')
                        Rdf1 = pd.read_sql_query(f"SELECT * FROM {'command_list'}", con)
                        con.close()
                        del Rdf1['id']
                        Rdf1 = Rdf1.drop_duplicates(subset=['command'], keep='first')
                        PID_LSTq = Rdf1['command'].tolist()
                        del Rdf1
                        final_node_lst = [value for value in PID_LSTq if value.strip()!= '']
                        retunmsg_cmd = 'doneee'
                        if not final_node_lst:
                            retunmsg_cmd = 'noooo'
                            showMessage_qt('Please add Node list before continue by clicking Node List button.', 10000)
                    except:
                        showMessage_qt('Please add Node list before continue by clicking Node List button.', 10000)
                        retunmsg = 'clsefunction'
                        retunmsg_cmd = 'noooo'
                        final_node_lst = []
                    SKPI_COMMAND = 'NO'
                    if retunmsg_cmd!= 'noooo':
                        try:
                            con = sqlite3.connect('./res/command_list.db')
                            Rdf1 = pd.read_sql_query(f'SELECT * FROM {self.tecno_mode_optionemenu.get()}', con)
                            con.close()
                            del Rdf1['id']
                            Rdf1 = Rdf1.drop_duplicates(subset=['command'], keep='first')
                            PID_LSTq = Rdf1['command'].tolist()
                            del Rdf1
                            final_cmd_lst2 = [value for value in PID_LSTq if value.strip()!= '']
                            final_cmd_lst = []
                            for ccnm in final_cmd_lst2:
                                SPLIT_txt = list(filter(None, ccnm.split(':')))
                                SPLIT_txt = SPLIT_txt[0].upper().strip()
                                if SPLIT_txt[:2] not in GSMCMD_lst:
                                    final_cmd_lst.append(ccnm)
                            if final_cmd_lst:
                                SKPI_COMMAND = 'NO'
                            else:  # inserted
                                SKPI_COMMAND = 'YES'
                            if not final_cmd_lst:
                                retunmsg_cmd = 'noooo'
                                showMessage('1-Please add command list before continue by clicking command List button..\n\n2-GSM Commands not executable right now will be available soon..')
                        except:
                            showMessage('1-Please add command list before continue by clicking command List button..\n\n2-GSM Commands not executable right now will be available soon..')
                            retunmsg_cmd = 'cmdfunction'
                            final_cmd_lst = []
                    if retunmsg!= 'clsefunction' and retunmsg_cmd!= 'cmdfunction':
                        Save_log_switch = self.swich_Sav_loss.get()
                        if Save_log_switch == 1:
                            Save_logfile = 'YES'
                        else:  # inserted
                            Save_logfile = 'NO'
                        paaaawe = 'OK'
                        try:
                            if SKPI_COMMAND == 'NO' and paaaawe:
                                threading.Thread(target=self.perform_long_running_task, args=(paaaawe, Save_logfile, final_node_lst, final_cmd_lst, RSG_LST_N, tecno_mode)).start()
                        except:
                            pass
            if Radio_option == 'KGET_CLICK':
                try:
                    con = sqlite3.connect('./res/' + str(SELECTED_PROJECT[0]) + '_RSG_CRED.db')
                    Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                    con.close()
                    del Rdf['indexxx']
                    ppPID_LST = Rdf['value'].tolist()
                    ppPID_LST2 = []
                    for ppp in ppPID_LST:
                        ppPID_LST2.append(decrypt(ppp, 3))
                    del Rdf['value']
                    Rdf['value'] = ppPID_LST2
                    PID_LST = Rdf['value'].tolist()
                    remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8', 'LINE9', 'LINE10']
                    RSG_LST_N = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                except:
                    RSG_LST_N = []
                if not RSG_LST_N:
                    showMessage_qt('Please update RSG details before continue by clicking RSG Login button.', 10000)
                    return
                retunmsg = 'noooo'
                try:
                    con = sqlite3.connect('./res/Node_list.db')
                    Rdf1 = pd.read_sql_query(f"SELECT * FROM {'command_list'}", con)
                    con.close()
                    del Rdf1['id']
                    Rdf1 = Rdf1.drop_duplicates(subset=['command'], keep='first')
                    PID_LSTq = Rdf1['command'].tolist()
                    del Rdf1
                    final_node_lst = [value for value in PID_LSTq if value.strip()!= '']
                    retunmsg_cmd = 'doneee'
                    if not final_node_lst:
                        retunmsg_cmd = 'noooo'
                        showMessage_qt('Please add Node list before continue by clicking Node List button.', 10000)
                except:
                    showMessage_qt('Please add Node list before continue by clicking Node List button.', 10000)
                    retunmsg = 'clsefunction'
                    retunmsg_cmd = 'noooo'
                    final_node_lst = []
                if retunmsg_cmd!= 'noooo':
                    try:
                        con = sqlite3.connect('./res/command_list.db')
                        Rdf1 = pd.read_sql_query(f'SELECT * FROM {self.tecno_mode_optionemenu.get()}', con)
                        con.close()
                        del Rdf1['id']
                        Rdf1 = Rdf1.drop_duplicates(subset=['command'], keep='first')
                        PID_LSTq = Rdf1['command'].tolist()
                        del Rdf1
                        final_cmd_lst = [value for value in PID_LSTq if value.strip()!= '']
                        if not final_cmd_lst:
                            retunmsg_cmd = 'noooo'
                            showMessage_qt('Please add command list before continue by clicking command List button.', 10000)
                    except:
                        showMessage_qt('Please add command list before continue by clicking command List button.', 10000)
                        retunmsg_cmd = 'cmdfunction'
                        final_cmd_lst = []
                if retunmsg!= 'clsefunction' and retunmsg_cmd!= 'cmdfunction':
                    Save_log_switch = self.swich_Sav_loss.get()
                    if Save_log_switch == 1:
                        Save_logfile = 'YES'
                    else:  # inserted
                        Save_logfile = 'NO'
                    paaaawe = 'ok'
                    try:
                        if paaaawe:
                            threading.Thread(target=self.perform_long_running_task_kget, args=(paaaawe, Save_logfile, final_node_lst, ['kget all'], RSG_LST_N, tecno_mode)).start()
                    except:
                        return

        def START_QUIT(self):
            LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
            if LIVECMD_TEXT.strip().endswith('>'):
                cmmmddd.clear()
                siteddd.clear()
                checknxtsit.clear()
                processdone.clear()
                threading.Thread(target=RRSG_enter_command, args=(self, [str('quit')], 'nochatbuddy')).start()
                self.after(0, self.textbox8.delete, '1.0', 'end')
                self.after(0, self.textbox8.insert, 'end', 'Press enter to excute command...')
                self.after(0, self.textbox8.see, 'end')

        def START_amos(self):
            threading.Thread(target=amoooss, args=(self,)).start()

        def START_bigwindow(self):
            self.withdraw()
            text_updater = TextUpdater(self)

        def START_bigwindow2(self):
            self.withdraw()
            text_updater = TextUpdater2(self)

        def START_command(self, event=None):
            def handle_selected_option(selected_option):
                con = sqlite3.connect(aql_path23)
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                Rdf['New_lst'] = Rdf['indexxx'].astype(str) + '_' + Rdf['value'].astype(str)
                COMMAND_LST = Rdf['New_lst'].unique()
                COMMAND_LST = list(COMMAND_LST)
                try:
                    del Rdf
                except:
                    pass
                for cmm in COMMAND_LST:
                    if str(selected_option) == cmm.split('_')[0]:
                        messaget_cmd = cmm.split('_')[(-1)]
                        break
                    messaget_cmd = '-WINDOW CLOSE ATTEMPTED-'

                def del_textbox(self):
                    self.after(0, self.textbox8.delete, '1.0', 'end')
                if str(messaget_cmd)!= '-WINDOW CLOSE ATTEMPTED-':
                    self.after(0, self.textbox8.delete, '1.0', 'end')
                    self.after(0, self.textbox8.insert, 'end', str(messaget_cmd))
                    self.after(0, self.textbox8.see, 'end')
                    threading.Thread(target=RRSG_enter_command, args=(self, [str(messaget_cmd)], 'nochatbuddy')).start()
                    threading.Thread(target=del_textbox, args=(self,)).start()
            selected_option = ui_commands()
            threading.Thread(target=disable_button, args=(self,)).start()
            self.after(0, handle_selected_option, selected_option)

        def START_enter(self, event=None):
            SKPI_COMMAND = 'NO'
            text_content = self.textbox8.get('1.0', 'end').replace('\n', '').strip()
            text_content = text_content.replace('Press enter to excute command...', '').strip()
            SPLIT_txt = list(filter(None, text_content.split(':')))
            SPLIT_txt = SPLIT_txt[0].upper().strip()
            if SPLIT_txt[:2] in GSMCMD_lst:
                SKPI_COMMAND = 'YES'
            if SKPI_COMMAND == 'NO':
                if text_content:
                    if 'updating' in check_textbox:
                        def print_update(self):
                            self.after(0, self.textbox8.delete, '1.0', 'end')
                            self.after(0, self.textbox8.insert, 'end', 'Please wait data fetching in progress...')
                            self.after(0, self.textbox8.see, 'end')
                            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Hey! Please wait data fetching in progress !\n')
                            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                            start_index = '1.0'
                            while True:
                                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                                if not start_index:
                                    break
                                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                                start_index = end_index
                            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
                        threading.Thread(target=print_update, args=(self,)).start()
                        return
                    if False:
                        pass  # postinserted
                    if SQL_METHOD == 'LOCAL':
                        user_signum = os.getlogin().lower()
                        command_to_barred_required = 'YES'
                        command_to_barred = text_content.lower().strip()
                        ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check(user_signum, command_to_barred_required, command_to_barred)
                    if SQL_METHOD == 'AZURE_CLOUDE':
                        if change_User:
                            user_signum = os.getlogin().lower()
                            command_to_barred_required = 'YES'
                            command_to_barred = text_content.lower().strip()
                            ucercmd_AUTHENTICATION = USER_AUTHENTICATION_check_AZURE(change_User, command_barred, user_signum, command_to_barred_required, command_to_barred)
                        else:  # inserted
                            if 'amos' in text_content.lower().strip():
                                ucercmd_AUTHENTICATION = 'User_Authenticated'
                            else:  # inserted
                                if 'rbs' in text_content.lower().strip():
                                    ucercmd_AUTHENTICATION = 'User_Authenticated'
                                else:  # inserted
                                    if 'lt all' in text_content.lower().strip():
                                        ucercmd_AUTHENTICATION = 'User_Authenticated'
                                    else:  # inserted
                                        ucercmd_AUTHENTICATION = 'NO'
                    if ucercmd_AUTHENTICATION == 'User_Authenticated':
                        threading.Thread(target=self.perform_enter_task, args=(text_content,)).start()
                    else:  # inserted
                        threading.Thread(target=unauthorized_cmd, args=(self,)).start()
                else:  # inserted
                    self.after(0, self.textbox8.delete, '1.0', 'end')
                    self.after(0, self.textbox8.insert, 'end', 'Press enter to excute command...')
                    self.after(0, self.textbox8.see, 'end')
            else:  # inserted
                self.after(0, self.textbox8.delete, '1.0', 'end')
                self.after(0, self.textbox8.insert, 'end', 'Press enter to excute command...')
                self.after(0, self.textbox8.see, 'end')
                threading.Thread(target=check_GSM_CMD, args=(self,)).start()

        def perform_enter_task(self, text_content_v1):
            threading.Thread(target=disable_button, args=(self,)).start()
            try:
                if 'amos' in str(text_content_v1).lower():
                    text_content = str(text_content_v1).lower().replace('amos', '').strip()
                    text_content = 'amos ' + text_content.upper()
                else:  # inserted
                    text_content = str(text_content_v1)
            except:
                text_content = str(text_content_v1)
            try:
                self.textbox8.delete(1.0, 'end')
            except:
                pass
            if 'amos' in text_content.lower():
                LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
                if LIVECMD_TEXT.strip().endswith('>'):
                    text_node = str(text_content_v1).lower().replace('amos', '').strip()
                    text_node = str(text_node) + '>'
                    if LIVECMD_TEXT.strip()!= str(text_node.upper()) + '>':
                        RRSG_enter_command(self, ['quit'], 'chatbuddy')
                        clear_all_lists()
                RRSG_enter_command(self, [text_content], 'nochatbuddy')
                RRSG_enter_command(self, ['lt all'], 'nochatbuddy')
                RBS_RBS_switch = self.swich_Sec_rbs.get()
                if RBS_RBS_switch == 1:
                    RBS_RBS_switch_loe = 'YES'
                else:  # inserted
                    RBS_RBS_switch_loe = 'NO'
                if RBS_RBS_switch_loe == 'YES':
                    if RBSPASS_CRED_CRREDD[0] == 'rbs':
                        user_rbs = 'rbs'
                    else:  # inserted
                        try:
                            user_rbs = RBSPASS_CRED_CRREDD[0]
                        except:
                            user_rbs = 'rbs'
                    if RBSPASS_CRED_CRREDD[1] == 'rbs':
                        pass_rbs = 'rbs'
                    else:  # inserted
                        try:
                            pass_rbs = RBSPASS_CRED_CRREDD[1]
                        except:
                            pass_rbs = 'rbs'
                    RRSG_enter_command(self, [str(user_rbs)], 'nochatbuddy')
                    RRSG_enter_command(self, [str(pass_rbs)], 'nochatbuddy')
            else:  # inserted
                RRSG_enter_command(self, [text_content], 'nochatbuddy')
            thread = threading.Thread(target=enabled_button, args=(self,))
            thread.start()
            thread.join()

        def perform_long_running_task(self, paaaawe, Save_logfile, node_lst, fnl_cmd_lst, RSG_LST_N, tecno_mode):
            def delayed_destroy():
                if self.winfo_exists():
                    try:
                        os.remove(REPORT_INPATH22 + 'User_access.db')
                    except:
                        pass
                    try:
                        loggs_up()
                    except:
                        pass
                    try:
                        delete_empty_folder(OUTPATH + 'OSS_logs_' + str(mfolder_create))
                    except:
                        pass
                    self.destroy()
            try:
                db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                con.close()
                Rdf = Rdf.drop_duplicates(keep='first')
                PID_LST = Rdf['value'].tolist()
                HOSTIP = PID_LST[0]
            except:
                HOSTIP = '148.135.15.71'
            login_command_lst = RSG_LST_N
            Node_cmd_lst = node_lst
            oss_cmd_lst = fnl_cmd_lst
            router = HOSTIP
            username = os.getlogin().lower()
            password = paaaawe
            MOTASK = 'MOBATCH'
            LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
            if amos_node and LIVECMD_TEXT.strip()!= '>':
                RRSG_enter_command(self, ['quit'], 'chatbuddy')
                clear_all_lists()
            PATHHHH_TIME_mokgt = time.strftime('%d%m%Y%H%M%S')
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Mobatch processing ongoing !\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            self.activity_BU.set(0)
            WCDMA_LST_RBS = []
            WCDMA_LST_RNC = []
            for item in oss_cmd_lst:
                try:
                    prefix, suffix = item.split('@')
                    if prefix == 'RBS':
                        WCDMA_LST_RBS.append(suffix)
                    else:  # inserted
                        if prefix == 'RNC':
                            WCDMA_LST_RNC.append(suffix)
                except:
                    WCDMA_LST_RBS.append(item)
            Node_cmd_RBS = []
            Node_cmd_RNC = []
            for item in Node_cmd_lst:
                if '@' not in item:
                    if item not in Node_cmd_RBS:
                        Node_cmd_RBS.append(item)
                else:  # inserted
                    if item not in Node_cmd_RNC:
                        Node_cmd_RNC.append(item)
            FINAL_RNC_COMMAND_LST = []
            FINAL_RNC_NODE_LST = []
            for item in Node_cmd_RNC:
                for item2 in WCDMA_LST_RNC:
                    item2 = item2.replace('CELL_NAME', str(item.split('@')[(-1)]).strip()).strip()
                    if str(item2) not in FINAL_RNC_COMMAND_LST:
                        FINAL_RNC_COMMAND_LST.append(str(item2))
                    if str(item.split('@')[0]).strip() not in FINAL_RNC_NODE_LST:
                        FINAL_RNC_NODE_LST.append(str(item.split('@')[0]).strip())
            if Node_cmd_RBS:
                RRSG(self, login_command_lst, Node_cmd_RBS, WCDMA_LST_RBS, router, username, password, Save_logfile, MOTASK, PATHHHH_TIME_mokgt, tecno_mode, 'RBS')
            if FINAL_RNC_NODE_LST:
                RRSG(self, login_command_lst, FINAL_RNC_NODE_LST, FINAL_RNC_COMMAND_LST, router, username, password, Save_logfile, MOTASK, PATHHHH_TIME_mokgt, tecno_mode, 'RNC')
            showMessage_qt('Mobatch completed successfully.', 5000)
            thread = threading.Thread(target=enabled_button, args=(self,))
            thread.start()
            thread.join()
            try:
                eyye = RRSG_auto_login.remote_conn
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Mobatch processing completed !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            except:
                pass
            if 'cancle' not in processdone:
                try:
                    ISF_STEP_CLOSE(ISF_API.API_Based, ISF_STEP_START_Task_Id_lst[0], ISF_API.woid, ISF_API.headers)
                except:
                    pass
                try:
                    ISF_WO_CLOSE(ISF_API.API_Based, ISF_API.headers, ISF_API.woid)
                except:
                    pass
                if MOTASK == 'MOBATCH':
                    try:
                        os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('MOBATCH_') + str(PATHHHH_TIME_mokgt))
                    except:
                        pass
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(tecno_mode) + '__' + str('MOBATCHWOClose WOID') + '__' + str(ISF_API.woid) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            loggs_up_audit()
            clear_all_lists()

        def perform_long_running_task_kget(self, paaaawe, Save_logfile, node_lst, fnl_cmd_lst, RSG_LST_N, tecno_mode):
            def delayed_destroy2():
                if self.winfo_exists():
                    try:
                        os.remove(REPORT_INPATH22 + 'User_access.db')
                    except:
                        pass
                    try:
                        loggs_up()
                    except:
                        pass
                    try:
                        delete_empty_folder(OUTPATH + 'OSS_logs_' + str(mfolder_create))
                    except:
                        pass
                    self.destroy()
            try:
                db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_Hostname.db')
                con = sqlite3.connect(db_path)
                cursor = con.cursor()
                cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
                con.commit()
                Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
                con.close()
                Rdf = Rdf.drop_duplicates(keep='first')
                PID_LST = Rdf['value'].tolist()
                HOSTIP = PID_LST[0]
            except:
                HOSTIP = '148.135.15.71'
            login_command_lst = RSG_LST_N
            Node_cmd_lst = node_lst
            oss_cmd_lst = fnl_cmd_lst
            router = HOSTIP
            username = os.getlogin().lower()
            password = paaaawe
            MOTASK = 'KGET'
            LIVECMD_TEXT = self.textbox_LIVECMD.get('1.0', 'end').replace('\n', '').strip()
            if amos_node and LIVECMD_TEXT.strip()!= '>':
                RRSG_enter_command(self, ['quit'], 'chatbuddy')
                clear_all_lists()
            PATHHHH_TIME_mokgt = time.strftime('%d%m%Y%H%M%S')
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Please wait Kget processing ongoing !\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            self.activity_BU.set(0)
            RRSG(self, login_command_lst, Node_cmd_lst, oss_cmd_lst, router, username, password, Save_logfile, MOTASK, PATHHHH_TIME_mokgt, tecno_mode, 'KGET')
            showMessage_qt('KGET completed successfully.', 5000)
            thread = threading.Thread(target=enabled_button, args=(self,))
            thread.start()
            thread.join()
            try:
                eyye = RRSG_auto_login.remote_conn
                self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Kget processing completed !\n')
                self.after(0, self.CHATBUDDY_textbox3.see, 'end')
                start_index = '1.0'
                while True:
                    start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                    if not start_index:
                        break
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
                self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            except:
                pass
            if 'cancle' not in processdone:
                try:
                    ISF_STEP_CLOSE(ISF_API.API_Based, ISF_STEP_START_Task_Id_lst[0], ISF_API.woid, ISF_API.headers)
                except:
                    pass
                try:
                    ISF_WO_CLOSE(ISF_API.API_Based, ISF_API.headers, ISF_API.woid)
                except:
                    pass
                if MOTASK == 'KGET':
                    try:
                        os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create) + '\\' + str('KGET_') + str(PATHHHH_TIME_mokgt))
                    except:
                        pass
                    with open('./res/CHATLOGS/' + str(os.getlogin().lower()) + '__Stats' + '__' + str(PATHHHH_TIME) + '.Log', 'a', encoding='utf-8') as f:
                        f.writelines(f"{str(SELECTED_PROJECT[0]) + '__' + str(os.getlogin().lower()) + '__' + str(tecno_mode) + '__' + str('KGETWOClose WOID') + '__' + str(ISF_API.woid) + '__' + str(time.strftime('%d%m%Y %H:%M:%S'))}\n")
            loggs_up_audit()
            clear_all_lists()

        def update_textbox(self, message):
            self.after(0, self.textbox_LIVECMD.insert, 'end', message)
            self.after(0, self.textbox_LIVECMD.see, 'end')
            lines = int(self.textbox_LIVECMD.index('end-1c').split('.')[0])
            if lines > 200:
                self.textbox_LIVECMD.delete('1.0', str(lines - 200) + '.0')

        def update_textbox_big_window_enter_cmd(self, message, iii):
            terminal_TEXT = self.textbox2.get('1.0', 'end').replace('\n', '').strip()
            self.after(0, self.textbox2.insert, 'end', message + '\n')
            self.after(0, self.textbox2.see, 'end')
            try:
                lines = int(self.textbox2.index('end-1c').split('.')[0])
                if lines > 400:
                    self.textbox2.delete('1.0', str(lines - 400) + '.0')
            except:
                return None

        def update_textbox_big_window_enter_cmd_summ(self, message, iii):
            self.after(0, self.textbox2.insert, 'end', message + '\n')
            self.after(0, self.textbox2.see, 'end')
            try:
                lines = int(self.textbox2.index('end-1c').split('.')[0])
                if lines > 400:
                    self.textbox2.delete('1.0', str(lines - 400) + '.0')
            except:
                return None

        def update_textbox_big_window(self, message, iii):
            self.after(0, self.textbox2.insert, 'end', message + '\n')
            self.after(0, self.textbox2.see, 'end')
            try:
                lines = int(self.textbox2.index('end-1c').split('.')[0])
                if lines > 400:
                    self.textbox2.delete('1.0', str(lines - 400) + '.0')
            except:
                return None
    app = App()
    app.mainloop()
INPATH = 'C:\\00_OSS_CHAT_BOT\\'
OUTPATH = INPATH + 'Output\\'
AUTOCLOSE_TIME = 1200
MODEEE = 'light'
REPORT_INPATH22 = './_internal/'
PID_ISF_FILE = './res/'
aql_path = REPORT_INPATH22 + 'process.db'
aql_path23 = REPORT_INPATH22 + 'CMD_SHORTCUT.db'
pathpp = REPORT_INPATH22 + 'authnnnti.db'
try:
    os.remove(REPORT_INPATH22 + 'Audit_command_list_default.db')
except:
    pass
try:
    os.remove(REPORT_INPATH22 + 'authnnnti.db')
except:
    pass
try:
    os.remove(REPORT_INPATH22 + 'authnnnti_2.db')
except:
    pass
try:
    os.remove(REPORT_INPATH22 + 'process.db')
except:
    pass
try:
    os.remove(REPORT_INPATH22 + 'conv.db')
except:
    pass
try:
    os.remove(REPORT_INPATH22 + 'User_access.db')
except:
    pass
try:
    os.remove(REPORT_INPATH22 + 'CMD_SHORTCUT.db')
except:
    pass
try:
    os.remove('./res/my_database.db')
except:
    pass
try:
    os.remove('C:\\Users\\' + os.getlogin().lower() + '\\AppData\\Local\\OSSchatguide\\' + 'ossfil.zip')
except:
    pass

def remove_files_in_folder(folder_path):
    try:
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
            except:
                continue
    except:
        return None
folder_path = './res/CHATLOGS/'
remove_files_in_folder(folder_path)
ISF_WORKORDER = 'Required'
LOG_UPLODER = 'Required'
ssh = None
if ISF_WORKORDER == 'NOTRequired':
    ISF_API.timestr = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
    ISF_API.urlcreateWO = 'Dummy work order created for testing purpose'
    ISF_API.API_Based = 'Test'
    ISF_STEP_START_Task_Id_lst = ['Test']
    ISF_API.woid = 'Test'
    ISF_API.headers = 'Test'

def check_internet_connection():
    try:
        socket.create_connection(('8.8.8.8', 53), timeout=5)
    except OSError:
        return 'No internet connection'

def nltkstopwords():
    nltk.download('stopwords')
    nltk.download('punkt')
PROJECT_MAPPING(REPORT_INPATH22)
PROJECT_M = ALL_PROJECT_MAPPING_LST2(REPORT_INPATH22 + 'PROJECT_MAPPING_SME.db')
try:
    PROJECT_LST.remove('0')
except:
    pass
try:
    VPN_select_METHOD_F = []
    db_path = os.path.join('./res/VPN_CRED_METHOD.db')
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS Credd (value TEXT)')
    con.commit()
    Rdf = pd.read_sql_query('SELECT * FROM Credd', con)
    ppPID_LST = Rdf['value'].tolist()
    VPN_select_METHOD_F.append(ppPID_LST[0])
except:
    pass
try:
    db_path = os.path.join('./res/' + str(SELECTED_PROJECT[0]) + '_RBSPASS_CRED.db')
    con = sqlite3.connect(db_path)
    cursor = con.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS Credd (indexxx TEXT, value TEXT)')
    con.commit()
    Rdf = pd.read_sql_query('SELECT * FROM Credd', con, index_col='indexxx')
    ppPID_LST = Rdf['value'].tolist()
    for ppp in ppPID_LST:
        RBSPASS_CRED_CRREDD.append(decrypt(ppp, 3))
    con.close()
except:
    pass

def newstartfun():
    GSMGMRRTool(INPATH, OUTPATH, MODEEE)

def check_prroject(self):
    pass
    thread = threading.Thread(target=enabled_button, args=(self,))
    thread.start()
    thread.join()

def decheck_prroject(self):
    pass
    threading.Thread(target=disable_button, args=(self,)).start()
CHKLST = []
CHKLST22 = []
USERCHKLST = []
anothercheck = []
CHKLST_again = []

def check_download_completion(self):
    CHK_AZURE = []
    while True:
        if not CHK_AZURE:
            try:
                con = sqlite3.connect(aql_path)
                Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                con.close()
                Rdf = Rdf.loc[Rdf['indexxx']!= 'ASK_ISF_CLOSE']
                SME_LST = Rdf['value'].tolist()
                if os.getlogin().lower() in SME_LST:
                    SQL_METHOD = 'AZURE_CLOUDE'
                else:  # inserted
                    SQL_METHOD = 'LOCAL'
                try:
                    del Rdf
                except:
                    pass
                if SQL_METHOD == 'AZURE_CLOUDE':
                    threading.Thread(target=AZURE_CLOUDE_login, args=(self,)).start()
                else:  # inserted
                    threading.Thread(target=conv_Sh2, args=(REPORT_INPATH22,)).start()
                CHK_AZURE.append('done')
            except:
                pass
        if CHK_TO_CONTinue and CHK_TO_CONTinue2:
            if CHKLST22:
                return
            self.after(0, self.CHATBUDDY_textbox3.insert, 'end', 'Buddy : Data loading completed. Please select a project before continuing...\n')
            self.after(0, self.CHATBUDDY_textbox3.see, 'end')
            start_index = '1.0'
            while True:
                start_index = self.CHATBUDDY_textbox3.search('Buddy :', start_index, 'end')
                if not start_index:
                    break
                else:  # inserted
                    end_index = self.CHATBUDDY_textbox3.index(f'{start_index} + 5 chars')
                    self.CHATBUDDY_textbox3.tag_add('Buddy', start_index, end_index)
                    start_index = end_index
            self.CHATBUDDY_textbox3.tag_config('Buddy', foreground='#43884e')
            self.sidebar_Project.configure(state='normal')
            CHKLST22.append('done')
            return
        time.sleep(0.1)

def check_expiration():
    current_date = datetime.now()
    if current_date > expiration_date:
        return False
    return True
expiration_date = datetime(2024, 12, 30)
if check_expiration():
    chk_int = check_internet_connection()
    if chk_int == 'No internet connection':
        showMessage_qt('No internet connection please try again.', 10000)
    Autenticate_Sh3(REPORT_INPATH22)
    check_user = checkuser_expired(REPORT_INPATH22 + 'authnnnti_2.db')
    if check_user == 'continue':
        threading.Thread(target=newstartfun).start()
        Autenticate_Sh2(REPORT_INPATH22)
        check_expiry = program_expired(pathpp)
        if check_expiry == 'continue':
            if check_user == 'continue':
                try:
                    con = sqlite3.connect(aql_path)
                    Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                    con.close()
                    ASK_ISF_CLOSE = Rdf.iloc[1, 0]
                    Rdf = Rdf.loc[Rdf['indexxx']!= 'ASK_ISF_CLOSE']
                    SME_LST = Rdf['value'].tolist()
                    if os.getlogin().lower() in SME_LST:
                        SQL_METHOD = 'AZURE_CLOUDE'
                    else:  # inserted
                        SQL_METHOD = 'LOCAL'
                    try:
                        del Rdf
                    except:
                        pass
                except:
                    ASK_ISF_CLOSE = 'YES'
                    SQL_METHOD = 'NO'
                if SQL_METHOD == 'AZURE_CLOUDE':
                    try:
                        os.remove(REPORT_INPATH22 + 'conv.db')
                    except:
                        pass
                if SQL_METHOD!= 'NO':
                    try:
                        con = sqlite3.connect('./res/ISFdetails.db')
                        Rdf = pd.read_sql_query(f"SELECT * FROM {'Credd'}", con)
                        con.close()
                        del Rdf['indexxx']
                        ppPID_LST = Rdf['value'].tolist()
                        ppPID_LST2 = []
                        for ppp in ppPID_LST:
                            ppPID_LST2.append(decrypt(ppp, 3))
                        del Rdf['value']
                        Rdf['value'] = ppPID_LST2
                        PID_LST = Rdf['value'].tolist()
                        remove_lst = ['LINE1', 'LINE2', 'LINE3', 'LINE4', 'LINE5', 'LINE6', 'LINE7', 'LINE8']
                        ISF_LST = [value.strip() for value in PID_LST if value.strip()!= '' and value not in remove_lst]
                        try:
                            del Rdf
                        except:
                            pass
                    except:
                        pass
                    try:
                        projectID_global = SELECTED_PROJECT[0]
                    except:
                        projectID_global = 'NA'
                    mfolder_create = time.strftime('%d%m%Y%H%M%S')
                    mfolder_create_lst.append(mfolder_create)
                    PATHHHH_TIME = time.strftime('%d%m%Y%H%M%S')
                    try:
                        os.mkdir(OUTPATH + 'OSS_logs_' + str(mfolder_create))
                    except:
                        pass
                    if SQL_METHOD == 'LOCAL':
                        Autenticate_Sh()
                    try:
                        baselllin_DF = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='Baseline', usecols='A,B,C,D,E,F', engine='openpyxl')
                    except:
                        baselllin_DF = pd.DataFrame()
                    try:
                        shmap_DF = pd.read_excel(INPATH + 'Baseline.xlsx', sheet_name='SH_Mapping', usecols='A,B', engine='openpyxl')
                        shmap_DF['sh_map'] = shmap_DF['Actual SH'].astype(str) + '@' + shmap_DF['Req SH'].astype(str)
                        sh_map_lst = shmap_DF['sh_map'].unique()
                        sh_map_lst = list(sh_map_lst)
                        del shmap_DF
                    except:
                        pass
                    GSMCMD_lst = ['RL', 'RX', 'SA', 'DB']
                    try:
                        os.mkdir('C:\\00_OSS_CHAT_BOT\\')
                    except:
                        pass
                    try:
                        os.mkdir('C:\\00_OSS_CHAT_BOT\\Output\\')
                    except:
                        pass
                    nltkstopwords()
                    CHK_TO_CONTinue.append('continuee')
                else:  # inserted
                    showMessage_qt('Required files \'not found please try again.', 30000)
else:  # inserted
    showMessage_qt('This executable has expired.', 30000)